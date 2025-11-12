from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File, Form, Header
from fastapi.responses import FileResponse, Response, StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.staticfiles import StaticFiles
from dotenv import load_dotenv
from pathlib import Path
import os

# CRITICAL: Load .env FIRST before importing any modules that use environment variables
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import logging
import traceback
from pydantic import BaseModel, Field
import httpx
import asyncio
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
from zoneinfo import ZoneInfo
from decimal import Decimal, ROUND_HALF_UP
import jwt
import hashlib
from io import BytesIO
import io
import base64
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.interval import IntervalTrigger
from reportlab.lib.pagesizes import A4, letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.colors import Color, black, blue, grey
from reportlab.lib.units import inch, cm
from reportlab.lib import colors
from gcs_storage import get_gcs_storage
from backup_service import (
    create_backup,
    get_backup_history,
    restore_backup,
    create_incremental_backup,
    cleanup_old_backups
)
from email_service import (
    send_welcome_client_email, 
    send_welcome_admin_email, 
    send_notification_email,
    send_client_wallet_transfer_approved_email,
    send_client_wallet_transfer_rejected_email,
    send_admin_wallet_transfer_request_email,
    send_client_account_request_approved_email,
    send_client_account_request_rejected_email,
    send_client_account_request_completed_email,
    send_admin_new_share_request_email,
    send_client_share_request_approved_email,
    send_client_share_request_rejected_email,
    # Phase 1 Critical Emails
    send_client_wallet_topup_approved_email,
    send_client_wallet_topup_rejected_email,
    send_admin_wallet_topup_proof_uploaded_email,
    send_client_topup_auto_cancelled_email,
    send_client_wallet_topup_auto_cancelled_email,
    send_client_super_admin_completion_email,
    # Phase 2 Important Emails
    send_client_transfer_request_success_email,
    send_client_transfer_request_failed_email,
    send_client_transfer_request_rejected_email,
    send_client_account_deleted_email,
    send_admin_transfer_request_created_email
)

# GCS Helper Function
async def upload_to_gcs(file: UploadFile, folder: str = "uploads") -> dict:
    """
    Upload file to Google Cloud Storage
    Returns dict with gcs_path, gcs_bucket, file_size, mime_type
    """
    try:
        from google.cloud import storage
        import os
        
        # Read file content
        content = await file.read()
        
        # Initialize GCS client
        os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = "/app/backend/gcs-service-account.json"
        storage_client = storage.Client()
        bucket_name = os.getenv("GCS_BUCKET_NAME", "rimuru-file-uploads")
        bucket = storage_client.bucket(bucket_name)
        
        # Generate unique filename for GCS
        file_extension = Path(file.filename).suffix
        gcs_filename = f"{folder}/{str(uuid.uuid4())[:12]}{file_extension}"
        
        # Upload to GCS
        blob = bucket.blob(gcs_filename)
        blob.upload_from_string(content, content_type=file.content_type)
        
        logger.info(f"✅ Uploaded to GCS: {gcs_filename}, size={len(content)}")
        
        return {
            "gcs_path": gcs_filename,
            "gcs_bucket": bucket_name,
            "storage_type": "gcs",
            "file_size": len(content),
            "mime_type": file.content_type,
            "file_name": file.filename
        }
    except Exception as e:
        logger.error(f"❌ GCS upload failed: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to upload to cloud storage: {str(e)}")

async def download_from_gcs(gcs_path: str) -> tuple:
    """
    Download file from Google Cloud Storage
    Returns tuple of (file_content, mime_type)
    """
    try:
        from google.cloud import storage
        import os
        
        # Initialize GCS client
        os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = "/app/backend/gcs-service-account.json"
        storage_client = storage.Client()
        bucket_name = os.getenv("GCS_BUCKET_NAME", "rimuru-file-uploads")
        bucket = storage_client.bucket(bucket_name)
        
        # Download from GCS
        blob = bucket.blob(gcs_path)
        
        if not blob.exists():
            raise HTTPException(status_code=404, detail="File not found in cloud storage")
        
        content = blob.download_as_bytes()
        mime_type = blob.content_type or "application/octet-stream"
        
        logger.info(f"✅ Downloaded from GCS: {gcs_path}, size={len(content)} bytes")
        
        return content, mime_type
    except Exception as e:
        logger.error(f"❌ GCS download failed for {gcs_path}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to download from cloud storage: {str(e)}")

# Notification translations
NOTIFICATION_TRANSLATIONS = {
    'en': {
        'new_user_registration': '👤 New User Registration',
        'user_registered': 'New user {username} has registered and needs approval.',
        'password_reset': '🔄 Password Reset',
        'password_reset_requested': 'User {username} has requested a password reset.',
        'payment_verified': '✅ Payment Verified',
        'payment_rejected': '❌ Payment Rejected',
        'payment_status_updated': 'Payment has been {status}.',
        'new_account_request': '🔔 New {platform} Request',
        'account_request_submitted': 'New {platform} account request from user {username}.',
        'new_topup_request': '💰 New Top-Up Request',
        'topup_request_submitted': 'User {username} submitted a top-up request of {amount}.',
        'wallet_transfer_success': '✅ Wallet Transfer Success',
        'wallet_transfer_completed': 'Successfully transferred {amount} from wallet to account {account_name}.',
        'wallet_topup_request': '🔔 New Wallet Top-Up Request',
        'wallet_transfer_request': '🔄 New Wallet Transfer Request',
        'wallet_transfer_needs_verification': 'User {username} requested wallet transfer of {amount} to {account_name}. Please verify.',
        'wallet_transfer_submitted': 'Wallet Transfer Submitted',
        'wallet_transfer_pending_admin': 'Your wallet transfer request of {amount} to {account_name} is pending admin verification.',
        'payment_proof_uploaded': '📸 Payment Proof Uploaded',
        'proof_uploaded_message': 'User {username} uploaded payment proof for request #{code}.',
        'account_request_approved': '🎉 {platform} Request Approved',
        'account_approved_message': 'Your {platform} request \'{account_name}\' has been approved! Your account is currently being shared and will be ready soon.',
        'account_status_changed': '📢 Account Status Changed',
        'account_status_message': 'Your {platform} account \'{account_name}\' has been {status}.',
        'account_deleted': '❌ Account Deleted',
        'account_deleted_message': 'Your {platform} account \'{account_name}\' has been permanently deleted.',
        'new_withdraw_request': '🏦 New Withdraw Request',
        'withdraw_request_message': 'User {username} requested withdraw of {currency} {amount} from {platform} account',
        'withdraw_approved': '✅ Withdraw Approved',
        'withdraw_approved_message': 'Your withdraw request of {currency} {amount} has been approved and processed.',
        'withdraw_rejected': '❌ Withdraw Rejected',
        'withdraw_rejected_message': 'Your withdraw request of {currency} {amount} has been rejected. {notes}'
    },
    'id': {
        'new_user_registration': '👤 Registrasi Pengguna Baru',
        'user_registered': 'Pengguna baru {username} telah mendaftar dan membutuhkan persetujuan.',
        'password_reset': '🔄 Reset Password',
        'password_reset_requested': 'Pengguna {username} meminta reset password.',
        'payment_verified': '✅ Pembayaran Diverifikasi',
        'payment_rejected': '❌ Pembayaran Ditolak',
        'payment_status_updated': 'Pembayaran telah {status}.',
        'new_account_request': '🔔 Permintaan {platform} Baru',
        'account_request_submitted': 'Permintaan akun {platform} baru dari pengguna {username}.',
        'new_topup_request': '💰 Permintaan Top-Up Baru',
        'topup_request_submitted': 'User {username} mengajukan permintaan top-up sebesar {amount}.',
        'wallet_transfer_success': '✅ Transfer Wallet Berhasil',
        'wallet_transfer_completed': 'Berhasil transfer {amount} dari wallet ke akun {account_name}.',
        'wallet_topup_request': '🔔 Permintaan Wallet Top-Up Baru',
        'wallet_transfer_request': '🔄 Permintaan Transfer Wallet Baru',
        'wallet_transfer_needs_verification': 'Pengguna {username} mengajukan transfer wallet sebesar {amount} ke {account_name}. Mohon verifikasi.',
        'wallet_transfer_submitted': 'Transfer Wallet Dikirim',
        'wallet_transfer_pending_admin': 'Permintaan transfer wallet Anda sebesar {amount} ke {account_name} sedang menunggu verifikasi admin.',
        'payment_proof_uploaded': '📸 Bukti Pembayaran Diupload',
        'proof_uploaded_message': 'Pengguna {username} mengupload bukti pembayaran untuk permintaan #{code}.',
        'account_request_approved': '🎉 Permintaan {platform} Disetujui',
        'account_approved_message': 'Permintaan {platform} Anda \'{account_name}\' telah disetujui! Akun Anda sedang dalam proses share.',
        'account_request_completed': '✅ Akun {platform} Siap Digunakan',
        'account_completed_message': 'Akun {platform} Anda \'{account_name}\' telah berhasil dibagikan dan sekarang aktif! Silakan login untuk mulai menggunakan.',
        'account_status_changed': '📢 Status Akun Berubah',
        'account_status_message': 'Akun {platform} Anda \'{account_name}\' telah {status}.',
        'account_deleted': '❌ Akun Dihapus',
        'account_deleted_message': 'Akun {platform} Anda \'{account_name}\' telah dihapus secara permanen.',
        'new_withdraw_request': '🏦 Permintaan Withdraw Baru',
        'withdraw_request_message': 'Pengguna {username} meminta withdraw {currency} {amount} dari akun {platform}',
        'withdraw_approved': '✅ Withdraw Disetujui',
        'withdraw_approved_message': 'Permintaan withdraw Anda sebesar {currency} {amount} telah disetujui dan diproses.',
        'withdraw_rejected': '❌ Withdraw Ditolak',
        'withdraw_rejected_message': 'Permintaan withdraw Anda sebesar {currency} {amount} telah ditolak. {notes}'
    }
}

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Security
security = HTTPBearer()
SECRET_KEY = os.environ.get("SECRET_KEY", "your-secret-key-here-change-in-production")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 1440  # Admin token: 24 hours (1440 minutes)
CLIENT_TOKEN_EXPIRE_MINUTES = 43200  # Client token: 30 days (43200 minutes)

# Create the main app without a prefix
app = FastAPI()

origins = [origin.strip() for origin in os.environ.get("CORS_ORIGINS", "").split(",") if origin.strip()]

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=origins,
    allow_methods=["*"],
    allow_headers=["*"],
)
# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Initialize APScheduler for auto-cancel tasks
scheduler = AsyncIOScheduler()

# Helper functions for precise financial calculations
def to_decimal(value):
    """Convert float to Decimal for precise financial calculations"""
    if value is None:
        return Decimal('0')
    return Decimal(str(value))

def decimal_add(a, b):
    """Precise addition of two numbers"""
    return to_decimal(a) + to_decimal(b)

def decimal_subtract(a, b):
    """Precise subtraction of two numbers"""
    return to_decimal(a) - to_decimal(b)

def decimal_round(value, places=2):
    """Round decimal to specified places"""
    return to_decimal(value).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

def to_float(decimal_value):
    """Convert Decimal back to float for database storage"""
    return float(decimal_value)

# Notification helper functions
def get_notification_text(key: str, lang: str = 'id', **kwargs):
    """Get notification text based on language"""
    translations = NOTIFICATION_TRANSLATIONS.get(lang, NOTIFICATION_TRANSLATIONS['id'])
    text = translations.get(key, key)
    
    # Format text with provided kwargs
    try:
        return text.format(**kwargs)
    except KeyError:
        return text

async def create_notification(title: str, message: str, notification_type: str, reference_id: str = None):
    """Create ONE notification (not per admin) with duplicate protection"""
    # FIXED: Create only 1 notification, not 1 per admin
    # All admins will see the same notification
    notification = {
        "id": str(uuid.uuid4()),
        "admin_id": None,  # Null = visible to all admins
        "title": title,
        "message": message,
        "type": notification_type,
        "reference_id": reference_id,
        "is_read": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    try:
        await db.notifications.insert_one(notification)
        logger.info(f"Notification created: {title}")
    except Exception as e:
        # Catch duplicate key error (code 11000) from unique index
        if "duplicate key" in str(e).lower() or "11000" in str(e):
            logger.warning(f"Duplicate notification blocked by database: {reference_id}")
        else:
            # Re-raise other errors
            raise
        
async def create_localized_notification(title_key: str, message_key: str, notification_type: str, 
                                       lang: str = 'id', reference_id: str = None, **kwargs):
    """Create localized admin notification with idempotency check"""
    # CRITICAL: Check for existing notification BEFORE creating
    # This prevents race conditions
    if reference_id:
        existing = await db.notifications.find_one({
            "reference_id": reference_id,
            "type": notification_type
        })
        if existing:
            logger.info(f"Notification already exists for {reference_id}, skipping duplicate")
            return  # STOP! Don't create duplicate
    
    title = get_notification_text(title_key, lang, **kwargs)
    message = get_notification_text(message_key, lang, **kwargs)
    await create_notification(title, message, notification_type, reference_id)

async def get_active_admin_emails() -> List[str]:
    """Get all active admin email addresses"""
    admins = await db.admin_users.find({"is_active": {"$ne": False}}).to_list(None)
    return [admin["email"] for admin in admins if admin.get("email")]

@app.on_event("startup")
async def startup_event():
    """Initialize database, auto-migrate payment proofs, and start scheduler on startup"""
    try:
        # Check if any admin users exist
        admin_count = await db.admin_users.count_documents({})
        
        if admin_count == 0:
            # Create default super admin
            default_admin = {
                "id": str(uuid.uuid4()),
                "username": "admin",
                "email": "admin@rimuru.com",
                "password_hash": get_password_hash("admin123"),
                "full_name": "System Administrator",
                "whatsapp_number": "+628123456789", 
                "is_super_admin": True,
                "created_at": datetime.now(timezone.utc)
            }
            
            await db.admin_users.insert_one(prepare_for_mongo(default_admin))
            logger.info("Default admin user created: username=admin, password=admin123")
        
        # Create database indexes for optimized queries
        logger.info("Creating database indexes...")
        try:
            # Index for ad_account_requests lookups by user_id
            await db.ad_account_requests.create_index([("user_id", 1)])
            # Index for transactions lookups by user_id, type, and status
            await db.transactions.create_index([("user_id", 1), ("type", 1), ("status", 1)])
            # Index for admin_users lookups by id
            await db.admin_users.create_index([("id", 1)])
            # Index for users by created_at for sorting
            await db.users.create_index([("created_at", -1)])
            logger.info("Database indexes created successfully")
        except Exception as idx_error:
            logger.warning(f"Index creation warning: {idx_error}")
        
        logger.info("Database initialization completed")
        
        # Auto-migrate payment proofs on startup
        logger.info("🔄 Running auto-migration for payment proofs...")
        await auto_migrate_payment_proofs()
        
        # Start the scheduler for auto-cancel tasks
        logger.info("🚀 Starting APScheduler for auto-cancel tasks...")
        
        # Add job to run auto-cancel every hour
        scheduler.add_job(
            auto_cancel_expired_topup_requests,
            IntervalTrigger(hours=1),
            id='auto_cancel_account_topups',
            name='Auto-cancel expired account top-up requests',
            replace_existing=True
        )
        
        scheduler.add_job(
            auto_cancel_expired_wallet_topup_requests,
            IntervalTrigger(hours=1),
            id='auto_cancel_wallet_topups',
            name='Auto-cancel expired wallet top-up requests',
            replace_existing=True
        )
        
        scheduler.start()
        logger.info("✅ Scheduler started successfully - auto-cancel will run every 1 hour")
        
    except Exception as e:
        logger.error(f"Startup failed: {e}")


@app.on_event("shutdown")
async def shutdown_event():
    """Shutdown scheduler on application shutdown"""
    try:
        scheduler.shutdown()
        logger.info("✅ Scheduler shutdown successfully")
    except Exception as e:
        logger.error(f"Shutdown failed: {e}")

async def auto_migrate_payment_proofs():
    """Auto-migrate filesystem payment proofs to database storage"""
    try:
        # Find proofs that need migration
        old_proofs = await db.payment_proofs.find({
            "$or": [
                {"storage_type": {"$exists": False}},
                {"storage_type": "local"}
            ]
        }).to_list(None)
        
        if len(old_proofs) == 0:
            logger.info("✅ No proofs need migration - all already in database")
            return
        
        logger.info(f"📦 Found {len(old_proofs)} proofs to migrate to database")
        
        migrated = 0
        failed = 0
        
        for proof in old_proofs:
            proof_id = proof.get("id")
            file_path = proof.get("file_path")
            
            if not file_path:
                failed += 1
                continue
            
            file_path_obj = Path(file_path)
            if not file_path_obj.is_absolute():
                file_path_obj = Path("/app") / file_path
            
            if not file_path_obj.exists():
                logger.warning(f"⚠️  Migration skipped - file not found: {file_path}")
                failed += 1
                continue
            
            try:
                with open(file_path_obj, 'rb') as f:
                    file_content = f.read()
                
                file_base64 = base64.b64encode(file_content).decode('utf-8')
                
                await db.payment_proofs.update_one(
                    {"id": proof_id},
                    {
                        "$set": {
                            "file_data": file_base64,
                            "storage_type": "database",
                            "migrated_from_filesystem": True,
                            "original_file_path": file_path
                        }
                    }
                )
                
                migrated += 1
                logger.info(f"✅ Migrated proof {proof_id}")
                
            except Exception as e:
                logger.error(f"❌ Failed to migrate {proof_id}: {e}")
                failed += 1
        
        logger.info(f"🎉 Auto-migration complete: {migrated} migrated, {failed} failed")
        
    except Exception as e:
        logger.error(f"❌ Auto-migration error: {e}")


def prepare_for_mongo(data):
    """Convert datetime objects to ISO strings for MongoDB with Z suffix for UTC"""
    if isinstance(data, dict):
        for key, value in data.items():
            if isinstance(value, datetime):
                # Convert to ISO format and replace timezone offset with Z
                iso_str = value.isoformat()
                # Remove timezone offset like +00:00 and replace with Z
                if '+00:00' in iso_str:
                    iso_str = iso_str.replace('+00:00', 'Z')
                elif not iso_str.endswith('Z'):
                    iso_str = iso_str + 'Z'
                data[key] = iso_str
    return data

def parse_from_mongo(item):
    """Convert ISO strings back to datetime objects and handle ObjectId"""
    if isinstance(item, dict):
        # Remove MongoDB's _id field to avoid ObjectId serialization issues
        if '_id' in item:
            del item['_id']
        
        for key, value in item.items():
            if isinstance(value, str) and 'T' in value and 'Z' in value:
                try:
                    item[key] = datetime.fromisoformat(value.replace('Z', '+00:00'))
                except ValueError:
                    pass
    return item

def generate_invoice_pdf(invoice_data: "InvoiceData") -> bytes:
    """Generate PDF invoice for top-up request"""
    buffer = BytesIO()
    
    # Create PDF document
    doc = SimpleDocTemplate(buffer, pagesize=A4, 
                          rightMargin=72, leftMargin=72,
                          topMargin=72, bottomMargin=18)
    
    # Get styles
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        alignment=1,  # Center
        textColor=colors.HexColor('#1f2937')
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=12,
        textColor=colors.HexColor('#374151')
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=6
    )
    
    # Story elements
    story = []
    
    # Add Rimuru Logo (if exists) 
    try:
        logo_path = "/app/frontend/public/images/rimuru-logo.png"
        import os
        if os.path.exists(logo_path):
            # Preserve aspect ratio for logo - calculate height based on width
            from PIL import Image as PILImage
            with PILImage.open(logo_path) as img:
                original_width, original_height = img.size
                desired_width = 2*inch
                aspect_ratio = original_height / original_width
                desired_height = desired_width * aspect_ratio
            
            logo = Image(logo_path, width=desired_width, height=desired_height)
            logo.hAlign = 'CENTER'
            story.append(logo)
            story.append(Spacer(1, 12))
    except Exception as e:
        logger.warning(f"Could not add logo to invoice: {e}")
    
    # Title
    story.append(Paragraph("RIMURU - INVOICE TOP UP SALDO", title_style))
    story.append(Spacer(1, 12))
    
    # Invoice info section
    invoice_info = [
        ["Invoice ID:", invoice_data.invoice_id],
        ["Tanggal:", invoice_data.created_at.astimezone(timezone(timedelta(hours=7))).strftime("%d %B %Y, %H:%M WIB")],
        ["Client:", invoice_data.user_name],
        ["Email:", invoice_data.user_email],
        ["Mata Uang:", invoice_data.currency],
        ["Status Pembayaran:", invoice_data.payment_status]
    ]
    
    invoice_table = Table(invoice_info, colWidths=[2*inch, 3*inch])
    invoice_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    
    story.append(invoice_table)
    story.append(Spacer(1, 20))
    
    # Account details section
    story.append(Paragraph("DETAIL AKUN TOP UP", heading_style))
    
    account_data = [["Platform", "Nama Akun", "Account ID", "Jumlah Top Up", "Fee", "Total"]]
    
    for account in invoice_data.accounts:
        platform = account['platform'].upper()
        name = account['account_name'][:25] + "..." if len(account['account_name']) > 25 else account['account_name']
        account_id = account['account_id'][:15] + "..." if len(account['account_id']) > 15 else account['account_id']
        amount = f"{invoice_data.currency} {account['amount']:,.2f}"
        fee = f"{invoice_data.currency} {account['fee']:,.2f}"
        total = f"{invoice_data.currency} {account['total']:,.2f}"
        
        account_data.append([platform, name, account_id, amount, fee, total])
    
    account_table = Table(account_data, colWidths=[0.8*inch, 1.5*inch, 1.2*inch, 1*inch, 0.8*inch, 1*inch])
    account_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f3f4f6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
    ]))
    
    story.append(account_table)
    story.append(Spacer(1, 20))
    
    # Summary section
    story.append(Paragraph("RINGKASAN PEMBAYARAN", heading_style))
    
    currency_symbol = "Rp" if invoice_data.currency == "IDR" else "$"
    
    summary_data = [
        ["Subtotal:", f"{currency_symbol} {invoice_data.subtotal:,.2f}"],
        ["Total Fee:", f"{currency_symbol} {invoice_data.fees:,.2f}"],
        ["Kode Unik:", f"{currency_symbol} {invoice_data.unique_code:.2f}"],
        ["TOTAL TRANSFER:", f"{currency_symbol} {invoice_data.total:,.2f}"]
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (0, -2), 'Helvetica'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -2), 10),
        ('FONTSIZE', (0, -1), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, -1), (-1, -1), 12),
        ('LINEABOVE', (0, -1), (-1, -1), 2, colors.black),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#dc2626')),
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 20))
    
    # Payment instructions
    story.append(Paragraph("INSTRUKSI PEMBAYARAN", heading_style))
    
    if invoice_data.currency == "IDR" and invoice_data.bank_details:
        bank_info = f"""
        <b>Transfer Bank BRI:</b><br/>
        Nama: {invoice_data.bank_details.get('account_name', 'N/A')}<br/>
        Nomor Rekening: {invoice_data.bank_details.get('account_number', 'N/A')}<br/>
        Bank: {invoice_data.bank_details.get('bank_name', 'BRI')}<br/><br/>
        <b>Jumlah yang harus ditransfer: {currency_symbol} {invoice_data.total:,.2f}</b>
        """
        story.append(Paragraph(bank_info, normal_style))
    elif invoice_data.currency == "USD" and invoice_data.crypto_wallet:
        crypto_info = f"""
        <b>Transfer USDT (TRC20):</b><br/>
        Wallet Address: {invoice_data.crypto_wallet}<br/>
        Network: TRC20<br/><br/>
        <b>Jumlah yang harus ditransfer: {currency_symbol} {invoice_data.total:.2f}</b>
        """
        story.append(Paragraph(crypto_info, normal_style))
    
    # Important notes
    story.append(Spacer(1, 15))
    notes = """
    <b>CATATAN PENTING:</b><br/>
    • Transfer sesuai dengan jumlah EXACT yang tertera di invoice ini<br/>
    • Kode unik membantu admin memverifikasi pembayaran dengan mudah<br/>
    • Upload bukti pembayaran setelah melakukan transfer<br/>
    • Saldo akan diproses dalam 1-24 jam setelah verifikasi pembayaran<br/>
    • Simpan invoice ini untuk referensi di kemudian hari
    """
    story.append(Paragraph(notes, normal_style))
    
    # Footer
    story.append(Spacer(1, 30))
    footer_text = f"Invoice generated on {datetime.now(timezone.utc).astimezone(timezone(timedelta(hours=7))).strftime('%d %B %Y, %H:%M')} WIB"
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=1  # Center
    )
    story.append(Paragraph(footer_text, footer_style))
    
    # Build PDF
    doc.build(story)
    
    # Get PDF bytes
    pdf_bytes = buffer.getvalue()
    buffer.close()
    
    return pdf_bytes

def generate_wallet_topup_invoice_pdf(invoice_data: "InvoiceData") -> bytes:
    """Generate PDF invoice for wallet top-up request"""
    buffer = BytesIO()
    
    # Create PDF document
    doc = SimpleDocTemplate(buffer, pagesize=A4, 
                          rightMargin=72, leftMargin=72,
                          topMargin=72, bottomMargin=18)
    
    # Get styles
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        alignment=1,  # Center
        textColor=colors.HexColor('#1f2937')
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=12,
        textColor=colors.HexColor('#374151')
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=6
    )
    
    # Story elements
    story = []
    
    # Add Rimuru Logo (if exists) 
    try:
        logo_path = "/app/frontend/public/images/rimuru-logo.png"
        import os
        if os.path.exists(logo_path):
            # Preserve aspect ratio for logo - calculate height based on width
            from PIL import Image as PILImage
            with PILImage.open(logo_path) as img:
                original_width, original_height = img.size
                desired_width = 2*inch
                aspect_ratio = original_height / original_width
                desired_height = desired_width * aspect_ratio
            
            logo = Image(logo_path, width=desired_width, height=desired_height)
            logo.hAlign = 'CENTER'
            story.append(logo)
            story.append(Spacer(1, 12))
    except Exception as e:
        logger.warning(f"Could not add logo to wallet invoice: {e}")
    
    # Title
    story.append(Paragraph("RIMURU - INVOICE WALLET TOP UP", title_style))
    story.append(Spacer(1, 12))
    
    # Invoice info section
    invoice_info = [
        ["Invoice ID:", invoice_data.invoice_id],
        ["Tanggal:", invoice_data.created_at.astimezone(timezone(timedelta(hours=7))).strftime("%d %B %Y, %H:%M WIB")],
        ["Client:", invoice_data.user_name],
        ["Email:", invoice_data.user_email],
        ["Wallet Type:", invoice_data.wallet_type.title()],
        ["Payment Method:", invoice_data.payment_method.upper()],
        ["Mata Uang:", invoice_data.currency],
        ["Status Pembayaran:", invoice_data.payment_status]  # Use payment_status directly (PAID/UNPAID)
    ]
    
    invoice_table = Table(invoice_info, colWidths=[2*inch, 3*inch])
    invoice_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    
    story.append(invoice_table)
    story.append(Spacer(1, 20))
    
    # Wallet top-up details section
    story.append(Paragraph("DETAIL WALLET TOP UP", heading_style))
    
    currency_symbol = "Rp" if invoice_data.currency == "IDR" else "$"
    
    wallet_data = [
        ["Jumlah Top Up:", f"{currency_symbol} {invoice_data.amount:,.2f}"],
        ["Kode Unik:", f"{currency_symbol} {invoice_data.unique_code:.2f}"],
        ["TOTAL TRANSFER:", f"{currency_symbol} {(invoice_data.amount + invoice_data.unique_code):,.2f}"]
    ]
    
    wallet_table = Table(wallet_data, colWidths=[3*inch, 2*inch])
    wallet_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (0, -2), 'Helvetica'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -2), 10),
        ('FONTSIZE', (0, -1), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, -1), (-1, -1), 12),
        ('LINEABOVE', (0, -1), (-1, -1), 2, colors.black),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#dc2626')),
    ]))
    
    story.append(wallet_table)
    story.append(Spacer(1, 20))
    
    # Payment instructions
    story.append(Paragraph("INSTRUKSI PEMBAYARAN", heading_style))
    
    if invoice_data.currency == "IDR" and invoice_data.bank_name:
        bank_info = f"""
        <b>Transfer Bank {invoice_data.bank_name}:</b><br/>
        Nama: {invoice_data.bank_holder or 'N/A'}<br/>
        Nomor Rekening: {invoice_data.bank_account or 'N/A'}<br/>
        Bank: {invoice_data.bank_name}<br/><br/>
        <b>Jumlah yang harus ditransfer: {currency_symbol} {(invoice_data.amount + invoice_data.unique_code):,.2f}</b>
        """
        story.append(Paragraph(bank_info, normal_style))
    elif invoice_data.currency == "USD" and invoice_data.crypto_wallet:
        crypto_info = f"""
        <b>Transfer USDT ({invoice_data.network or 'TRC20'}):</b><br/>
        Wallet Address: {invoice_data.crypto_wallet}<br/>
        Network: {invoice_data.network or 'TRC20'}<br/><br/>
        <b>Jumlah yang harus ditransfer: {currency_symbol} {invoice_data.amount:.2f}</b>
        """
        story.append(Paragraph(crypto_info, normal_style))
    
    # Important notes
    story.append(Spacer(1, 15))
    notes = """
    <b>CATATAN PENTING:</b><br/>
    • Transfer sesuai dengan jumlah EXACT yang tertera di invoice ini<br/>
    • Kode unik membantu admin memverifikasi pembayaran dengan mudah<br/>
    • Upload bukti pembayaran setelah melakukan transfer<br/>
    • Saldo wallet akan diproses dalam 1-24 jam setelah verifikasi pembayaran<br/>
    • Simpan invoice ini untuk referensi di kemudian hari
    """
    story.append(Paragraph(notes, normal_style))
    
    # Footer
    story.append(Spacer(1, 30))
    footer_text = f"Wallet Invoice generated on {datetime.now(timezone.utc).astimezone(timezone(timedelta(hours=7))).strftime('%d %B %Y, %H:%M')} WIB"
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=1  # Center
    )
    story.append(Paragraph(footer_text, footer_style))
    
    # Build PDF
    doc.build(story)
    
    # Get PDF bytes
    pdf_bytes = buffer.getvalue()
    buffer.close()
    
    return pdf_bytes


def generate_wallet_transfer_invoice_pdf(invoice_data: "InvoiceData", target_account_name: str, target_platform: str) -> bytes:
    """Generate PDF invoice for wallet transfer request"""
    buffer = BytesIO()
    
    # Create PDF document
    doc = SimpleDocTemplate(buffer, pagesize=A4, 
                          rightMargin=72, leftMargin=72,
                          topMargin=72, bottomMargin=18)
    
    # Get styles
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        alignment=1,  # Center
        textColor=colors.HexColor('#1f2937')
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=12,
        textColor=colors.HexColor('#374151')
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=6
    )
    
    # Story elements
    story = []
    
    # Add Rimuru Logo (if exists) 
    try:
        logo_path = "/app/frontend/public/images/rimuru-logo.png"
        import os
        if os.path.exists(logo_path):
            from PIL import Image as PILImage
            with PILImage.open(logo_path) as img:
                original_width, original_height = img.size
                desired_width = 2*inch
                aspect_ratio = original_height / original_width
                desired_height = desired_width * aspect_ratio
            
            logo = Image(logo_path, width=desired_width, height=desired_height)
            logo.hAlign = 'CENTER'
            story.append(logo)
            story.append(Spacer(1, 12))
    except Exception as e:
        logger.warning(f"Could not add logo to transfer invoice: {e}")
    
    # Title
    story.append(Paragraph("RIMURU - INVOICE WALLET TRANSFER", title_style))
    story.append(Spacer(1, 12))
    
    # Invoice info section
    invoice_info = [
        ["Invoice ID:", invoice_data.invoice_id],
        ["Tanggal:", invoice_data.created_at.astimezone(timezone(timedelta(hours=7))).strftime("%d %B %Y, %H:%M WIB")],
        ["Client:", invoice_data.user_name],
        ["Email:", invoice_data.user_email],
        ["Wallet Type:", invoice_data.wallet_type.title() if invoice_data.wallet_type else "N/A"],
        ["Target Account:", f"{target_account_name} ({target_platform})"],
        ["Mata Uang:", invoice_data.currency],
        ["Status:", invoice_data.payment_status]
    ]
    
    if invoice_data.verified_at:
        invoice_info.insert(-1, ["Diproses pada:", invoice_data.verified_at.astimezone(timezone(timedelta(hours=7))).strftime("%d %B %Y, %H:%M WIB")])
    
    invoice_table = Table(invoice_info, colWidths=[2*inch, 3*inch])
    invoice_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    
    story.append(invoice_table)
    story.append(Spacer(1, 20))
    
    # Transfer details section
    story.append(Paragraph("DETAIL TRANSFER", heading_style))
    
    currency_symbol = "Rp" if invoice_data.currency == "IDR" else "$"
    
    transfer_data = [
        ["Jumlah Transfer:", f"{currency_symbol} {invoice_data.amount:,.2f}"],
        ["Biaya Admin (5%):", f"{currency_symbol} {invoice_data.fees:,.2f}"],
        ["TOTAL DIKURANGI DARI WALLET:", f"{currency_symbol} {invoice_data.total:,.2f}"]
    ]
    
    transfer_table = Table(transfer_data, colWidths=[3*inch, 2*inch])
    transfer_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (0, -2), 'Helvetica'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -2), 10),
        ('FONTSIZE', (0, -1), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, -1), (-1, -1), 12),
        ('LINEABOVE', (0, -1), (-1, -1), 2, colors.black),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#dc2626')),
    ]))
    
    story.append(transfer_table)
    story.append(Spacer(1, 20))
    
    # Admin notes if available
    if invoice_data.admin_notes:
        story.append(Paragraph("CATATAN ADMIN", heading_style))
        story.append(Paragraph(invoice_data.admin_notes, normal_style))
        story.append(Spacer(1, 15))
    
    # Important notes
    story.append(Paragraph("CATATAN PENTING:", heading_style))
    notes = """
    • Transfer dari wallet ke akun iklan telah diproses<br/>
    • Saldo akun iklan akan diupdate sesuai jumlah transfer<br/>
    • Simpan invoice ini untuk referensi di kemudian hari<br/>
    • Untuk pertanyaan lebih lanjut, silakan hubungi admin
    """
    story.append(Paragraph(notes, normal_style))
    
    # Footer
    story.append(Spacer(1, 30))
    footer_text = f"Transfer Invoice generated on {datetime.now(timezone.utc).astimezone(timezone(timedelta(hours=7))).strftime('%d %B %Y, %H:%M')} WIB"
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=1  # Center
    )
    story.append(Paragraph(footer_text, footer_style))
    
    # Build PDF
    doc.build(story)
    
    # Get PDF bytes
    pdf_bytes = buffer.getvalue()
    buffer.close()
    
    return pdf_bytes

# Exchange Rate Functions
async def get_exchange_rate(from_currency: str, to_currency: str) -> float:
    """Fetch real-time exchange rate from API"""
    try:
        if from_currency == to_currency:
            return 1.0
        
        # Using ExchangeRate-API.com (free tier)
        async with httpx.AsyncClient() as client:
            response = await client.get(f"https://api.exchangerate-api.com/v4/latest/{from_currency}")
            if response.status_code == 200:
                data = response.json()
                if to_currency in data.get("rates", {}):
                    return float(data["rates"][to_currency])
            
            # Fallback: if API fails, use approximate rate (should be replaced with cached rates in production)
            if from_currency == "IDR" and to_currency == "USD":
                return 0.000067  # Approximate IDR to USD
            elif from_currency == "USD" and to_currency == "IDR":
                return 15000.0   # Approximate USD to IDR
            
    except Exception as e:
        logger.error(f"Error fetching exchange rate: {e}")
        # Fallback rates
        if from_currency == "IDR" and to_currency == "USD":
            return 0.000067
        elif from_currency == "USD" and to_currency == "IDR":
            return 15000.0
    
    return 1.0

# Models
class User(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    username: str
    email: str
    password_hash: str
    wallet_balance_idr: float = 0.0  # Legacy field - keep for backward compatibility
    wallet_balance_usd: float = 0.0  # Legacy field - keep for backward compatibility
    
    # New wallet system - separate main and withdrawal wallets
    main_wallet_idr: float = 0.0
    main_wallet_usd: float = 0.0
    withdrawal_wallet_idr: float = 0.0
    withdrawal_wallet_usd: float = 0.0
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    # Profile fields
    name: Optional[str] = None  # Full name
    display_name: Optional[str] = None
    phone_number: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    province: Optional[str] = None
    company_name: Optional[str] = None
    profile_picture: Optional[str] = None
    updated_at: Optional[datetime] = None

class AdminUser(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    username: str
    email: str
    password_hash: str
    full_name: str
    whatsapp_number: Optional[str] = None
    is_super_admin: bool = False
    profile_picture: Optional[str] = None
    last_login: Optional[datetime] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: Optional[datetime] = None

class AdminUserCreate(BaseModel):
    username: str
    email: str
    password: str
    full_name: str
    whatsapp_number: Optional[str] = None
    is_super_admin: bool = False

class AdminUserLogin(BaseModel):
    username: str
    password: str

class AdminUserProfile(BaseModel):
    id: str
    username: str
    email: str
    full_name: str
    whatsapp_number: Optional[str] = None
    is_super_admin: bool
    profile_picture: Optional[str] = None
    last_login: Optional[datetime] = None
    created_at: datetime

class AdminUserUpdate(BaseModel):
    username: Optional[str] = None
    email: Optional[str] = None
    full_name: Optional[str] = None
    whatsapp_number: Optional[str] = None
    current_password: str

class AdminUserEditRequest(BaseModel):
    username: str
    email: str
    full_name: Optional[str] = None
    whatsapp_number: Optional[str] = None
    is_super_admin: bool = False

class UserCreate(BaseModel):
    username: str
    name: str  # Full name - required
    company_name: Optional[str] = None  # Business/Company name - optional
    phone_number: str  # Phone number - required
    address: str  # Address - required
    city: str  # City - required
    province: str  # Province - required
    email: str
    password: str

class UserLogin(BaseModel):
    username: str
    password: str

class UserProfile(BaseModel):
    username: str
    email: str
    name: Optional[str] = None
    display_name: Optional[str] = None
    phone_number: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    province: Optional[str] = None
    company_name: Optional[str] = None
    profile_picture: Optional[str] = None

class UserProfileUpdate(BaseModel):
    username: Optional[str] = None
    email: Optional[str] = None
    display_name: Optional[str] = None
    phone_number: Optional[str] = None
    address: Optional[str] = None
    company_name: Optional[str] = None
    current_password: str

class PasswordChange(BaseModel):
    current_password: str
    new_password: str
    confirm_password: str

class CurrencyExchange(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    from_currency: str  # "IDR" or "USD"
    to_currency: str    # "USD" or "IDR"
    from_amount: float
    to_amount: float
    exchange_rate: float
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ExchangeRequest(BaseModel):
    from_currency: str
    to_currency: str
    amount: float

class Group(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    user_id: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    
class GroupCreate(BaseModel):
    name: str

class AdAccount(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    platform: str  # facebook, google, tiktok
    account_name: str
    account_id: str
    balance: float = 0.0
    status: str = "pending"  # pending, active, suspended
    fee_percentage: Optional[float] = None  # Fee percentage for top-ups (e.g., 5.0 for 5%)
    group_id: Optional[str] = None  # Group ID for organization
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    # Facebook-specific fields
    gmt: Optional[str] = None
    currency: Optional[str] = None
    delivery_method: Optional[str] = None
    bm_id_or_email: Optional[str] = None
    # Google Ads specific fields
    email: Optional[str] = None  # Google Ads email
    website: Optional[str] = None  # Website URL (used by both Google Ads and TikTok Ads)
    # TikTok Ads specific fields  
    bc_id: Optional[str] = None  # TikTok Business Center ID
    # Common fields
    notes: Optional[str] = None

class AdAccountRequest(BaseModel):
    platform: str
    account_name: str
    # Facebook-specific fields
    gmt: Optional[str] = None
    currency: Optional[str] = None  # IDR or USD
    delivery_method: Optional[str] = None  # BM_ID or EMAIL
    bm_id_or_email: Optional[str] = None  # Single BM/Email (legacy)
    bm_ids: Optional[List[str]] = None  # Multiple BM IDs for 1 account (NEW)
    
    # Google Ads specific fields  
    email: Optional[str] = None  # Google Ads email
    website: Optional[str] = None  # Google Ads website URL
    
    # TikTok Ads specific fields
    bc_id: Optional[str] = None  # TikTok Business Center ID
    website: Optional[str] = None  # TikTok Ads website URL (same field name as Google for consistency)
    
    # Common fields
    notes: Optional[str] = None
    group_id: Optional[str] = None  # Optional group association
    
    # Admin fields
    status: str = "pending"  # pending, approved, rejected, processing, completed, failed
    admin_notes: Optional[str] = None
    admin_id: Optional[str] = None  # Admin who processed the request
    processed_at: Optional[datetime] = None

class RequestStatusUpdate(BaseModel):
    status: str  # approved, rejected, processing, completed, failed
    admin_notes: Optional[str] = None
    account_id: Optional[str] = None  # Required for Facebook Ads when status is approved
    account_name: Optional[str] = None  # Optional: Admin can edit account name during approval
    fee_percentage: Optional[float] = None  # Required fee percentage when status is approved

class ClientStatusUpdate(BaseModel):
    is_active: bool

class AccountStatusUpdate(BaseModel):
    status: str  # active, suspended, disabled

class AccountFeeUpdate(BaseModel):
    fee_percentage: float  # Fee percentage (0-100)

class AccountGroup(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    name: str
    description: Optional[str] = None
    account_ids: List[str] = []
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class AccountGroupCreate(BaseModel):
    name: str
    description: Optional[str] = None
    accounts: List[str] = []  # Account IDs to add to group

class AccountGroupUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    accounts: Optional[List[str]] = None

class AddAccountsToGroup(BaseModel):
    account_ids: List[str]

class BulkRequestUpdate(BaseModel):
    request_ids: List[str]
    status: str  # approved, rejected, processing, completed, failed  
    admin_notes: Optional[str] = None
    fee_percentage: Optional[float] = None  # For bulk approvals

class InvoiceData(BaseModel):
    invoice_id: str
    user_name: str
    user_email: str
    currency: str
    accounts: Optional[List[Dict]] = []  # List of account topup details (optional for wallet)
    subtotal: Optional[float] = 0.0
    fees: Optional[float] = 0.0
    unique_code: float
    total: Optional[float] = 0.0
    amount: Optional[float] = 0.0  # For wallet topups
    bank_details: Optional[Dict] = None
    crypto_wallet: Optional[str] = None
    created_at: datetime
    verified_at: Optional[datetime] = None
    payment_status: str = "NON PAID"  # PAID or NON PAID
    # Wallet-specific fields
    wallet_type: Optional[str] = None
    payment_method: Optional[str] = None
    reference_code: Optional[str] = None
    bank_name: Optional[str] = None
    bank_account: Optional[str] = None
    bank_holder: Optional[str] = None
    network: Optional[str] = None
    admin_notes: Optional[str] = None

# ===== LANDING PAGE MODELS =====
class PricingPackage(BaseModel):
    name: str
    price: float
    original_price: float = 0
    description: str = ""
    features: List[str] = []
    badge: str = ""
    is_highlighted: bool = False
    cta_text: str = "Beli Sekarang"

class WhatsAppCS(BaseModel):
    number: str
    name: str = "CS"
    percentage: int = 100  # Distribution percentage (must total 100% across all CS)

class LandingPageCreate(BaseModel):
    template_id: str = "modern_gradient"  # Template selection
    product_name: str
    product_description: str
    pricing_mode: str = "single"  # "single" or "multiple"
    product_price: Optional[float] = None
    product_original_price: Optional[float] = None
    currency: str = "IDR"
    pricing_packages: List[Dict[str, Any]] = []
    benefits: List[str] = []
    hero_image: str = ""
    gallery_images: List[str] = []
    testimonials: List[Dict[str, str]] = []
    primary_color: str = "#0EA5E9"
    accent_color: str = "#F59E0B"
    font_heading: str = "Inter"
    font_body: str = "Inter"
    facebook_pixel_id: str = ""
    tiktok_pixel_id: str = ""
    ga_measurement_id: str = ""
    whatsapp_numbers: List[Dict[str, Any]] = []  # Multiple WhatsApp CS with rotation
    whatsapp_number: str = ""  # Deprecated - kept for backward compatibility
    cta_event_name: str = "Contact"
    seo_title: str = ""
    seo_description: str = ""
    seo_keywords: List[str] = []
    slug: str
    product_details: Optional[Dict[str, Any]] = None  # CRITICAL FIX: Add product_details field

class LandingPageUpdate(BaseModel):
    template_id: Optional[str] = None
    product_name: Optional[str] = None
    product_description: Optional[str] = None
    pricing_mode: Optional[str] = None
    product_price: Optional[float] = None
    product_original_price: Optional[float] = None
    currency: Optional[str] = None
    pricing_packages: Optional[List[Dict[str, Any]]] = None
    benefits: Optional[List[str]] = None
    hero_image: Optional[str] = None
    gallery_images: Optional[List[str]] = None
    testimonials: Optional[List[Dict[str, str]]] = None
    primary_color: Optional[str] = None
    accent_color: Optional[str] = None
    font_heading: Optional[str] = None
    font_body: Optional[str] = None
    facebook_pixel_id: Optional[str] = None
    tiktok_pixel_id: Optional[str] = None
    ga_measurement_id: Optional[str] = None
    whatsapp_numbers: Optional[List[Dict[str, Any]]] = None
    whatsapp_number: Optional[str] = None
    cta_event_name: Optional[str] = None
    seo_title: Optional[str] = None
    seo_description: Optional[str] = None
    seo_keywords: Optional[List[str]] = None
    product_details: Optional[Dict[str, Any]] = None  # CRITICAL FIX: Add product_details field

class LandingPage(BaseModel):
    id: str
    user_id: str
    username: str
    template_id: str = "modern_gradient"
    product_name: str
    product_description: str
    pricing_mode: str = "single"
    product_price: Optional[float] = None
    product_original_price: Optional[float] = None
    currency: str = "IDR"
    pricing_packages: List[Dict[str, Any]] = []
    benefits: List[str] = []
    hero_image: str = ""
    gallery_images: List[str] = []
    testimonials: List[Dict[str, str]] = []
    primary_color: str = "#0EA5E9"
    accent_color: str = "#F59E0B"
    font_heading: str = "Inter"
    font_body: str = "Inter"
    facebook_pixel_id: str = ""
    tiktok_pixel_id: str = ""
    ga_measurement_id: str = ""
    whatsapp_numbers: List[Dict[str, Any]] = []
    whatsapp_number: str = ""
    cta_event_name: str = "Contact"
    seo_title: str = ""
    seo_description: str = ""
    seo_keywords: List[str] = []
    slug: str
    copy_blocks: Dict[str, Any] = {}
    layout_map: Dict[str, Any] = {}
    status: str = "draft"
    created_at: str
    updated_at: str


class Notification(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    message: str
    type: str  # new_request, new_topup, new_user, status_change
    reference_id: Optional[str] = None  # ID of related request/transaction/user
    is_read: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class NotificationResponse(BaseModel):
    id: str
    title: str
    message: str
    type: str
    reference_id: Optional[str] = None
    admin_id: Optional[str] = None  # Admin ID for admin notifications
    is_read: bool = False  # Default to False for backward compatibility
    created_at: datetime

class ClientNotification(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str  # Client user ID
    title: str
    message: str
    type: str  # approval, rejection, completed, info
    reference_id: Optional[str] = None  # ID of related request/transaction
    is_read: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ClientNotificationResponse(BaseModel):
    id: str
    title: str
    message: str
    type: str
    reference_id: Optional[str] = None
    is_read: bool = False  # Default to False for backward compatibility
    created_at: datetime

class ShareRequest(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    account_id: str  # The ad account being shared
    platform: str  # facebook, google, tiktok
    account_name: str  # Name of the account being shared
    # Share targets based on platform - Changed to List to support multiple recipients
    target_bm_email: Optional[List[str]] = None  # Facebook: List of BM IDs or Emails
    target_email: Optional[List[str]] = None     # Google: List of Emails
    target_bc_id: Optional[List[str]] = None     # TikTok: List of BC IDs
    notes: Optional[str] = None
    status: str = "pending"  # pending, approved, rejected, completed, failed
    admin_notes: Optional[str] = None
    processed_by: Optional[str] = None  # Admin ID who processed
    processed_at: Optional[datetime] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ShareRequestCreate(BaseModel):
    account_id: str
    target_bm_email: Optional[List[str]] = None  # For Facebook - List of BM IDs/Emails
    target_email: Optional[List[str]] = None     # For Google - List of Emails
    target_bc_id: Optional[List[str]] = None     # For TikTok - List of BC IDs
    notes: Optional[str] = None

class ShareRequestResponse(BaseModel):
    id: str
    account_id: str
    platform: str
    account_name: str
    target_bm_email: Optional[List[str]] = None  # List of BM IDs/Emails
    target_email: Optional[List[str]] = None     # List of Emails
    target_bc_id: Optional[List[str]] = None     # List of BC IDs
    notes: Optional[str] = None
    status: str
    admin_notes: Optional[str] = None
    created_at: datetime

class Transaction(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    type: str  # topup, withdraw, account_request, wallet_to_account_transfer
    amount: float
    currency: str = "IDR"  # IDR, USD - default to IDR for backward compatibility
    description: str
    status: str = "pending"  # pending, completed, failed
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    # Additional fields for wallet transfers
    account_id: Optional[str] = None
    account_name: Optional[str] = None
    reference_id: Optional[str] = None  # Links to wallet_transfers collection
    fee: Optional[float] = None
    total_amount: Optional[float] = None
    updated_at: Optional[datetime] = None

class TopUpAccount(BaseModel):
    account_id: str
    amount: float
    fee_percentage: float = 0.0
    fee_amount: float = 0.0

class TopUpRequest(BaseModel):
    currency: str  # IDR or USD
    accounts: List[TopUpAccount]
    total_amount: float
    total_fee: float
    unique_code: Optional[int] = None  # Accept unique code from frontend

class PaymentProof(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    topup_request_id: str
    user_id: str
    file_name: str
    file_path: Optional[str] = None  # Optional - for legacy filesystem storage
    gcs_path: Optional[str] = None  # GCS storage path
    gcs_bucket: Optional[str] = None  # GCS bucket name
    storage_type: Optional[str] = "gcs"  # "gcs", "database", or "local"
    file_size: int
    mime_type: str
    uploaded_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class TopUpRequestRecord(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    currency: str
    accounts: List[dict]  # Store account details with account_name, account_platform, and proofs
    total_amount: float
    total_fee: float
    unique_code: int  # 3 digit unique code (100-999)
    total_with_unique_code: float  # total_amount + unique_code
    reference_code: Optional[str] = None  # Reference code for tracking (RMR + 8 chars)
    status: str = "pending"  # pending, proof_uploaded, verified, completed, rejected
    payment_proof_id: Optional[str] = None  # Client's payment proof (shared for all accounts)
    admin_id: Optional[str] = None
    admin_notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    verified_at: Optional[datetime] = None
    
    # Bank/Wallet details for transfer (set dynamically based on currency)
    bank_name: Optional[str] = None
    bank_account: Optional[str] = None
    bank_holder: Optional[str] = None
    wallet_address: Optional[str] = None
    wallet_name: Optional[str] = None
    network: Optional[str] = None
    
# Note: Each account in accounts array can have:
# - account_id, account_name, account_platform, amount, fee_amount, fee_percentage
# - spend_limit_proof_url: Optional[str] (admin uploads this)
# - budget_aspire_proof_url: Optional[str] (admin uploads this)

# Wallet models
class WalletTopUpRequest(BaseModel):
    wallet_type: str = "main"  # main or withdrawal
    currency: str  # IDR or USD
    amount: float
    payment_method: str  # bank_bca, bank_mandiri, usdt_trc20, etc.
    notes: Optional[str] = None

class WalletTopUpRecord(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    wallet_type: str  # main or withdrawal
    currency: str
    amount: float
    payment_method: str
    notes: Optional[str] = None
    unique_code: int  # 3 digit unique code (100-999)
    total_with_unique_code: float  # amount + unique_code (for IDR only)
    status: str = "pending"  # pending, proof_uploaded, verified, completed, rejected
    payment_proof_id: Optional[str] = None
    admin_id: Optional[str] = None
    admin_notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    verified_at: Optional[datetime] = None
    
    # Bank/Wallet details for transfer
    bank_name: Optional[str] = None
    bank_account: Optional[str] = None
    bank_holder: Optional[str] = None
    wallet_address: Optional[str] = None
    wallet_name: Optional[str] = None
    network: Optional[str] = None
    reference_code: Optional[str] = None

# Wallet Deduction Request Models
class WalletDeductionRequest(BaseModel):
    wallet_type: str  # main_idr, main_usd, withdrawal_idr, withdrawal_usd
    amount: float
    reason: str
    # proof_file_id will be stored separately after upload

class WalletDeductionRecord(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    client_id: str
    client_name: str
    admin_id: str
    admin_username: str
    wallet_type: str  # main_idr, main_usd, withdrawal_idr, withdrawal_usd
    amount: float
    reason: str
    proof_file_id: Optional[str] = None
    proof_file_url: Optional[str] = None
    status: str = "pending"  # pending, approved, rejected
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    processed_at: Optional[datetime] = None
    super_admin_id: Optional[str] = None
    super_admin_username: Optional[str] = None
    approval_notes: Optional[str] = None

class WithdrawRequest(BaseModel):
    amount: float
    account_id: str  # ad account id to withdraw from

class WithdrawRequestNew(BaseModel):
    account_id: str  # ad account id to withdraw from
    currency: str  # IDR or USD (for wallet matching)

class WithdrawRequestRecord(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    account_id: str  # Source ad account
    platform: str  # facebook, google, tiktok
    account_name: str  # Name of the ad account
    requested_amount: float  # Amount requested by client
    admin_verified_amount: Optional[float] = None  # Real balance verified by admin
    currency: str  # IDR or USD
    status: str = "pending"  # pending, approved, rejected, completed
    admin_id: Optional[str] = None
    admin_notes: Optional[str] = None
    actual_balance_proof_url: Optional[str] = None  # Bukti saldo aktual
    after_withdrawal_proof_url: Optional[str] = None  # Bukti saldo setelah ditarik
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    processed_at: Optional[datetime] = None

class AdminWithdrawUpdate(BaseModel):
    status: str  # approved, rejected
    verified_amount: Optional[float] = None  # Actual balance verified by admin
    admin_notes: Optional[str] = None
    actual_balance_proof_url: Optional[str] = None  # Bukti saldo aktual
    after_withdrawal_proof_url: Optional[str] = None  # Bukti saldo setelah ditarik

class BalanceTransfer(BaseModel):
    from_type: str  # "wallet" or "account"
    to_type: str  # "wallet" or "account" 
    account_id: Optional[str] = None  # Required if from_type or to_type is "account"
    amount: float

class TransferRequestRecord(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    account_id: str  # Target account to transfer to
    amount: float
    currency: str = "IDR"  # IDR or USD
    status: str = "pending"  # pending, approved, rejected, completed
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    processed_at: Optional[datetime] = None
    admin_id: Optional[str] = None
    admin_notes: Optional[str] = None

# Admin Actions Models (for Super Admin Approval)
class AdminActionTopUpWallet(BaseModel):
    client_id: str
    wallet_type: str  # "main_idr", "main_usd", "withdrawal_idr", "withdrawal_usd"
    amount: float
    notes: Optional[str] = None
    
class AdminActionWithdrawAccount(BaseModel):
    client_id: str
    account_id: str
    amount: float
    currency: str  # IDR or USD
    notes: Optional[str] = None

class AdminActionTransferWalletToAccount(BaseModel):
    client_id: str
    from_wallet: str  # "main_idr", "main_usd", "withdrawal_idr", "withdrawal_usd"
    to_account_id: str
    amount: float
    currency: str  # IDR or USD
    notes: Optional[str] = None

class AdminActionRecord(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    action_type: str  # "topup_wallet", "withdraw_account", "transfer_wallet_to_account"
    client_id: str
    client_username: str
    admin_id: str  # Admin who created the request
    admin_username: str
    
    # Action Details
    wallet_type: Optional[str] = None
    account_id: Optional[str] = None
    from_wallet: Optional[str] = None
    to_account_id: Optional[str] = None
    amount: float
    currency: str = "IDR"
    notes: Optional[str] = None
    
    # Proof Files (GCS paths)
    payment_proof_gcs: Optional[str] = None  # For topup_wallet
    real_balance_proof_gcs: Optional[str] = None  # For withdraw_account
    spending_limit_proof_gcs: Optional[str] = None  # For transfer_wallet_to_account
    budget_aspire_proof_gcs: Optional[str] = None  # For transfer_wallet_to_account
    
    # Status & Approval
    status: str = "pending"  # pending, approved, rejected
    super_admin_id: Optional[str] = None
    super_admin_username: Optional[str] = None
    approval_notes: Optional[str] = None
    
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    processed_at: Optional[datetime] = None

class AdminActionApproval(BaseModel):
    action: str  # "approve" or "reject"
    notes: Optional[str] = None

class Token(BaseModel):
    access_token: str
    token_type: str

# Auth functions
def verify_password(plain_password, hashed_password):
    # Simple SHA-256 based password verification
    password_hash = hashlib.sha256(plain_password.encode()).hexdigest()
    return password_hash == hashed_password

def get_password_hash(password):
    # Simple SHA-256 based password hashing
    return hashlib.sha256(password.encode()).hexdigest()

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.now(timezone.utc) + expires_delta
    else:
        expire = datetime.now(timezone.utc) + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Could not validate credentials",
                headers={"WWW-Authenticate": "Bearer"},
            )
        user = await db.users.find_one({"username": username})
        if user is None:
            raise HTTPException(status_code=404, detail="User not found")
        return User(**parse_from_mongo(user))
    except jwt.PyJWTError:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Could not validate credentials",
            headers={"WWW-Authenticate": "Bearer"},
        )

async def get_current_admin(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        logger.info(f"[get_current_admin] Starting authentication check")
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        user_type: str = payload.get("user_type")
        
        logger.info(f"[get_current_admin] JWT decoded - username={username}, user_type={user_type}")
        
        if username is None or user_type != "admin":
            logger.error(f"[get_current_admin] Auth failed - username={username}, user_type={user_type}")
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Could not validate admin credentials",
                headers={"WWW-Authenticate": "Bearer"},
            )
        
        admin = await db.admin_users.find_one({"username": username})
        if admin is None:
            logger.error(f"[get_current_admin] Admin not found in database - username={username}")
            raise HTTPException(status_code=404, detail="Admin not found")
        
        logger.info(f"[get_current_admin] Auth successful for admin: {username}")
        return AdminUser(**parse_from_mongo(admin))
    except jwt.PyJWTError as e:
        logger.error(f"[get_current_admin] JWT decode error: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Could not validate admin credentials",
            headers={"WWW-Authenticate": "Bearer"},
        )
    except Exception as e:
        logger.error(f"[get_current_admin] Unexpected error: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Authentication error: {str(e)}"
        )

async def get_current_super_admin(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Verify that current admin is a super admin"""
    admin = await get_current_admin(credentials)
    
    # Check if admin is super admin
    admin_data = await db.admin_users.find_one({"id": admin.id})
    if not admin_data or not admin_data.get("is_super_admin", False):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Super admin access required"
        )
    
    return admin

def require_super_admin(current_admin: AdminUser = Depends(get_current_admin)):
    if not current_admin.is_super_admin:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Super admin access required"
        )
    return current_admin

# Test endpoint
@api_router.get("/")
async def root():
    return {"message": "Rimuru API is running"}

# Auth endpoints
@api_router.post("/auth/register", response_model=dict)
async def register(user: UserCreate):
    # Check if user exists
    existing_user = await db.users.find_one({"$or": [{"username": user.username}, {"email": user.email}]})
    if existing_user:
        raise HTTPException(status_code=400, detail="Username or email already registered")
    
    # Create user
    hashed_password = get_password_hash(user.password)
    new_user = User(
        username=user.username,
        email=user.email,
        password_hash=hashed_password,
        name=user.name,
        company_name=user.company_name,
        phone_number=user.phone_number,
        address=user.address,
        city=user.city,
        province=user.province
    )
    user_dict = prepare_for_mongo(new_user.dict())
    await db.users.insert_one(user_dict)
    
    # Create notification for admin
    await create_localized_notification(
        title_key="new_user_registration",
        message_key="user_registered",
        notification_type="user_registration",
        lang="id",  # Default to Indonesian
        reference_id=new_user.id,
        username=user.username
    )
    
    # Send welcome email to new client (async, don't wait for result)
    try:
        send_welcome_client_email(
            user_email=user.email,
            user_name=user.name or user.username,
            username=user.username
        )
        logger.info(f"📧 Welcome email sent to {user.email}")
    except Exception as e:
        logger.error(f"❌ Failed to send welcome email: {e}")
        # Don't fail registration if email fails
    
    # Send notification email to all active admins
    try:
        admin_emails = await get_active_admin_emails()
        if admin_emails:
            from email_service import send_admin_new_client_email
            send_admin_new_client_email(
                admin_emails=admin_emails,
                client_name=user.name or user.username,
                client_username=user.username,
                client_email=user.email
            )
            logger.info(f"📧 Admin notification emails sent to {len(admin_emails)} admins")
    except Exception as e:
        logger.error(f"❌ Failed to send admin notification emails: {e}")
    
    return {"message": "User registered successfully"}

@api_router.post("/auth/login", response_model=Token)
async def login(user: UserLogin):
    # Check maintenance mode
    maintenance_settings = await db.admin_settings.find_one({"setting_key": "maintenance_mode"})
    if maintenance_settings and maintenance_settings.get("enabled", False):
        message = maintenance_settings.get("message", "System sedang dalam maintenance.")
        raise HTTPException(
            status_code=503,
            detail=f"MAINTENANCE_MODE:{message}"
        )
    
    db_user = await db.users.find_one({"username": user.username})
    if not db_user or not verify_password(user.password, db_user["password_hash"]):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    # Client tokens expire in 7 days for better user experience
    access_token_expires = timedelta(minutes=CLIENT_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user.username, "user_type": "client"}, expires_delta=access_token_expires
    )
    return {"access_token": access_token, "token_type": "bearer"}

@api_router.get("/auth/me", response_model=dict)
async def get_me(current_user: User = Depends(get_current_user)):
    # Handle profile_picture path format for backward compatibility
    profile_picture = getattr(current_user, 'profile_picture', None)
    if profile_picture:
        # If it's a GCS path without /files/ prefix, add it
        if not profile_picture.startswith('/files/') and not profile_picture.startswith('http'):
            profile_picture = f"/files/{profile_picture}"
    
    return {
        "id": current_user.id,
        "username": current_user.username,
        "email": current_user.email,
        "name": current_user.name,
        "company_name": getattr(current_user, 'company_name', None),
        "phone_number": getattr(current_user, 'phone_number', None),
        "address": getattr(current_user, 'address', None),
        "city": getattr(current_user, 'city', None),
        "province": getattr(current_user, 'province', None),
        "profile_picture": profile_picture,
        "wallet_balance_idr": current_user.wallet_balance_idr,  # Legacy
        "wallet_balance_usd": current_user.wallet_balance_usd,  # Legacy
        "main_wallet_idr": getattr(current_user, 'main_wallet_idr', 0.0),
        "main_wallet_usd": getattr(current_user, 'main_wallet_usd', 0.0),
        "withdrawal_wallet_idr": getattr(current_user, 'withdrawal_wallet_idr', 0.0),
        "withdrawal_wallet_usd": getattr(current_user, 'withdrawal_wallet_usd', 0.0)
    }

# Admin Auth endpoints
@api_router.post("/admin/auth/login")
async def admin_login(admin: AdminUserLogin):
    db_admin = await db.admin_users.find_one({"username": admin.username})
    if not db_admin or not verify_password(admin.password, db_admin["password_hash"]):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    # Update last login
    await db.admin_users.update_one(
        {"id": db_admin["id"]},
        {"$set": {"last_login": datetime.now(timezone.utc)}}
    )
    
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": admin.username, "user_type": "admin"}, expires_delta=access_token_expires
    )
    return {
        "access_token": access_token, 
        "token_type": "bearer",
        "role": db_admin.get("role", "admin"),
        "username": db_admin.get("username"),
        "is_super_admin": db_admin.get("is_super_admin", False)
    }

@api_router.get("/admin/auth/me", response_model=AdminUserProfile)
async def get_admin_me(current_admin: AdminUser = Depends(get_current_admin)):
    return AdminUserProfile(
        id=current_admin.id,
        username=current_admin.username,
        email=current_admin.email,
        full_name=current_admin.full_name,
        whatsapp_number=current_admin.whatsapp_number,
        is_super_admin=current_admin.is_super_admin,
        profile_picture=current_admin.profile_picture,
        last_login=current_admin.last_login,
        created_at=current_admin.created_at
    )

# Profile endpoints
@api_router.get("/profile", response_model=UserProfile)
async def get_profile(current_user: User = Depends(get_current_user)):
    # Handle profile_picture path format for backward compatibility
    profile_picture = current_user.profile_picture
    if profile_picture:
        # If it's a GCS path without /files/ prefix, add it
        if not profile_picture.startswith('/files/') and not profile_picture.startswith('http'):
            profile_picture = f"/files/{profile_picture}"
    
    return UserProfile(
        username=current_user.username,
        email=current_user.email,
        display_name=current_user.display_name,
        phone_number=current_user.phone_number,
        address=current_user.address,
        company_name=current_user.company_name,
        profile_picture=profile_picture
    )

@api_router.put("/profile", response_model=dict)
async def update_profile(profile_data: UserProfileUpdate, current_user: User = Depends(get_current_user)):
    # Verify current password
    if not verify_password(profile_data.current_password, current_user.password_hash):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Current password is incorrect"
        )
    
    # Check if username/email already exists (if changed)
    if profile_data.username and profile_data.username != current_user.username:
        existing_user = await db.users.find_one({"username": profile_data.username})
        if existing_user:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Username already exists"
            )
    
    if profile_data.email and profile_data.email != current_user.email:
        existing_user = await db.users.find_one({"email": profile_data.email})
        if existing_user:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Email already exists"
            )
    
    # Prepare update data
    update_data = {
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Only update fields that are provided
    if profile_data.username:
        update_data["username"] = profile_data.username
    if profile_data.email:
        update_data["email"] = profile_data.email
    if profile_data.display_name is not None:
        update_data["display_name"] = profile_data.display_name
    if profile_data.phone_number is not None:
        update_data["phone_number"] = profile_data.phone_number
    if profile_data.address is not None:
        update_data["address"] = profile_data.address
    if profile_data.company_name is not None:
        update_data["company_name"] = profile_data.company_name
    
    # Update user in database
    result = await db.users.update_one(
        {"id": current_user.id},
        {"$set": update_data}
    )
    
    if result.modified_count == 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Failed to update profile"
        )
    
    return {"message": "Profile updated successfully"}

@api_router.put("/profile/password", response_model=dict)
async def change_password(password_data: PasswordChange, current_user: User = Depends(get_current_user)):
    # Verify current password
    if not verify_password(password_data.current_password, current_user.password_hash):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Current password is incorrect"
        )
    
    # Validate new password confirmation
    if password_data.new_password != password_data.confirm_password:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="New password and confirmation do not match"
        )
    
    # Validate password strength (minimum 6 characters)
    if len(password_data.new_password) < 6:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Password must be at least 6 characters long"
        )
    
    # Hash new password
    new_password_hash = get_password_hash(password_data.new_password)
    
    # Update password in database
    result = await db.users.update_one(
        {"id": current_user.id},
        {"$set": {
            "password_hash": new_password_hash,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    if result.modified_count == 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Failed to update password"
        )
    
    return {"message": "Password changed successfully"}

# Admin endpoints
@api_router.get("/admin/profile", response_model=AdminUserProfile)
async def get_admin_profile(current_admin: AdminUser = Depends(get_current_admin)):
    return AdminUserProfile(
        id=current_admin.id,
        username=current_admin.username,
        email=current_admin.email,
        full_name=current_admin.full_name,
        whatsapp_number=current_admin.whatsapp_number,
        is_super_admin=current_admin.is_super_admin,
        profile_picture=current_admin.profile_picture,
        last_login=current_admin.last_login,
        created_at=current_admin.created_at
    )

@api_router.put("/admin/profile", response_model=dict)
async def update_admin_profile(profile_data: AdminUserUpdate, current_admin: AdminUser = Depends(get_current_admin)):
    # Verify current password
    if not verify_password(profile_data.current_password, current_admin.password_hash):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Current password is incorrect"
        )
    
    # Build update data
    update_data = {}
    if profile_data.username and profile_data.username != current_admin.username:
        # Check if username already exists
        existing_admin = await db.admin_users.find_one({"username": profile_data.username, "id": {"$ne": current_admin.id}})
        if existing_admin:
            raise HTTPException(status_code=400, detail="Username already exists")
        update_data["username"] = profile_data.username
    
    if profile_data.email and profile_data.email != current_admin.email:
        # Check if email already exists
        existing_admin = await db.admin_users.find_one({"email": profile_data.email, "id": {"$ne": current_admin.id}})
        if existing_admin:
            raise HTTPException(status_code=400, detail="Email already exists")
        update_data["email"] = profile_data.email
    
    if profile_data.full_name:
        update_data["full_name"] = profile_data.full_name
    
    if profile_data.whatsapp_number is not None:
        update_data["whatsapp_number"] = profile_data.whatsapp_number
    
    if update_data:
        update_data["updated_at"] = datetime.now(timezone.utc)
        await db.admin_users.update_one({"id": current_admin.id}, {"$set": update_data})
    
    return {"message": "Profile updated successfully"}

@api_router.put("/admin/profile/password", response_model=dict)
async def change_admin_password(password_data: PasswordChange, current_admin: AdminUser = Depends(get_current_admin)):
    # Verify current password
    if not verify_password(password_data.current_password, current_admin.password_hash):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Current password is incorrect"
        )
    
    # Validate new password
    if password_data.new_password != password_data.confirm_password:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="New passwords do not match"
        )
    
    if len(password_data.new_password) < 6:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="New password must be at least 6 characters long"
        )
    
    # Update password
    new_password_hash = get_password_hash(password_data.new_password)
    await db.admin_users.update_one(
        {"id": current_admin.id},
        {"$set": {"password_hash": new_password_hash, "updated_at": datetime.now(timezone.utc)}}
    )
    
    return {"message": "Password changed successfully"}

# Admin Management - Super Admin Only
@api_router.get("/admin/admins", response_model=List[AdminUserProfile])
async def get_all_admins(current_admin: AdminUser = Depends(require_super_admin)):
    admins_cursor = db.admin_users.find()
    admins = await admins_cursor.to_list(length=None)
    return [AdminUserProfile(**parse_from_mongo(admin)) for admin in admins]

@api_router.post("/admin/admins", response_model=dict)
async def create_admin(admin_data: AdminUserCreate, current_admin: AdminUser = Depends(require_super_admin)):
    # Check if admin exists
    existing_admin = await db.admin_users.find_one({"$or": [{"username": admin_data.username}, {"email": admin_data.email}]})
    if existing_admin:
        raise HTTPException(status_code=400, detail="Admin with this username or email already exists")
    
    # Hash password
    password_hash = get_password_hash(admin_data.password)
    
    # Create admin
    admin_dict = admin_data.dict()
    admin_dict["password_hash"] = password_hash
    del admin_dict["password"]
    admin_dict["id"] = str(uuid.uuid4())
    admin_dict["created_at"] = datetime.now(timezone.utc)
    
    await db.admin_users.insert_one(prepare_for_mongo(admin_dict))
    
    # Send welcome email to new admin (async, don't wait for result)
    try:
        send_welcome_admin_email(
            admin_email=admin_data.email,
            admin_name=admin_data.full_name or admin_data.username,
            username=admin_data.username,
            is_super_admin=admin_data.is_super_admin
        )
        logger.info(f"📧 Welcome email sent to admin {admin_data.email}")
    except Exception as e:
        logger.error(f"❌ Failed to send admin welcome email: {e}")
        # Don't fail admin creation if email fails
    
    return {"message": "Admin created successfully"}

@api_router.put("/admin/admins/{admin_id}", response_model=dict)
async def update_admin(admin_id: str, admin_data: AdminUserEditRequest, current_admin: AdminUser = Depends(require_super_admin)):
    # Check if admin exists
    admin = await db.admin_users.find_one({"id": admin_id})
    if not admin:
        raise HTTPException(status_code=404, detail="Admin not found")
    
    # Check for duplicate username or email (excluding current admin)
    existing_admin = await db.admin_users.find_one({
        "$or": [{"username": admin_data.username}, {"email": admin_data.email}],
        "id": {"$ne": admin_id}
    })
    if existing_admin:
        raise HTTPException(status_code=400, detail="Admin with this username or email already exists")
    
    # Build update data
    update_data = {
        "username": admin_data.username,
        "email": admin_data.email,
        "full_name": admin_data.full_name,
        "whatsapp_number": admin_data.whatsapp_number,
        "is_super_admin": admin_data.is_super_admin,
        "updated_at": datetime.now(timezone.utc)
    }
    
    await db.admin_users.update_one({"id": admin_id}, {"$set": update_data})
    
    return {"message": "Admin updated successfully"}

@api_router.put("/admin/admins/{admin_id}/status", response_model=dict)
async def toggle_admin_status(admin_id: str, status_data: dict, current_admin: AdminUser = Depends(require_super_admin)):
    # Check if admin exists
    admin = await db.admin_users.find_one({"id": admin_id})
    if not admin:
        raise HTTPException(status_code=404, detail="Admin not found")
    
    # Prevent deactivating self
    if admin_id == current_admin.id:
        raise HTTPException(status_code=400, detail="Cannot change your own status")
    
    # Update status
    await db.admin_users.update_one(
        {"id": admin_id}, 
        {"$set": {"is_active": status_data.get("is_active", True), "updated_at": datetime.now(timezone.utc)}}
    )
    
    return {"message": "Admin status updated successfully"}

@api_router.delete("/admin/admins/{admin_id}", response_model=dict)
async def delete_admin(admin_id: str, current_admin: AdminUser = Depends(require_super_admin)):
    # Check if admin exists
    admin = await db.admin_users.find_one({"id": admin_id})
    if not admin:
        raise HTTPException(status_code=404, detail="Admin not found")
    
    # Prevent deleting self
    if admin_id == current_admin.id:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    
    # Delete admin
    await db.admin_users.delete_one({"id": admin_id})
    
    return {"message": "Admin deleted successfully"}

@api_router.put("/admin/change-password", response_model=dict)
async def change_admin_password_alt(password_data: PasswordChange, current_admin: AdminUser = Depends(get_current_admin)):
    # Verify current password
    if not verify_password(password_data.current_password, current_admin.password_hash):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Current password is incorrect"
        )
    
    # Update password
    new_password_hash = get_password_hash(password_data.new_password)
    await db.admin_users.update_one(
        {"id": current_admin.id},
        {"$set": {"password_hash": new_password_hash, "updated_at": datetime.now(timezone.utc)}}
    )
    
    return {"message": "Password changed successfully"}

# Client Management
@api_router.get("/admin/clients", response_model=List[dict])
async def get_all_clients(current_admin: AdminUser = Depends(get_current_admin)):
    # Optimized aggregation pipeline to reduce database queries
    # Instead of N+1 queries (1 for users + N for each user's data), 
    # we use a single aggregation pipeline
    pipeline = [
        # Sort by newest first
        {"$sort": {"created_at": -1}},
        
        # Lookup total requests count for each user
        {
            "$lookup": {
                "from": "ad_account_requests",
                "localField": "id",
                "foreignField": "user_id",
                "as": "requests"
            }
        },
        
        # Lookup completed topup transactions for each user
        {
            "$lookup": {
                "from": "transactions",
                "let": {"user_id": "$id"},
                "pipeline": [
                    {
                        "$match": {
                            "$expr": {
                                "$and": [
                                    {"$eq": ["$user_id", "$$user_id"]},
                                    {"$eq": ["$type", "topup"]},
                                    {"$eq": ["$status", "completed"]}
                                ]
                            }
                        }
                    }
                ],
                "as": "topup_transactions"
            }
        },
        
        # Lookup admin info if updated_by exists
        {
            "$lookup": {
                "from": "admin_users",
                "localField": "updated_by",
                "foreignField": "id",
                "as": "admin_info"
            }
        },
        
        # Add computed fields
        {
            "$addFields": {
                "total_requests": {"$size": "$requests"},
                "total_topup": {
                    "$sum": "$topup_transactions.amount"
                },
                "updated_by_admin": {
                    "$cond": {
                        "if": {"$gt": [{"$size": "$admin_info"}, 0]},
                        "then": {
                            "$let": {
                                "vars": {"admin": {"$arrayElemAt": ["$admin_info", 0]}},
                                "in": {
                                    "id": "$$admin.id",
                                    "username": "$$admin.username",
                                    "name": {
                                        "$ifNull": ["$$admin.name", "$$admin.username"]
                                    }
                                }
                            }
                        },
                        "else": None
                    }
                }
            }
        },
        
        # Remove temporary fields
        {
            "$project": {
                "requests": 0,
                "topup_transactions": 0,
                "admin_info": 0
            }
        }
    ]
    
    # Execute aggregation pipeline
    users_cursor = db.users.aggregate(pipeline)
    users = await users_cursor.to_list(length=None)
    
    result = []
    for user in users:
        user_data = parse_from_mongo(user)
        
        # Handle profile_picture path format for backward compatibility
        if user_data.get("profile_picture"):
            profile_picture = user_data["profile_picture"]
            # If it's a GCS path without /files/ prefix, add it
            if not profile_picture.startswith('/files/') and not profile_picture.startswith('http'):
                user_data["profile_picture"] = f"/files/{profile_picture}"
        
        # Ensure is_active field exists
        user_data["is_active"] = user_data.get("is_active", True)
        
        # Handle None case for updated_by_admin
        if user_data.get("updated_by_admin") is None:
            user_data.pop("updated_by_admin", None)
        
        result.append(user_data)
    
    return result

@api_router.get("/admin/clients/{client_id}", response_model=dict)
async def get_client_detail(
    client_id: str, 
    start_date: str = None,
    end_date: str = None,
    current_admin: AdminUser = Depends(get_current_admin)
):
    user = await db.users.find_one({"id": client_id})
    if not user:
        raise HTTPException(status_code=404, detail="Client not found")
    
    user_data = parse_from_mongo(user)
    
    # Handle profile_picture path format for backward compatibility
    if user_data.get("profile_picture"):
        profile_picture = user_data["profile_picture"]
        # If it's a GCS path without /api/files/ prefix, add it
        if not profile_picture.startswith('/api/files/') and not profile_picture.startswith('http'):
            # Also handle legacy /files/ format
            if profile_picture.startswith('/files/'):
                user_data["profile_picture"] = profile_picture.replace('/files/', '/api/files/', 1)
            else:
                user_data["profile_picture"] = f"/api/files/{profile_picture}"
    
    # Build date filter if provided
    date_filter = {}
    if start_date and end_date:
        try:
            # Handle URL encoding where + becomes space
            start_date_fixed = start_date.replace(' ', '+').replace('Z', '+00:00')
            end_date_fixed = end_date.replace(' ', '+').replace('Z', '+00:00')
            
            start = datetime.fromisoformat(start_date_fixed)
            end = datetime.fromisoformat(end_date_fixed)
            
            # Convert to ISO strings for comparison (dates stored as strings in DB)
            start_iso = start.isoformat()
            end_iso = end.isoformat()
            
            print(f"🔍 Date filter applied: {start_iso} to {end_iso}")
            
            date_filter = {
                "created_at": {
                    "$gte": start_iso,
                    "$lte": end_iso
                }
            }
        except Exception as e:
            print(f"❌ Date parsing error: {e}")
            pass  # If date parsing fails, don't apply filter
    
    # Build query filters
    request_query = {"user_id": client_id}
    transaction_query = {"user_id": client_id}
    
    # Apply date filter if provided
    if date_filter:
        request_query.update(date_filter)
        transaction_query.update(date_filter)
        print(f"🔍 Request query: {request_query}")
        print(f"🔍 Transaction query: {transaction_query}")
    
    # Get requests
    requests_cursor = db.ad_account_requests.find(request_query)
    requests = await requests_cursor.to_list(length=None)
    print(f"🔍 Found {len(requests)} requests for client {client_id}")
    user_data["requests"] = [parse_from_mongo(req) for req in requests]
    
    # Get ad accounts
    accounts_cursor = db.ad_accounts.find({"user_id": client_id})
    accounts = await accounts_cursor.to_list(length=None)
    user_data["accounts"] = [parse_from_mongo(acc) for acc in accounts]
    
    # Get transactions with proof paths
    transactions_cursor = db.transactions.find(transaction_query)
    transactions = await transactions_cursor.to_list(length=None)
    print(f"🔍 Found {len(transactions)} transactions for client {client_id}")
    
    # Enrich transactions with proof paths
    enriched_transactions = []
    for trans in transactions:
        trans_data = parse_from_mongo(trans)
        
        # Add proof path if reference_id exists
        if trans_data.get('reference_id'):
            ref_id = trans_data['reference_id']
            trans_type = trans_data.get('type')
            
            # Determine proof collection and construct path
            proof_path = None
            
            if trans_type in ['wallet_topup']:
                # Check wallet_payment_proofs collection
                proof = await db.wallet_payment_proofs.find_one({'topup_id': ref_id})
                if proof and proof.get('gcs_path'):
                    proof_path = proof['gcs_path']
            elif trans_type in ['topup', 'account_topup']:
                # Check payment_proofs collection
                proof = await db.payment_proofs.find_one({'topup_id': ref_id})
                if proof and proof.get('gcs_path'):
                    proof_path = proof['gcs_path']
            elif trans_type == 'wallet_to_account_transfer':
                # Check wallet_payment_proofs for transfer
                proof = await db.wallet_payment_proofs.find_one({'topup_id': ref_id})
                if proof and proof.get('gcs_path'):
                    proof_path = proof['gcs_path']
            
            if proof_path:
                trans_data['proof_path'] = proof_path
        
        enriched_transactions.append(trans_data)
    
    user_data["transactions"] = enriched_transactions
    
    # Calculate Total Top Up (same logic as Financial Reports, but for this client only)
    # Build match filters for top-up calculations
    wallet_topup_idr_match = {"user_id": client_id, "status": {"$in": ["verified", "approved", "completed"]}, "currency": "IDR"}
    wallet_topup_usd_match = {"user_id": client_id, "status": {"$in": ["verified", "approved", "completed"]}, "currency": "USD"}
    account_topup_idr_match = {"user_id": client_id, "status": {"$in": ["verified", "approved", "completed"]}, "currency": "IDR"}
    account_topup_usd_match = {"user_id": client_id, "status": {"$in": ["verified", "approved", "completed"]}, "currency": "USD"}
    
    # Apply date filter to top-up calculations if provided
    if date_filter:
        wallet_topup_idr_match.update(date_filter)
        wallet_topup_usd_match.update(date_filter)
        account_topup_idr_match.update(date_filter)
        account_topup_usd_match.update(date_filter)
        print(f"🔍 Applied date filter to top-up calculations")
    
    # 1. Wallet Top-Up (verified/approved/completed)
    wallet_topup_idr_pipeline = [
        {"$match": wallet_topup_idr_match},
        {"$group": {"_id": None, "total_amount": {"$sum": "$amount"}}}
    ]
    wallet_topup_idr = await db.wallet_topup_requests.aggregate(wallet_topup_idr_pipeline).to_list(1)
    wallet_topup_idr_amount = wallet_topup_idr[0]["total_amount"] if wallet_topup_idr else 0
    
    wallet_topup_usd_pipeline = [
        {"$match": wallet_topup_usd_match},
        {"$group": {"_id": None, "total_amount": {"$sum": "$amount"}}}
    ]
    wallet_topup_usd = await db.wallet_topup_requests.aggregate(wallet_topup_usd_pipeline).to_list(1)
    wallet_topup_usd_amount = wallet_topup_usd[0]["total_amount"] if wallet_topup_usd else 0
    
    # 2. Ad Account Top-Up (verified/approved/completed)
    account_topup_idr_pipeline = [
        {"$match": account_topup_idr_match},
        {"$group": {"_id": None, "total_amount": {"$sum": "$total_amount"}}}
    ]
    account_topup_idr = await db.topup_requests.aggregate(account_topup_idr_pipeline).to_list(1)
    account_topup_idr_amount = account_topup_idr[0]["total_amount"] if account_topup_idr else 0
    
    account_topup_usd_pipeline = [
        {"$match": account_topup_usd_match},
        {"$group": {"_id": None, "total_amount": {"$sum": "$total_amount"}}}
    ]
    account_topup_usd = await db.topup_requests.aggregate(account_topup_usd_pipeline).to_list(1)
    account_topup_usd_amount = account_topup_usd[0]["total_amount"] if account_topup_usd else 0
    
    # Calculate Total Top Up
    user_data["total_topup_idr"] = wallet_topup_idr_amount + account_topup_idr_amount
    user_data["total_topup_usd"] = wallet_topup_usd_amount + account_topup_usd_amount
    
    print(f"🔍 Total Top Up IDR: {user_data['total_topup_idr']} (Wallet: {wallet_topup_idr_amount}, Account: {account_topup_idr_amount})")
    print(f"🔍 Total Top Up USD: {user_data['total_topup_usd']} (Wallet: {wallet_topup_usd_amount}, Account: {account_topup_usd_amount})")
    
    return user_data

@api_router.put("/admin/clients/{client_id}/status", response_model=dict)
async def update_client_status(client_id: str, status_data: ClientStatusUpdate, current_admin: AdminUser = Depends(get_current_admin)):
    user = await db.users.find_one({"id": client_id})
    if not user:
        raise HTTPException(status_code=404, detail="Client not found")
    
    await db.users.update_one(
        {"id": client_id},
        {"$set": {"is_active": status_data.is_active, "updated_at": datetime.now(timezone.utc)}}
    )
    
    return {"message": "Client status updated successfully"}

@api_router.put("/admin/clients/{client_id}", response_model=dict)
async def update_client(client_id: str, client_data: dict, current_admin: AdminUser = Depends(get_current_admin)):
    user = await db.users.find_one({"id": client_id})
    if not user:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Validate email uniqueness if email is being updated
    if "email" in client_data and client_data["email"] != user.get("email"):
        existing_user = await db.users.find_one({"email": client_data["email"], "id": {"$ne": client_id}})
        if existing_user:
            raise HTTPException(status_code=400, detail="Email already exists")
    
    # Build update data
    update_data = {}
    allowed_fields = ["display_name", "email", "phone_number", "company_name", "address"]
    
    for field in allowed_fields:
        if field in client_data:
            update_data[field] = client_data[field]
    
    if update_data:
        update_data["updated_at"] = datetime.now(timezone.utc)
        update_data["updated_by"] = current_admin.id
        await db.users.update_one({"id": client_id}, {"$set": update_data})
    
    return {"message": "Client updated successfully"}

@api_router.post("/admin/clients/{client_id}/reset-password", response_model=dict)
async def reset_client_password(client_id: str, current_admin: AdminUser = Depends(get_current_admin)):
    user = await db.users.find_one({"id": client_id})
    if not user:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Generate new temporary password
    import secrets
    import string
    
    new_password = ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(12))
    hashed_password = get_password_hash(new_password)
    
    # Update user password
    await db.users.update_one(
        {"id": client_id},
        {"$set": {
            "password_hash": hashed_password,
            "updated_at": datetime.now(timezone.utc),
            "password_reset_required": True  # Force user to change password on next login
        }}
    )
    
    # Create admin notification
    await create_localized_notification(
        title_key="password_reset",
        message_key="password_reset_requested",
        notification_type="password_reset",
        lang="id",
        reference_id=user["id"],
        username=user["username"]
    )
    
    return {
        "message": "Password reset successfully", 
        "new_password": new_password,
        "username": user["username"]
    }

# Request Management
@api_router.get("/admin/requests", response_model=List[dict])
async def get_all_requests(
    status: Optional[str] = None,
    platform: Optional[str] = None,
    current_admin: AdminUser = Depends(get_current_admin)
):
    filter_query = {}
    if status:
        filter_query["status"] = status
    if platform:
        filter_query["platform"] = platform
    
    requests_cursor = db.ad_account_requests.find(filter_query).sort("created_at", -1)
    requests = await requests_cursor.to_list(length=None)
    
    result = []
    for request in requests:
        request_data = parse_from_mongo(request)
        # Get user info
        user = await db.users.find_one({"id": request_data["user_id"]})
        if user:
            request_data["user"] = {"username": user["username"], "email": user["email"]}
        
        # Get admin info if processed
        if request_data.get("verified_by"):
            admin = await db.admin_users.find_one({"id": request_data["verified_by"]})
            if admin:
                admin = parse_from_mongo(admin)
                request_data["verified_by_admin"] = {
                    "id": admin.get("id"),
                    "username": admin.get("username"),
                    "name": admin.get("name", admin.get("username"))
                }
        
        result.append(request_data)
    
    return result

@api_router.put("/admin/requests/{request_id}/status", response_model=dict)
async def update_request_status(
    request_id: str, 
    status_data: RequestStatusUpdate, 
    current_admin: AdminUser = Depends(get_current_admin)
):
    request = await db.ad_account_requests.find_one({"id": request_id})
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    
    # Validate Facebook Ads approval requires account_id
    if (status_data.status == "approved" and 
        request.get("platform") == "facebook" and 
        not status_data.account_id):
        raise HTTPException(
            status_code=400, 
            detail="Account ID is required for Facebook Ads approval"
        )
    
    # Validate approval requires fee_percentage
    if (status_data.status == "approved" and 
        status_data.fee_percentage is None):
        raise HTTPException(
            status_code=400, 
            detail="Fee percentage is required for approval"
        )
    
    # Validate fee_percentage range
    if (status_data.fee_percentage is not None and 
        (status_data.fee_percentage < 0 or status_data.fee_percentage > 100)):
        raise HTTPException(
            status_code=400, 
            detail="Fee percentage must be between 0 and 100"
        )
    
    update_data = {
        "status": status_data.status,
        "verified_by": current_admin.id,
        "processed_at": datetime.now(timezone.utc).isoformat()
    }
    
    if status_data.admin_notes:
        update_data["admin_notes"] = status_data.admin_notes
        
    if status_data.account_id:
        update_data["account_id"] = status_data.account_id
    
    if status_data.account_name:
        update_data["account_name"] = status_data.account_name
        
    if status_data.fee_percentage is not None:
        update_data["fee_percentage"] = status_data.fee_percentage
    
    # If request is approved, create actual ad account (only if not already created)
    if status_data.status == "approved":
        # Check if account already exists for this request
        existing_account = None
        if "account_id" in request and request["account_id"]:
            existing_account = await db.ad_accounts.find_one({
                "user_id": request["user_id"],
                "account_id": request["account_id"]
            })
        
        # Only create new account if one doesn't exist
        if not existing_account:
            # Validate account_id is unique (anti-duplicate)
            if status_data.account_id:
                duplicate_check = await db.ad_accounts.find_one({"account_id": status_data.account_id})
                if duplicate_check:
                    raise HTTPException(
                        status_code=400, 
                        detail=f"Account ID '{status_data.account_id}' already exists. Please use a unique Account ID."
                    )
            
            ad_account = AdAccount(
                user_id=request["user_id"],
                platform=request["platform"],
                account_name=status_data.account_name if status_data.account_name else request["account_name"],
                account_id=status_data.account_id if status_data.account_id else f"{request['platform']}_{str(uuid.uuid4())[:8]}",
                status="sharing",  # Start with sharing status instead of active
                fee_percentage=status_data.fee_percentage,
                group_id=request.get("group_id"),  # Copy group_id from request
                # Facebook-specific fields
                gmt=request.get("gmt"),
                currency=request.get("currency"),
                delivery_method=request.get("delivery_method"),
                bm_id_or_email=request.get("bm_id_or_email"),
                # Google Ads specific fields
                email=request.get("email"),
                website=request.get("website"),
                # TikTok Ads specific fields
                bc_id=request.get("bc_id"),
                # Common fields
                notes=request.get("notes")
            )
            
            account_dict = prepare_for_mongo(ad_account.dict())
            await db.ad_accounts.insert_one(account_dict)
            
            # Store the account_id in the request record
            update_data["account_id"] = ad_account.account_id
    
    # Update existing ad account status based on request status (if account exists)
    if "account_id" in request and request["account_id"]:
        ad_account_status = None
        
        # Map request status to ad account status with new sharing flow
        if status_data.status == "completed":
            ad_account_status = "active"  # Completed = active (final status)
        elif status_data.status == "approved":
            ad_account_status = "sharing"  # Approved = sharing (intermediate status)
        elif status_data.status == "disabled":
            ad_account_status = "disabled"
        elif status_data.status == "pending":
            ad_account_status = "pending"
        elif status_data.status == "processing":
            ad_account_status = "processing" 
        elif status_data.status in ["rejected", "failed"]:
            ad_account_status = "suspended"
        
        if ad_account_status:
            try:
                await db.ad_accounts.update_one(
                    {
                        "user_id": request["user_id"],
                        "account_id": request["account_id"]
                    },
                    {"$set": {"status": ad_account_status}}
                )
                logger.info(f"Updated ad account {request['account_id']} status to {ad_account_status}")
            except Exception as e:
                logger.error(f"Failed to update ad account status: {str(e)}")
    
    await db.ad_account_requests.update_one(
        {"id": request_id},
        {"$set": update_data}
    )
    
    # Create notification for client when request is approved
    if status_data.status == "approved":
        platform_name = request["platform"].title()
        if request["platform"] == "facebook":
            platform_name = "Facebook Ads"
        elif request["platform"] == "google":
            platform_name = "Google Ads"
        elif request["platform"] == "tiktok":
            platform_name = "TikTok Ads"
        
        # Use updated account name (from admin) if provided, otherwise use original
        final_account_name = status_data.account_name if status_data.account_name else request['account_name']
        
        # Create client notification
        client_notification = {
            "id": str(uuid.uuid4()),
            "user_id": request["user_id"], 
            "title": get_notification_text("account_request_approved", "id", platform=platform_name),
            "message": get_notification_text("account_approved_message", "id", platform=platform_name, account_name=final_account_name),
            "type": "approval",
            "reference_id": request_id,
            "is_read": False,
            "created_at": datetime.now(timezone.utc)
        }
        
        client_notification_dict = prepare_for_mongo(client_notification)
        await db.client_notifications.insert_one(client_notification_dict)
        
        # Send email notification to client about account request approval
        try:
            user = await db.users.find_one({"id": request["user_id"]})
            if user and user.get("email"):
                send_client_account_request_approved_email(
                    client_email=user["email"],
                    client_name=user.get("full_name") or user.get("username"),
                    platform=request["platform"],
                    account_name=final_account_name
                )
                logger.info(f"📧 Account request approval email sent to {user['email']}")
        except Exception as e:
            logger.error(f"Failed to send account request approval email: {e}")
    
    # Create notification for client when request is processing
    if status_data.status == "processing":
        platform_name = request["platform"].title()
        if request["platform"] == "facebook":
            platform_name = "Facebook Ads"
        elif request["platform"] == "google":
            platform_name = "Google Ads"
        elif request["platform"] == "tiktok":
            platform_name = "TikTok Ads"
        
        # Create client notification for processing status
        client_notification = {
            "id": str(uuid.uuid4()),
            "user_id": request["user_id"], 
            "title": f"📋 {platform_name} - Status Diperbarui",
            "message": f"Permintaan akun {platform_name} Anda '{request['account_name']}' sedang diproses oleh tim admin.",
            "type": "status_update",
            "reference_id": request_id,
            "is_read": False,
            "created_at": datetime.now(timezone.utc)
        }
        
        client_notification_dict = prepare_for_mongo(client_notification)
        await db.client_notifications.insert_one(client_notification_dict)

    # Create notification for client when request is completed (sharing -> active)
    if status_data.status == "completed":
        platform_name = request["platform"].title()
        if request["platform"] == "facebook":
            platform_name = "Facebook Ads"
        elif request["platform"] == "google":
            platform_name = "Google Ads"
        elif request["platform"] == "tiktok":
            platform_name = "TikTok Ads"
        
        # Create client notification for completion
        client_notification = {
            "id": str(uuid.uuid4()),
            "user_id": request["user_id"], 
            "title": get_notification_text("account_request_completed", "id", platform=platform_name),
            "message": get_notification_text("account_completed_message", "id", platform=platform_name, account_name=request['account_name']),
            "type": "account_completed",
            "reference_id": request_id,
            "is_read": False,
            "created_at": datetime.now(timezone.utc)
        }
        
        client_notification_dict = prepare_for_mongo(client_notification)
        await db.client_notifications.insert_one(client_notification_dict)
        
        # Send email notification to client about account request completion
        try:
            user = await db.users.find_one({"id": request["user_id"]})
            if user and user.get("email"):
                send_client_account_request_completed_email(
                    client_email=user["email"],
                    client_name=user.get("full_name") or user.get("username"),
                    platform=request["platform"],
                    account_name=request["account_name"]
                )
                logger.info(f"📧 Account request completed email sent to {user['email']}")
        except Exception as e:
            logger.error(f"Failed to send account request completed email: {e}")

    # Create notification for client when request is rejected
    if status_data.status in ["rejected", "failed"]:
        platform_name = request["platform"].title()
        if request["platform"] == "facebook":
            platform_name = "Facebook Ads"
        elif request["platform"] == "google":
            platform_name = "Google Ads"
        elif request["platform"] == "tiktok":
            platform_name = "TikTok Ads"
        
        # Create client notification for rejection
        client_notification = {
            "id": str(uuid.uuid4()),
            "user_id": request["user_id"], 
            "title": f"❌ {platform_name} - Status Diperbarui",
            "message": f"Permintaan akun {platform_name} Anda '{request['account_name']}' telah ditolak. {status_data.admin_notes or 'Silakan hubungi admin untuk info lebih lanjut.'}",
            "type": "rejection",
            "reference_id": request_id,
            "is_read": False,
            "created_at": datetime.now(timezone.utc)
        }
        
        client_notification_dict = prepare_for_mongo(client_notification)
        await db.client_notifications.insert_one(client_notification_dict)
        
        # Send email notification to client about account request rejection
        try:
            user = await db.users.find_one({"id": request["user_id"]})
            if user and user.get("email"):
                send_client_account_request_rejected_email(
                    client_email=user["email"],
                    client_name=user.get("full_name") or user.get("username"),
                    platform=request["platform"],
                    account_name=request["account_name"],
                    reason=status_data.admin_notes or ""
                )
                logger.info(f"📧 Account request rejection email sent to {user['email']}")
        except Exception as e:
            logger.error(f"Failed to send account request rejection email: {e}")
    
    # Update transaction status when request status changes - CONSISTENT mapping per user requirements
    if status_data.status == "completed":
        transaction_status = "completed"
    elif status_data.status in ["rejected", "failed"]:
        transaction_status = "failed"
    elif status_data.status == "pending":
        transaction_status = "pending"
    elif status_data.status == "approved":
        transaction_status = "sharing"  # CRITICAL FIX: approved request = sharing account = sharing transaction
    elif status_data.status == "processing":
        transaction_status = "processing"
    elif status_data.status == "disabled":
        transaction_status = "disabled"
    else:  # fallback
        transaction_status = "pending"
    
    # Use updated account name (from admin) if provided, otherwise use original
    final_account_name = status_data.account_name if status_data.account_name else request['account_name']
    
    # Build the correct description pattern based on platform - use ORIGINAL name to find transaction
    if request['platform'] == "facebook":
        # Match both simple pattern and pattern with GMT info
        old_description_pattern = f"Request Facebook ads account: {request['account_name']}"
        new_description = f"Request Facebook ads account: {final_account_name} (GMT: {request.get('gmt', 'N/A')}, Currency: {request.get('currency', 'N/A')})"
    else:
        old_description_pattern = f"Request {request['platform']} ads account: {request['account_name']}"
        new_description = f"Request {request['platform']} ads account: {final_account_name}"
    
    # Update transaction with new status AND new description - use more flexible regex
    result = await db.transactions.update_one(
        {
            "user_id": request["user_id"],
            "type": "account_request", 
            "description": {"$regex": f"Request.*account: {request['account_name']}"}
        },
        {
            "$set": {
                "status": transaction_status,
                "description": new_description  # Update description with new account name
            }
        }
    )
    
    logger.info(f"Updated transaction for user {request['user_id']}: status={transaction_status}, new_name={final_account_name}, modified={result.modified_count}")
    
    return {"message": "Request status updated successfully"}

# Bulk update requests endpoint  
@api_router.put("/admin/requests/bulk-update", response_model=dict)
async def bulk_update_request_status(
    bulk_data: BulkRequestUpdate,
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Bulk update multiple request statuses"""
    if not bulk_data.request_ids:
        raise HTTPException(status_code=400, detail="No request IDs provided")
    
    # Validate approval requirements
    if bulk_data.status == "approved" and bulk_data.fee_percentage is None:
        raise HTTPException(
            status_code=400,
            detail="Fee percentage is required for bulk approval"
        )
    
    successful_updates = 0
    failed_updates = []
    
    for request_id in bulk_data.request_ids:
        try:
            request = await db.ad_account_requests.find_one({"id": request_id})
            if not request:
                failed_updates.append(f"Request {request_id} not found")
                continue
                
            # Validate Facebook Ads approval requirements
            if (bulk_data.status == "approved" and 
                request.get("platform") == "facebook"):
                # For Facebook, we'll skip if no account_id is in the original request
                # Admin should handle Facebook accounts individually
                failed_updates.append(f"Request {request_id}: Facebook Ads requires individual processing with Account ID")
                continue
            
            update_data = {
                "status": bulk_data.status,
                "admin_id": current_admin.id,
                "processed_at": datetime.now(timezone.utc)
            }
            
            if bulk_data.admin_notes:
                update_data["admin_notes"] = bulk_data.admin_notes
                
            if bulk_data.fee_percentage is not None:
                update_data["fee_percentage"] = bulk_data.fee_percentage
            
            # If request is approved, create actual ad account (only if not already created)
            if bulk_data.status == "approved":
                # Check if account already exists for this request
                existing_account = None
                if "account_id" in request and request["account_id"]:
                    existing_account = await db.ad_accounts.find_one({
                        "user_id": request["user_id"],
                        "account_id": request["account_id"]
                    })
                
                # Only create new account if one doesn't exist
                if not existing_account:
                    ad_account = AdAccount(
                        user_id=request["user_id"],
                        platform=request["platform"],
                        account_name=request["account_name"],
                        account_id=f"{request['platform']}_{str(uuid.uuid4())[:8]}",
                        status="active",
                        fee_percentage=bulk_data.fee_percentage,
                        group_id=request.get("group_id"),  # Copy group_id from request
                        # Platform-specific fields
                        gmt=request.get("gmt"),
                        currency=request.get("currency"),
                        delivery_method=request.get("delivery_method"),
                        bm_id_or_email=request.get("bm_id_or_email"),
                        email=request.get("email"),
                        website=request.get("website"),
                        bc_id=request.get("bc_id"),
                        notes=request.get("notes")
                    )
                    
                    account_dict = prepare_for_mongo(ad_account.dict())
                    await db.ad_accounts.insert_one(account_dict)
                    update_data["account_id"] = ad_account.account_id
            
            # Update existing ad account status based on request status (if account exists)
            if "account_id" in request and request["account_id"]:
                ad_account_status = None
                
                # Map request status to ad account status - each status should match exactly per user requirements
                if bulk_data.status == "completed":
                    ad_account_status = "active"  # Completed requests = active accounts (special case)
                elif bulk_data.status == "disabled":
                    ad_account_status = "disabled"
                elif bulk_data.status == "pending":
                    ad_account_status = "pending"  # Pending should stay pending
                elif bulk_data.status == "approved":
                    ad_account_status = "approved"  # Approved should stay approved
                elif bulk_data.status == "processing":
                    ad_account_status = "processing"  # Processing should stay processing
                elif bulk_data.status in ["rejected", "failed"]:
                    ad_account_status = "suspended"  # Failed/rejected = suspended
                
                if ad_account_status:
                    try:
                        await db.ad_accounts.update_one(
                            {
                                "user_id": request["user_id"],
                                "account_id": request["account_id"]
                            },
                            {"$set": {"status": ad_account_status}}
                        )
                        logger.info(f"Bulk update: Updated ad account {request['account_id']} status to {ad_account_status}")
                    except Exception as e:
                        logger.error(f"Bulk update: Failed to update ad account status: {str(e)}")
            
            await db.ad_account_requests.update_one(
                {"id": request_id},
                {"$set": update_data}
            )
            
            # Create notification for client when request is approved
            if bulk_data.status == "approved":
                platform_name = request["platform"].title()
                if request["platform"] == "facebook":
                    platform_name = "Facebook Ads"
                elif request["platform"] == "google":
                    platform_name = "Google Ads"
                elif request["platform"] == "tiktok":
                    platform_name = "TikTok Ads"
                
                # Use account name from request
                final_account_name = request['account_name']
                
                client_notification = {
                    "id": str(uuid.uuid4()),
                    "user_id": request["user_id"], 
                    "title": get_notification_text("account_request_approved", "id", platform=platform_name),
                    "message": get_notification_text("account_approved_message", "id", platform=platform_name, account_name=final_account_name),
                    "type": "approval",
                    "reference_id": request_id,
                    "is_read": False,
                    "created_at": datetime.now(timezone.utc)
                }
                
                client_notification_dict = prepare_for_mongo(client_notification)
                await db.client_notifications.insert_one(client_notification_dict)
                
                # Send email notification to client about account request approval
                try:
                    user = await db.users.find_one({"id": request["user_id"]})
                    if user and user.get("email"):
                        send_client_account_request_approved_email(
                            client_email=user["email"],
                            client_name=user.get("full_name") or user.get("username"),
                            platform=request["platform"],
                            account_name=final_account_name
                        )
                        logger.info(f"📧 Bulk update: Account request approval email sent to {user['email']}")
                except Exception as e:
                    logger.error(f"Bulk update: Failed to send account request approval email: {e}")
            
            # Create notification for client when request is completed
            elif bulk_data.status == "completed":
                platform_name = request["platform"].title()
                if request["platform"] == "facebook":
                    platform_name = "Facebook Ads"
                elif request["platform"] == "google":
                    platform_name = "Google Ads"
                elif request["platform"] == "tiktok":
                    platform_name = "TikTok Ads"
                
                # Use account name from request
                final_account_name = request['account_name']
                
                client_notification = {
                    "id": str(uuid.uuid4()),
                    "user_id": request["user_id"], 
                    "title": get_notification_text("account_request_completed", "id", platform=platform_name),
                    "message": get_notification_text("account_completed_message", "id", platform=platform_name, account_name=final_account_name),
                    "type": "account_completed",
                    "reference_id": request_id,
                    "is_read": False,
                    "created_at": datetime.now(timezone.utc)
                }
                
                client_notification_dict = prepare_for_mongo(client_notification)
                await db.client_notifications.insert_one(client_notification_dict)
                
                # Send email notification to client about account request completion
                try:
                    user = await db.users.find_one({"id": request["user_id"]})
                    if user and user.get("email"):
                        send_client_account_request_completed_email(
                            client_email=user["email"],
                            client_name=user.get("full_name") or user.get("username"),
                            platform=request["platform"],
                            account_name=final_account_name
                        )
                        logger.info(f"📧 Bulk update: Account request completion email sent to {user['email']}")
                except Exception as e:
                    logger.error(f"Bulk update: Failed to send account request completion email: {e}")
            
            # Create notification for client when request is processing
            elif bulk_data.status == "processing":
                platform_name = request["platform"].title()
                if request["platform"] == "facebook":
                    platform_name = "Facebook Ads"
                elif request["platform"] == "google":
                    platform_name = "Google Ads"
                elif request["platform"] == "tiktok":
                    platform_name = "TikTok Ads"
                
                client_notification = {
                    "id": str(uuid.uuid4()),
                    "user_id": request["user_id"], 
                    "title": f"📋 {platform_name} - Status Diperbarui",
                    "message": f"Permintaan akun {platform_name} Anda '{request['account_name']}' sedang diproses oleh tim admin.",
                    "type": "status_update",
                    "reference_id": request_id,
                    "is_read": False,
                    "created_at": datetime.now(timezone.utc)
                }
                
                client_notification_dict = prepare_for_mongo(client_notification)
                await db.client_notifications.insert_one(client_notification_dict)
            
            # Update transaction status when request status changes - CONSISTENT mapping per user requirements
            if bulk_data.status == "completed":
                transaction_status = "completed"
            elif bulk_data.status in ["rejected", "failed"]:
                transaction_status = "failed"
            elif bulk_data.status == "pending":
                transaction_status = "pending"
            elif bulk_data.status == "approved":
                transaction_status = "sharing"  # CRITICAL FIX: approved request = sharing account = sharing transaction
            elif bulk_data.status == "processing":
                transaction_status = "processing"
            elif bulk_data.status == "disabled":
                transaction_status = "disabled"
            else:  # fallback
                transaction_status = "pending"
            
            # Use account name from request
            final_account_name = request['account_name']
            
            # Build the correct description pattern based on platform
            if request['platform'] == "facebook":
                new_description = f"Request Facebook ads account: {final_account_name} (GMT: {request.get('gmt', 'N/A')}, Currency: {request.get('currency', 'N/A')})"
            else:
                new_description = f"Request {request['platform']} ads account: {final_account_name}"
            
            result = await db.transactions.update_one(
                {
                    "user_id": request["user_id"],
                    "type": "account_request", 
                    "description": {"$regex": f"Request.*account: {request['account_name']}"}
                },
                {
                    "$set": {
                        "status": transaction_status,
                        "description": new_description  # Update description with new account name
                    }
                }
            )
            
            logger.info(f"Bulk update - Updated transaction for user {request['user_id']}: status={transaction_status}, new_name={final_account_name}, modified={result.modified_count}")
            
            successful_updates += 1
            
        except Exception as e:
            failed_updates.append(f"Request {request_id}: {str(e)}")
    
    return {
        "message": f"Bulk update completed. {successful_updates} successful, {len(failed_updates)} failed.",
        "successful_updates": successful_updates,
        "failed_updates": failed_updates
    }

@api_router.post("/admin/payments/{payment_id}/upload-verification-files")
async def upload_verification_files(
    payment_id: str,
    file: UploadFile = File(...),
    type: str = Form(...)  # "spend_limit_proof" or "budget_aspire_proof"
):
    """Upload verification files for payment approval - NO AUTH for stability"""
    try:
        # Validate file type
        if file.content_type not in ['image/jpeg', 'image/png', 'image/jpg', 'application/pdf']:
            raise HTTPException(status_code=400, detail="Invalid file type. Only JPEG, PNG, and PDF files are allowed")
        
        # Create directories if they don't exist - use absolute path
        upload_dir = Path("/app/uploads/verification_files")
        upload_dir.mkdir(parents=True, exist_ok=True)
        
        # Generate unique filename
        file_extension = Path(file.filename).suffix
        unique_filename = f"{payment_id}_{type}_{uuid.uuid4().hex[:8]}{file_extension}"
        file_path = upload_dir / unique_filename
        
        # Save file
        with open(file_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)
        
        # Return file path for storage in database - use relative path for consistency
        relative_path = f"uploads/verification_files/{unique_filename}"
        return {
            "success": True,
            "file_path": relative_path,
            "original_filename": file.filename,
            "type": type
        }
        
    except Exception as e:
        logger.error(f"Error uploading verification file: {e}")
        raise HTTPException(status_code=500, detail="Failed to upload verification file")

@api_router.post("/admin/payments/{payment_id}/upload-account-proof")
async def upload_account_proof(
    payment_id: str,
    account_id: str = Form(...),
    file: UploadFile = File(...),
    proof_type: str = Form(...)  # "spend_limit_proof" or "budget_aspire_proof"
):
    """Upload verification proof for specific account within a top-up request to GCS"""
    try:
        # Get the top-up request
        topup_request = await db.topup_requests.find_one({"id": payment_id})
        if not topup_request:
            raise HTTPException(status_code=404, detail="Top-up request not found")
        
        # Validate file type
        if file.content_type not in ['image/jpeg', 'image/png', 'image/jpg', 'application/pdf']:
            raise HTTPException(status_code=400, detail="Invalid file type. Only JPEG, PNG, and PDF files are allowed")
        
        # Generate unique filename
        file_extension = Path(file.filename).suffix
        unique_filename = f"{payment_id}_{account_id}_{proof_type}_{uuid.uuid4().hex[:12]}{file_extension}"
        
        # Upload to GCS
        gcs_path = f"account_topup_proofs/{unique_filename}"
        gcs = get_gcs_storage()
        
        try:
            # Read file content
            file_content = await file.read()
            file_obj = io.BytesIO(file_content)
            
            # Upload to GCS
            gcs.upload_file(
                file_obj=file_obj,
                destination_path=gcs_path,
                content_type=file.content_type,
                metadata={
                    "payment_id": payment_id,
                    "account_id": account_id,
                    "proof_type": proof_type,
                    "original_filename": file.filename
                }
            )
            
            logger.info(f"Uploaded account proof to GCS: {gcs_path}")
            
            # Store with /files/ prefix for consistent serving
            stored_path = f"/files/account_topup_proofs/{unique_filename}"
            
        except Exception as gcs_error:
            logger.error(f"Failed to upload to GCS, falling back to filesystem: {gcs_error}")
            # Fallback to filesystem
            upload_dir = Path("/app/uploads/verification_files")
            upload_dir.mkdir(parents=True, exist_ok=True)
            file_path = upload_dir / unique_filename
            with open(file_path, "wb") as buffer:
                buffer.write(file_content)
            stored_path = f"uploads/verification_files/{unique_filename}"
        
        # Find and update the account in the accounts array
        accounts = topup_request.get("accounts", [])
        account_updated = False
        
        for account in accounts:
            if account.get("account_id") == account_id:
                if proof_type == "spend_limit_proof":
                    account["spend_limit_proof_url"] = stored_path
                elif proof_type == "budget_aspire_proof":
                    account["budget_aspire_proof_url"] = stored_path
                account_updated = True
                break
        
        if not account_updated:
            raise HTTPException(status_code=404, detail=f"Account {account_id} not found in this request")
        
        # Update the request in database
        await db.topup_requests.update_one(
            {"id": payment_id},
            {"$set": {"accounts": accounts}}
        )
        
        return {
            "success": True,
            "file_path": stored_path,
            "storage": "gcs",
            "account_id": account_id,
            "proof_type": proof_type,
            "original_filename": file.filename
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error uploading account proof: {e}")
        raise HTTPException(status_code=500, detail="Failed to upload account proof")
# Ad Account Management endpoints
@api_router.get("/admin/accounts", response_model=List[dict])
async def get_all_ad_accounts(
    platform: Optional[str] = None,
    status: Optional[str] = None,
    no_topup_days: Optional[int] = None,  # Filter accounts with no topup for X days
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get all ad accounts with optional filters"""
    try:
        query = {}
        if platform:
            query["platform"] = platform
        if status:
            query["status"] = status
            
        accounts = await db.ad_accounts.find(query).sort("created_at", -1).to_list(1000)
        
        result = []
        current_time = datetime.now(timezone.utc)
        
        for account in accounts:
            account = parse_from_mongo(account)
            # Get user info
            user = await db.users.find_one({"id": account["user_id"]})
            if user:
                account["user_name"] = user.get("name", "Unknown")
                account["user_username"] = user.get("username", "Unknown")
            
            # Get admin info if fee was updated by admin
            if account.get("fee_updated_by"):
                admin = await db.admin_users.find_one({"id": account["fee_updated_by"]})
                if admin:
                    admin = parse_from_mongo(admin)
                    account["fee_updated_by_admin"] = {
                        "id": admin.get("id"),
                        "username": admin.get("username"),
                        "name": admin.get("name", admin.get("username"))
                    }
            
            # Calculate last topup date and days since last topup
            # Check BOTH topup_requests (bank/crypto) AND wallet_transfers (wallet top-ups)
            
            # 1. Check topup_requests for verified requests that include this account
            last_topup_from_bank = await db.topup_requests.find_one(
                {
                    "accounts.account_id": account["id"],
                    "status": "verified"
                },
                sort=[("verified_at", -1), ("created_at", -1)]  # Fallback to created_at if verified_at not exists
            )
            
            # 2. Check wallet_transfers for completed/approved transfers to this account
            last_topup_from_wallet = await db.wallet_transfers.find_one(
                {
                    "target_account_id": account["id"],
                    "status": {"$in": ["completed", "approved"]}
                },
                sort=[("processed_at", -1), ("verified_at", -1), ("created_at", -1)]  # Prioritize processed_at > verified_at > created_at
            )
            
            # Determine the most recent top-up from both sources
            last_topup_at = None
            
            if last_topup_from_bank:
                last_topup_from_bank = parse_from_mongo(last_topup_from_bank)
                bank_topup_at = last_topup_from_bank.get("verified_at") or last_topup_from_bank.get("created_at")
                if isinstance(bank_topup_at, str):
                    bank_topup_at = datetime.fromisoformat(bank_topup_at.replace('Z', '+00:00'))
                if bank_topup_at.tzinfo is None:
                    bank_topup_at = bank_topup_at.replace(tzinfo=timezone.utc)
                last_topup_at = bank_topup_at
            
            if last_topup_from_wallet:
                last_topup_from_wallet = parse_from_mongo(last_topup_from_wallet)
                # For wallet transfers, prioritize processed_at > verified_at > created_at
                wallet_topup_at = (
                    last_topup_from_wallet.get("processed_at") or 
                    last_topup_from_wallet.get("verified_at") or 
                    last_topup_from_wallet.get("created_at")
                )
                if isinstance(wallet_topup_at, str):
                    wallet_topup_at = datetime.fromisoformat(wallet_topup_at.replace('Z', '+00:00'))
                if wallet_topup_at.tzinfo is None:
                    wallet_topup_at = wallet_topup_at.replace(tzinfo=timezone.utc)
                
                # Compare with bank top-up and take the most recent
                if last_topup_at is None or wallet_topup_at > last_topup_at:
                    last_topup_at = wallet_topup_at
            
            if last_topup_at:
                days_since_last_topup = (current_time - last_topup_at).days
                account["last_topup_at"] = last_topup_at.isoformat()
                account["days_since_last_topup"] = days_since_last_topup
                account["never_topped_up"] = False  # Explicitly set to False
            else:
                # No topup ever (neither bank/crypto nor wallet)
                created_at = account.get("created_at")
                if isinstance(created_at, str):
                    created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                if created_at.tzinfo is None:
                    created_at = created_at.replace(tzinfo=timezone.utc)
                
                days_since_creation = (current_time - created_at).days
                account["last_topup_at"] = None
                account["days_since_last_topup"] = days_since_creation
                account["never_topped_up"] = True
            
            # Apply no_topup_days filter
            if no_topup_days is not None:
                if account["days_since_last_topup"] < no_topup_days:
                    continue  # Skip accounts that had recent topup
            
            result.append(account)
        
        # Sort by days_since_last_topup descending if no_topup_days filter is active
        if no_topup_days is not None:
            result.sort(key=lambda x: x["days_since_last_topup"], reverse=True)
        
        return result
    except Exception as e:
        logger.error(f"Error fetching ad accounts: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch ad accounts")

@api_router.put("/admin/accounts/{account_id}/status", response_model=dict)
async def update_account_status(
    account_id: str,
    status_data: AccountStatusUpdate,
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Update ad account status (active, suspended, disabled)"""
    account = await db.ad_accounts.find_one({"id": account_id})
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    
    if status_data.status not in ["active", "suspended", "disabled"]:
        raise HTTPException(status_code=400, detail="Invalid status. Use: active, suspended, or disabled")
    
    await db.ad_accounts.update_one(
        {"id": account_id},
        {"$set": {"status": status_data.status, "updated_at": datetime.now(timezone.utc)}}
    )
    
    # Create notification for client
    user = await db.users.find_one({"id": account["user_id"]})
    if user:
        status_message = {
            "active": "activated",
            "suspended": "suspended",
            "disabled": "disabled"
        }
        
        client_notification = {
            "id": str(uuid.uuid4()),
            "user_id": account["user_id"], 
            "title": get_notification_text("account_status_changed", "id"),
            "message": get_notification_text("account_status_message", "id", 
                                           platform=account['platform'].title(), 
                                           account_name=account['account_name'],
                                           status=status_message[status_data.status]),
            "type": "status_change",
            "reference_id": account_id,
            "is_read": False,
            "created_at": datetime.now(timezone.utc)
        }
        
        client_notification_dict = prepare_for_mongo(client_notification)
        await db.client_notifications.insert_one(client_notification_dict)
    
    return {"message": f"Account status updated to {status_data.status}"}

@api_router.put("/admin/accounts/{account_id}/fee", response_model=dict)
async def update_account_fee_percentage(
    account_id: str,
    fee_data: AccountFeeUpdate,
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Update ad account fee percentage"""
    account = await db.ad_accounts.find_one({"id": account_id})
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    
    # Validate fee percentage range
    if fee_data.fee_percentage < 0 or fee_data.fee_percentage > 100:
        raise HTTPException(status_code=400, detail="Fee percentage must be between 0 and 100")
    
    await db.ad_accounts.update_one(
        {"id": account_id},
        {"$set": {
            "fee_percentage": fee_data.fee_percentage, 
            "fee_updated_by": current_admin.id,
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    return {"message": f"Account fee percentage updated to {fee_data.fee_percentage}%"}

@api_router.put("/admin/accounts/{account_id}/name", response_model=dict)
async def update_account_name(
    account_id: str,
    name_data: dict,
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Update ad account name"""
    account = await db.ad_accounts.find_one({"id": account_id})
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    
    account_name = name_data.get('account_name', '').strip()
    if not account_name:
        raise HTTPException(status_code=400, detail="Account name cannot be empty")
    
    # Update account name in ad_accounts collection
    await db.ad_accounts.update_one(
        {"id": account_id},
        {"$set": {
            "account_name": account_name,
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    # Also update account_name in related ad_account_requests if exists
    if account.get("platform") and account.get("user_id"):
        await db.ad_account_requests.update_many(
            {
                "user_id": account["user_id"],
                "account_id": account.get("account_id"),  # Match by platform account_id
                "status": "approved"
            },
            {"$set": {
                "account_name": account_name,
                "updated_at": datetime.now(timezone.utc)
            }}
        )
    
    return {"message": f"Account name updated to '{account_name}'"}

@api_router.put("/admin/accounts/{account_id}/account-id", response_model=dict)
async def update_account_id(
    account_id: str,
    account_id_data: dict,
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Update ad account ID (platform-specific ID like Facebook Account ID, TikTok BC ID, etc)"""
    account = await db.ad_accounts.find_one({"id": account_id})
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    
    new_account_id = account_id_data.get('new_account_id', '').strip()
    if not new_account_id:
        raise HTTPException(status_code=400, detail="Account ID cannot be empty")
    
    # Check if new account_id already exists (avoid duplicates)
    existing = await db.ad_accounts.find_one({
        "account_id": new_account_id,
        "id": {"$ne": account_id}  # Exclude current account
    })
    if existing:
        raise HTTPException(
            status_code=400, 
            detail=f"Account ID '{new_account_id}' already exists. Please use a unique ID."
        )
    
    old_account_id = account.get("account_id")
    
    # Update account_id in ad_accounts collection
    await db.ad_accounts.update_one(
        {"id": account_id},
        {"$set": {
            "account_id": new_account_id,
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    # Update account_id in related ad_account_requests
    if account.get("user_id") and old_account_id:
        await db.ad_account_requests.update_many(
            {
                "user_id": account["user_id"],
                "account_id": old_account_id
            },
            {"$set": {
                "account_id": new_account_id,
                "updated_at": datetime.now(timezone.utc)
            }}
        )
    
    logger.info(f"Admin {current_admin.username} updated account ID: {old_account_id} → {new_account_id}")
    
    return {
        "message": f"Account ID updated from '{old_account_id}' to '{new_account_id}'",
        "old_account_id": old_account_id,
        "new_account_id": new_account_id
    }

@api_router.delete("/admin/accounts/{account_id}", response_model=dict)
async def delete_account(
    account_id: str,
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Delete an ad account"""
    account = await db.ad_accounts.find_one({"id": account_id})
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    
    # Check if account has transactions or active balance
    transactions = await db.transactions.count_documents({"account_id": account_id})
    if transactions > 0:
        raise HTTPException(
            status_code=400, 
            detail="Cannot delete account with transaction history. Consider disabling instead."
        )
    
    if account.get("balance", 0) > 0:
        raise HTTPException(
            status_code=400, 
            detail="Cannot delete account with positive balance. Please withdraw funds first."
        )
    
    await db.ad_accounts.delete_one({"id": account_id})
    
    # Create notification for client
    user = await db.users.find_one({"id": account["user_id"]})
    if user:
        client_notification = {
            "id": str(uuid.uuid4()),
            "user_id": account["user_id"], 
            "title": get_notification_text("account_deleted", "id"),
            "message": get_notification_text("account_deleted_message", "id", 
                                           platform=account['platform'].title(), 
                                           account_name=account['account_name']),
            "type": "account_deleted",
            "reference_id": account_id,
            "is_read": False,
            "created_at": datetime.now(timezone.utc)
        }
        
        client_notification_dict = prepare_for_mongo(client_notification)
        await db.client_notifications.insert_one(client_notification_dict)
        
        # Send email notification to client about account deletion
        try:
            if user.get("email"):
                send_client_account_deleted_email(
                    client_email=user["email"],
                    client_name=user.get("name") or user.get("username"),
                    account_name=account['account_name'],
                    platform=account['platform'].title(),
                    balance_transferred=False,
                    target_account=""
                )
                logger.info(f"📧 Account deletion email sent to {user['email']}")
        except Exception as e:
            logger.error(f"Failed to send account deletion email: {e}")
    
    return {"message": "Account deleted successfully"}

@api_router.delete("/admin/accounts/{account_id}/with-balance-transfer", response_model=dict)
async def delete_account_with_balance_transfer(
    account_id: str,
    balance_amount: float = Form(...),
    balance_proof: UploadFile = File(...),
    withdraw_proof: UploadFile = File(...),
    aspire_proof: UploadFile = File(...),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Delete an ad account and transfer balance to client's withdrawal wallet"""
    account = await db.ad_accounts.find_one({"id": account_id})
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    
    account = parse_from_mongo(account)
    user_id = account["user_id"]
    
    # Get user
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Validate balance amount
    if balance_amount <= 0:
        raise HTTPException(status_code=400, detail="Balance amount must be greater than 0")
    
    # Save proof files
    upload_dir = "/app/uploads/account_deletion_proofs"
    os.makedirs(upload_dir, exist_ok=True)
    
    proof_files = {}
    for file, file_type in [
        (balance_proof, "balance_proof"),
        (withdraw_proof, "withdraw_proof"),
        (aspire_proof, "aspire_proof")
    ]:
        if file:
            file_ext = os.path.splitext(file.filename)[1]
            unique_filename = f"{account_id}_{file_type}_{str(uuid.uuid4())[:8]}{file_ext}"
            file_path = os.path.join(upload_dir, unique_filename)
            
            with open(file_path, "wb") as f:
                content = await file.read()
                f.write(content)
            
            proof_files[file_type] = {
                "filename": unique_filename,
                "path": file_path,
                "original_name": file.filename
            }
    
    # Transfer balance to withdrawal wallet
    currency = account.get("currency", "IDR")
    if currency == "IDR":
        current_balance = user.get("withdrawal_wallet_idr", 0)
        new_balance = current_balance + balance_amount
        await db.users.update_one(
            {"id": user_id},
            {"$set": {"withdrawal_wallet_idr": new_balance}}
        )
    else:  # USD
        current_balance = user.get("withdrawal_wallet_usd", 0)
        new_balance = current_balance + balance_amount
        await db.users.update_one(
            {"id": user_id},
            {"$set": {"withdrawal_wallet_usd": new_balance}}
        )
    
    # Create transaction record
    transaction = {
        "id": str(uuid.uuid4()),
        "user_id": user_id,
        "account_id": account_id,
        "account_name": account.get("account_name"),
        "type": "account_deletion_balance_transfer",
        "amount": balance_amount,
        "currency": currency,
        "status": "completed",
        "description": f"Transfer saldo dari akun {account.get('account_name')} ke withdrawal wallet (akun dihapus)",
        "admin_id": current_admin.id,
        "balance_before": current_balance,
        "balance_after": new_balance,
        "proof_files": proof_files,
        "created_at": datetime.now(timezone.utc)
    }
    
    transaction_dict = prepare_for_mongo(transaction)
    await db.transactions.insert_one(transaction_dict)
    
    # Delete the account
    await db.ad_accounts.delete_one({"id": account_id})
    
    # Create notification for client
    balance_str = f"Rp {balance_amount:,.0f}" if currency == "IDR" else f"${balance_amount:,.2f}"
    client_notification = {
        "id": str(uuid.uuid4()),
        "user_id": user_id,
        "title": "Akun Dihapus & Saldo Ditransfer",
        "message": f"Akun {account.get('account_name')} telah dihapus. Saldo sebesar {balance_str} telah ditransfer ke withdrawal wallet Anda.",
        "type": "account_deleted_with_balance",
        "reference_id": account_id,
        "is_read": False,
        "created_at": datetime.now(timezone.utc)
    }
    
    client_notification_dict = prepare_for_mongo(client_notification)
    await db.client_notifications.insert_one(client_notification_dict)
    
    # Send email notification to client about account deletion with balance transfer
    try:
        if user.get("email"):
            send_client_account_deleted_email(
                client_email=user["email"],
                client_name=user.get("name") or user.get("username"),
                account_name=account.get('account_name'),
                platform=account.get('platform', '').title(),
                balance_transferred=True,
                target_account="Withdrawal Wallet"
            )
            logger.info(f"📧 Account deletion with balance transfer email sent to {user['email']}")
    except Exception as e:
        logger.error(f"Failed to send account deletion email: {e}")
    
    return {
        "message": "Account deleted successfully and balance transferred to withdrawal wallet",
        "balance_transferred": balance_amount,
        "currency": currency,
        "new_withdrawal_balance": new_balance
    }

# Admin Notification endpoints
@api_router.get("/admin/notifications", response_model=List[NotificationResponse])
async def get_notifications(
    unread_only: bool = False,
    limit: int = 50,
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get notifications for admin"""
    filter_query = {}
    if unread_only:
        filter_query["is_read"] = False
    
    notifications_cursor = db.notifications.find(filter_query).sort("created_at", -1).limit(limit)
    notifications = await notifications_cursor.to_list(length=None)
    
    # Handle missing is_read field for backward compatibility
    parsed_notifications = []
    for notification in notifications:
        parsed_notif = parse_from_mongo(notification)
        if 'is_read' not in parsed_notif:
            parsed_notif['is_read'] = False
        parsed_notifications.append(NotificationResponse(**parsed_notif))
    
    return parsed_notifications

@api_router.get("/admin/notifications/unread-count", response_model=dict)
async def get_unread_notification_count(current_admin: AdminUser = Depends(get_current_admin)):
    """Get count of unread notifications"""
    count = await db.notifications.count_documents({"is_read": False})
    return {"count": count}

@api_router.put("/admin/notifications/{notification_id}/read", response_model=dict)
async def mark_notification_read(
    notification_id: str,
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Mark notification as read"""
    result = await db.notifications.update_one(
        {"id": notification_id},
        {"$set": {"is_read": True}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Notification not found")
    
    return {"message": "Notification marked as read"}

@api_router.put("/admin/notifications/mark-all-read", response_model=dict)
async def mark_all_notifications_read(current_admin: AdminUser = Depends(get_current_admin)):
    """Mark all notifications as read"""
    await db.notifications.update_many(
        {"is_read": False},
        {"$set": {"is_read": True}}
    )
    
    return {"message": "All notifications marked as read"}

# Payment Verification Admin Endpoints
@api_router.get("/admin/payments", response_model=List[dict])
async def get_payment_requests(
    status: Optional[str] = None,
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get all payment requests for admin review"""
    query = {}
    if status:
        query["status"] = status
    
    requests = await db.topup_requests.find(query).sort("created_at", -1).to_list(1000)
    
    result = []
    for req in requests:
        req = parse_from_mongo(req)
        
        # Get user info
        user = await db.users.find_one({"id": req["user_id"]})
        user = parse_from_mongo(user) if user else {}
        
        # Get payment proof if exists
        payment_proof = None
        if req.get("payment_proof_id"):
            proof = await db.payment_proofs.find_one({"id": req["payment_proof_id"]})
            if proof:
                payment_proof = parse_from_mongo(proof)
        
        # Get admin info if verified
        verified_by_admin = None
        if req.get("verified_by"):
            admin = await db.admin_users.find_one({"id": req["verified_by"]})
            if admin:
                admin = parse_from_mongo(admin)
                verified_by_admin = {
                    "id": admin.get("id"),
                    "username": admin.get("username"),
                    "name": admin.get("name", admin.get("username"))
                }
        
        # Get first account details (for single account top-ups)
        account_name = None
        account_type = None
        account_id = None
        
        # IMPORTANT: Enrich accounts with proof edit status BEFORE formatting
        enriched_accounts = await enrich_accounts_with_proof_status(req.get("accounts", []), req["id"])
        
        # Format proof URLs in accounts array and populate platform account ID
        formatted_accounts = []
        for acc in enriched_accounts:
            # Lookup ad_account to get platform ID and other details
            internal_account_id = acc.get("account_id")  # This is internal UUID
            ad_account = await db.ad_accounts.find_one({"id": internal_account_id})
            
            if ad_account:
                ad_account = parse_from_mongo(ad_account)
                platform_account_id = ad_account.get("account_id")  # Platform ID (FB/Google/TikTok)
                account_name_from_db = ad_account.get("account_name")
                platform = ad_account.get("platform")
            else:
                platform_account_id = None
                account_name_from_db = acc.get("account_name", "Unknown")
                platform = acc.get("account_platform", "Unknown")
            
            # Format proof URLs with /files/ prefix if not already present
            spend_limit_url = acc.get("spend_limit_proof_url")
            if spend_limit_url and not spend_limit_url.startswith("/files/"):
                # Check if it's old format (uploads/verification_files/...)
                if spend_limit_url.startswith("uploads/"):
                    spend_limit_url = f"/{spend_limit_url}"
            
            budget_aspire_url = acc.get("budget_aspire_proof_url")
            if budget_aspire_url and not budget_aspire_url.startswith("/files/"):
                # Check if it's old format
                if budget_aspire_url.startswith("uploads/"):
                    budget_aspire_url = f"/{budget_aspire_url}"
            
            # Build formatted account with platform ID
            formatted_acc = {
                **acc,
                "account_id": internal_account_id,  # Keep internal ID as account_id
                "platform_account_id": platform_account_id,  # Add platform ID separately
                "account_name": account_name_from_db,
                "account_platform": platform,
                # Explicitly preserve pending edit flags
                "spend_limit_proof_pending_edit": acc.get("spend_limit_proof_pending_edit", False),
                "budget_aspire_proof_pending_edit": acc.get("budget_aspire_proof_pending_edit", False)
            }
            
            if spend_limit_url:
                formatted_acc["spend_limit_proof_url"] = spend_limit_url
            if budget_aspire_url:
                formatted_acc["budget_aspire_proof_url"] = budget_aspire_url
            
            formatted_accounts.append(formatted_acc)
        
        # Get first account details for table display
        if len(formatted_accounts) > 0:
            first_acc = formatted_accounts[0]
            account_name = first_acc.get("account_name")
            account_type = first_acc.get("account_platform")
            account_id = first_acc.get("platform_account_id")  # Use platform ID for display
        
        try:
            result.append({
                "id": req.get("id"),
                "reference_code": req.get("reference_code", "N/A"),
                "user": {
                    "id": user.get("id"),
                    "username": user.get("username"),
                    "email": user.get("email"),
                    "name": user.get("name")
                },
                "account_name": account_name,  # First account name for display
                "account_type": account_type,  # First account platform
                "account_id": account_id,  # First account platform ID
                "amount": formatted_accounts[0].get("amount") if len(formatted_accounts) > 0 else 0,
                "currency": req.get("currency", "IDR"),
                "total_amount": req.get("total_amount", 0),
                "total_fee": req.get("total_fee", 0),
                "unique_code": req.get("unique_code"),
                "total_with_unique_code": req.get("total_with_unique_code"),
                "accounts": formatted_accounts,  # IMPORTANT: Include formatted accounts array
                "accounts_count": len(formatted_accounts),
                "status": req.get("status", "pending"),
                "created_at": req.get("created_at"),
                "verified_at": req.get("verified_at"),
                "verified_by": verified_by_admin,
                "admin_notes": req.get("admin_notes"),
                "claimed_by": req.get("claimed_by"),
                "claimed_by_username": req.get("claimed_by_username"),
                "claimed_at": req.get("claimed_at"),
                "spend_limit_proof_url": req.get("spend_limit_proof_path"),
                "budget_aspire_proof_url": req.get("budget_aspire_proof_path"),
                "payment_proof": {
                    "uploaded": payment_proof is not None,
                    "uploaded_at": payment_proof.get("uploaded_at") if payment_proof else None,
                    "file_name": payment_proof.get("file_name") if payment_proof else None,
                    "file_path": payment_proof.get("file_path") if payment_proof else None
                }
            })
        except Exception as e:
            logger.error(f"Error processing payment request {req.get('id')}: {str(e)}")
            continue
    
    return result

@api_router.get("/admin/payments/{request_id}", response_model=dict)
async def get_payment_request_detail(
    request_id: str,
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get detailed payment request for admin review"""
    topup_request = await db.topup_requests.find_one({"id": request_id})
    if not topup_request:
        raise HTTPException(status_code=404, detail="Payment request not found")
    
    topup_request = parse_from_mongo(topup_request)
    
    # Get user info
    user = await db.users.find_one({"id": topup_request["user_id"]})
    user = parse_from_mongo(user) if user else {}
    
    # Get payment proof if exists
    payment_proof = None
    if topup_request.get("payment_proof_id"):
        proof = await db.payment_proofs.find_one({"id": topup_request["payment_proof_id"]})
        if proof:
            payment_proof = parse_from_mongo(proof)
    
    # Get account details
    account_details = []
    for acc in topup_request["accounts"]:
        ad_account = await db.ad_accounts.find_one({"id": acc["account_id"]})
        if ad_account:
            ad_account = parse_from_mongo(ad_account)
            account_details.append({
                "account_id": acc["account_id"],
                "account_name": ad_account["account_name"],
                "platform": ad_account["platform"],
                "amount": acc["amount"],
                "fee_percentage": acc["fee_percentage"],
                "fee_amount": acc["fee_amount"],
                "total": acc["amount"] + acc["fee_amount"]
            })
    
    return {
        "id": topup_request["id"],
        "reference_code": topup_request.get("reference_code", "N/A"),
        "user": user,
        "currency": topup_request["currency"],
        "total_amount": topup_request["total_amount"],
        "total_fee": topup_request["total_fee"],
        "accounts": account_details,
        "status": topup_request["status"],
        "created_at": topup_request["created_at"],
        "verified_at": topup_request.get("verified_at"),
        "admin_notes": topup_request.get("admin_notes"),
        "claimed_by": topup_request.get("claimed_by"),
        "claimed_by_username": topup_request.get("claimed_by_username"),
        "claimed_at": topup_request.get("claimed_at"),
        "transfer_details": {
            "type": "bank_transfer" if topup_request["currency"] == "IDR" else "crypto_wallet",
            "bank_name": topup_request.get("bank_name"),
            "account_number": topup_request.get("bank_account"),
            "account_holder": topup_request.get("bank_holder"),
            "wallet_address": topup_request.get("wallet_address"),
            "wallet_name": topup_request.get("wallet_name"),
            "network": topup_request.get("network"),
            "subtotal": topup_request.get("total_amount") if topup_request["currency"] == "IDR" else None,
            "unique_code": topup_request.get("unique_code", 0) if topup_request["currency"] == "IDR" else None,
            "total_transfer": topup_request.get("total_with_unique_code", topup_request.get("total_amount")) if topup_request["currency"] == "IDR" else topup_request.get("total_amount")
        },
        "payment_proof": {
            "uploaded": payment_proof is not None,
            "uploaded_at": payment_proof["uploaded_at"] if payment_proof else None,
            "file_name": payment_proof["file_name"] if payment_proof else None,
            "file_path": payment_proof["file_path"] if payment_proof else None
        }
    }

@api_router.put("/admin/payments/{request_id}/verify", response_model=dict)
async def verify_payment_request(
    request_id: str,
    verification_data: dict,
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Verify payment request (approve/reject)"""
    topup_request = await db.topup_requests.find_one({"id": request_id})
    if not topup_request:
        raise HTTPException(status_code=404, detail="Payment request not found")
    
    status = verification_data.get("status")  # "verified" or "rejected"
    admin_notes = verification_data.get("admin_notes", "")
    
    if status not in ["verified", "rejected"]:
        raise HTTPException(status_code=400, detail="Status must be 'verified' or 'rejected'")
    
    # For verification (approved), check that proofs are uploaded (after redesign: single account per request)
    if status == "verified":
        topup_request_parsed = parse_from_mongo(topup_request)
        
        # Check if this is new single-account structure or old accounts array structure
        if "accounts" in topup_request_parsed and isinstance(topup_request_parsed["accounts"], list):
            # OLD STRUCTURE: accounts array
            accounts = topup_request_parsed.get("accounts", [])
            
            missing_proofs = []
            for acc in accounts:
                account_name = acc.get("account_name", acc.get("account_id", "Unknown"))
                account_platform = acc.get("account_platform", "").lower()
                
                # Spend limit proof required for ALL platforms
                if not acc.get("spend_limit_proof_url"):
                    missing_proofs.append(f"{account_name}: Spend limit proof missing")
                
                # Budget aspire proof ONLY required for Facebook
                if account_platform == "facebook" and not acc.get("budget_aspire_proof_url"):
                    missing_proofs.append(f"{account_name}: Budget aspire proof missing")
            
            if missing_proofs:
                raise HTTPException(
                    status_code=400, 
                    detail=f"Missing proofs: {', '.join(missing_proofs)}"
                )
        else:
            # NEW STRUCTURE: single account with account_id field
            account_name = topup_request_parsed.get("account_name", "Unknown")
            account_platform = topup_request_parsed.get("account_platform", "").lower()
            
            missing_proofs = []
            
            # Spend limit proof required for ALL platforms
            if not topup_request_parsed.get("spend_limit_proof_url"):
                missing_proofs.append(f"{account_name}: Spend limit proof missing")
            
            # Budget aspire proof ONLY required for Facebook
            if account_platform == "facebook" and not topup_request_parsed.get("budget_aspire_proof_url"):
                missing_proofs.append(f"{account_name}: Budget aspire proof missing")
            
            if missing_proofs:
                raise HTTPException(
                    status_code=400, 
                    detail=f"Missing proofs: {', '.join(missing_proofs)}"
                )
    
    update_data = {
        "status": status,
        "verified_by": current_admin.id,
        "admin_notes": admin_notes,
        "verified_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.topup_requests.update_one(
        {"id": request_id},
        {"$set": update_data}
    )
    
    # If verified, update account balances
    if status == "verified":
        topup_request = parse_from_mongo(topup_request)
        
        # Support both old (accounts array) and new (single account) structures
        if "accounts" in topup_request and isinstance(topup_request["accounts"], list):
            # OLD STRUCTURE: accounts array
            for acc in topup_request["accounts"]:
                # Add balance to ad account and update last topup date
                await db.ad_accounts.update_one(
                    {"id": acc["account_id"]},
                    {
                        "$inc": {"balance": acc["amount"]},
                        "$set": {"last_topup_date": datetime.now(timezone.utc)}
                    }
                )
            
            # Create transaction record
            accounts_desc = ", ".join([f"{acc.get('account_name', 'Unknown')}" for acc in topup_request["accounts"]])
            transaction = Transaction(
                user_id=topup_request["user_id"],
                type="topup",
                amount=topup_request["total_amount"],
                currency=topup_request.get("currency", "IDR"),
                description=f"Top-up verified - {topup_request.get('reference_code', 'N/A')} ({len(topup_request['accounts'])} accounts: {accounts_desc})",
                status="completed"
            )
        else:
            # NEW STRUCTURE: single account per request
            # Add balance to ad account and update last topup date
            await db.ad_accounts.update_one(
                {"id": topup_request["account_id"]},
                {
                    "$inc": {"balance": topup_request["amount"]},
                    "$set": {"last_topup_date": datetime.now(timezone.utc)}
                }
            )
            
            # Create transaction record
            account_name = topup_request.get("account_name", "Unknown")
            transaction = Transaction(
                user_id=topup_request["user_id"],
                type="topup",
                amount=topup_request["total_amount"],
                currency=topup_request.get("currency", "IDR"),
                description=f"Top-up verified - {topup_request.get('reference_code', 'N/A')} ({account_name})",
                status="completed"
            )
        
        transaction_dict = prepare_for_mongo(transaction.dict())
        await db.transactions.insert_one(transaction_dict)
        
        # Create notification for client based on status
        notification_title_key = "payment_verified"
    else:
        # Create notification for client (rejected)
        notification_title_key = "payment_rejected"
        
        # Create transaction record for rejected account top-up
        if "accounts" in topup_request and isinstance(topup_request["accounts"], list):
            # OLD STRUCTURE
            accounts_desc = ", ".join([f"{acc.get('account_name', 'Unknown')}" for acc in topup_request["accounts"]])
            description = f"Top-up rejected - {topup_request.get('reference_code', 'N/A')} ({len(topup_request['accounts'])} accounts: {accounts_desc})"
        else:
            # NEW STRUCTURE
            account_name = topup_request.get("account_name", "Unknown")
            description = f"Top-up rejected - {topup_request.get('reference_code', 'N/A')} ({account_name})"
        
        transaction = Transaction(
            user_id=topup_request["user_id"],
            type="topup",
            amount=topup_request["total_amount"],
            currency=topup_request.get("currency", "IDR"),
            description=description,
            status="rejected"
        )
        
        transaction_dict = prepare_for_mongo(transaction.dict())
        # Add admin_notes to transaction dict
        transaction_dict["admin_notes"] = admin_notes or ""
        await db.transactions.insert_one(transaction_dict)
    
    # Create client notification
    if status == "verified":
        notification_type = "payment_verified"
        notification_title = "Top-Up Disetujui"
        if "accounts" in topup_request and isinstance(topup_request["accounts"], list):
            # OLD STRUCTURE
            accounts_desc = ", ".join([f"{acc.get('account_name', 'Unknown')}" for acc in topup_request["accounts"]])
            notification_message = f"Top-up untuk {len(topup_request['accounts'])} akun ({accounts_desc}) telah disetujui. Saldo sudah ditambahkan ke akun Anda."
        else:
            # NEW STRUCTURE
            account_name = topup_request.get("account_name", "Unknown")
            notification_message = f"Top-up untuk akun {account_name} telah disetujui. Saldo sudah ditambahkan ke akun Anda."
    else:
        notification_type = "payment_rejected"
        notification_title = "Top-Up Ditolak"
        notification_message = f"Top-up request ditolak. Alasan: {admin_notes if admin_notes else 'Tidak ada catatan'}"
    
    client_notification = {
        "id": str(uuid.uuid4()),
        "user_id": topup_request["user_id"],
        "title": notification_title,
        "message": notification_message,
        "type": notification_type,
        "reference_id": request_id,
        "is_read": False,
        "created_at": datetime.now(timezone.utc)
    }
    
    client_notification_dict = prepare_for_mongo(client_notification)
    await db.client_notifications.insert_one(client_notification_dict)
    
    # Send email notification to client
    try:
        user = await db.users.find_one({"id": topup_request["user_id"]})
        if user and user.get("email"):
            from email_service import send_client_topup_approved_email, send_client_topup_rejected_email
            
            if status == "verified":
                # Get first account details for email
                first_account = topup_request["accounts"][0]
                send_client_topup_approved_email(
                    client_email=user["email"],
                    client_name=user.get("name") or user.get("username"),
                    amount=topup_request["total_amount"],
                    currency=topup_request.get("currency", "IDR"),
                    account_name=first_account.get("account_name", "Unknown"),
                    admin_notes=admin_notes
                )
                logger.info(f"📧 Top-up approved email sent to {user['email']}")
            else:
                # Rejected
                first_account = topup_request["accounts"][0]
                send_client_topup_rejected_email(
                    client_email=user["email"],
                    client_name=user.get("name") or user.get("username"),
                    amount=topup_request["total_amount"],
                    currency=topup_request.get("currency", "IDR"),
                    account_name=first_account.get("account_name", "Unknown"),
                    reason=admin_notes or "Mohon hubungi admin untuk informasi lebih lanjut"
                )
                logger.info(f"📧 Top-up rejected email sent to {user['email']}")
    except Exception as e:
        logger.error(f"❌ Failed to send client top-up email: {e}")
    
    return {
        "message": f"Payment request {status} successfully",
        "status": status
    }

@api_router.get("/admin/payments/{request_id}/account-proof/{account_id}/{proof_type}")
async def get_account_proof(
    request_id: str,
    account_id: str,
    proof_type: str,  # "spend_limit" or "budget_aspire"
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get account top-up proof file from GCS with filesystem fallback"""
    from fastapi.responses import Response
    
    logger.info(f"📥 Account proof request: request_id={request_id}, account_id={account_id}, proof_type={proof_type}")
    
    try:
        # Get top-up request
        topup = await db.topup_requests.find_one({"id": request_id})
        if not topup:
            raise HTTPException(status_code=404, detail="Top-up request not found")
        
        # Find account in accounts array
        account = None
        for acc in topup.get("accounts", []):
            if acc.get("account_id") == account_id:
                account = acc
                break
        
        if not account:
            raise HTTPException(status_code=404, detail="Account not found in request")
        
        # Get proof URL based on type
        if proof_type == "spend_limit":
            proof_url = account.get("spend_limit_proof_url")
        elif proof_type == "budget_aspire":
            proof_url = account.get("budget_aspire_proof_url")
        else:
            raise HTTPException(status_code=400, detail="Invalid proof type")
        
        if not proof_url:
            raise HTTPException(status_code=404, detail=f"{proof_type} proof not found")
        
        # Extract GCS path from proof_url (format: /files/account_topup_proofs/filename.jpg)
        if proof_url.startswith("/files/"):
            gcs_path = proof_url[7:]  # Remove "/files/" prefix
        else:
            gcs_path = proof_url
        
        try:
            # Try GCS first
            content, mime_type = await download_from_gcs(gcs_path)
            logger.info(f"✅ Served account proof from GCS: {gcs_path}")
            return Response(
                content=content,
                media_type=mime_type,
                headers={
                    "Access-Control-Allow-Origin": "*",
                    "Access-Control-Allow-Methods": "GET",
                    "Access-Control-Allow-Headers": "Authorization",
                    "Cache-Control": "public, max-age=3600"
                }
            )
        except HTTPException:
            # Filesystem fallback
            logger.warning(f"⚠️ File not in GCS, trying filesystem: {gcs_path}")
            
            file_path = Path(f"/app/{proof_url}") if not proof_url.startswith("/") else Path(proof_url)
            
            if not file_path.exists():
                file_path = Path(f"/app/uploads/{gcs_path}")
            
            if not file_path.exists():
                logger.error(f"❌ Account proof not found: {proof_url}")
                raise HTTPException(status_code=404, detail="Proof file not found")
            
            # Determine content type
            extension = file_path.suffix.lower()
            mime_type_map = {
                '.jpg': 'image/jpeg',
                '.jpeg': 'image/jpeg',
                '.png': 'image/png',
                '.pdf': 'application/pdf',
                '.webp': 'image/webp'
            }
            mime_type = mime_type_map.get(extension, 'image/jpeg')
            
            logger.info(f"✅ Served account proof from filesystem: {file_path}")
            return FileResponse(
                path=str(file_path),
                media_type=mime_type,
                headers={
                    "Access-Control-Allow-Origin": "*",
                    "Access-Control-Allow-Methods": "GET",
                    "Access-Control-Allow-Headers": "Authorization",
                    "Cache-Control": "public, max-age=3600"
                }
            )
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error serving account proof: {e}")
        raise HTTPException(status_code=500, detail="Error serving proof file")

@api_router.get("/admin/payments/{request_id}/payment-proof")
async def get_account_topup_payment_proof(
    request_id: str,
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get account top-up payment proof file from GCS with filesystem fallback"""
    from fastapi.responses import Response
    
    logger.info(f"📥 Account top-up proof request: request_id={request_id}, admin={current_admin.username}")
    
    try:
        # Get top-up request
        topup = await db.topup_requests.find_one({"id": request_id})
        if not topup:
            logger.error(f"❌ Account top-up request not found: request_id={request_id}")
            raise HTTPException(status_code=404, detail="Top-up request not found")
        
        # Get payment proof
        payment_proof_id = topup.get("payment_proof_id")
        if not payment_proof_id:
            logger.error(f"❌ Payment proof ID missing: request_id={request_id}, topup_data={topup}")
            raise HTTPException(status_code=404, detail="Payment proof not found")
        
        logger.info(f"🔍 Looking for proof: proof_id={payment_proof_id}, request_status={topup.get('status')}, proof_uploaded_at={topup.get('proof_uploaded_at')}")
        
        proof = await db.payment_proofs.find_one({"id": payment_proof_id})
        if not proof:
            logger.error(f"❌ Payment proof not found in DB: proof_id={payment_proof_id}")
            raise HTTPException(status_code=404, detail="Payment proof not found")
        
        storage_type = proof.get("storage_type", "local")
        content_type = proof.get("mime_type", "image/jpeg")
        file_name = proof.get("file_name", "proof.jpg")
        
        # Serve from GCS
        if storage_type == "gcs":
            gcs_path = proof.get("gcs_path")
            if not gcs_path:
                raise HTTPException(status_code=404, detail="GCS path not found")
            
            try:
                # Use download_from_gcs helper
                file_content, mime_type = await download_from_gcs(gcs_path)
                
                logger.info(f"✅ Served account top-up proof from GCS: proof_id={payment_proof_id}, path={gcs_path}, size={len(file_content)} bytes")
                
                return Response(
                    content=file_content,
                    media_type=mime_type,
                    headers={
                        "Access-Control-Allow-Origin": "*",
                        "Access-Control-Allow-Methods": "GET",
                        "Access-Control-Allow-Headers": "Authorization",
                        "Content-Disposition": f'inline; filename="{file_name}"',
                        "Cache-Control": "no-cache, must-revalidate"  # Force reload, no caching
                    }
                )
            except HTTPException:
                logger.warning(f"⚠️ File not in GCS, trying filesystem fallback")
                # Continue to filesystem fallback
        
        # Filesystem fallback
        file_path_str = proof.get("file_path", proof.get("original_file_path"))
        if not file_path_str:
            raise HTTPException(status_code=404, detail="File path not found")
        
        file_path = Path(file_path_str)
        if not file_path.is_absolute():
            file_path = Path("/app") / file_path_str
        
        if not file_path.exists():
            logger.error(f"❌ File not found in filesystem: {file_path}")
            raise HTTPException(status_code=404, detail="Proof file not found")
        
        logger.info(f"✅ Served account top-up proof from filesystem: {file_path}")
        
        return FileResponse(
            path=str(file_path),
            media_type=content_type,
            headers={
                "Access-Control-Allow-Origin": "*",
                "Access-Control-Allow-Methods": "GET",
                "Access-Control-Allow-Headers": "Authorization",
                "Content-Disposition": f'inline; filename="{file_name}"',
                "Cache-Control": "public, max-age=3600"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error serving account top-up proof: {e}")
        raise HTTPException(status_code=500, detail="Error serving payment proof")

@api_router.get("/admin/payments/{request_id}/proof-file")
async def get_payment_proof_file(
    request_id: str,
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Serve payment proof file for admin review"""
    from fastapi.responses import FileResponse
    
    # Get payment proof
    topup_request = await db.topup_requests.find_one({"id": request_id})
    if not topup_request or not topup_request.get("payment_proof_id"):
        raise HTTPException(status_code=404, detail="Payment proof not found")
    
    payment_proof = await db.payment_proofs.find_one({"id": topup_request["payment_proof_id"]})
    if not payment_proof:
        raise HTTPException(status_code=404, detail="Payment proof not found")
    
    file_path = payment_proof["file_path"]
    if not Path(file_path).exists():
        raise HTTPException(status_code=404, detail="File not found on disk")
    
    return FileResponse(
        path=file_path,
        filename=payment_proof["file_name"],
        media_type=payment_proof["mime_type"],
        headers={
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "GET",
            "Access-Control-Allow-Headers": "Authorization"
        }
    )

@api_router.get("/admin/payments/{request_id}/spend_limit_proof")
async def get_payment_spend_limit_proof_file(
    request_id: str,
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Serve spend limit proof file for admin review"""
    from fastapi.responses import FileResponse
    
    topup_request = await db.topup_requests.find_one({"id": request_id})
    if not topup_request or not topup_request.get("spend_limit_proof_url"):
        raise HTTPException(status_code=404, detail="Spend limit proof not found")
    
    file_path = Path("/app") / topup_request["spend_limit_proof_url"]
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found on disk")
    
    return FileResponse(
        path=str(file_path),
        media_type="image/jpeg",
        headers={
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "GET",
            "Access-Control-Allow-Headers": "Authorization"
        }
    )

@api_router.get("/admin/payments/{request_id}/budget_aspire_proof")
async def get_payment_budget_aspire_proof_file(
    request_id: str,
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Serve budget aspire proof file for admin review"""
    from fastapi.responses import FileResponse
    
    topup_request = await db.topup_requests.find_one({"id": request_id})
    if not topup_request or not topup_request.get("budget_aspire_proof_url"):
        raise HTTPException(status_code=404, detail="Budget aspire proof not found")
    
    file_path = Path("/app") / topup_request["budget_aspire_proof_url"]
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found on disk")
    
    return FileResponse(
        path=str(file_path),
        media_type="image/jpeg",
        headers={
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "GET",
            "Access-Control-Allow-Headers": "Authorization"
        }
    )

@api_router.get("/admin/payments/{request_id}/account/{account_id}/spend_limit_proof")
async def get_account_spend_limit_proof(
    request_id: str,
    account_id: str
):
    """Serve spend limit proof file for specific account - PUBLIC"""
    from fastapi.responses import FileResponse
    
    topup_request = await db.topup_requests.find_one({"id": request_id})
    if not topup_request:
        raise HTTPException(status_code=404, detail="Top-up request not found")
    
    # Find the account in accounts array
    account_found = None
    for acc in topup_request.get("accounts", []):
        if acc.get("account_id") == account_id:
            account_found = acc
            break
    
    if not account_found or not account_found.get("spend_limit_proof_url"):
        raise HTTPException(status_code=404, detail="Spend limit proof not found for this account")
    
    file_path = Path("/app") / account_found["spend_limit_proof_url"]
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found on disk")
    
    return FileResponse(
        path=str(file_path),
        media_type="image/jpeg",
        headers={
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "GET, OPTIONS",
            "Access-Control-Allow-Headers": "*"
        }
    )

@api_router.get("/admin/payments/{request_id}/account/{account_id}/budget_aspire_proof")
async def get_account_budget_aspire_proof(
    request_id: str,
    account_id: str
):
    """Serve budget aspire proof file for specific account - PUBLIC"""
    from fastapi.responses import FileResponse
    
    topup_request = await db.topup_requests.find_one({"id": request_id})
    if not topup_request:
        raise HTTPException(status_code=404, detail="Top-up request not found")
    
    # Find the account in accounts array
    account_found = None
    for acc in topup_request.get("accounts", []):
        if acc.get("account_id") == account_id:
            account_found = acc
            break
    
    if not account_found or not account_found.get("budget_aspire_proof_url"):
        raise HTTPException(status_code=404, detail="Budget aspire proof not found for this account")
    
    file_path = Path("/app") / account_found["budget_aspire_proof_url"]
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found on disk")
    
    return FileResponse(
        path=str(file_path),
        media_type="image/jpeg",
        headers={
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "GET, OPTIONS",
            "Access-Control-Allow-Headers": "*"
        }
    )

# Invoice endpoints
@api_router.get("/topup-request/{request_id}/invoice")
async def download_invoice(
    request_id: str,
    current_user: User = Depends(get_current_user)
):
    """Download invoice PDF for top-up request"""
    try:
        # Get topup request record
        topup_record = await db.topup_requests.find_one({"id": request_id, "user_id": current_user.id})
        
        if not topup_record:
            raise HTTPException(status_code=404, detail="Top-up request not found")
        
        # Get user info
        user = await db.users.find_one({"id": current_user.id})
        
        # Parse the topup record to handle datetime conversion
        topup_record = parse_from_mongo(topup_record)
        
        # Prepare invoice data
        accounts_data = []
        subtotal = 0
        fees = 0
        
        for account_topup in topup_record.get("accounts", []):
            account = await db.ad_accounts.find_one({"id": account_topup["account_id"]})
            if account:
                amount = account_topup["amount"]
                fee = account_topup.get("fee_amount", 0)
                total = amount + fee
                
                accounts_data.append({
                    "platform": account["platform"],
                    "account_name": account["account_name"],
                    "account_id": account["account_id"],
                    "amount": amount,
                    "fee": fee,
                    "total": total
                })
                
                subtotal += amount
                fees += fee
        
        # Determine payment status based on verification status
        payment_status = "PAID" if topup_record.get("status") == "verified" else "UNPAID"
        
        invoice_data = InvoiceData(
            invoice_id=f"INV-{topup_record['unique_code']}-{datetime.now().strftime('%Y%m%d')}",
            user_name=user.get("name", "N/A"),
            user_email=user.get("email", "N/A"),
            currency=topup_record["currency"],
            accounts=accounts_data,
            subtotal=subtotal,
            fees=fees,
            unique_code=topup_record["unique_code"],
            total=subtotal + fees + topup_record["unique_code"],  # Correct total calculation: subtotal + fees + unique_code
            bank_details=topup_record.get("bank_details"),
            crypto_wallet=topup_record.get("crypto_wallet"),
            created_at=datetime.now(timezone.utc).astimezone(timezone(timedelta(hours=7))),
            payment_status=payment_status
        )
        
        # Generate PDF
        pdf_bytes = generate_invoice_pdf(invoice_data)
        
        # Return PDF as response
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=invoice_{invoice_data.invoice_id}.pdf",
                "Access-Control-Allow-Origin": "*"
            }
        )
        
    except Exception as e:
        logger.error(f"Error generating invoice: {e}")
        raise HTTPException(status_code=500, detail="Error generating invoice")

@api_router.get("/wallet-topup-request/{request_id}/invoice")
async def generate_wallet_topup_invoice(request_id: str, current_user: User = Depends(get_current_user)):
    """Generate invoice PDF for a wallet top-up request (client version)"""
    topup_record = await db.wallet_topup_requests.find_one({"id": request_id, "user_id": current_user.id})
    
    if not topup_record:
        raise HTTPException(status_code=404, detail="Wallet top-up request not found")
    
    # Allow invoice generation for pending, proof_uploaded, and verified status
    # Only reject for cancelled/rejected status
    if topup_record["status"] in ["cancelled", "rejected"]:
        raise HTTPException(status_code=400, detail="Invoice cannot be generated for cancelled/rejected requests")
    
    topup_record = parse_from_mongo(topup_record)
    
    # Set payment status based on verification status
    payment_status = "PAID" if topup_record["status"] == "verified" else "UNPAID"
    
    invoice_data = InvoiceData(
        invoice_id=f"WALLET-INV-{topup_record.get('unique_code', '000')}-{datetime.now().strftime('%Y%m%d')}",
        user_name=current_user.name or current_user.username,
        user_email=current_user.email,
        amount=topup_record["amount"],
        currency=topup_record["currency"],
        created_at=topup_record["created_at"],
        verified_at=topup_record.get("verified_at"),
        reference_code=topup_record.get("reference_code", ""),
        wallet_type=topup_record["wallet_type"],
        payment_method=topup_record["payment_method"],
        unique_code=topup_record.get("unique_code", 0),
        bank_name=topup_record.get("bank_name"),
        bank_account=topup_record.get("bank_account"),
        bank_holder=topup_record.get("bank_holder"),
        crypto_wallet=topup_record.get("wallet_address"),
        network=topup_record.get("network"),
        admin_notes=topup_record.get("admin_notes", ""),
        payment_status=payment_status  # Set based on verification status
    )
    
    pdf_buffer = generate_wallet_topup_invoice_pdf(invoice_data)
    
    return StreamingResponse(
        io.BytesIO(pdf_buffer),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=wallet_invoice_{request_id}.pdf"}
    )

@api_router.get("/admin/wallet-topup-request/{request_id}/invoice")
async def admin_generate_wallet_topup_invoice(request_id: str, current_admin: AdminUser = Depends(get_current_admin)):
    """Generate invoice PDF for a wallet top-up request (admin version)"""
    topup_record = await db.wallet_topup_requests.find_one({"id": request_id})
    
    if not topup_record:
        raise HTTPException(status_code=404, detail="Wallet top-up request not found")
    
    # Allow invoice generation for any status except cancelled/rejected
    if topup_record["status"] in ["cancelled", "rejected"]:
        raise HTTPException(status_code=400, detail="Invoice cannot be generated for cancelled/rejected requests")
    
    topup_record = parse_from_mongo(topup_record)
    
    # Get user info
    user = await db.users.find_one({"id": topup_record["user_id"]})
    
    # Set payment status based on verification status
    payment_status = "PAID" if topup_record["status"] == "verified" else "UNPAID"
    
    invoice_data = InvoiceData(
        invoice_id=f"WALLET-INV-{topup_record.get('unique_code', '000')}-{datetime.now().strftime('%Y%m%d')}",
        user_name=user.get("name") or user.get("username", "N/A"),
        user_email=user.get("email", "N/A"),
        amount=topup_record["amount"],
        currency=topup_record["currency"],
        created_at=topup_record["created_at"],
        verified_at=topup_record.get("verified_at"),
        reference_code=topup_record.get("reference_code", ""),
        wallet_type=topup_record["wallet_type"],
        payment_method=topup_record["payment_method"],
        unique_code=topup_record.get("unique_code", 0),
        bank_name=topup_record.get("bank_name"),
        bank_account=topup_record.get("bank_account"),
        bank_holder=topup_record.get("bank_holder"),
        crypto_wallet=topup_record.get("wallet_address"),
        network=topup_record.get("network"),
        admin_notes=topup_record.get("admin_notes", ""),
        payment_status=payment_status
    )
    
    pdf_buffer = generate_wallet_topup_invoice_pdf(invoice_data)
    
    return StreamingResponse(
        io.BytesIO(pdf_buffer),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=wallet_invoice_{request_id}.pdf"}
    )


@api_router.get("/admin/wallet-transfer-request/{request_id}/invoice")
async def admin_generate_wallet_transfer_invoice(request_id: str, current_admin: AdminUser = Depends(get_current_admin)):
    """Generate invoice PDF for a wallet transfer request (admin version)"""
    transfer_record = await db.wallet_transfers.find_one({"id": request_id})
    
    if not transfer_record:
        raise HTTPException(status_code=404, detail="Wallet transfer request not found")
    
    # Allow invoice generation for any status except rejected
    if transfer_record["status"] == "rejected":
        raise HTTPException(status_code=400, detail="Invoice cannot be generated for rejected requests")
    
    transfer_record = parse_from_mongo(transfer_record)
    
    # Get user info
    user = await db.users.find_one({"id": transfer_record["user_id"]})
    
    # Get target account info
    account = await db.ad_accounts.find_one({"id": transfer_record["target_account_id"]})
    
    # Set payment status based on transfer status
    payment_status = "COMPLETED" if transfer_record["status"] == "approved" else "PENDING"
    
    # Calculate fee - use from transfer record, or calculate from account fee_percentage
    amount = transfer_record["amount"]
    fee = transfer_record.get("fee")
    
    # If fee not in transfer record, calculate from account's fee_percentage
    if fee is None or fee == 0:
        if account and account.get("fee_percentage"):
            fee = amount * (account["fee_percentage"] / 100)
        else:
            # Default fallback if no fee info available
            fee = 0
    
    total = amount + fee
    
    invoice_data = InvoiceData(
        invoice_id=f"TRANSFER-INV-{request_id[:8].upper()}-{datetime.now().strftime('%Y%m%d')}",
        user_name=user.get("name") or user.get("username", "N/A"),
        user_email=user.get("email", "N/A"),
        amount=amount,
        fees=fee,
        total=total,
        currency=transfer_record["currency"],
        created_at=transfer_record["created_at"],
        verified_at=transfer_record.get("processed_at"),
        wallet_type=transfer_record["source_wallet_type"],
        admin_notes=transfer_record.get("admin_notes", ""),
        payment_status=payment_status,
        unique_code=0  # Not applicable for transfers
    )
    
    target_account_name = transfer_record["target_account_name"]
    target_platform = account.get("platform", "Unknown") if account else "Unknown"
    
    pdf_buffer = generate_wallet_transfer_invoice_pdf(invoice_data, target_account_name, target_platform)
    
    return StreamingResponse(
        io.BytesIO(pdf_buffer),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=transfer_invoice_{request_id}.pdf"}
    )

@api_router.get("/admin/topup-request/{request_id}/invoice")
async def admin_download_invoice(
    request_id: str,
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Admin download invoice PDF for top-up request"""
    try:
        # Get topup request record
        topup_record = await db.topup_requests.find_one({"id": request_id})
        
        if not topup_record:
            raise HTTPException(status_code=404, detail="Top-up request not found")
        
        # Get user info
        user = await db.users.find_one({"id": topup_record["user_id"]})
        
        # Parse the topup record to handle datetime conversion
        topup_record = parse_from_mongo(topup_record)
        
        # Prepare invoice data
        accounts_data = []
        subtotal = 0
        fees = 0
        
        for account_topup in topup_record.get("accounts", []):
            account = await db.ad_accounts.find_one({"id": account_topup["account_id"]})
            if account:
                amount = account_topup["amount"]
                fee = account_topup.get("fee_amount", 0)
                total = amount + fee
                
                accounts_data.append({
                    "platform": account["platform"],
                    "account_name": account["account_name"],
                    "account_id": account["account_id"],
                    "amount": amount,
                    "fee": fee,
                    "total": total
                })
                
                subtotal += amount
                fees += fee
        
        # Determine payment status based on verification status
        payment_status = "PAID" if topup_record.get("status") == "verified" else "UNPAID"
        
        invoice_data = InvoiceData(
            invoice_id=f"INV-{topup_record['unique_code']}-{datetime.now().strftime('%Y%m%d')}",
            user_name=user.get("name", "N/A"),
            user_email=user.get("email", "N/A"),
            currency=topup_record["currency"],
            accounts=accounts_data,
            subtotal=subtotal,
            fees=fees,
            unique_code=topup_record["unique_code"],
            total=subtotal + fees + topup_record["unique_code"],  # Correct total calculation: subtotal + fees + unique_code
            bank_details=topup_record.get("bank_details"),
            crypto_wallet=topup_record.get("crypto_wallet"),
            created_at=datetime.now(timezone.utc).astimezone(timezone(timedelta(hours=7))),
            payment_status=payment_status
        )
        
        # Generate PDF
        pdf_bytes = generate_invoice_pdf(invoice_data)
        
        # Return PDF as response
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename=invoice_{invoice_data.invoice_id}.pdf",
                "Access-Control-Allow-Origin": "*"
            }
        )
        
    except Exception as e:
        logger.error(f"Error generating invoice: {e}")
        raise HTTPException(status_code=500, detail="Error generating invoice")

# Client Withdraw endpoints
@api_router.get("/withdrawals", response_model=List[dict])
async def get_user_withdrawals(current_user: User = Depends(get_current_user)):
    """Get current user's withdrawal requests"""
    try:
        # Get withdrawals for current user
        withdrawals = await db.withdraw_requests.find({"user_id": current_user.id}).sort("created_at", -1).to_list(100)
        
        result = []
        for withdraw in withdrawals:
            withdraw = parse_from_mongo(withdraw)
            
            # Get account details
            account = await db.ad_accounts.find_one({"id": withdraw["account_id"]})
            if account:
                withdraw["account"] = {
                    "account_name": account["account_name"],
                    "platform": account["platform"],
                    "account_id": account["account_id"]
                }
            
            # Format for frontend
            formatted_withdraw = {
                "id": withdraw["id"],
                "account": withdraw.get("account"),
                "requested_amount": withdraw.get("requested_amount"),
                "actual_amount": withdraw.get("admin_verified_amount"),
                "currency": withdraw["currency"],
                "status": withdraw["status"],
                "admin_notes": withdraw.get("admin_notes"),
                "proof_image": withdraw.get("actual_balance_proof_url"),  # Fixed: use correct field name
                "created_at": withdraw["created_at"].isoformat() if isinstance(withdraw["created_at"], datetime) else withdraw["created_at"],
                "processed_at": withdraw["processed_at"].isoformat() if withdraw.get("processed_at") and isinstance(withdraw["processed_at"], datetime) else None
            }
            
            result.append(formatted_withdraw)
        
        return result
        
    except Exception as e:
        logging.error(f"Error fetching user withdrawals: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch withdrawal requests")

@api_router.post("/withdrawals", response_model=dict)
async def create_withdrawal_request(request: WithdrawRequestNew, current_user: User = Depends(get_current_user)):
    """Create new withdrawal request (no amount - admin will verify actual balance)"""
    try:
        # Check ad account exists and belongs to user
        ad_account = await db.ad_accounts.find_one({"id": request.account_id, "user_id": current_user.id})
        if not ad_account:
            raise HTTPException(status_code=404, detail="Ad account not found")
        
        # Check if account is active
        if ad_account.get("status") != "active":
            raise HTTPException(status_code=400, detail="Akun harus dalam status aktif untuk melakukan penarikan")
        
        # Check if there's already a pending withdrawal request for this account
        existing_pending = await db.withdraw_requests.find_one({
            "user_id": current_user.id,
            "account_id": request.account_id,
            "status": {"$in": ["pending", "approved", "processing"]}
        })
        
        if existing_pending:
            raise HTTPException(status_code=400, detail="Sudah ada permintaan penarikan yang sedang diproses untuk akun ini. Tunggu hingga selesai atau lakukan top-up untuk mengaktifkan kembali akun.")
        
        # SIMPLIFIED LOGIC: Just check if account has balance > 0
        if ad_account.get("balance", 0) <= 0:
            raise HTTPException(
                status_code=400, 
                detail="Akun tidak memiliki saldo yang cukup untuk ditarik."
            )
        
        # Create withdraw request record (amount will be verified by admin)
        # Use account's actual currency, not the request currency
        account_currency = ad_account.get("currency", "IDR")  # Default to IDR if not specified
        
        withdraw_record = WithdrawRequestRecord(
            user_id=current_user.id,
            account_id=request.account_id,
            platform=ad_account["platform"],
            account_name=ad_account["account_name"],
            requested_amount=0.0,  # No amount requested - admin will verify actual balance
            currency=account_currency,  # Use account's currency, not request currency
            status="pending"
        )
        
        withdraw_dict = prepare_for_mongo(withdraw_record.dict())
        await db.withdraw_requests.insert_one(withdraw_dict)
        
        # Create notification for admins
        admin_notification = {
            "id": str(uuid.uuid4()),
            "title": get_notification_text("new_withdraw_request", "id"),
            "message": get_notification_text("withdraw_request_message", "id",
                                           username=current_user.username,
                                           currency=account_currency,
                                           amount="(Admin akan verifikasi)",
                                           platform=ad_account["platform"]),
            "type": "new_withdraw_request",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "read": False,
            "user_id": current_user.id
        }
        
        await db.notifications.insert_one(admin_notification)
        
        # Send email notification to admins about new withdraw request
        try:
            admin_emails = await get_active_admin_emails()
            if admin_emails:
                from email_service import send_admin_new_withdraw_request_email
                send_admin_new_withdraw_request_email(
                    admin_emails=admin_emails,
                    client_name=current_user.name or current_user.username,
                    amount=0.0,  # Amount will be verified by admin
                    currency=account_currency,
                    account_name=ad_account['account_name']
                )
                logger.info(f"📧 New withdraw request email sent to {len(admin_emails)} admins")
        except Exception as e:
            logger.error(f"Failed to send withdraw request email: {e}")
        
        # Create transaction record
        transaction_record = {
            "id": str(uuid.uuid4()),
            "user_id": current_user.id,
            "type": "withdraw_request",
            "description": f"Permintaan penarikan saldo akun {ad_account['platform']} - {ad_account['account_name']} ({account_currency})",
            "amount": 0.0,  # Amount will be determined by admin
            "currency": account_currency,  # Use account's currency
            "status": "pending",
            "reference_id": withdraw_record.id,
            "reference_type": "withdraw_request",
            "created_at": datetime.now(timezone.utc),
            "platform": ad_account["platform"],
            "account_name": ad_account["account_name"],
            "account_id": request.account_id
        }
        
        transaction_dict = prepare_for_mongo(transaction_record)
        await db.transactions.insert_one(transaction_dict)
        
        return {
            "message": "Permintaan penarikan berhasil dibuat",
            "withdrawal_id": withdraw_record.id,
            "status": "pending"
        }
        
    except HTTPException as e:
        raise e
    except Exception as e:
        logging.error(f"Error creating withdrawal request: {str(e)}")
        raise HTTPException(status_code=500, detail="Gagal membuat permintaan penarikan")

# Admin Withdraw Management endpoints  
@api_router.get("/admin/withdraws", response_model=List[dict])
async def get_withdraw_requests(
    status: Optional[str] = None,
    platform: Optional[str] = None,
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get all withdraw requests with filters"""
    try:
        query = {}
        if status:
            query["status"] = status
        if platform:
            query["platform"] = platform
            
        withdraws = await db.withdraw_requests.find(query).sort("created_at", -1).to_list(1000)
        
        # Enrich with user info
        result = []
        for withdraw in withdraws:
            withdraw = parse_from_mongo(withdraw)
            
            # Handle proof file paths for backward compatibility
            if withdraw.get("actual_balance_proof_url"):
                proof_path = withdraw["actual_balance_proof_url"]
                if not proof_path.startswith('/files/') and not proof_path.startswith('http'):
                    # Old filesystem path, convert to API format
                    if 'balance_proofs' in proof_path:
                        filename = Path(proof_path).name
                        withdraw["actual_balance_proof_url"] = f"/files/balance_proofs/{filename}"
                    else:
                        withdraw["actual_balance_proof_url"] = f"/files/{proof_path}"
            
            if withdraw.get("after_withdrawal_proof_url"):
                proof_path = withdraw["after_withdrawal_proof_url"]
                if not proof_path.startswith('/files/') and not proof_path.startswith('http'):
                    # Old filesystem path, convert to API format
                    if 'balance_proofs' in proof_path:
                        filename = Path(proof_path).name
                        withdraw["after_withdrawal_proof_url"] = f"/files/balance_proofs/{filename}"
                    else:
                        withdraw["after_withdrawal_proof_url"] = f"/files/{proof_path}"
            
            # Get user info
            user = await db.users.find_one({"id": withdraw["user_id"]})
            if user:
                withdraw["user"] = {
                    "name": user.get("name", "Unknown"),
                    "username": user.get("username", "Unknown"),
                    "email": user.get("email", "Unknown")
                }
            
            # Get account info
            account = await db.ad_accounts.find_one({"id": withdraw["account_id"]})
            if account:
                withdraw["account_balance"] = account.get("balance", 0)
                withdraw["account_external_id"] = account.get("account_id", "")  # Facebook/Google/TikTok ID
            
            # Get admin info if verified
            if withdraw.get("verified_by"):
                admin = await db.admin_users.find_one({"id": withdraw["verified_by"]})
                if admin:
                    admin = parse_from_mongo(admin)
                    withdraw["verified_by_admin"] = {
                        "id": admin.get("id"),
                        "username": admin.get("username"),
                        "name": admin.get("name", admin.get("username"))
                    }
            
            # Add claim/lock fields - let FastAPI serialize datetime
            withdraw["claimed_by"] = withdraw.get("claimed_by")
            withdraw["claimed_by_username"] = withdraw.get("claimed_by_username")
            withdraw["claimed_at"] = withdraw.get("claimed_at")
            
            # Keep datetime objects for FastAPI to serialize
            # No manual conversion needed
                
            result.append(withdraw)
        
        return result
    except Exception as e:
        logger.error(f"Error fetching withdraw requests: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch withdraw requests")


@api_router.get("/admin/withdraws/{withdraw_id}/actual-balance-proof")
async def get_withdraw_actual_balance_proof(
    withdraw_id: str,
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get actual balance proof file for withdraw request"""
    try:
        withdraw_request = await db.withdraw_requests.find_one({"id": withdraw_id})
        if not withdraw_request:
            raise HTTPException(status_code=404, detail="Withdraw request not found")
        
        proof_url = withdraw_request.get("actual_balance_proof_url")
        if not proof_url:
            raise HTTPException(status_code=404, detail="Actual balance proof not found")
        
        # Extract GCS path from URL format
        if proof_url.startswith('/files/'):
            gcs_path = proof_url.replace('/files/', '')
        else:
            gcs_path = proof_url
        
        try:
            # Try GCS first
            content, mime_type = await download_from_gcs(gcs_path)
            logger.info(f"✅ Served actual balance proof from GCS: {gcs_path}")
            return Response(content=content, media_type=mime_type)
        except:
            # Fallback to filesystem
            file_path = Path(f"/app/uploads/{gcs_path}")
            if file_path.exists():
                return FileResponse(path=str(file_path))
            raise HTTPException(status_code=404, detail="Proof file not found")
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error serving actual balance proof: {e}")
        raise HTTPException(status_code=500, detail="Failed to serve proof")

@api_router.get("/admin/withdraws/{withdraw_id}/after-withdrawal-proof")
async def get_withdraw_after_withdrawal_proof(
    withdraw_id: str,
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get after withdrawal proof file for withdraw request"""
    try:
        withdraw_request = await db.withdraw_requests.find_one({"id": withdraw_id})
        if not withdraw_request:
            raise HTTPException(status_code=404, detail="Withdraw request not found")
        
        proof_url = withdraw_request.get("after_withdrawal_proof_url")
        if not proof_url:
            raise HTTPException(status_code=404, detail="After withdrawal proof not found")
        
        # Extract GCS path from URL format
        if proof_url.startswith('/files/'):
            gcs_path = proof_url.replace('/files/', '')
        else:
            gcs_path = proof_url
        
        try:
            # Try GCS first
            content, mime_type = await download_from_gcs(gcs_path)
            logger.info(f"✅ Served after withdrawal proof from GCS: {gcs_path}")
            return Response(content=content, media_type=mime_type)
        except:
            # Fallback to filesystem
            file_path = Path(f"/app/uploads/{gcs_path}")
            if file_path.exists():
                return FileResponse(path=str(file_path))
            raise HTTPException(status_code=404, detail="Proof file not found")
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error serving after withdrawal proof: {e}")
        raise HTTPException(status_code=500, detail="Failed to serve proof")

@api_router.put("/admin/withdraws/{withdraw_id}/status", response_model=dict)
async def update_withdraw_status(
    withdraw_id: str,
    update_data: AdminWithdrawUpdate,
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Admin update withdraw request status"""
    withdraw_request = await db.withdraw_requests.find_one({"id": withdraw_id})
    if not withdraw_request:
        raise HTTPException(status_code=404, detail="Withdraw request not found")
        
    # Allow transitions: pending → approved/rejected, approved → completed/rejected, processing → approved/rejected
    valid_transitions = {
        "pending": ["approved", "rejected"],
        "approved": ["completed", "rejected"],
        "processing": ["approved", "rejected"]
    }
    
    current_status = withdraw_request["status"]
    if current_status not in valid_transitions or update_data.status not in valid_transitions[current_status]:
        raise HTTPException(
            status_code=400, 
            detail=f"Invalid status transition from {current_status} to {update_data.status}"
        )
    
    now = datetime.now(timezone.utc)
    update_fields = {
        "status": update_data.status,
        "verified_by": current_admin.id,
        "verified_at": now.isoformat(),
        "processed_at": now.isoformat()
    }
    
    if update_data.admin_notes:
        update_fields["admin_notes"] = update_data.admin_notes
    
    # Save proof files if provided
    if update_data.actual_balance_proof_url:
        update_fields["actual_balance_proof_url"] = update_data.actual_balance_proof_url
    if update_data.after_withdrawal_proof_url:
        update_fields["after_withdrawal_proof_url"] = update_data.after_withdrawal_proof_url
        
    if update_data.status == "approved":
        # Admin approves the withdrawal - transfer balance from account to wallet immediately
        if update_data.verified_amount is None:
            raise HTTPException(status_code=400, detail="Verified amount is required for approval")
        
        update_fields["admin_verified_amount"] = update_data.verified_amount
        
        # Process the withdrawal - move money from account to wallet
        account = await db.ad_accounts.find_one({"id": withdraw_request["account_id"]})
        if not account:
            raise HTTPException(status_code=404, detail="Account not found")
        
        # Add to user WITHDRAWAL wallet using precise decimal calculations AND set account balance to 0 (complete withdrawal)
        user = await db.users.find_one({"id": withdraw_request["user_id"]})
        if user:
            # Determine withdrawal wallet field based on currency
            currency = withdraw_request.get("currency", "IDR")
            wallet_field = f"withdrawal_wallet_{currency.lower()}"  # withdrawal_wallet_idr or withdrawal_wallet_usd
            current_wallet_balance = user.get(wallet_field, 0)
            
            # Use precise decimal calculations for withdrawal wallet balance update
            current_wallet_decimal = to_decimal(current_wallet_balance)
            verified_amount_decimal = to_decimal(update_data.verified_amount)
            new_wallet_decimal = decimal_add(current_wallet_decimal, verified_amount_decimal)
            new_wallet_balance = to_float(decimal_round(new_wallet_decimal))
            
            await db.users.update_one(
                {"id": withdraw_request["user_id"]},
                {"$set": {wallet_field: new_wallet_balance}}
            )
        
        # CRITICAL FIX: Set account balance to 0 after withdrawal approval
        try:
            result = await db.ad_accounts.update_one(
                {"id": withdraw_request["account_id"]},
                {"$set": {"balance": 0}}
            )
            logger.info(f"Balance update result for account {withdraw_request['account_id']}: matched={result.matched_count}, modified={result.modified_count}")
            
            if result.matched_count == 0:
                logger.error(f"Account not found for balance update: {withdraw_request['account_id']}")
            elif result.modified_count == 0:
                logger.warning(f"Account balance was already 0 for account: {withdraw_request['account_id']}")
        except Exception as e:
            logger.error(f"Error updating account balance to 0: {e}")
            # Don't raise exception to not block the withdrawal approval
        
        # Update transaction status to completed
        await db.transactions.update_one(
            {"reference_id": withdraw_id, "type": "withdraw_request"},  # FIX: use correct type
            {"$set": {
                "status": "completed", 
                "amount": update_data.verified_amount,
                "currency": withdraw_request.get("currency", "IDR")  # Ensure currency is set
            }}
        )
        
        # Create success notification for client
        currency_symbol = "Rp " if withdraw_request.get('currency', 'IDR') == "IDR" else "$"
        
        # Format amount based on currency
        if withdraw_request.get('currency', 'IDR') == 'IDR':
            formatted_amount = f"{update_data.verified_amount:,.0f}"
        else:
            formatted_amount = f"{update_data.verified_amount:.2f}"
        
        client_notification = {
            "id": str(uuid.uuid4()),
            "user_id": withdraw_request["user_id"], 
            "title": "✅ Penarikan Selesai",
            "message": f"Penarikan {currency_symbol}{formatted_amount} telah selesai diproses.",
            "type": "completion",
            "reference_id": withdraw_id,
            "is_read": False,
            "created_at": datetime.now(timezone.utc)
        }
        
        client_notification_dict = prepare_for_mongo(client_notification)
        await db.client_notifications.insert_one(client_notification_dict)
        
        # Send email notification to client
        try:
            user = await db.users.find_one({"id": withdraw_request["user_id"]})
            if user and user.get("email"):
                from email_service import send_client_withdraw_approved_email
                send_client_withdraw_approved_email(
                    client_email=user["email"],
                    client_name=user.get("name") or user.get("username"),
                    amount=update_data.verified_amount,
                    currency=withdraw_request.get("currency", "IDR"),
                    account_name=withdraw_request.get("account_name", "Unknown")
                )
                logger.info(f"📧 Withdraw approved email sent to {user['email']}")
        except Exception as e:
            logger.error(f"❌ Failed to send withdraw approved email: {e}")
        
        # Auto-complete the withdrawal after approval processing
        update_fields["status"] = "completed"
        
    elif update_data.status == "completed":
        # Admin marks as completed - ensure balance is 0 (safety net in case admin skipped 'approved')
        try:
            result = await db.ad_accounts.update_one(
                {"id": withdraw_request["account_id"]},
                {"$set": {"balance": 0}}
            )
            logger.info(f"Completed status - Balance update result for account {withdraw_request['account_id']}: matched={result.matched_count}, modified={result.modified_count}")
        except Exception as e:
            logger.error(f"Error updating account balance to 0 on completed: {e}")
        
        # If verified_amount exists, ensure wallet is updated using precise calculations
        if update_data.verified_amount:
            user = await db.users.find_one({"id": withdraw_request["user_id"]})
            if user:
                wallet_field = "wallet_balance_idr" if withdraw_request.get("currency", "IDR") == "IDR" else "wallet_balance_usd"
                current_wallet_balance = user.get(wallet_field, 0)
                
                # Use precise decimal calculations for wallet balance update
                current_wallet_decimal = to_decimal(current_wallet_balance)
                verified_amount_decimal = to_decimal(update_data.verified_amount)
                new_wallet_decimal = decimal_add(current_wallet_decimal, verified_amount_decimal)
                new_wallet_balance = to_float(decimal_round(new_wallet_decimal))
                
                await db.users.update_one(
                    {"id": withdraw_request["user_id"]},
                    {"$set": {wallet_field: new_wallet_balance}}
                )
        
        # Create completion notification for client
        currency_symbol = "Rp " if withdraw_request.get('currency', 'IDR') == "IDR" else "$"
        verified_amount = withdraw_request.get("admin_verified_amount", withdraw_request.get("requested_amount", 0))
        
        client_notification = {
            "id": str(uuid.uuid4()),
            "user_id": withdraw_request["user_id"], 
            "title": get_notification_text("withdraw_approved", "id"),
            "message": get_notification_text("withdraw_approved_message", "id",
                                           currency=currency_symbol,
                                           amount=f"{verified_amount:,.2f}"),
            "type": "withdraw_approved",
            "reference_id": withdraw_id,
            "is_read": False,
            "created_at": datetime.now(timezone.utc)
        }
        
        client_notification_dict = prepare_for_mongo(client_notification)
        await db.client_notifications.insert_one(client_notification_dict)
        
    elif update_data.status == "rejected":
        # Update transaction status to failed
        await db.transactions.update_one(
            {"reference_id": withdraw_id, "type": "withdraw_request"},  # FIX: use correct type
            {"$set": {"status": "failed"}}
        )
        
        # Create rejection notification for client
        currency_symbol = "Rp " if withdraw_request.get('currency', 'IDR') == "IDR" else "$"
        
        client_notification = {
            "id": str(uuid.uuid4()),
            "user_id": withdraw_request["user_id"], 
            "title": get_notification_text("withdraw_rejected", "id"),
            "message": get_notification_text("withdraw_rejected_message", "id",
                                           currency=currency_symbol,
                                           amount=f"{withdraw_request['requested_amount']:,.2f}",
                                           notes=update_data.admin_notes or ''),
            "type": "withdraw_rejected",
            "reference_id": withdraw_id,
            "is_read": False,
            "created_at": datetime.now(timezone.utc)
        }
        
        client_notification_dict = prepare_for_mongo(client_notification)
        await db.client_notifications.insert_one(client_notification_dict)
    
    await db.withdraw_requests.update_one(
        {"id": withdraw_id},
        {"$set": update_fields}
    )
    
    return {"message": f"Withdraw request {update_data.status} successfully"}

@api_router.post("/admin/withdraws/{withdraw_id}/upload-proof")
async def upload_balance_proof(
    withdraw_id: str,
    file: UploadFile = File(...),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Admin upload proof of real account balance"""
    withdraw_request = await db.withdraw_requests.find_one({"id": withdraw_id})
    if not withdraw_request:
        raise HTTPException(status_code=404, detail="Withdraw request not found")
    
    # Validate file type
    allowed_types = ["image/jpeg", "image/png", "image/jpg", "image/webp", "application/pdf"]
    if file.content_type not in allowed_types:
        raise HTTPException(
            status_code=400, 
            detail="Invalid file type. Only JPEG, PNG, WebP, and PDF files are allowed"
        )
    
    # Upload to GCS
    gcs_data = await upload_to_gcs(file, folder="balance_proofs")
    balance_proof_path = f"/files/{gcs_data['gcs_path']}"
    
    # Update withdraw request with proof URL
    await db.withdraw_requests.update_one(
        {"id": withdraw_id},
        {"$set": {
            "actual_balance_proof_url": balance_proof_path,
            "actual_balance_proof_gcs": gcs_data['gcs_path']  # Store raw GCS path
        }}
    )
    
    return {
        "message": "Balance proof uploaded successfully", 
        "file_path": balance_proof_path,
        "storage": "gcs"
    }

@api_router.post("/admin/withdraws/{withdraw_id}/upload-after-proof")
async def upload_after_withdrawal_proof(
    withdraw_id: str,
    file: UploadFile = File(...),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Admin upload proof of balance after withdrawal"""
    withdraw_request = await db.withdraw_requests.find_one({"id": withdraw_id})
    if not withdraw_request:
        raise HTTPException(status_code=404, detail="Withdraw request not found")
    
    # Validate file type
    allowed_types = ["image/jpeg", "image/png", "image/jpg", "image/webp", "application/pdf"]
    if file.content_type not in allowed_types:
        raise HTTPException(
            status_code=400, 
            detail="Invalid file type. Only JPEG, PNG, WebP, and PDF files are allowed"
        )
    
    # Upload to GCS
    gcs_data = await upload_to_gcs(file, folder="balance_proofs")
    after_proof_path = f"/files/{gcs_data['gcs_path']}"
    
    # Update withdraw request with proof URL
    await db.withdraw_requests.update_one(
        {"id": withdraw_id},
        {"$set": {
            "after_withdrawal_proof_url": after_proof_path,
            "after_withdrawal_proof_gcs": gcs_data['gcs_path']  # Store raw GCS path
        }}
    )
    
    return {
        "message": "After withdrawal proof uploaded successfully", 
        "file_path": after_proof_path,
        "storage": "gcs"
    }

@api_router.get("/files/balance-proof/{file_name}")
async def get_balance_proof_file(
    file_name: str,
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get balance proof file (admin only)"""
    file_path = Path("/app/uploads/balance_proofs") / file_name
    
    # Handle relative path - prepend current working directory
    if not file_path.is_absolute():
        file_path = Path("/app") / file_path
    
    if not file_path.exists():
        logger.error(f"Admin proof file not found at: {file_path}")
        raise HTTPException(status_code=404, detail="File not found")
    
    # Determine content type
    file_name = file_path.name
    content_type = "image/jpeg"
    if file_name.lower().endswith('.png'):
        content_type = "image/png"
    elif file_name.lower().endswith('.pdf'):
        content_type = "application/pdf"
    elif file_name.lower().endswith('.webp'):
        content_type = "image/webp"
    
    return FileResponse(
        path=file_path,
        media_type=content_type,
        headers={
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "GET",
            "Access-Control-Allow-Headers": "Authorization"
        }
    )

@api_router.get("/client/balance-proof/{withdraw_id}")
async def get_client_balance_proof(
    withdraw_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get actual balance proof for client's own withdrawal request"""
    try:
        # Verify this withdrawal belongs to the user
        withdraw_request = await db.withdraw_requests.find_one({
            "id": withdraw_id,
            "user_id": current_user.id
        })
        
        if not withdraw_request:
            raise HTTPException(status_code=404, detail="Withdrawal request not found")
        
        # Only show proof for approved/completed withdrawals
        if withdraw_request.get("status") not in ["approved", "completed"]:
            raise HTTPException(status_code=403, detail="Proof not available for this withdrawal status")
        
        # Get actual_balance_proof_url (not after_withdrawal_proof)
        proof_url = withdraw_request.get("actual_balance_proof_url")
        if not proof_url:
            raise HTTPException(status_code=404, detail="No actual balance proof uploaded")
        
        # Extract GCS path from URL format
        if proof_url.startswith('/files/'):
            gcs_path = proof_url.replace('/files/', '')
        else:
            gcs_path = proof_url
        
        try:
            # Try GCS first
            content, mime_type = await download_from_gcs(gcs_path)
            logger.info(f"✅ Served client actual balance proof from GCS: {gcs_path}")
            return Response(content=content, media_type=mime_type)
        except:
            # Fallback to filesystem
            file_path = Path(f"/app/uploads/{gcs_path}")
            if file_path.exists():
                return FileResponse(path=str(file_path))
            raise HTTPException(status_code=404, detail="Actual balance proof file not found")
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error serving client actual balance proof: {e}")
        raise HTTPException(status_code=500, detail="Failed to serve proof")

# Admin Wallet Management endpoints
@api_router.get("/admin/wallet-topup-requests", response_model=List[dict])
async def get_wallet_topup_requests(
    status: Optional[str] = None,
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get all wallet top-up requests for admin review"""
    query = {}
    if status:
        query["status"] = status
    
    requests = await db.wallet_topup_requests.find(query).sort("created_at", -1).to_list(1000)
    
    result = []
    for req in requests:
        req = parse_from_mongo(req)
        
        # Get user info
        user = await db.users.find_one({"id": req["user_id"]})
        user = parse_from_mongo(user) if user else {}
        
        # Get payment proof if exists
        payment_proof = None
        if req.get("payment_proof_id"):
            proof = await db.payment_proofs.find_one({"id": req["payment_proof_id"]})
            if proof:
                payment_proof = parse_from_mongo(proof)
        
        # Get admin info if verified (check both verified_by and admin_id for backward compatibility)
        verified_by_admin = None
        admin_id = req.get("verified_by") or req.get("admin_id")
        if admin_id:
            admin = await db.admin_users.find_one({"id": admin_id})
            if admin:
                admin = parse_from_mongo(admin)
                verified_by_admin = {
                    "id": admin.get("id"),
                    "username": admin.get("username"),
                    "name": admin.get("name", admin.get("username"))
                }
        
        result.append({
            "id": req["id"],
            "reference_code": req.get("reference_code", ""),
            "user": {
                "id": user.get("id"),
                "username": user.get("username"),
                "email": user.get("email"),
                "name": user.get("name")
            },
            "wallet_type": req["wallet_type"],
            "currency": req["currency"],
            "amount": req["amount"],
            "payment_method": req["payment_method"],
            "unique_code": req.get("unique_code", 0),
            "total_with_unique_code": req.get("total_with_unique_code", req["amount"]),
            "bank_name": req.get("bank_name"),
            "bank_account": req.get("bank_account"),
            "bank_holder": req.get("bank_holder"),
            "wallet_address": req.get("wallet_address"),
            "wallet_name": req.get("wallet_name"),
            "network": req.get("network"),
            "status": req["status"],
            "notes": req.get("notes"),
            "admin_notes": req.get("admin_notes"),
            "claimed_by": req.get("claimed_by"),
            "claimed_by_username": req.get("claimed_by_username"),
            "claimed_at": req.get("claimed_at"),
            "created_at": req["created_at"],
            "verified_at": req.get("verified_at"),
            "verified_by": verified_by_admin,
            "payment_proof": {
                "uploaded": payment_proof is not None,
                "uploaded_at": payment_proof["uploaded_at"] if payment_proof else None,
                "file_name": payment_proof["file_name"] if payment_proof else None,
                "file_path": payment_proof["file_path"] if payment_proof else None
            },
            "type": "wallet_topup"  # Add type identifier
        })
    
    return result


async def enrich_accounts_with_proof_status(accounts, request_id):
    """Enrich accounts array with proof edit pending status and fee_percentage"""
    enriched = []
    for acc in accounts:
        acc_copy = acc.copy()
        account_id = acc.get("account_id")
        
        # Get account details including fee_percentage
        account_details = await db.ad_accounts.find_one({"id": account_id})
        if account_details:
            acc_copy["fee_percentage"] = account_details.get("fee_percentage", 0)
            acc_copy["platform"] = account_details.get("platform", "Unknown")
        
        # Check if spend_limit_proof has pending edit using tracking_id
        if acc.get("spend_limit_proof_url"):
            tracking_id = f"proof_tracking_{request_id}_{account_id}_spend_limit"
            proof = await db.payment_proofs.find_one({"tracking_id": tracking_id})
            if proof:
                acc_copy["spend_limit_proof_pending_edit"] = proof.get("pending_edit", False)
        
        # Check if budget_aspire_proof has pending edit using tracking_id
        if acc.get("budget_aspire_proof_url"):
            tracking_id = f"proof_tracking_{request_id}_{account_id}_budget_aspire"
            proof = await db.payment_proofs.find_one({"tracking_id": tracking_id})
            if proof:
                acc_copy["budget_aspire_proof_pending_edit"] = proof.get("pending_edit", False)
        
        enriched.append(acc_copy)
    
    return enriched

@api_router.get("/admin/topup-requests", response_model=List[dict])
async def get_account_topup_requests(
    status: Optional[str] = None,
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get all account top-up requests for admin review"""
    query = {}
    if status:
        query["status"] = status
    
    requests = await db.topup_requests.find(query).sort("created_at", -1).to_list(1000)
    
    result = []
    for req in requests:
        req = parse_from_mongo(req)
        
        # Get user info
        user = await db.users.find_one({"id": req["user_id"]})
        user = parse_from_mongo(user) if user else {}
        
        # Get payment proof if exists
        payment_proof = None
        if req.get("payment_proof_id"):
            proof = await db.payment_proofs.find_one({"id": req["payment_proof_id"]})
            if proof:
                payment_proof = parse_from_mongo(proof)
        
        # Get admin info if verified
        verified_by_admin = None
        admin_id = req.get("verified_by") or req.get("admin_id")
        if admin_id:
            admin = await db.admin_users.find_one({"id": admin_id})
            if admin:
                admin = parse_from_mongo(admin)
                verified_by_admin = {
                    "id": admin.get("id"),
                    "username": admin.get("username"),
                    "name": admin.get("name", admin.get("username"))
                }
        
        # Get claimed_by admin info if request is claimed
        claimed_by_admin = None
        claimed_by_id = req.get("claimed_by")
        if claimed_by_id:
            admin = await db.admin_users.find_one({"id": claimed_by_id})
            if admin:
                admin = parse_from_mongo(admin)
                claimed_by_admin = {
                    "id": admin.get("id"),
                    "username": admin.get("username"),
                    "name": admin.get("name", admin.get("username"))
                }
        
        result.append({
            "id": req["id"],
            "reference_code": req.get("reference_code", "N/A"),
            "user": {
                "id": user.get("id"),
                "username": user.get("username"),
                "email": user.get("email"),
                "name": user.get("name")
            },
            "currency": req["currency"],
            "total_amount": req["total_amount"],
            "total_fee": req.get("total_fee", 0),
            "accounts": await enrich_accounts_with_proof_status(req.get("accounts", []), req["id"]),  # Pass request_id
            "unique_code": req.get("unique_code", 0),
            "total_with_unique_code": req.get("total_with_unique_code", req["total_amount"]),
            "bank_name": req.get("bank_name"),
            "bank_account": req.get("bank_account"),
            "bank_holder": req.get("bank_holder"),
            "wallet_address": req.get("wallet_address"),
            "wallet_name": req.get("wallet_name"),
            "network": req.get("network"),
            "status": req["status"],
            "admin_notes": req.get("admin_notes"),
            "created_at": req["created_at"].isoformat() if isinstance(req["created_at"], datetime) else req["created_at"],
            "verified_at": req.get("verified_at").isoformat() if req.get("verified_at") and isinstance(req.get("verified_at"), datetime) else req.get("verified_at"),
            "verified_by": verified_by_admin,
            "claimed_by": req.get("claimed_by"),
            "claimed_by_username": claimed_by_admin["username"] if claimed_by_admin else None,
            "claimed_at": req.get("claimed_at").isoformat() if req.get("claimed_at") and isinstance(req.get("claimed_at"), datetime) else req.get("claimed_at"),
            "payment_proof": {
                "uploaded": payment_proof is not None,
                "uploaded_at": payment_proof["uploaded_at"] if payment_proof else None,
                "file_name": payment_proof["file_name"] if payment_proof else None,
                "file_path": payment_proof["file_path"] if payment_proof else None
            },
            "type": "account_topup"  # Add type identifier
        })
    
    return result

@api_router.get("/admin/transactions/{transaction_id}/payment-proof")
async def get_transaction_payment_proof(
    transaction_id: str,
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get transaction payment proof file"""
    from fastapi.responses import Response
    
    logger.info(f"📥 Transaction proof request: transaction_id={transaction_id}")
    
    try:
        # Get transaction
        transaction = await db.transactions.find_one({"id": transaction_id})
        if not transaction:
            raise HTTPException(status_code=404, detail="Transaction not found")
        
        reference_id = transaction.get("reference_id")
        if not reference_id:
            raise HTTPException(status_code=404, detail="No reference ID for this transaction")
        
        trans_type = transaction.get("type")
        
        # Find proof based on transaction type
        proof = None
        
        if trans_type == "wallet_topup":
            # Check payment_proofs (wallet topup juga pakai payment_proofs collection!)
            topup_req = await db.wallet_topup_requests.find_one({"id": reference_id})
            if topup_req and topup_req.get("payment_proof_id"):
                proof = await db.payment_proofs.find_one({"id": topup_req["payment_proof_id"]})
        
        elif trans_type in ["topup", "account_topup"]:
            # Check payment_proofs
            topup_req = await db.topup_requests.find_one({"id": reference_id})
            if topup_req and topup_req.get("payment_proof_id"):
                proof = await db.payment_proofs.find_one({"id": topup_req["payment_proof_id"]})
        
        elif trans_type == "wallet_to_account_transfer":
            # Check payment_proofs for transfer (also uses payment_proofs!)
            transfer_req = await db.wallet_transfer_requests.find_one({"id": reference_id})
            if transfer_req and transfer_req.get("payment_proof_id"):
                proof = await db.payment_proofs.find_one({"id": transfer_req["payment_proof_id"]})
        
        if not proof:
            raise HTTPException(status_code=404, detail="Payment proof not found")
        
        # Serve from GCS
        storage_type = proof.get("storage_type", "gcs")
        
        if storage_type == "gcs":
            gcs_path = proof.get("gcs_path")
            if not gcs_path:
                raise HTTPException(status_code=404, detail="GCS path not found")
            
            # Download from GCS
            content, mime_type = await download_from_gcs(gcs_path)
            logger.info(f"✅ Served transaction proof from GCS: {gcs_path}")
            
            return Response(
                content=content,
                media_type=mime_type,
                headers={
                    "Cache-Control": "public, max-age=3600",
                    "Content-Disposition": f'inline; filename="{proof.get("file_name", "proof.jpg")}"'
                }
            )
        else:
            # Local file fallback
            file_path = proof.get("file_path")
            if not file_path or not Path(file_path).exists():
                raise HTTPException(status_code=404, detail="File not found")
            
            return FileResponse(
                path=file_path,
                media_type=proof.get("mime_type", "image/jpeg"),
                filename=proof.get("file_name", "proof.jpg")
            )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error serving transaction proof: {e}")
        raise HTTPException(status_code=500, detail="Failed to serve proof")

@api_router.get("/admin/wallet-topup-requests/{topup_id}/payment-proof")
async def get_wallet_topup_payment_proof(
    topup_id: str,
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get wallet top-up payment proof file from database"""
    from fastapi.responses import Response
    
    logger.info(f"📥 Proof request received: topup_id={topup_id}, admin={current_admin.username}")
    
    try:
        # Get topup request
        topup = await db.wallet_topup_requests.find_one({"id": topup_id})
        if not topup:
            logger.error(f"❌ Top-up request not found: topup_id={topup_id}")
            raise HTTPException(status_code=404, detail="Top-up request not found")
        
        # Get payment proof
        payment_proof_id = topup.get("payment_proof_id")
        if not payment_proof_id:
            logger.error(f"❌ Payment proof ID missing: topup_id={topup_id}, topup_data={topup}")
            raise HTTPException(status_code=404, detail="Payment proof not found - no proof_id in request")
        
        logger.info(f"🔍 Looking for proof: proof_id={payment_proof_id}")
        
        proof = await db.payment_proofs.find_one({"id": payment_proof_id})
        if not proof:
            logger.error(f"❌ Payment proof not found in DB: proof_id={payment_proof_id}")
            raise HTTPException(status_code=404, detail="Payment proof not found")
        
        logger.info(f"✅ Found proof in DB: proof_id={payment_proof_id}, file_name={proof.get('file_name')}, uploaded_at={proof.get('uploaded_at')}, gcs_path={proof.get('gcs_path')}")
        
        storage_type = proof.get("storage_type", "local")  # Default to local for backward compatibility
        content_type = proof.get("mime_type", "image/jpeg")
        file_name = proof.get("file_name", "proof.jpg")
        
        # Check storage type and serve accordingly
        if storage_type == "gcs":
            # Serve from Google Cloud Storage
            gcs_path = proof.get("gcs_path")
            if not gcs_path:
                raise HTTPException(status_code=404, detail="GCS path not found")
            
            try:
                from google.cloud import storage
                import os
                
                os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = "/app/backend/gcs-service-account.json"
                storage_client = storage.Client()
                bucket_name = proof.get("gcs_bucket", os.getenv("GCS_BUCKET_NAME"))
                bucket = storage_client.bucket(bucket_name)
                blob = bucket.blob(gcs_path)
                
                # Download file content
                file_content = blob.download_as_bytes()
                
                logger.info(f"✅ Served from GCS: proof_id={payment_proof_id}, path={gcs_path}, size={len(file_content)} bytes")
                
                return Response(
                    content=file_content,
                    media_type=content_type,
                    headers={
                        "Access-Control-Allow-Origin": "*",
                        "Access-Control-Allow-Methods": "GET",
                        "Access-Control-Allow-Headers": "Authorization",
                        "Content-Disposition": f'inline; filename="{file_name}"',
                        "Cache-Control": "no-cache, must-revalidate"  # Force reload, no caching
                    }
                )
            except Exception as e:
                logger.error(f"Error downloading from GCS: {e}")
                raise HTTPException(status_code=500, detail="Error retrieving file from GCS")
        
        elif storage_type == "database":
            # Serve from database (base64 stored)
            file_data_base64 = proof.get("file_data")
            if not file_data_base64:
                raise HTTPException(status_code=404, detail="Payment proof file data not found")
            
            try:
                # Decode base64 to binary
                file_content = base64.b64decode(file_data_base64)
                
                logger.info(f"Served wallet topup proof from database: proof_id={payment_proof_id}")
                
                return Response(
                    content=file_content,
                    media_type=content_type,
                    headers={
                        "Access-Control-Allow-Origin": "*",
                        "Access-Control-Allow-Methods": "GET",
                        "Access-Control-Allow-Headers": "Authorization",
                        "Content-Disposition": f'inline; filename="{file_name}"',
                        "Cache-Control": "no-cache, must-revalidate"  # Force reload, no caching
                    }
                )
            except Exception as e:
                logger.error(f"Error decoding base64 file data: {e}")
                raise HTTPException(status_code=500, detail="Error retrieving file from database")
        
        elif storage_type == "gcs":
            # Legacy GCS files (if any)
            logger.error(f"❌ GCS storage attempted: proof_id={payment_proof_id}")
            raise HTTPException(
                status_code=404,
                detail="GCS storage no longer supported. Please re-upload proof."
            )
        
        elif storage_type == "gcs":
            # Already handled above
            logger.error(f"❌ GCS fallback reached - should not happen")
            raise HTTPException(status_code=500, detail="GCS handling error")
        
        else:
            # Legacy: Local filesystem (old files)
            # TRY filesystem first, if not found, try to migrate on-the-fly
            logger.info(f"📁 Attempting filesystem: storage_type={storage_type}, proof_id={payment_proof_id}")
            
            file_path = proof.get("file_path")
            if not file_path:
                logger.error(f"❌ File path missing: proof_id={payment_proof_id}")
                raise HTTPException(status_code=404, detail="File path not found")
            
            # Handle both relative and absolute paths
            from pathlib import Path
            from fastapi.responses import FileResponse
            
            file_path_obj = Path(file_path)
            if not file_path_obj.is_absolute():
                file_path_obj = Path("/app") / file_path
            
            logger.info(f"🔍 Checking file exists: {file_path_obj}")
            
            if file_path_obj.exists():
                # File exists on filesystem - serve it directly
                logger.info(f"✅ Served wallet topup proof from filesystem: proof_id={payment_proof_id}, path={file_path_obj}, size={file_path_obj.stat().st_size} bytes")
                
                return FileResponse(
                    path=file_path_obj,
                    media_type=content_type,
                    headers={
                        "Access-Control-Allow-Origin": "*",
                        "Access-Control-Allow-Methods": "GET",
                        "Access-Control-Allow-Headers": "Authorization",
                        "Content-Disposition": f'inline; filename="{file_name}"'
                    }
                )
            else:
                # File NOT found on filesystem - Try to migrate from any available source
                logger.warning(f"⚠️ File not found on filesystem: {file_path_obj}")
                logger.info(f"🔄 Attempting on-the-fly migration for proof_id={payment_proof_id}")
                
                # Check if we already have file_data from a previous migration attempt
                if proof.get("file_data"):
                    logger.info(f"✅ Found file_data in database, updating storage_type")
                    # Update storage type to database
                    await db.payment_proofs.update_one(
                        {"id": payment_proof_id},
                        {"$set": {"storage_type": "database"}}
                    )
                    
                    # Serve from database
                    file_content = base64.b64decode(proof["file_data"])
                    return Response(
                        content=file_content,
                        media_type=content_type,
                        headers={
                            "Access-Control-Allow-Origin": "*",
                            "Access-Control-Allow-Methods": "GET",
                            "Access-Control-Allow-Headers": "Authorization",
                            "Content-Disposition": f'inline; filename="{file_name}"',
                            "Cache-Control": "public, max-age=3600"
                        }
                    )
                else:
                    # No file_data and no file - cannot serve
                    logger.error(f"❌ Cannot serve proof: No file on disk and no file_data in database")
                    raise HTTPException(
                        status_code=404,
                        detail=f"Payment proof file not available. File may have been lost during deployment. Please ask user to re-upload."
                    )
            
            logger.info(f"✅ Served wallet topup proof: proof_id={payment_proof_id}, storage={storage_type}")
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error serving wallet topup proof: {e}")
        raise HTTPException(status_code=500, detail="Error serving file")


@api_router.get("/admin/wallet-transfer-requests", response_model=List[dict])
async def get_wallet_transfer_requests(
    status: Optional[str] = None,
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get all wallet-to-account transfer requests for admin review"""
    query = {}
    if status:
        query["status"] = status
    
    requests = await db.wallet_transfers.find(query).sort("created_at", -1).to_list(1000)
    
    result = []
    for req in requests:
        req = parse_from_mongo(req)
        
        # Get user info
        user = await db.users.find_one({"id": req["user_id"]})
        user = parse_from_mongo(user) if user else {}
        
        # Get target account info
        account = await db.ad_accounts.find_one({"id": req["target_account_id"]})
        account = parse_from_mongo(account) if account else {}
        
        # Get admin info if verified
        verified_by_admin = None
        if req.get("verified_by"):
            admin = await db.admin_users.find_one({"id": req["verified_by"]})
            if admin:
                admin = parse_from_mongo(admin)
                verified_by_admin = {
                    "id": admin.get("id"),
                    "username": admin.get("username"),
                    "name": admin.get("name", admin.get("username"))
                }
        
        # Check for pending proof edits
        transfer_id = req["id"]
        target_account_id = req["target_account_id"]
        
        # Check spend_limit proof pending edit
        spend_limit_tracking_id = f"proof_tracking_{transfer_id}_{target_account_id}_spend_limit"
        spend_limit_pending = await db.payment_proofs.find_one({
            "tracking_id": spend_limit_tracking_id,
            "pending_edit": True
        })
        
        # Check budget_aspire proof pending edit
        budget_aspire_tracking_id = f"proof_tracking_{transfer_id}_{target_account_id}_budget_aspire"
        budget_aspire_pending = await db.payment_proofs.find_one({
            "tracking_id": budget_aspire_tracking_id,
            "pending_edit": True
        })
        
        result.append({
            "id": req["id"],
            "user": {
                "id": user.get("id"),
                "username": user.get("username"),
                "email": user.get("email"),
                "name": user.get("name")
            },
            "source_wallet_type": req["source_wallet_type"],
            "target_account_id": req["target_account_id"],  # Internal UUID
            "target_account_name": req["target_account_name"],
            "target_account": {
                "id": req["target_account_id"],
                "name": req["target_account_name"],
                "platform": account.get("platform", "Unknown"),
                "account_id": account.get("account_id", "N/A"),  # The actual ad account ID
                "fee_percentage": account.get("fee_percentage", 0)  # CRITICAL FIX: Add fee_percentage
            },
            "currency": req["currency"],
            "amount": req["amount"],
            "fee": req.get("fee", 0),  # CRITICAL FIX: Add fee field
            "total": req.get("total", req["amount"]),  # CRITICAL FIX: Add total field
            "status": req["status"],
            "notes": req.get("notes"),
            "admin_notes": req.get("admin_notes", ""),
            "verified_by": verified_by_admin,
            "verified_at": req.get("verified_at"),
            "claimed_by": req.get("claimed_by"),
            "claimed_by_username": req.get("claimed_by_username"),
            "claimed_at": req.get("claimed_at"),
            "spend_limit_proof_url": req.get("spend_limit_proof_url"),
            "budget_aspire_proof_url": req.get("budget_aspire_proof_url"),
            "spend_limit_proof_pending_edit": spend_limit_pending is not None,
            "budget_aspire_proof_pending_edit": budget_aspire_pending is not None,
            "created_at": req["created_at"],
            "processed_at": req.get("processed_at")
        })
    
    return result

class WalletTopUpStatusUpdate(BaseModel):
    status: str  # verified, rejected
    admin_notes: Optional[str] = None

@api_router.put("/admin/wallet-topup-requests/{request_id}/status", response_model=dict)
async def update_wallet_topup_status(
    request_id: str,
    update_data: WalletTopUpStatusUpdate,
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Update wallet top-up request status"""
    if update_data.status not in ["verified", "rejected"]:
        raise HTTPException(status_code=400, detail="Invalid status. Must be 'verified' or 'rejected'")
    
    # Get the request
    wallet_request = await db.wallet_topup_requests.find_one({"id": request_id})
    if not wallet_request:
        raise HTTPException(status_code=404, detail="Wallet top-up request not found")
    
    # Update status - set timestamp for both verified and rejected
    update_fields = {
        "status": update_data.status,
        "verified_by": current_admin.id,
        "admin_notes": update_data.admin_notes or "",
        "verified_at": datetime.now(timezone.utc).isoformat()  # Set timestamp for both verified and rejected
    }
    
    # If verified, update user wallet balance
    if update_data.status == "verified":
        user_id = wallet_request["user_id"]
        currency_original = wallet_request["currency"]  # Keep original uppercase (IDR/USD)
        currency_lower = currency_original.lower()  # For wallet field names
        amount = wallet_request["amount"]
        wallet_type = wallet_request["wallet_type"]
        
        # Update user wallet balance
        wallet_field = f"{wallet_type}_wallet_{currency_lower}"  # main_wallet_idr or withdrawal_wallet_usd
        
        await db.users.update_one(
            {"id": user_id},
            {"$inc": {wallet_field: amount}}
        )
        
        # Get user for notification
        user = await db.users.find_one({"id": user_id})
        
        # Create transaction record for wallet top-up
        transaction = {
            "id": str(uuid.uuid4()),
            "user_id": user_id,
            "type": "wallet_topup",
            "amount": amount,
            "currency": currency_original,  # Use uppercase currency (IDR/USD)
            "status": "completed",
            "description": f"Wallet Top-Up - {wallet_type.title()} Wallet",
            "reference_id": request_id,
            "account_name": f"{wallet_type.title()} Wallet",
            "platform": "wallet",
            "admin_notes": update_data.admin_notes or "",
            "created_at": datetime.now(timezone.utc)
        }
        
        transaction_dict = prepare_for_mongo(transaction)
        await db.transactions.insert_one(transaction_dict)
        
        # Create notification for client
        notification = {
            "id": str(uuid.uuid4()),
            "user_id": user_id,
            "title": "✅ Wallet Top-Up Berhasil",
            "message": f"Top-up {wallet_request['currency']} {amount:,.2f} ke {wallet_type} wallet telah berhasil diverifikasi.",
            "type": "wallet_topup_success",
            "reference_id": request_id,
            "is_read": False,
            "created_at": datetime.now(timezone.utc)
        }
        
        notification_dict = prepare_for_mongo(notification)
        await db.client_notifications.insert_one(notification_dict)
        
        # Send email notification to client about wallet top-up approval
        try:
            user = await db.users.find_one({"id": user_id})
            if user and user.get("email"):
                send_client_wallet_topup_approved_email(
                    client_email=user["email"],
                    client_name=user.get("full_name") or user.get("username"),
                    amount=amount,
                    currency=wallet_request["currency"],
                    wallet_type=wallet_type
                )
                logger.info(f"📧 Wallet top-up approved email sent to {user['email']}")
        except Exception as e:
            logger.error(f"Failed to send wallet top-up approved email: {e}")
    
    elif update_data.status == "rejected":
        # Get user for notification
        user = await db.users.find_one({"id": wallet_request["user_id"]})
        
        # Create rejection notification for client
        notification = {
            "id": str(uuid.uuid4()),
            "user_id": wallet_request["user_id"],
            "title": "❌ Wallet Top-Up Ditolak",
            "message": f"Top-up {wallet_request['currency']} {wallet_request['amount']:,.2f} ditolak. {update_data.admin_notes}",
            "type": "wallet_topup_rejected",
            "reference_id": request_id,
            "is_read": False,
            "created_at": datetime.now(timezone.utc)
        }
        
        notification_dict = prepare_for_mongo(notification)
        await db.client_notifications.insert_one(notification_dict)
        
        # Send email notification to client about wallet top-up rejection
        try:
            if user and user.get("email"):
                send_client_wallet_topup_rejected_email(
                    client_email=user["email"],
                    client_name=user.get("full_name") or user.get("username"),
                    amount=wallet_request["amount"],
                    currency=wallet_request["currency"],
                    wallet_type=wallet_request["wallet_type"],
                    reason=update_data.admin_notes or ""
                )
                logger.info(f"📧 Wallet top-up rejected email sent to {user['email']}")
        except Exception as e:
            logger.error(f"Failed to send wallet top-up rejected email: {e}")
        
        # Create transaction record for rejected wallet top-up
        user_id = wallet_request["user_id"]
        currency_original = wallet_request["currency"]
        amount = wallet_request["amount"]
        wallet_type = wallet_request["wallet_type"]
        
        transaction = {
            "id": str(uuid.uuid4()),
            "user_id": user_id,
            "type": "wallet_topup",
            "amount": amount,
            "currency": currency_original,
            "status": "rejected",
            "description": f"Wallet Top-Up - {wallet_type.title()} Wallet (Ditolak)",
            "reference_id": request_id,
            "account_name": f"{wallet_type.title()} Wallet",
            "platform": "wallet",
            "admin_notes": update_data.admin_notes or "",
            "created_at": datetime.now(timezone.utc)
        }
        
        transaction_dict = prepare_for_mongo(transaction)
        await db.transactions.insert_one(transaction_dict)
    
    await db.wallet_topup_requests.update_one(
        {"id": request_id},
        {"$set": update_fields}
    )
    
    # Send email notification to client
    try:
        user = await db.users.find_one({"id": wallet_request["user_id"]})
        if user and user.get("email"):
            from email_service import send_client_topup_approved_email, send_client_topup_rejected_email
            
            wallet_type = wallet_request["wallet_type"]
            account_display = f"{wallet_type.capitalize()} Wallet"
            
            if update_data.status == "verified":
                send_client_topup_approved_email(
                    client_email=user["email"],
                    client_name=user.get("name") or user.get("username"),
                    amount=wallet_request["amount"],
                    currency=wallet_request["currency"],
                    account_name=account_display,
                    admin_notes=update_data.admin_notes or ""
                )
                logger.info(f"📧 Wallet top-up approved email sent to {user['email']}")
            else:
                send_client_topup_rejected_email(
                    client_email=user["email"],
                    client_name=user.get("name") or user.get("username"),
                    amount=wallet_request["amount"],
                    currency=wallet_request["currency"],
                    account_name=account_display,
                    reason=update_data.admin_notes or "Mohon hubungi admin untuk informasi lebih lanjut"
                )
                logger.info(f"📧 Wallet top-up rejected email sent to {user['email']}")
    except Exception as e:
        logger.error(f"❌ Failed to send wallet top-up email: {e}")
    
    return {"message": f"Wallet top-up request {update_data.status} successfully"}

@api_router.get("/admin/wallet-topup-requests/{request_id}/proof-file")
async def get_wallet_topup_proof_file(
    request_id: str,
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get wallet top-up payment proof file"""
    wallet_request = await db.wallet_topup_requests.find_one({"id": request_id})
    if not wallet_request:
        raise HTTPException(status_code=404, detail="Wallet request not found")
    
    payment_proof_id = wallet_request.get("payment_proof_id")
    if not payment_proof_id:
        raise HTTPException(status_code=404, detail="No payment proof found")
    
    proof = await db.payment_proofs.find_one({"id": payment_proof_id})
    if not proof:
        raise HTTPException(status_code=404, detail="Payment proof not found")
    
    # Handle both relative and absolute paths
    file_path = Path(proof["file_path"])
    if not file_path.is_absolute():
        file_path = Path("/app") / file_path
    
    if not file_path.exists():
        logger.error(f"Wallet payment proof not found at: {file_path}")
        raise HTTPException(status_code=404, detail="Proof file not found")
    
    return FileResponse(
        path=file_path,
        media_type=proof.get("mime_type", "image/jpeg"),
        headers={
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "GET",
            "Access-Control-Allow-Headers": "Authorization"
        }
    )

class WalletTransferStatusUpdate(BaseModel):
    status: str  # approved, rejected  
    admin_notes: Optional[str] = None
    spend_limit_proof_path: Optional[str] = None
    budget_aspire_proof_path: Optional[str] = None

@api_router.get("/admin/wallet-transfers/{transfer_id}/proof/{proof_type}")
async def get_wallet_transfer_proof(
    transfer_id: str,
    proof_type: str,  # "spend_limit" or "budget_aspire"
):
    """Get wallet transfer verification proof files from GCS - PUBLIC for admin use (no auth to avoid CORS)"""
    try:
        # Get transfer request
        transfer = await db.wallet_transfers.find_one({"id": transfer_id})
        if not transfer:
            raise HTTPException(status_code=404, detail="Transfer request not found")
        
        # Get the proof URL based on type
        if proof_type == "spend_limit":
            proof_url = transfer.get("spend_limit_proof_url")
        elif proof_type == "budget_aspire":
            proof_url = transfer.get("budget_aspire_proof_url")
        else:
            raise HTTPException(status_code=400, detail="Invalid proof type. Use 'spend_limit' or 'budget_aspire'")
        
        if not proof_url:
            raise HTTPException(status_code=404, detail=f"{proof_type} proof not found")
        
        # Extract GCS path from proof_url (format: /files/wallet_transfer_proofs/filename.jpg)
        # Remove /files/ prefix to get GCS path
        if proof_url.startswith("/files/"):
            gcs_path = proof_url[7:]  # Remove "/files/" prefix
        else:
            gcs_path = proof_url
        
        try:
            # Try to download from GCS first
            content, mime_type = await download_from_gcs(gcs_path)
            logger.info(f"✅ Served wallet transfer proof from GCS: {gcs_path}")
            return Response(
                content=content,
                media_type=mime_type,
                headers={
                    "Access-Control-Allow-Origin": "*",
                    "Access-Control-Allow-Methods": "GET, OPTIONS",
                    "Access-Control-Allow-Headers": "*",
                    "Cache-Control": "public, max-age=3600"
                }
            )
        except HTTPException:
            # If not in GCS, try filesystem fallback
            logger.warning(f"⚠️ File not in GCS, trying filesystem: {gcs_path}")
            
            file_path = Path(f"/app/{proof_url}") if not proof_url.startswith("/") else Path(proof_url)
            
            if not file_path.exists():
                # Try without /files/ prefix
                file_path = Path(f"/app/uploads/{gcs_path}")
            
            if not file_path.exists():
                logger.error(f"❌ Wallet transfer proof not found in GCS or filesystem: {proof_url}")
                raise HTTPException(status_code=404, detail="Proof file not found")
            
            # Determine content type
            file_name = file_path.name
            content_type = "image/jpeg"
            if file_name.lower().endswith('.png'):
                content_type = "image/png"
            elif file_name.lower().endswith('.pdf'):
                content_type = "application/pdf"
            elif file_name.lower().endswith('.webp'):
                content_type = "image/webp"
            
            logger.info(f"✅ Served wallet transfer proof from filesystem: {file_path}")
            return FileResponse(
                path=file_path,
                media_type=content_type,
                headers={
                    "Access-Control-Allow-Origin": "*",
                    "Access-Control-Allow-Methods": "GET, OPTIONS",
                    "Access-Control-Allow-Headers": "*",
                    "Cache-Control": "public, max-age=3600"
                }
            )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error serving wallet transfer proof: {e}")
        raise HTTPException(status_code=500, detail="Error serving file")

@api_router.delete("/admin/cleanup-wallet-transfer-transactions", response_model=dict)
async def cleanup_wallet_transfer_transactions(current_admin: AdminUser = Depends(get_current_admin)):
    """Delete all wallet_to_account_transfer transactions for clean testing"""
    try:
        # Delete all wallet_to_account_transfer transactions
        result = await db.transactions.delete_many({"type": "wallet_to_account_transfer"})
        
        # Delete pending wallet_transfers
        pending_result = await db.wallet_transfers.delete_many({"status": "pending"})
        
        return {
            "success": True,
            "message": "Database cleaned successfully",
            "transactions_deleted": result.deleted_count,
            "pending_requests_deleted": pending_result.deleted_count
        }
    except Exception as e:
        logger.error(f"Error cleaning database: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to clean database: {str(e)}")

print("=== REGISTERING ENDPOINT: PUT /admin/wallet-transfer-requests/{request_id}/status ===")

@api_router.put("/admin/wallet-transfer-requests/{request_id}/status", response_model=dict)
async def update_wallet_transfer_status(
    request_id: str,
    update_data: WalletTransferStatusUpdate,
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Update wallet-to-account transfer request status"""
    try:
        print(f"=== FUNCTION CALLED: update_wallet_transfer_status, request_id={request_id}, status={update_data.status} ===")
        logger.info(f"=== FUNCTION CALLED: update_wallet_transfer_status, request_id={request_id}, status={update_data.status} ===")
        logger.info(f"[update_wallet_transfer_status] Admin user: {current_admin.username if current_admin else 'None'}")
        logger.info(f"[update_wallet_transfer_status] Request data: status={update_data.status}, admin_notes={update_data.admin_notes}")
        
        if update_data.status not in ["approved", "rejected"]:
            raise HTTPException(status_code=400, detail="Invalid status. Must be 'approved' or 'rejected'")
        
        # Get the transfer request
        transfer_request = await db.wallet_transfers.find_one({"id": request_id})
        if not transfer_request:
            raise HTTPException(status_code=404, detail="Wallet transfer request not found")
        
        # Update status
        now = datetime.now(timezone.utc)
        update_fields = {
            "status": update_data.status,
            "verified_by": current_admin.id,
            "verified_at": now.isoformat(),
            "admin_notes": update_data.admin_notes or "",
            "processed_at": now.isoformat()
        }
        
        # Add file paths if provided
        if update_data.spend_limit_proof_path:
            update_fields["spend_limit_proof_url"] = update_data.spend_limit_proof_path
        if update_data.budget_aspire_proof_path:
            update_fields["budget_aspire_proof_url"] = update_data.budget_aspire_proof_path
        
        logger.info(f"UPDATE WALLET TRANSFER STATUS: request_id={request_id}, status={update_data.status}")
        
        if update_data.status == "approved":
            # Get user and account for notification
            user = await db.users.find_one({"id": transfer_request["user_id"]})
            
            # Deduct from wallet NOW (when approved)
            wallet_field = f"{transfer_request['source_wallet_type']}_wallet_{transfer_request['currency'].lower()}"
            current_wallet_balance = user.get(wallet_field, 0)
            
            # Calculate total deduction: amount + fee
            transfer_amount = transfer_request["amount"]
            transfer_fee = transfer_request.get("fee", 0)
            
            # If fee not in field, calculate from total
            if transfer_fee == 0 and "total" in transfer_request:
                transfer_fee = transfer_request["total"] - transfer_amount
            
            deduct_amount = transfer_amount + transfer_fee
            
            logger.info(f"Approving transfer {request_id}: amount={transfer_amount}, fee={transfer_fee}, total_deduct={deduct_amount}")
            
            if current_wallet_balance < deduct_amount:
                raise HTTPException(
                    status_code=400,
                    detail=f"Saldo wallet tidak mencukupi saat approve. Dibutuhkan: {deduct_amount:,.2f}, Tersedia: {current_wallet_balance:,.2f}"
                )
            
            new_wallet_balance = current_wallet_balance - deduct_amount
            await db.users.update_one(
                {"id": transfer_request["user_id"]},
                {"$set": {wallet_field: new_wallet_balance}}
            )
            
            logger.info(f"Wallet deducted: {current_wallet_balance} -> {new_wallet_balance} (deducted {deduct_amount})")
            
            # Update target account balance
            target_account = await db.ad_accounts.find_one({
                "id": transfer_request["target_account_id"],
                "user_id": transfer_request["user_id"]
            })
            
            if target_account:
                current_account_balance = target_account.get("balance", 0)
                new_account_balance = current_account_balance + transfer_request["amount"]
                
                await db.ad_accounts.update_one(
                    {"id": transfer_request["target_account_id"]},
                    {"$set": {"balance": new_account_balance}}
                )
                
                # Try to find existing pending transaction first
                existing_transaction = await db.transactions.find_one({"reference_id": request_id})
                logger.info(f"Looking for transaction with reference_id={request_id}: {'FOUND' if existing_transaction else 'NOT FOUND'}")
                
                if existing_transaction:
                    logger.info(f"Found transaction: id={existing_transaction.get('id')}, status={existing_transaction.get('status')}, type={existing_transaction.get('type')}")
                
                # Update existing transaction record status to completed
                update_result = await db.transactions.update_one(
                    {"reference_id": request_id},
                    {"$set": {
                        "status": "completed",
                        "description": f"Transfer dari {transfer_request['source_wallet_type']} wallet ke akun {transfer_request['target_account_name']} (Disetujui)",
                        "admin_notes": update_data.admin_notes or "",
                        "updated_at": datetime.now(timezone.utc)
                    }}
                )
                
                logger.info(f"Update by reference_id matched_count: {update_result.matched_count}")
                
                # If no transaction found by reference_id, try alternative search
                if update_result.matched_count == 0:
                    logger.warning(f"Transaction not found by reference_id: {request_id}, trying alternative search")
                    
                    # Search for pending transaction
                    alt_search = {
                        "user_id": transfer_request["user_id"],
                        "type": "wallet_to_account_transfer",
                        "account_id": transfer_request["target_account_id"],
                        "amount": transfer_request["amount"],
                        "status": "pending"
                    }
                    logger.info(f"Alternative search criteria: {alt_search}")
                    
                    pending_tx = await db.transactions.find_one(alt_search)
                    logger.info(f"Alternative search result: {'FOUND' if pending_tx else 'NOT FOUND'}")
                    
                    if pending_tx:
                        logger.info(f"Found pending tx: id={pending_tx.get('id')}, reference_id={pending_tx.get('reference_id')}")
                    
                    update_result = await db.transactions.update_one(
                        alt_search,
                        {"$set": {
                            "status": "completed",
                            "description": f"Transfer dari {transfer_request['source_wallet_type']} wallet ke akun {transfer_request['target_account_name']} (Disetujui)",
                            "reference_id": request_id,  # Add reference_id for future tracking
                            "admin_notes": update_data.admin_notes or "",
                            "updated_at": datetime.now(timezone.utc)
                        }}
                    )
                    logger.info(f"Alternative update matched_count: {update_result.matched_count}")
                
                # If still no transaction found, create one
                if update_result.matched_count == 0:
                    logger.warning(f"No transaction found for reference_id: {request_id}, creating new one")
                    transaction = {
                        "id": str(uuid.uuid4()),
                        "user_id": transfer_request["user_id"],
                        "type": "wallet_to_account_transfer",
                        "amount": transfer_request["amount"],
                        "currency": transfer_request["currency"],
                        "status": "completed",
                        "description": f"Transfer dari {transfer_request['source_wallet_type']} wallet ke akun {transfer_request['target_account_name']} (Disetujui)",
                        "account_id": transfer_request["target_account_id"],
                        "account_name": transfer_request["target_account_name"],
                        "reference_id": request_id,
                        "created_at": datetime.now(timezone.utc),
                        "updated_at": datetime.now(timezone.utc)
                    }
                    transaction_dict = prepare_for_mongo(transaction)
                    await db.transactions.insert_one(transaction_dict)
                    logger.info(f"Created new transaction for approved transfer: {request_id}")
            
            # Create success notification for client
            notification = {
                "id": str(uuid.uuid4()),
                "user_id": transfer_request["user_id"],
                "title": "✅ Transfer Wallet Berhasil",
                "message": f"Transfer {transfer_request['currency']} {transfer_request['amount']:,.2f} dari {transfer_request['source_wallet_type']} wallet ke akun {transfer_request['target_account_name']} telah disetujui.",
                "type": "wallet_transfer_success",
                "reference_id": request_id,
                "is_read": False,
                "created_at": datetime.now(timezone.utc)
            }
            
            notification_dict = prepare_for_mongo(notification)
            await db.client_notifications.insert_one(notification_dict)
            
            # Send email notification to client about wallet transfer approval
            try:
                user = await db.users.find_one({"id": transfer_request["user_id"]})
                if user and user.get("email"):
                    from_wallet = f"{transfer_request['source_wallet_type'].title()} Wallet"
                    to_account = transfer_request['target_account_name']
                    send_client_wallet_transfer_approved_email(
                        client_email=user["email"],
                        client_name=user.get("full_name") or user.get("username"),
                        amount=transfer_request["amount"],
                        currency=transfer_request["currency"],
                        from_wallet=from_wallet,
                        to_account=to_account
                    )
                    logger.info(f"📧 Wallet transfer approval email sent to {user['email']}")
            except Exception as e:
                logger.error(f"Failed to send wallet transfer approval email: {e}")
        
        elif update_data.status == "rejected":
            # NOTE: No need to refund wallet because it was not deducted on submit
            # Wallet is only deducted when transfer is approved
            
            # Update existing transaction record status to rejected
            update_result = await db.transactions.update_one(
                {"reference_id": request_id, "type": "wallet_to_account_transfer"},
                {"$set": {
                    "status": "rejected",
                    "description": f"Transfer dari {transfer_request['source_wallet_type']} wallet ke akun {transfer_request['target_account_name']} (Ditolak)",
                    "admin_notes": update_data.admin_notes or "",
                    "updated_at": datetime.now(timezone.utc)
                }}
            )
            
            # If no transaction found by reference_id, try alternative search
            if update_result.matched_count == 0:
                logger.warning(f"Transaction not found by reference_id: {request_id}, trying alternative search")
                update_result = await db.transactions.update_one(
                    {
                        "user_id": transfer_request["user_id"],
                        "type": "wallet_to_account_transfer",
                        "account_id": transfer_request["target_account_id"],
                        "amount": transfer_request["amount"],
                        "status": "pending"
                    },
                    {"$set": {
                        "status": "rejected",
                        "description": f"Transfer dari {transfer_request['source_wallet_type']} wallet ke akun {transfer_request['target_account_name']} (Ditolak)",
                        "admin_notes": update_data.admin_notes or "",
                        "reference_id": request_id,
                        "updated_at": datetime.now(timezone.utc)
                    }}
                )
                if update_result.matched_count > 0:
                    logger.info(f"Transaction updated using alternative search")
            
            # If still no transaction found, create one
            if update_result.matched_count == 0:
                logger.warning(f"No transaction found for reference_id: {request_id}, creating new one")
                transaction = {
                    "id": str(uuid.uuid4()),
                    "user_id": transfer_request["user_id"],
                    "type": "wallet_to_account_transfer",
                    "amount": transfer_request["amount"],
                    "currency": transfer_request["currency"],
                    "status": "rejected",
                    "description": f"Transfer dari {transfer_request['source_wallet_type']} wallet ke akun {transfer_request['target_account_name']} (Ditolak)",
                    "admin_notes": update_data.admin_notes or "",
                    "account_id": transfer_request["target_account_id"],
                    "account_name": transfer_request["target_account_name"],
                    "reference_id": request_id,
                    "created_at": datetime.now(timezone.utc),
                    "updated_at": datetime.now(timezone.utc)
                }
                transaction_dict = prepare_for_mongo(transaction)
                await db.transactions.insert_one(transaction_dict)
                logger.info(f"Created new transaction for rejected transfer: {request_id}")
            
            # Create rejection notification for client
            notification = {
                "id": str(uuid.uuid4()),
                "user_id": transfer_request["user_id"],
                "title": "❌ Transfer Wallet Ditolak",
                "message": f"Transfer {transfer_request['currency']} {transfer_request['amount']:,.2f} dari {transfer_request['source_wallet_type']} wallet ditolak. Saldo dikembalikan. {update_data.admin_notes}",
                "type": "wallet_transfer_rejected",
                "reference_id": request_id,
                "is_read": False,
                "created_at": datetime.now(timezone.utc)
            }
            
            notification_dict = prepare_for_mongo(notification)
            await db.client_notifications.insert_one(notification_dict)
            
            # Send email notification to client about wallet transfer rejection
            try:
                user = await db.users.find_one({"id": transfer_request["user_id"]})
                if user and user.get("email"):
                    from_wallet = f"{transfer_request['source_wallet_type'].title()} Wallet"
                    to_account = transfer_request['target_account_name']
                    send_client_wallet_transfer_rejected_email(
                        client_email=user["email"],
                        client_name=user.get("full_name") or user.get("username"),
                        amount=transfer_request["amount"],
                        currency=transfer_request["currency"],
                        from_wallet=from_wallet,
                        to_account=to_account,
                        reason=update_data.admin_notes or ""
                    )
                    logger.info(f"📧 Wallet transfer rejection email sent to {user['email']}")
            except Exception as e:
                logger.error(f"Failed to send wallet transfer rejection email: {e}")
        
        await db.wallet_transfers.update_one(
            {"id": request_id},
            {"$set": update_fields}
        )
        
        logger.info(f"[update_wallet_transfer_status] Successfully completed for request_id={request_id}")
        return {"message": f"Wallet transfer request {update_data.status} successfully"}
    
    except HTTPException as he:
        logger.error(f"[update_wallet_transfer_status] HTTP Exception: {he.detail}")
        raise
    except Exception as e:
        logger.error(f"[update_wallet_transfer_status] Unexpected error: {str(e)}")
        logger.error(f"[update_wallet_transfer_status] Error type: {type(e)}")
        import traceback
        logger.error(f"[update_wallet_transfer_status] Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Failed to update wallet transfer status: {str(e)}")

@api_router.post("/admin/wallet-transfers/{transfer_id}/upload-verification-files")
async def upload_wallet_transfer_verification_files(
    transfer_id: str,
    file: UploadFile = File(...),
    type: str = Form(...)  # "spend_limit_proof" or "budget_aspire_proof"
):
    """Upload verification files for wallet transfer approval - NO AUTH for stability"""
    try:
        # Get the wallet transfer to ensure it exists
        wallet_transfer = await db.wallet_transfers.find_one({"id": transfer_id})
        if not wallet_transfer:
            raise HTTPException(status_code=404, detail="Wallet transfer not found")
        
        # Validate file type
        if file.content_type not in ['image/jpeg', 'image/png', 'image/jpg', 'application/pdf']:
            raise HTTPException(status_code=400, detail="Invalid file type. Only JPEG, PNG, and PDF files are allowed")
        
        # Create directories if they don't exist - use absolute path
        upload_dir = Path("/app/uploads/verification_files")
        upload_dir.mkdir(parents=True, exist_ok=True)
        
        # Generate unique filename
        file_extension = Path(file.filename).suffix
        unique_filename = f"{transfer_id}_{type}_{uuid.uuid4().hex[:8]}{file_extension}"
        file_path = upload_dir / unique_filename
        
        # Save file
        with open(file_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)
        
        # Return file path for storage in database - use relative path for consistency
        relative_path = f"uploads/verification_files/{unique_filename}"
        
        # CRITICAL FIX: UPDATE DATABASE WITH FILE PATH IMMEDIATELY
        update_field = {}
        if type == "spend_limit_proof":
            update_field["spend_limit_proof_url"] = relative_path
        elif type == "budget_aspire_proof":
            update_field["budget_aspire_proof_url"] = relative_path
        
        await db.wallet_transfers.update_one(
            {"id": transfer_id},
            {"$set": update_field}
        )
        
        logger.info(f"Uploaded and saved {type} for wallet_transfer {transfer_id}: {relative_path}")
        
        return {
            "success": True,
            "file_path": relative_path,
            "original_filename": file.filename,
            "type": type
        }
        
    except Exception as e:
        logger.error(f"Error uploading wallet transfer verification file: {e}")
        raise HTTPException(status_code=500, detail="Failed to upload verification file")

@api_router.post("/admin/wallet-transfers/{transfer_id}/edit-proof")
async def edit_wallet_transfer_proof(
    transfer_id: str,
    file: UploadFile = File(...),
    transfer_id_form: str = Form(..., alias="transfer_id"),
    target_account_id: str = Form(...),
    proof_type: str = Form(...),  # "spend_limit" or "budget_aspire"
    notes: str = Form(""),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Admin submits edit request for wallet transfer proof (requires Super Admin approval)"""
    try:
        # Get wallet transfer
        transfer = await db.wallet_transfers.find_one({"id": transfer_id})
        if not transfer:
            raise HTTPException(status_code=404, detail="Wallet transfer not found")
        
        # Upload new file to GCS
        gcs = get_gcs_storage()
        file_content = await file.read()
        file_extension = Path(file.filename).suffix or '.jpg'
        unique_filename = f"{transfer_id}_{proof_type}_{uuid.uuid4().hex[:8]}{file_extension}"
        gcs_path = f"payment_proofs/pending_edits/{unique_filename}"
        
        # Convert bytes to file-like object for GCS upload
        file_obj = BytesIO(file_content)
        gcs.upload_file(file_obj, gcs_path, content_type=file.content_type)
        logger.info(f"📤 Uploaded new proof to GCS: {gcs_path}")
        
        # Create/update payment_proofs document with pending edit
        proof_field = "spend_limit_proof_url" if proof_type == "spend_limit" else "budget_aspire_proof_url"
        current_proof_url = transfer.get(proof_field, "")
        logger.info(f"📝 Current proof URL from wallet_transfer: {current_proof_url}")
        
        # Generate tracking_id for this proof
        tracking_id = f"proof_tracking_{transfer_id}_{target_account_id}_{proof_type}"
        
        # Check if payment_proofs document exists
        existing_proof = await db.payment_proofs.find_one({"tracking_id": tracking_id})
        
        if existing_proof:
            # Update existing document with pending edit
            # Preserve existing gcs_path or use current_proof_url if not set
            existing_gcs_path = existing_proof.get("gcs_path", current_proof_url)
            
            await db.payment_proofs.update_one(
                {"tracking_id": tracking_id},
                {"$set": {
                    "pending_edit": True,
                    "gcs_path": existing_gcs_path,  # Ensure current proof path is preserved
                    "file_url": f"/files/{existing_gcs_path}" if existing_gcs_path else "",
                    "new_gcs_path": gcs_path,
                    "new_file_url": f"/files/{gcs_path}",
                    "edit_requested_by": current_admin.id,
                    "edit_requested_by_username": current_admin.username,
                    "edit_requested_at": datetime.now(timezone.utc).isoformat(),
                    "edit_notes": notes,
                    "transfer_id": transfer_id,
                    "target_account_id": target_account_id,
                    "proof_type": proof_type
                }}
            )
            proof_id = existing_proof["id"]
            logger.info(f"✅ Updated existing payment_proofs with pending edit: {proof_id}")
        else:
            # Create new payment_proofs document
            proof_id = str(uuid.uuid4())
            proof_doc = {
                "id": proof_id,
                "tracking_id": tracking_id,
                "transfer_id": transfer_id,
                "target_account_id": target_account_id,
                "proof_type": proof_type,
                "gcs_path": current_proof_url,
                "file_url": f"/files/{current_proof_url}" if current_proof_url else "",
                "pending_edit": True,
                "new_gcs_path": gcs_path,
                "new_file_url": f"/files/{gcs_path}",
                "edit_requested_by": current_admin.id,
                "edit_requested_by_username": current_admin.username,
                "edit_requested_at": datetime.now(timezone.utc).isoformat(),
                "edit_notes": notes,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.payment_proofs.insert_one(proof_doc)
            logger.info(f"✅ Created new payment_proofs with pending edit: {proof_id}")
        
        # Notify super admins
        super_admins = await db.admin_users.find({"role": "super_admin"}).to_list(None)
        for super_admin in super_admins:
            notification = {
                "id": str(uuid.uuid4()),
                "title": "🔄 New Proof Edit Request",
                "message": f"Admin {current_admin.username} requested to edit {proof_type} proof for wallet transfer",
                "type": "proof_edit_request",
                "reference_id": proof_id,
                "is_read": False,
                "created_at": datetime.now(timezone.utc)
            }
            notification_dict = prepare_for_mongo(notification)
            await db.notifications.insert_one(notification_dict)
        
        logger.info(f"✅ Admin {current_admin.username} submitted proof edit for transfer {transfer_id}")
        
        return {
            "success": True,
            "proof_id": proof_id,
            "proof_type": proof_type,
            "message": "Proof edit request submitted successfully"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Error submitting proof edit: {e}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail="Error submitting proof edit")

@api_router.post("/admin/rebuild-wallet-transactions")
async def rebuild_wallet_transactions(current_admin: AdminUser = Depends(get_current_admin)):
    """DELETE all wallet_to_account_transfer transactions and REBUILD from wallet_transfers"""
    try:
        # Step 1: DELETE all wallet_to_account_transfer transactions
        delete_result = await db.transactions.delete_many({
            "type": "wallet_to_account_transfer"
        })
        deleted_count = delete_result.deleted_count
        logger.info(f"Deleted {deleted_count} wallet_to_account_transfer transactions")
        
        # Step 2: REBUILD from wallet_transfers (source of truth)
        wallet_transfers = await db.wallet_transfers.find().to_list(10000)
        created_count = 0
        
        for wt in wallet_transfers:
            wt = parse_from_mongo(wt)
            
            # Determine transaction status based on wallet_transfer status
            if wt['status'] == 'approved':
                tx_status = 'completed'
                description = f"Transfer dari {wt['source_wallet_type']} wallet ke akun {wt['target_account_name']} (Disetujui)"
            elif wt['status'] == 'rejected':
                tx_status = 'rejected'
                description = f"Transfer dari {wt['source_wallet_type']} wallet ke akun {wt['target_account_name']} (Ditolak)"
            else:  # pending
                tx_status = 'pending'
                description = f"Transfer dari {wt['source_wallet_type']} wallet ke akun {wt['target_account_name']} (Menunggu Verifikasi Admin)"
            
            # Calculate total (amount + fee)
            amount = wt.get('amount', 0)
            fee = wt.get('fee', 0)
            total = wt.get('total', amount + fee)
            
            # Create transaction
            transaction = {
                "id": str(uuid.uuid4()),
                "user_id": wt['user_id'],
                "type": "wallet_to_account_transfer",
                "amount": total,  # Total = amount + fee (what was deducted from wallet)
                "currency": wt['currency'],
                "status": tx_status,
                "description": description,
                "account_id": wt.get('target_account_id'),
                "account_name": wt.get('target_account_name'),
                "reference_id": wt['id'],
                "fee": fee,
                "created_at": wt.get('created_at', datetime.now(timezone.utc)),
                "updated_at": wt.get('processed_at', wt.get('created_at', datetime.now(timezone.utc)))
            }
            
            transaction_dict = prepare_for_mongo(transaction)
            await db.transactions.insert_one(transaction_dict)
            created_count += 1
        
        logger.info(f"Created {created_count} transactions from wallet_transfers")
        
        return {
            "success": True,
            "deleted": deleted_count,
            "created": created_count,
            "message": f"Deleted {deleted_count} old transactions, created {created_count} new transactions from wallet_transfers"
        }
        
    except Exception as e:
        logger.error(f"Error rebuilding transactions: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/uploads/{folder}/{filename}")
async def serve_uploaded_file(
    folder: str,
    filename: str
):
    """Serve uploaded files (payment proofs, verification files, etc.)"""
    try:
        # Construct file path using absolute path
        file_path = Path(f"/app/uploads/{folder}/{filename}")
        
        # Check if file exists
        if not file_path.exists():
            raise HTTPException(status_code=404, detail="File not found")
        
        # Determine media type based on file extension
        extension = file_path.suffix.lower()
        media_type_map = {
            '.jpg': 'image/jpeg',
            '.jpeg': 'image/jpeg',
            '.png': 'image/png',
            '.webp': 'image/webp',
            '.pdf': 'application/pdf',
            '.gif': 'image/gif'
        }
        media_type = media_type_map.get(extension, 'application/octet-stream')
        
        return FileResponse(
            path=str(file_path),
            media_type=media_type,
            filename=filename
        )
        
    except Exception as e:
        logger.error(f"Error serving file: {e}")
        raise HTTPException(status_code=500, detail="Failed to serve file")

# Balance Transfer endpoints
@api_router.post("/balance-transfer", response_model=dict)
async def create_transfer_request_via_balance_transfer(
    transfer_data: BalanceTransfer,
    current_user: User = Depends(get_current_user)
):
    """Create transfer request (requires admin approval) - replaces direct transfer"""
    
    # Validate transfer types
    if transfer_data.from_type not in ["wallet", "account"] or transfer_data.to_type not in ["wallet", "account"]:
        raise HTTPException(status_code=400, detail="Invalid transfer type")
    
    if transfer_data.from_type == transfer_data.to_type == "account":
        raise HTTPException(status_code=400, detail="Cannot transfer between two accounts directly")
        
    if transfer_data.amount <= 0:
        raise HTTPException(status_code=400, detail="Transfer amount must be greater than 0")
    
    # Only support wallet to account transfer for now
    if not (transfer_data.from_type == "wallet" and transfer_data.to_type == "account"):
        raise HTTPException(status_code=400, detail="Only wallet to account transfers are supported")
    
    if not transfer_data.account_id:
        raise HTTPException(status_code=400, detail="Account ID required for account transfer")
    
    user = await db.users.find_one({"id": current_user.id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
        
    account = await db.ad_accounts.find_one({"id": transfer_data.account_id, "user_id": current_user.id})
    if not account:
        raise HTTPException(status_code=404, detail="Ad account not found")
    
    # Determine currency and wallet field
    currency = account.get("currency", "IDR")
    wallet_field = "wallet_balance_idr" if currency == "IDR" else "wallet_balance_usd"
    current_wallet_balance = user.get(wallet_field, 0)
    
    # Check wallet balance
    if current_wallet_balance < transfer_data.amount:
        raise HTTPException(status_code=400, detail="Insufficient wallet balance")
    
    # Create transfer request (pending admin approval)
    transfer_request = TransferRequestRecord(
        user_id=current_user.id,
        account_id=transfer_data.account_id,
        amount=transfer_data.amount,
        currency=currency
    )
    
    # Save transfer request to database
    await db.transfer_requests.insert_one(prepare_for_mongo(transfer_request.dict()))
    
    # Create notifications (simplified - no duplicates)
    currency_symbol = "Rp " if currency == "IDR" else "$"
    formatted_amount = f"{transfer_data.amount:,.0f}" if currency == "IDR" else f"{transfer_data.amount:.2f}"
    
    # Create notification for client
    client_notification = {
        "id": str(uuid.uuid4()),
        "user_id": current_user.id, 
        "title": "Transfer Request Dibuat",
        "message": f"Transfer request {currency_symbol}{formatted_amount} ke {account.get('account_name', 'akun')} telah dibuat dan menunggu approval admin.",
        "type": "transfer_created",
        "reference_id": transfer_request.id,
        "is_read": False,
        "created_at": datetime.now(timezone.utc)
    }
    
    client_notification_dict = prepare_for_mongo(client_notification)
    await db.client_notifications.insert_one(client_notification_dict)
    
    # Create only ONE admin notification (not per admin to avoid confusion)
    admin_notification = {
        "id": str(uuid.uuid4()),
        "title": "Transfer Request Baru",
        "message": f"Transfer request baru {currency_symbol}{formatted_amount} dari {current_user.username} ke {account.get('account_name', 'akun')} perlu diproses.",
        "type": "transfer_request",
        "reference_id": transfer_request.id,
        "is_read": False,
        "created_at": datetime.now(timezone.utc)
    }
    
    admin_notification_dict = prepare_for_mongo(admin_notification)
    await db.notifications.insert_one(admin_notification_dict)
    
    return {
        "message": "Transfer request created successfully. Waiting for admin approval.",
        "transfer_request_id": transfer_request.id,
        "amount": transfer_data.amount,
        "currency": currency,
        "account_name": account.get("account_name"),
        "status": "pending",
        "note": "Your transfer request has been submitted and is pending admin approval. You will be notified once it's processed."
    }

# Financial Reports endpoints
@api_router.get("/admin/financial-reports/summary")
async def get_financial_summary(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    period: Optional[str] = "all",  # all, today, week, month, year
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get financial summary for admin dashboard"""
    try:
        # Use GMT+7 (Asia/Jakarta) timezone
        jakarta_tz = ZoneInfo("Asia/Jakarta")
        
        # Parse date filters
        date_filter = {}
        if start_date and end_date:
            try:
                start_dt = datetime.fromisoformat(start_date).replace(hour=0, minute=0, second=0, microsecond=0, tzinfo=jakarta_tz)
                end_dt = datetime.fromisoformat(end_date).replace(hour=23, minute=59, second=59, microsecond=999999, tzinfo=jakarta_tz)
                # Convert to UTC for database comparison
                start_dt_utc = start_dt.astimezone(timezone.utc)
                end_dt_utc = end_dt.astimezone(timezone.utc)
                date_filter = {"created_at": {"$gte": start_dt_utc.isoformat(), "$lte": end_dt_utc.isoformat()}}
            except ValueError:
                pass
        elif period != "all":
            now = datetime.now(jakarta_tz)  # Use Jakarta time
            if period == "today":
                start_of_day = now.replace(hour=0, minute=0, second=0, microsecond=0)
                start_of_day_utc = start_of_day.astimezone(timezone.utc)
                date_filter = {"created_at": {"$gte": start_of_day_utc.isoformat()}}
            elif period == "yesterday":
                yesterday = now - timedelta(days=1)
                start_of_yesterday = yesterday.replace(hour=0, minute=0, second=0, microsecond=0)
                end_of_yesterday = yesterday.replace(hour=23, minute=59, second=59, microsecond=999999)
                start_of_yesterday_utc = start_of_yesterday.astimezone(timezone.utc)
                end_of_yesterday_utc = end_of_yesterday.astimezone(timezone.utc)
                date_filter = {"created_at": {"$gte": start_of_yesterday_utc.isoformat(), "$lte": end_of_yesterday_utc.isoformat()}}
            elif period == "week":
                start_of_week = now - timedelta(days=7)
                start_of_week_utc = start_of_week.astimezone(timezone.utc)
                date_filter = {"created_at": {"$gte": start_of_week_utc.isoformat()}}
            elif period == "month":
                start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                start_of_month_utc = start_of_month.astimezone(timezone.utc)
                date_filter = {"created_at": {"$gte": start_of_month_utc.isoformat()}}
            elif period == "year":
                start_of_year = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
                # Convert datetime object to ISO string for MongoDB comparison
                date_filter = {"created_at": {"$gte": start_of_year.isoformat()}}

        # ============================================
        # REVENUE CALCULATION (Fee-based transactions)
        # ============================================
        
        # 1. Revenue from Ad Account Top-Up (topup_requests with bank/crypto payment)
        # - IDR: bank transfer (bank_bca, bank_mandiri, etc.)
        # - USD: crypto (usdt_trc20, etc.)
        # Note: topup_requests use "verified" (admin approval), "approved", or "completed" status
        topup_valid_filter = {**date_filter, "status": {"$in": ["verified", "approved", "completed"]}}
        
        # Get IDR revenue from bank transfers
        # Note: Include empty/null payment_method as bank transfer for IDR (backward compatibility)
        topup_idr_revenue_pipeline = [
            {"$match": {
                **topup_valid_filter,
                "currency": "IDR",
                "$or": [
                    {"payment_method": {"$regex": "^bank_", "$options": "i"}},
                    {"payment_method": {"$in": [None, ""]}},  # Handle empty/null payment_method
                    {"payment_method": {"$exists": False}}
                ]
            }},
            {"$project": {
                "total_fee": 1,
                "total_amount": 1,
                "pure_amount": {"$subtract": ["$total_amount", "$total_fee"]}  # Amount without fee
            }},
            {"$group": {
                "_id": None,
                "total_fee": {"$sum": "$total_fee"},
                "pure_amount": {"$sum": "$pure_amount"},  # Sum of amounts without fees
                "count": {"$sum": 1}
            }}
        ]
        topup_idr_revenue = await db.topup_requests.aggregate(topup_idr_revenue_pipeline).to_list(1)
        topup_idr_fee = topup_idr_revenue[0]["total_fee"] if topup_idr_revenue else 0
        topup_idr_amount = topup_idr_revenue[0]["pure_amount"] if topup_idr_revenue else 0  # Pure amount without fee
        
        # Get USD revenue from crypto
        # Note: Include empty/null payment_method as crypto for USD (backward compatibility)
        topup_usd_revenue_pipeline = [
            {"$match": {
                **topup_valid_filter,
                "currency": "USD",
                "$or": [
                    {"payment_method": {"$regex": "usdt|crypto", "$options": "i"}},
                    {"payment_method": {"$in": [None, ""]}},  # Handle empty/null payment_method
                    {"payment_method": {"$exists": False}}
                ]
            }},
            {"$project": {
                "total_fee": 1,
                "total_amount": 1,
                "pure_amount": {"$subtract": ["$total_amount", "$total_fee"]}  # Amount without fee
            }},
            {"$group": {
                "_id": None,
                "total_fee": {"$sum": "$total_fee"},
                "pure_amount": {"$sum": "$pure_amount"},  # Sum of amounts without fees
                "count": {"$sum": 1}
            }}
        ]
        topup_usd_revenue = await db.topup_requests.aggregate(topup_usd_revenue_pipeline).to_list(1)
        topup_usd_fee = topup_usd_revenue[0]["total_fee"] if topup_usd_revenue else 0
        topup_usd_amount = topup_usd_revenue[0]["pure_amount"] if topup_usd_revenue else 0  # Pure amount without fee
        
        # 2. Revenue from Wallet to Account Transfer (wallet_transfers)
        # Note: wallet_transfers use "completed" (direct transfers) or "approved" (admin-managed transfers)
        wallet_transfer_valid_filter = {**date_filter, "status": {"$in": ["completed", "approved"]}}
        
        # Get IDR revenue from wallet transfers
        wallet_transfer_idr_pipeline = [
            {"$match": {**wallet_transfer_valid_filter, "currency": "IDR"}},
            {"$group": {
                "_id": None,
                "total_fee": {"$sum": "$fee"},
                "count": {"$sum": 1}
            }}
        ]
        wallet_transfer_idr = await db.wallet_transfers.aggregate(wallet_transfer_idr_pipeline).to_list(1)
        wallet_transfer_idr_fee = wallet_transfer_idr[0]["total_fee"] if wallet_transfer_idr else 0
        
        # Get USD revenue from wallet transfers
        wallet_transfer_usd_pipeline = [
            {"$match": {**wallet_transfer_valid_filter, "currency": "USD"}},
            {"$group": {
                "_id": None,
                "total_fee": {"$sum": "$fee"},
                "count": {"$sum": 1}
            }}
        ]
        wallet_transfer_usd = await db.wallet_transfers.aggregate(wallet_transfer_usd_pipeline).to_list(1)
        wallet_transfer_usd_fee = wallet_transfer_usd[0]["total_fee"] if wallet_transfer_usd else 0
        
        # Calculate Total Revenue (from fees)
        total_revenue_idr = topup_idr_fee + wallet_transfer_idr_fee
        total_revenue_usd = topup_usd_fee + wallet_transfer_usd_fee
        
        # ============================================
        # TOP-UP VOLUME CALCULATION (Amount-based, not fee)
        # ============================================
        
        # 1. Wallet Top-Up (wallet_topup_requests)
        # Note: wallet_topup_requests use "verified" (admin approval), "approved", or "completed" status
        wallet_topup_valid_filter = {**date_filter, "status": {"$in": ["verified", "approved", "completed"]}}
        
        # Get IDR wallet top-ups
        wallet_topup_idr_pipeline = [
            {"$match": {**wallet_topup_valid_filter, "currency": "IDR"}},
            {"$group": {
                "_id": None,
                "total_amount": {"$sum": "$amount"},
                "count": {"$sum": 1}
            }}
        ]
        wallet_topup_idr = await db.wallet_topup_requests.aggregate(wallet_topup_idr_pipeline).to_list(1)
        wallet_topup_idr_amount = wallet_topup_idr[0]["total_amount"] if wallet_topup_idr else 0
        
        # Get USD wallet top-ups
        wallet_topup_usd_pipeline = [
            {"$match": {**wallet_topup_valid_filter, "currency": "USD"}},
            {"$group": {
                "_id": None,
                "total_amount": {"$sum": "$amount"},
                "count": {"$sum": 1}
            }}
        ]
        wallet_topup_usd = await db.wallet_topup_requests.aggregate(wallet_topup_usd_pipeline).to_list(1)
        wallet_topup_usd_amount = wallet_topup_usd[0]["total_amount"] if wallet_topup_usd else 0
        
        # 2. Ad Account Top-Up already calculated above (topup_idr_amount, topup_usd_amount)
        
        # Calculate Total Top-Up Volume
        total_topup_idr = wallet_topup_idr_amount + topup_idr_amount
        total_topup_usd = wallet_topup_usd_amount + topup_usd_amount
        
        # ============================================
        # WITHDRAWAL DATA (for completeness)
        # ============================================
        withdraw_pipeline = [
            {"$match": date_filter},
            {"$group": {
                "_id": {"currency": "$currency", "status": "$status"},
                "total_amount": {"$sum": "$requested_amount"},
                "count": {"$sum": 1}
            }}
        ]
        withdraw_data = await db.withdraw_requests.aggregate(withdraw_pipeline).to_list(length=None)
        
        withdraw_summary = {"IDR": {}, "USD": {}}
        for item in withdraw_data:
            currency = item["_id"]["currency"]
            status = item["_id"]["status"]
            
            if currency not in withdraw_summary:
                withdraw_summary[currency] = {}
            if status not in withdraw_summary[currency]:
                withdraw_summary[currency][status] = {
                    "amount": 0, "count": 0
                }
            
            withdraw_summary[currency][status]["amount"] += item["total_amount"]
            withdraw_summary[currency][status]["count"] += item["count"]

        # Build final summary
        summary = {
            "revenue": {
                "total_revenue_idr": total_revenue_idr,
                "total_revenue_usd": total_revenue_usd,
                "breakdown_idr": {
                    "ad_account_topup_fee": topup_idr_fee,
                    "wallet_transfer_fee": wallet_transfer_idr_fee
                },
                "breakdown_usd": {
                    "ad_account_topup_fee": topup_usd_fee,
                    "wallet_transfer_fee": wallet_transfer_usd_fee
                }
            },
            "topup_volume": {
                "total_topup_idr": total_topup_idr,
                "total_topup_usd": total_topup_usd,
                "breakdown_idr": {
                    "wallet_topup": wallet_topup_idr_amount,
                    "ad_account_topup": topup_idr_amount
                },
                "breakdown_usd": {
                    "wallet_topup": wallet_topup_usd_amount,
                    "ad_account_topup": topup_usd_amount
                }
            },
            "withdraw_summary": withdraw_summary,
            "period": period,
            "date_range": {
                "start": start_date,
                "end": end_date
            }
        }

        return summary

    except Exception as e:
        logger.error(f"Error generating financial summary: {e}")
        raise HTTPException(status_code=500, detail="Error generating financial report")

@api_router.get("/admin/financial-reports/growth")
async def get_financial_growth(
    period: str = "month",  # day, week, month
    months_back: int = 12,
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get financial growth data for charts - MUST match summary calculation exactly"""
    try:
        now = datetime.now(timezone.utc)
        
        # Calculate start date based on period
        if period == "day":
            start_date = now - timedelta(days=30)  # Last 30 days
            group_format = "%Y-%m-%d"
        elif period == "week":
            start_date = now - timedelta(weeks=months_back)  # Last N weeks
            group_format = "%Y-W%U"
        else:  # month
            start_date = now - timedelta(days=months_back * 30)  # Last N months
            group_format = "%Y-%m"

        # ============================================
        # 1. ACCOUNT TOP-UP (topup_requests)
        # Same filter as summary: "verified", "approved", "completed"
        # Revenue = total_fee
        # Top-up amount = total_amount - total_fee (pure amount)
        # ============================================
        account_topup_pipeline = [
            {"$match": {
                "created_at": {"$gte": start_date.isoformat()},
                "status": {"$in": ["verified", "approved", "completed"]}
            }},
            {"$project": {
                "period": {"$dateToString": {"format": group_format, "date": {"$dateFromString": {"dateString": "$created_at"}}}},
                "currency": 1,
                "total_amount": 1,
                "total_fee": {"$ifNull": ["$total_fee", 0]},
                "pure_amount": {"$subtract": ["$total_amount", {"$ifNull": ["$total_fee", 0]}]}
            }},
            {"$group": {
                "_id": {
                    "period": "$period",
                    "currency": "$currency"
                },
                "topup_amount": {"$sum": "$pure_amount"},  # Pure amount without fee
                "revenue": {"$sum": "$total_fee"},  # Fee only
                "count": {"$sum": 1}
            }},
            {"$sort": {"_id.period": 1}}
        ]
        
        # ============================================
        # 2. WALLET TOP-UP (wallet_topup_requests)
        # Same filter as summary: "verified", "approved", "completed"
        # Revenue = total_fee (if exists) or calculated from total_amount
        # Top-up amount = amount field (NOT total_amount!)
        # ============================================
        wallet_topup_pipeline = [
            {"$match": {
                "created_at": {"$gte": start_date.isoformat()},
                "status": {"$in": ["verified", "approved", "completed"]}
            }},
            {"$project": {
                "period": {"$dateToString": {"format": group_format, "date": {"$dateFromString": {"dateString": "$created_at"}}}},
                "currency": 1,
                "amount": 1,  # Use 'amount' field, not 'total_amount'
                "total_amount": {"$ifNull": ["$total_amount", "$amount"]},  # Fallback to amount
                "total_fee": {"$ifNull": ["$total_fee", 0]}
            }},
            {"$project": {
                "period": 1,
                "currency": 1,
                "topup_amount": "$amount",  # Direct amount (already pure)
                "revenue": "$total_fee"  # Fee if available
            }},
            {"$group": {
                "_id": {
                    "period": "$period",
                    "currency": "$currency"
                },
                "topup_amount": {"$sum": "$topup_amount"},
                "revenue": {"$sum": "$revenue"},
                "count": {"$sum": 1}
            }},
            {"$sort": {"_id.period": 1}}
        ]
        
        # ============================================
        # 3. WALLET TRANSFER FEE (wallet_transfers)
        # For revenue calculation only
        # ============================================
        wallet_transfer_pipeline = [
            {"$match": {
                "created_at": {"$gte": start_date.isoformat()},
                "status": {"$in": ["completed", "approved"]}
            }},
            {"$project": {
                "period": {"$dateToString": {"format": group_format, "date": {"$dateFromString": {"dateString": "$created_at"}}}},
                "currency": 1,
                "fee": {"$ifNull": ["$fee", 0]}
            }},
            {"$group": {
                "_id": {
                    "period": "$period",
                    "currency": "$currency"
                },
                "revenue": {"$sum": "$fee"}
            }},
            {"$sort": {"_id.period": 1}}
        ]
        
        # ============================================
        # 4. WITHDRAWAL (withdraw_requests)
        # ============================================
        withdraw_pipeline = [
            {"$match": {
                "created_at": {"$gte": start_date.isoformat()},
                "status": {"$in": ["approved", "completed"]}
            }},
            {"$group": {
                "_id": {
                    "period": {"$dateToString": {"format": group_format, "date": {"$dateFromString": {"dateString": "$created_at"}}}},
                    "currency": "$currency"
                },
                "amount": {"$sum": "$requested_amount"},
                "count": {"$sum": 1}
            }},
            {"$sort": {"_id.period": 1}}
        ]

        # Execute all pipelines
        account_topup_results = await db.topup_requests.aggregate(account_topup_pipeline).to_list(length=None)
        wallet_topup_results = await db.wallet_topup_requests.aggregate(wallet_topup_pipeline).to_list(length=None)
        wallet_transfer_results = await db.wallet_transfers.aggregate(wallet_transfer_pipeline).to_list(length=None)
        withdraw_results = await db.withdraw_requests.aggregate(withdraw_pipeline).to_list(length=None)

        # Initialize growth data structure
        growth_data = {
            "topup": {"IDR": [], "USD": []},
            "withdraw": {"IDR": [], "USD": []},
            "revenue": {"IDR": [], "USD": []}
        }

        # Helper to aggregate by period + currency
        def aggregate_data(results_list):
            aggregated = {}
            for item in results_list:
                currency = item["_id"]["currency"]
                period = item["_id"]["period"]
                key = f"{period}_{currency}"
                
                if key not in aggregated:
                    aggregated[key] = {
                        "period": period,
                        "currency": currency,
                        "topup_amount": 0,
                        "revenue": 0,
                        "withdraw_amount": 0,
                        "count": 0
                    }
                
                aggregated[key]["topup_amount"] += item.get("topup_amount", 0)
                aggregated[key]["revenue"] += item.get("revenue", 0)
                aggregated[key]["withdraw_amount"] += item.get("amount", 0)
                aggregated[key]["count"] += item.get("count", 0)
            
            return aggregated

        # Aggregate account + wallet top-ups
        combined = aggregate_data(account_topup_results + wallet_topup_results)
        
        # Add wallet transfer fees to revenue
        for item in wallet_transfer_results:
            currency = item["_id"]["currency"]
            period = item["_id"]["period"]
            key = f"{period}_{currency}"
            
            if key in combined:
                combined[key]["revenue"] += item["revenue"]
            else:
                combined[key] = {
                    "period": period,
                    "currency": currency,
                    "topup_amount": 0,
                    "revenue": item["revenue"],
                    "withdraw_amount": 0,
                    "count": 0
                }

        # Build top-up and revenue growth data
        for key, data in combined.items():
            currency = data["currency"]
            
            growth_data["topup"][currency].append({
                "period": data["period"],
                "amount": data["topup_amount"],
                "count": data["count"]
            })
            
            growth_data["revenue"][currency].append({
                "period": data["period"],
                "amount": data["revenue"]
            })

        # Build withdrawal growth data
        for item in withdraw_results:
            currency = item["_id"]["currency"]
            
            growth_data["withdraw"][currency].append({
                "period": item["_id"]["period"],
                "amount": item["amount"],
                "count": item["count"]
            })

        return {
            "growth_data": growth_data,
            "period": period,
            "start_date": start_date.isoformat(),
            "end_date": now.isoformat()
        }

    except Exception as e:
        logger.error(f"Error generating growth data: {e}")
        raise HTTPException(status_code=500, detail="Error generating growth report")

@api_router.get("/admin/financial-reports/export")
async def export_financial_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    period: Optional[str] = "all",
    format: str = "pdf",  # pdf or xlsx
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Export financial report to PDF or Excel"""
    try:
        # Get financial data first
        financial_data = await get_financial_summary(start_date, end_date, period, current_admin)
        
        if format == "pdf":
            return export_financial_pdf(financial_data)
        elif format == "xlsx":
            return export_financial_excel(financial_data)
        else:
            raise HTTPException(status_code=400, detail="Unsupported format. Use 'pdf' or 'xlsx'")
    
    except Exception as e:
        logger.error(f"Error exporting financial report: {e}")
        raise HTTPException(status_code=500, detail="Error exporting report")

def export_financial_pdf(financial_data: dict) -> Response:
    """Export financial report as PDF"""
    buffer = BytesIO()
    
    # Create PDF document
    doc = SimpleDocTemplate(buffer, pagesize=A4, 
                          rightMargin=72, leftMargin=72,
                          topMargin=72, bottomMargin=18)
    
    # Get styles
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        alignment=1,  # Center
        textColor=colors.HexColor('#1f2937')
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=12,
        textColor=colors.HexColor('#374151')
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=6
    )
    
    # Story elements
    story = []
    
    # Add Rimuru Logo (if exists)
    try:
        logo_path = "/app/frontend/public/images/rimuru-logo.png"
        import os
        if os.path.exists(logo_path):
            # Preserve aspect ratio for logo - calculate height based on width
            from PIL import Image as PILImage
            with PILImage.open(logo_path) as img:
                original_width, original_height = img.size
                desired_width = 2*inch
                aspect_ratio = original_height / original_width
                desired_height = desired_width * aspect_ratio
            
            logo = Image(logo_path, width=desired_width, height=desired_height)
            logo.hAlign = 'CENTER'
            story.append(logo)
            story.append(Spacer(1, 12))
    except Exception as e:
        logger.warning(f"Could not add logo to report: {e}")
    
    # Title
    story.append(Paragraph("RIMURU - LAPORAN KEUANGAN", title_style))
    story.append(Spacer(1, 12))
    
    # Period info
    period_info = f"Periode: {financial_data.get('period', 'N/A')}"
    if financial_data.get('date_range', {}).get('start'):
        period_info = f"Periode: {financial_data['date_range']['start']} - {financial_data['date_range']['end']}"
    
    story.append(Paragraph(period_info, normal_style))
    story.append(Paragraph(f"Tanggal Generate: {datetime.now(timezone.utc).astimezone(timezone(timedelta(hours=7))).strftime('%d %B %Y, %H:%M WIB')}", normal_style))
    story.append(Spacer(1, 20))
    
    # Revenue Summary
    story.append(Paragraph("RINGKASAN REVENUE", heading_style))
    
    revenue_data = [
        ["Mata Uang", "Total Fee", "Total Top-up", "Total Withdrawal"],
        ["IDR", f"Rp {financial_data['revenue']['total_fees_idr']:,.2f}", 
         f"Rp {financial_data['revenue']['total_topup_idr']:,.2f}",
         f"Rp {financial_data['revenue']['total_withdraw_idr']:,.2f}"],
        ["USD", f"$ {financial_data['revenue']['total_fees_usd']:,.2f}", 
         f"$ {financial_data['revenue']['total_topup_usd']:,.2f}",
         f"$ {financial_data['revenue']['total_withdraw_usd']:,.2f}"]
    ]
    
    revenue_table = Table(revenue_data, colWidths=[1*inch, 1.5*inch, 1.5*inch, 1.5*inch])
    revenue_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f3f4f6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
    ]))
    
    story.append(revenue_table)
    story.append(Spacer(1, 20))
    
    # Top-up Breakdown
    story.append(Paragraph("BREAKDOWN TOP-UP", heading_style))
    
    for currency, statuses in financial_data.get('topup_summary', {}).items():
        if not statuses:
            continue
            
        story.append(Paragraph(f"Top-up {currency}:", normal_style))
        
        topup_breakdown_data = [["Status", "Jumlah", "Fee", "Count"]]
        
        for status, data in statuses.items():
            symbol = "Rp" if currency == "IDR" else "$"
            topup_breakdown_data.append([
                status.title(),
                f"{symbol} {data['amount']:,.2f}",
                f"{symbol} {data['fee']:,.2f}",
                str(data['count'])
            ])
        
        topup_table = Table(topup_breakdown_data, colWidths=[1.2*inch, 1.5*inch, 1.2*inch, 0.8*inch])
        topup_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f3f4f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        story.append(topup_table)
        story.append(Spacer(1, 10))
    
    # Withdrawal Breakdown
    story.append(Spacer(1, 10))
    story.append(Paragraph("BREAKDOWN WITHDRAWAL", heading_style))
    
    for currency, statuses in financial_data.get('withdraw_summary', {}).items():
        if not statuses:
            continue
            
        story.append(Paragraph(f"Withdrawal {currency}:", normal_style))
        
        withdraw_breakdown_data = [["Status", "Jumlah", "Count"]]
        
        for status, data in statuses.items():
            symbol = "Rp" if currency == "IDR" else "$"
            withdraw_breakdown_data.append([
                status.title(),
                f"{symbol} {data['amount']:,.2f}",
                str(data['count'])
            ])
        
        withdraw_table = Table(withdraw_breakdown_data, colWidths=[1.5*inch, 2*inch, 1*inch])
        withdraw_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f3f4f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        story.append(withdraw_table)
        story.append(Spacer(1, 10))
    
    # Footer
    story.append(Spacer(1, 30))
    footer_text = f"Laporan dihasilkan pada {datetime.now(timezone.utc).astimezone(timezone(timedelta(hours=7))).strftime('%d %B %Y, %H:%M')} WIB"
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=1  # Center
    )
    story.append(Paragraph(footer_text, footer_style))
    
    # Build PDF
    doc.build(story)
    
    # Get PDF bytes
    pdf_bytes = buffer.getvalue()
    buffer.close()
    
    # Return PDF as response
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=financial_report_{datetime.now().strftime('%Y%m%d')}.pdf",
            "Access-Control-Allow-Origin": "*"
        }
    )

def export_financial_excel(financial_data: dict) -> Response:
    """Export financial report as Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
        import pandas as pd
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Laporan Keuangan"
        
        # Title
        ws['A1'] = "RIMURU - LAPORAN KEUANGAN"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:E1')
        
        # Period info
        period_info = f"Periode: {financial_data.get('period', 'N/A')}"
        if financial_data.get('date_range', {}).get('start'):
            period_info = f"Periode: {financial_data['date_range']['start']} - {financial_data['date_range']['end']}"
        
        ws['A3'] = period_info
        ws['A4'] = f"Tanggal Generate: {datetime.now(timezone.utc).astimezone(timezone(timedelta(hours=7))).strftime('%d %B %Y, %H:%M WIB')}"
        
        # Revenue Summary
        row_num = 6
        ws[f'A{row_num}'] = "RINGKASAN REVENUE"
        ws[f'A{row_num}'].font = Font(bold=True, size=14)
        row_num += 1
        
        # Revenue headers
        revenue_headers = ["Mata Uang", "Total Fee", "Total Top-up", "Total Withdrawal"]
        for col, header in enumerate(revenue_headers, 1):
            cell = ws.cell(row=row_num, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        
        row_num += 1
        
        # IDR Row
        ws.cell(row=row_num, column=1, value="IDR")
        ws.cell(row=row_num, column=2, value=financial_data['revenue']['total_fees_idr'])
        ws.cell(row=row_num, column=3, value=financial_data['revenue']['total_topup_idr'])
        ws.cell(row=row_num, column=4, value=financial_data['revenue']['total_withdraw_idr'])
        
        row_num += 1
        
        # USD Row
        ws.cell(row=row_num, column=1, value="USD")
        ws.cell(row=row_num, column=2, value=financial_data['revenue']['total_fees_usd'])
        ws.cell(row=row_num, column=3, value=financial_data['revenue']['total_topup_usd'])
        ws.cell(row=row_num, column=4, value=financial_data['revenue']['total_withdraw_usd'])
        
        # Add more detailed breakdowns for top-up and withdrawals
        row_num += 3
        
        # Top-up Breakdown
        ws[f'A{row_num}'] = "BREAKDOWN TOP-UP"
        ws[f'A{row_num}'].font = Font(bold=True, size=14)
        row_num += 2
        
        for currency, statuses in financial_data.get('topup_summary', {}).items():
            if not statuses:
                continue
                
            ws[f'A{row_num}'] = f"Top-up {currency}"
            ws[f'A{row_num}'].font = Font(bold=True)
            row_num += 1
            
            # Headers
            headers = ["Status", "Jumlah", "Fee", "Count"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
            row_num += 1
            
            # Data
            for status, data in statuses.items():
                ws.cell(row=row_num, column=1, value=status.title())
                ws.cell(row=row_num, column=2, value=data['amount'])
                ws.cell(row=row_num, column=3, value=data['fee'])
                ws.cell(row=row_num, column=4, value=data['count'])
                row_num += 1
            
            row_num += 1
        
        # Withdrawal Breakdown
        ws[f'A{row_num}'] = "BREAKDOWN WITHDRAWAL"
        ws[f'A{row_num}'].font = Font(bold=True, size=14)
        row_num += 2
        
        for currency, statuses in financial_data.get('withdraw_summary', {}).items():
            if not statuses:
                continue
                
            ws[f'A{row_num}'] = f"Withdrawal {currency}"
            ws[f'A{row_num}'].font = Font(bold=True)
            row_num += 1
            
            # Headers
            headers = ["Status", "Jumlah", "Count"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
            row_num += 1
            
            # Data
            for status, data in statuses.items():
                ws.cell(row=row_num, column=1, value=status.title())
                ws.cell(row=row_num, column=2, value=data['amount'])
                ws.cell(row=row_num, column=3, value=data['count'])
                row_num += 1
            
            row_num += 1
        
        # Save to bytes
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return Response(
            content=buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=financial_report_{datetime.now().strftime('%Y%m%d')}.xlsx",
                "Access-Control-Allow-Origin": "*"
            }
        )
    
    except ImportError:
        # If openpyxl not installed, fall back to CSV
        import csv
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Write headers and data
        writer.writerow(["RIMURU - LAPORAN KEUANGAN"])
        writer.writerow([])
        writer.writerow(["RINGKASAN REVENUE"])
        writer.writerow(["Mata Uang", "Total Fee", "Total Top-up", "Total Withdrawal"])
        writer.writerow(["IDR", financial_data['revenue']['total_fees_idr'], 
                        financial_data['revenue']['total_topup_idr'],
                        financial_data['revenue']['total_withdraw_idr']])
        writer.writerow(["USD", financial_data['revenue']['total_fees_usd'], 
                        financial_data['revenue']['total_topup_usd'],
                        financial_data['revenue']['total_withdraw_usd']])
        
        # Return as CSV
        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename=financial_report_{datetime.now().strftime('%Y%m%d')}.csv",
                "Access-Control-Allow-Origin": "*"
            }
        )

# Groups endpoints
@api_router.get("/groups", response_model=List[Group])
async def get_groups(current_user: User = Depends(get_current_user)):
    groups_cursor = db.groups.find({"user_id": current_user.id})
    groups = await groups_cursor.to_list(length=None)
    return [Group(**parse_from_mongo(group)) for group in groups]

@api_router.post("/groups", response_model=Group)
async def create_group(group_data: GroupCreate, current_user: User = Depends(get_current_user)):
    # Check if group name already exists for this user
    existing_group = await db.groups.find_one({
        "user_id": current_user.id,
        "name": group_data.name
    })
    
    if existing_group:
        raise HTTPException(
            status_code=400,
            detail="Group with this name already exists"
        )
    
    # Create new group
    new_group = Group(
        name=group_data.name,
        user_id=current_user.id
    )
    
    # Save to database
    group_dict = new_group.dict()
    group_dict["created_at"] = group_dict["created_at"].isoformat()
    await db.groups.insert_one(group_dict)
    
    return new_group

# Transfer Request endpoints
@api_router.post("/transfer-request", response_model=dict)
async def create_transfer_request(
    account_id: str,
    amount: float,
    current_user: User = Depends(get_current_user)
):
    """Create a new transfer request from wallet to account (requires admin approval)"""
    try:
        # Validate amount
        if amount <= 0:
            raise HTTPException(status_code=400, detail="Transfer amount must be greater than 0")
        
        # Get account details
        account = await db.ad_accounts.find_one({"id": account_id, "user_id": current_user.id})
        if not account:
            raise HTTPException(status_code=404, detail="Ad account not found")
        
        # Get user wallet balance
        user = await db.users.find_one({"id": current_user.id})
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        
        # Determine currency and wallet field
        currency = account.get("currency", "IDR")
        wallet_field = "wallet_balance_idr" if currency == "IDR" else "wallet_balance_usd"
        wallet_balance = user.get(wallet_field, 0)
        
        # Check wallet balance
        if wallet_balance < amount:
            raise HTTPException(status_code=400, detail="Insufficient wallet balance")
        
        # Create transfer request
        transfer_request = TransferRequestRecord(
            user_id=current_user.id,
            account_id=account_id,
            amount=amount,
            currency=currency
        )
        
        # Save to database
        await db.transfer_requests.insert_one(prepare_for_mongo(transfer_request.dict()))
        
        # Note: Notifications are handled by /api/balance-transfer endpoint used by frontend
        # This endpoint is for alternative access and doesn't create notifications to avoid duplicates
        
        return {
            "message": "Transfer request created successfully",
            "transfer_id": transfer_request.id,
            "amount": amount,
            "currency": currency,
            "account_name": account.get("account_name"),
            "status": "pending"
        }
        
    except Exception as e:
        logger.error(f"Error creating transfer request: {e}")
        raise HTTPException(status_code=500, detail="Error creating transfer request")

@api_router.get("/transfer-requests", response_model=List[dict])
async def get_user_transfer_requests(
    current_user: User = Depends(get_current_user)
):
    """Get transfer requests for current user"""
    try:
        # Get transfer requests with account details
        pipeline = [
            {"$match": {"user_id": current_user.id}},
            {"$lookup": {
                "from": "ad_accounts",
                "localField": "account_id", 
                "foreignField": "id",
                "as": "account_details"
            }},
            {"$sort": {"created_at": -1}}
        ]
        
        transfer_requests = await db.transfer_requests.aggregate(pipeline).to_list(length=None)
        
        # Process results
        result = []
        for request in transfer_requests:
            account_details = request.get("account_details", [{}])[0]
            result.append({
                "id": request["id"],
                "account_id": request["account_id"],
                "account_name": account_details.get("account_name", "N/A"),
                "platform": account_details.get("platform", "N/A"),
                "amount": request["amount"],
                "currency": request["currency"],
                "status": request["status"],
                "created_at": request["created_at"],
                "processed_at": request.get("processed_at"),
                "admin_notes": request.get("admin_notes")
            })
            
        return result
        
    except Exception as e:
        logger.error(f"Error fetching user transfer requests: {e}")
        raise HTTPException(status_code=500, detail="Error fetching transfer requests")

# Admin Transfer Request Management endpoints
@api_router.get("/admin/transfer-requests", response_model=List[dict])
async def get_all_transfer_requests(
    status: Optional[str] = None,
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get all transfer requests for admin management"""
    try:
        # Build match criteria
        match_criteria = {}
        if status:
            match_criteria["status"] = status
            
        # Get transfer requests with user and account details
        pipeline = [
            {"$match": match_criteria},
            {"$lookup": {
                "from": "users",
                "localField": "user_id",
                "foreignField": "id", 
                "as": "user_details"
            }},
            {"$lookup": {
                "from": "ad_accounts",
                "localField": "account_id",
                "foreignField": "id",
                "as": "account_details"
            }},
            {"$lookup": {
                "from": "admin_users",
                "localField": "admin_id",
                "foreignField": "id",
                "as": "admin_details"
            }},
            {"$sort": {"created_at": -1}}
        ]
        
        transfer_requests = await db.transfer_requests.aggregate(pipeline).to_list(length=None)
        
        # Process results
        result = []
        for request in transfer_requests:
            user_details_list = request.get("user_details", [])
            account_details_list = request.get("account_details", [])
            admin_details_list = request.get("admin_details", [])
            
            user_details = user_details_list[0] if user_details_list else {}
            account_details = account_details_list[0] if account_details_list else {}
            admin_details = admin_details_list[0] if admin_details_list else {}
            
            result.append({
                "id": request["id"],
                "user_id": request["user_id"],
                "account_id": request["account_id"],
                "user": {
                    "username": user_details.get("username", "N/A"),
                    "email": user_details.get("email", "N/A")
                },
                "account": {
                    "account_name": account_details.get("account_name", "N/A"),
                    "platform": account_details.get("platform", "N/A"),
                    "currency": account_details.get("currency", "IDR"),
                    "real_account_id": account_details.get("account_id", "N/A")  # The actual ad account ID
                },
                "amount": request["amount"],
                "currency": request["currency"],
                "status": request["status"],
                "created_at": request["created_at"],
                "processed_at": request.get("processed_at"),
                "admin_id": request.get("admin_id"),
                "admin_username": admin_details.get("username"),
                "admin_notes": request.get("admin_notes")
            })
            
        return result
        
    except Exception as e:
        logger.error(f"Error fetching transfer requests for admin: {e}")
        raise HTTPException(status_code=500, detail="Error fetching transfer requests")

@api_router.post("/admin/upload-proof")
async def upload_proof_file(
    file: UploadFile = File(...),
    type: str = Form(...),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Upload proof files for transfer request approval to GCS"""
    try:
        # Validate file type
        allowed_extensions = {'.jpg', '.jpeg', '.png', '.pdf', '.webp'}
        file_extension = os.path.splitext(file.filename)[1].lower()
        
        if file_extension not in allowed_extensions:
            raise HTTPException(status_code=400, detail="File type not allowed. Please upload JPG, PNG, PDF, or WEBP files.")
        
        # Generate unique filename
        unique_filename = f"{type}_{uuid.uuid4().hex[:12]}_{int(datetime.now().timestamp())}{file_extension}"
        gcs_path = f"wallet_transfer_proofs/{unique_filename}"
        
        # Get GCS storage instance
        gcs = get_gcs_storage()
        
        try:
            # Read file content
            file_content = await file.read()
            file_obj = io.BytesIO(file_content)
            
            # Upload to GCS
            gcs.upload_file(
                file_obj=file_obj,
                destination_path=gcs_path,
                content_type=file.content_type,
                metadata={
                    "type": type,
                    "uploaded_by": current_admin.username,
                    "original_filename": file.filename
                }
            )
            
            logger.info(f"Uploaded wallet transfer proof to GCS: {gcs_path}")
            
            # Store with /files/ prefix for consistent serving
            stored_path = f"/files/wallet_transfer_proofs/{unique_filename}"
            
            return {
                "success": True,
                "file_url": stored_path,
                "storage": "gcs",
                "message": "File uploaded successfully"
            }
            
        except Exception as gcs_error:
            logger.error(f"Failed to upload to GCS, falling back to filesystem: {gcs_error}")
            # Fallback to filesystem
            upload_dir = os.path.join("/app/uploads", "transfer_proofs")
            os.makedirs(upload_dir, exist_ok=True)
            
            file_path = os.path.join(upload_dir, unique_filename)
            
            # Save file to filesystem
            with open(file_path, "wb") as buffer:
                buffer.write(file_content)
            
            # Return relative path for filesystem
            relative_path = f"transfer_proofs/{unique_filename}"
            
            return {
                "success": True,
                "file_url": relative_path,
                "storage": "filesystem",
                "message": "File uploaded successfully (fallback to filesystem)"
            }
        
    except Exception as e:
        logger.error(f"Error uploading proof file: {e}")
        raise HTTPException(status_code=500, detail="Error uploading file")

# ==================== PROOF EDIT MANAGEMENT ENDPOINTS ====================

@api_router.post("/admin/topup-requests/{request_id}/edit-proof")
async def request_account_proof_edit(
    request_id: str,
    file: UploadFile = File(...),
    account_id: str = Form(...),
    proof_type: str = Form(...),  # 'spend_limit' or 'budget_aspire'
    notes: str = Form(""),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Admin request to edit/replace account proof (spend_limit or budget_aspire) - requires super admin approval"""
    try:
        # Find the topup request
        topup_request = await db.topup_requests.find_one({"id": request_id})
        if not topup_request:
            raise HTTPException(status_code=404, detail="Top-up request not found")
        
        # Find the specific account in accounts array
        accounts = topup_request.get("accounts", [])
        target_account = None
        target_account_index = None
        
        for idx, acc in enumerate(accounts):
            if acc.get("account_id") == account_id:
                target_account = acc
                target_account_index = idx
                break
        
        if not target_account:
            raise HTTPException(status_code=404, detail="Account not found in request")
        
        # Determine which proof_url to check
        if proof_type == "spend_limit":
            current_proof_url = target_account.get("spend_limit_proof_url")
            if not current_proof_url:
                raise HTTPException(status_code=400, detail="No existing spend limit proof to edit")
            proof_field = "spend_limit_proof_url"
        elif proof_type == "budget_aspire":
            current_proof_url = target_account.get("budget_aspire_proof_url")
            if not current_proof_url:
                raise HTTPException(status_code=400, detail="No existing budget aspire proof to edit")
            proof_field = "budget_aspire_proof_url"
        else:
            raise HTTPException(status_code=400, detail="Invalid proof_type. Must be 'spend_limit' or 'budget_aspire'")
        
        # Create or get payment_proofs document for this account proof
        # Since accounts only store proof_url, we need to create a tracking document
        proof_tracking_id = f"proof_tracking_{request_id}_{account_id}_{proof_type}"
        existing_proof_doc = await db.payment_proofs.find_one({"tracking_id": proof_tracking_id})
        
        if not existing_proof_doc:
            # Create new tracking document
            proof_doc = {
                "id": str(uuid.uuid4()),
                "tracking_id": proof_tracking_id,
                "request_id": request_id,
                "account_id": account_id,
                "proof_type": proof_type,
                "gcs_path": current_proof_url,
                "file_url": current_proof_url,
                "storage_type": "gcs",
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.payment_proofs.insert_one(proof_doc)
            proof_id = proof_doc["id"]
            logger.info(f"✅ Created tracking document for {proof_type}: {proof_id}")
        else:
            proof_id = existing_proof_doc["id"]
            logger.info(f"✅ Using existing tracking document: {proof_id}")
        
        # Validate file type
        allowed_extensions = {'.jpg', '.jpeg', '.png', '.pdf', '.webp'}
        file_extension = os.path.splitext(file.filename)[1].lower()
        if file_extension not in allowed_extensions:
            raise HTTPException(status_code=400, detail="File type not allowed")
        
        # Upload new file to GCS
        unique_filename = f"edit_{proof_type}_{uuid.uuid4().hex[:12]}_{int(datetime.now().timestamp())}{file_extension}"
        gcs_path = f"payment_proofs/pending_edits/{unique_filename}"
        
        gcs = get_gcs_storage()
        file_content = await file.read()
        file_obj = io.BytesIO(file_content)
        
        gcs.upload_file(
            file_obj=file_obj,
            destination_path=gcs_path,
            content_type=file.content_type,
            metadata={
                "type": f"proof_edit_{proof_type}",
                "request_id": request_id,
                "account_id": account_id,
                "uploaded_by": current_admin.username,
                "original_filename": file.filename
            }
        )
        
        logger.info(f"✅ Uploaded new {proof_type} proof for edit approval to GCS: {gcs_path}")
        
        # Update payment proof with pending edit
        edit_data = {
            "pending_edit": True,
            "new_gcs_path": gcs_path,
            "new_file_url": f"/files/{gcs_path}",
            "edit_requested_by": current_admin.id,
            "edit_requested_by_username": current_admin.username,
            "edit_requested_at": datetime.now(timezone.utc).isoformat(),
            "edit_notes": notes,
            "proof_type": proof_type,
            "account_id": account_id,
            "request_id": request_id
        }
        
        await db.payment_proofs.update_one(
            {"id": proof_id},
            {"$set": edit_data}
        )
        
        logger.info(f"✅ Set pending_edit flag for proof {proof_id}")
        
        # Create notification for super admin
        super_admins = await db.admin_users.find({"role": "super_admin"}).to_list(None)
        for super_admin in super_admins:
            notification = {
                "id": str(uuid.uuid4()),
                "title": "🔄 Account Proof Edit Request",
                "message": f"Admin {current_admin.username} requested to edit {proof_type.replace('_', ' ').title()} proof for account {target_account.get('account_name', account_id)}",
                "type": "proof_edit_request",
                "reference_id": proof_id,
                "is_read": False,
                "created_at": datetime.now(timezone.utc)
            }
            notification_dict = prepare_for_mongo(notification)
            await db.notifications.insert_one(notification_dict)
        
        logger.info(f"📧 Notified {len(super_admins)} super admins about {proof_type} proof edit request")
        
        return {
            "success": True,
            "message": "Edit request submitted. Waiting for super admin approval.",
            "proof_id": proof_id,
            "proof_type": proof_type
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Error requesting account proof edit: {e}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error requesting proof edit: {str(e)}")

@api_router.get("/super-admin/pending-proof-edits", response_model=List[dict])
async def get_pending_proof_edits(
    current_super_admin: AdminUser = Depends(get_current_super_admin)
):
    """Get all pending proof edit requests for super admin approval"""
    try:
        # Find all payment proofs with pending edits
        pending_proofs = await db.payment_proofs.find({"pending_edit": True}).to_list(None)
        
        result = []
        for proof in pending_proofs:
            proof_data = parse_from_mongo(proof)
            
            # Get request_id, account_id, and proof_type from proof document
            request_id = proof.get("request_id")
            transfer_id = proof.get("transfer_id")  # Check for wallet transfer
            account_id = proof.get("account_id")
            target_account_id = proof.get("target_account_id")  # For wallet transfers
            proof_type = proof.get("proof_type")
            
            # Get admin who requested edit
            admin = await db.admin_users.find_one({"id": proof.get("edit_requested_by")})
            admin_data = parse_from_mongo(admin) if admin else {}
            
            # Determine if this is a topup request or wallet transfer
            if transfer_id:
                # This is a wallet transfer proof edit
                wallet_transfer = await db.wallet_transfers.find_one({"id": transfer_id})
                if not wallet_transfer:
                    logger.warning(f"⚠️ Wallet transfer {transfer_id} not found for proof {proof.get('id')}, skipping")
                    continue
                
                transfer_data = parse_from_mongo(wallet_transfer)
                
                # Get user info
                user = await db.users.find_one({"id": transfer_data["user_id"]})
                user_data = parse_from_mongo(user) if user else {}
                
                # Get target account info - directly from transfer_data
                account_name = transfer_data.get("target_account_name", f"Account {target_account_id[:8] if target_account_id else 'N/A'}")
                actual_account_id = target_account_id or transfer_data.get("target_account_id", account_id)
                
                # Get current proof path - fallback to wallet_transfer if not in payment_proofs
                current_proof_path = proof.get("gcs_path", "")
                if not current_proof_path:
                    # Fallback: get from wallet_transfer
                    proof_field = "spend_limit_proof_url" if proof_type == "spend_limit" else "budget_aspire_proof_url"
                    current_proof_path = transfer_data.get(proof_field, "")
                    logger.info(f"📝 Using fallback current proof from wallet_transfer: {current_proof_path}")
                
                # Remove /files/ prefix if it exists to avoid duplication
                if current_proof_path.startswith('/files/'):
                    current_proof_path = current_proof_path[7:]  # Remove '/files/' prefix
                elif current_proof_path.startswith('files/'):
                    current_proof_path = current_proof_path[6:]  # Remove 'files/' prefix
                
                result.append({
                    "proof_id": proof["id"],
                    "request_id": transfer_id,
                    "account_id": actual_account_id,
                    "account_name": account_name,
                    "proof_type": proof_type,
                    "request_code": transfer_data.get("transfer_code", transfer_data["id"][:8]),
                    "client_name": user_data.get("name") or user_data.get("username"),
                    "client_username": user_data.get("username"),
                    "current_gcs_path": current_proof_path,
                    "current_file_url": current_proof_path,  # Just use the path directly, frontend will add /api/files/
                    "new_gcs_path": proof.get("new_gcs_path", ""),
                    "new_file_url": proof.get("new_file_url") or proof.get("new_gcs_path"),
                    "requested_by_username": admin_data.get("username") or admin_data.get("name"),
                    "requested_at": proof.get("edit_requested_at"),
                    "edit_notes": proof.get("edit_notes", ""),
                    "amount": transfer_data.get("amount", 0),
                    "currency": transfer_data.get("currency", "IDR"),
                    "request_type": "wallet_transfer"
                })
                
            elif request_id:
                # This is a topup request proof edit
                if not account_id or not proof_type:
                    logger.warning(f"⚠️ Proof {proof.get('id')} missing required fields, skipping")
                    continue
                
                # Find associated topup request by request_id
                topup_request = await db.topup_requests.find_one({"id": request_id})
                if not topup_request:
                    logger.warning(f"⚠️ Topup request {request_id} not found for proof {proof.get('id')}, skipping")
                    continue
                    
                topup_data = parse_from_mongo(topup_request)
                
                # Get user info
                user = await db.users.find_one({"id": topup_data["user_id"]})
                user_data = parse_from_mongo(user) if user else {}
                
                # Find the account in the accounts array
                account_name = None
                accounts = topup_data.get("accounts", [])
                for acc in accounts:
                    if acc.get("account_id") == account_id:
                        account_name = acc.get("account_name")
                        break
                
                # Get current proof path and clean it
                current_proof_path = proof.get("gcs_path", "")
                # Remove /files/ prefix if it exists to avoid duplication
                if current_proof_path.startswith('/files/'):
                    current_proof_path = current_proof_path[7:]
                elif current_proof_path.startswith('files/'):
                    current_proof_path = current_proof_path[6:]
                
                # Get new proof path and clean it
                new_proof_path = proof.get("new_gcs_path", "")
                if new_proof_path.startswith('/files/'):
                    new_proof_path = new_proof_path[7:]
                elif new_proof_path.startswith('files/'):
                    new_proof_path = new_proof_path[6:]
                
                result.append({
                    "proof_id": proof["id"],
                    "request_id": request_id,
                    "account_id": account_id,
                    "account_name": account_name or f"Account {account_id[:8]}",
                    "proof_type": proof_type,
                    "request_code": topup_data.get("request_code", topup_data["id"][:8]),
                    "client_name": user_data.get("name") or user_data.get("username"),
                    "client_username": user_data.get("username"),
                    "current_gcs_path": current_proof_path,
                    "current_file_url": current_proof_path,  # Just use the path directly
                    "new_gcs_path": new_proof_path,
                    "new_file_url": new_proof_path,  # Just use the path directly
                    "requested_by_username": admin_data.get("username") or admin_data.get("name"),
                    "requested_at": proof.get("edit_requested_at"),
                    "edit_notes": proof.get("edit_notes", ""),
                    "amount": topup_data.get("amount", 0),
                    "currency": topup_data.get("currency", "IDR"),
                    "request_type": "topup_request"
                })
            else:
                logger.warning(f"⚠️ Proof {proof.get('id')} has no request_id or transfer_id, skipping")
        
        logger.info(f"✅ Found {len(result)} pending proof edits")
        return result
        
    except Exception as e:
        logger.error(f"❌ Error fetching pending proof edits: {e}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail="Error fetching pending edits")

@api_router.put("/super-admin/proof-edits/{proof_id}/approve")
async def approve_proof_edit(
    proof_id: str,
    notes: str = "",
    current_super_admin: AdminUser = Depends(get_current_super_admin)
):
    """Super admin approves proof edit request"""
    try:
        # Get proof with pending edit
        proof = await db.payment_proofs.find_one({"id": proof_id, "pending_edit": True})
        if not proof:
            raise HTTPException(status_code=404, detail="Pending edit not found")
        
        # Save current file to history
        proof_history = proof.get("proof_history", [])
        proof_history.append({
            "file_path": proof.get("gcs_path"),
            "file_url": proof.get("file_url"),
            "edited_by": proof.get("edit_requested_by"),
            "edited_by_username": proof.get("edit_requested_by_username"),
            "edited_at": proof.get("edit_requested_at"),
            "approved_by": current_super_admin.id,
            "approved_by_username": current_super_admin.username,
            "approved_at": datetime.now(timezone.utc).isoformat(),
            "status": "approved",
            "notes": notes
        })
        
        # Replace current file with new file
        await db.payment_proofs.update_one(
            {"id": proof_id},
            {"$set": {
                "gcs_path": proof["new_gcs_path"],
                "file_url": proof["new_file_url"],
                "pending_edit": False,
                "proof_history": proof_history,
                "last_edited_at": datetime.now(timezone.utc).isoformat(),
                "last_edited_by": current_super_admin.id
            },
            "$unset": {
                "new_gcs_path": "",
                "new_file_url": "",
                "edit_requested_by": "",
                "edit_requested_by_username": "",
                "edit_requested_at": "",
                "edit_notes": ""
            }}
        )
        
        # IMPORTANT: Update the proof_url in source collection (topup_requests or wallet_transfers)
        request_id = proof.get("request_id")
        transfer_id = proof.get("transfer_id")
        account_id = proof.get("account_id")
        target_account_id = proof.get("target_account_id")
        proof_type = proof.get("proof_type")
        
        # Check if this is a wallet transfer or topup request
        if transfer_id and proof_type:
            # This is a wallet transfer proof edit
            proof_field = f"{proof_type}_proof_url"  # spend_limit_proof_url or budget_aspire_proof_url
            
            await db.wallet_transfers.update_one(
                {"id": transfer_id},
                {"$set": {
                    proof_field: proof["new_gcs_path"],
                    f"{proof_type}_proof_pending_edit": False
                }}
            )
            logger.info(f"✅ Updated {proof_field} in wallet_transfers for transfer {transfer_id}")
            
        elif request_id and account_id and proof_type:
            # This is a topup request proof edit
            # Update the specific account in accounts array
            field_to_update = f"accounts.$.{proof_type}_proof_url" if proof_type in ["spend_limit", "budget_aspire"] else None
            
            if field_to_update:
                await db.topup_requests.update_one(
                    {"id": request_id, "accounts.account_id": account_id},
                    {"$set": {field_to_update: proof["new_gcs_path"]}}
                )
                logger.info(f"✅ Updated {proof_type}_proof_url in topup_requests for account {account_id}")
        
        # Notify admin who requested edit
        admin_id = proof.get("edit_requested_by")
        if admin_id:
            notification = {
                "id": str(uuid.uuid4()),
                "title": "✅ Proof Edit Approved",
                "message": f"Your proof edit request has been approved by {current_super_admin.username}",
                "type": "proof_edit_approved",
                "reference_id": proof_id,
                "is_read": False,
                "created_at": datetime.now(timezone.utc)
            }
            notification_dict = prepare_for_mongo(notification)
            await db.notifications.insert_one(notification_dict)
        
        # IMPORTANT: Save to admin_actions for history tracking
        action_record = {
            "id": str(uuid.uuid4()),
            "action_type": "proof_edit",
            "proof_type": proof_type,
            "proof_id": proof_id,
            "request_id": request_id,
            "account_id": account_id,
            "client_id": proof.get("client_id"),  # We need to get this from topup_request
            "edited_by": proof.get("edit_requested_by"),
            "edited_by_username": proof.get("edit_requested_by_username"),
            "super_admin_id": current_super_admin.id,
            "super_admin_username": current_super_admin.username,
            "status": "approved",
            "notes": notes or "Approved by Super Admin",
            "requested_at": proof.get("edit_requested_at"),
            "processed_at": datetime.now(timezone.utc),
            "created_at": datetime.now(timezone.utc)
        }
        
        # Get client_id from topup_request or wallet_transfer
        if transfer_id:
            wallet_transfer = await db.wallet_transfers.find_one({"id": transfer_id})
            if wallet_transfer:
                action_record["client_id"] = wallet_transfer.get("user_id")
        elif request_id:
            topup_request = await db.topup_requests.find_one({"id": request_id})
            if topup_request:
                action_record["client_id"] = topup_request.get("user_id")
        
        action_dict = prepare_for_mongo(action_record)
        await db.admin_actions.insert_one(action_dict)
        logger.info(f"✅ Saved proof edit approval to admin_actions history")
        
        logger.info(f"✅ Super admin {current_super_admin.username} approved proof edit for {proof_id}")
        
        return {"success": True, "message": "Proof edit approved successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Error approving proof edit: {e}")
        raise HTTPException(status_code=500, detail="Error approving proof edit")

@api_router.put("/super-admin/proof-edits/{proof_id}/reject")
async def reject_proof_edit(
    proof_id: str,
    notes: str = "",
    current_super_admin: AdminUser = Depends(get_current_super_admin)
):
    """Super admin rejects proof edit request"""
    try:
        # Get proof with pending edit
        proof = await db.payment_proofs.find_one({"id": proof_id, "pending_edit": True})
        if not proof:
            raise HTTPException(status_code=404, detail="Pending edit not found")
        
        # Add to history as rejected
        proof_history = proof.get("proof_history", [])
        proof_history.append({
            "file_path": proof.get("new_gcs_path"),
            "file_url": proof.get("new_file_url"),
            "edited_by": proof.get("edit_requested_by"),
            "edited_by_username": proof.get("edit_requested_by_username"),
            "edited_at": proof.get("edit_requested_at"),
            "rejected_by": current_super_admin.id,
            "rejected_by_username": current_super_admin.username,
            "rejected_at": datetime.now(timezone.utc).isoformat(),
            "status": "rejected",
            "notes": notes
        })
        
        # Delete new file from GCS
        try:
            gcs = get_gcs_storage()
            gcs_path = proof.get("new_gcs_path", "").replace("/files/", "")
            if gcs_path:
                gcs.delete_file(gcs_path)
                logger.info(f"🗑️ Deleted rejected proof file from GCS: {gcs_path}")
        except Exception as e:
            logger.warning(f"⚠️ Could not delete rejected file: {e}")
        
        # Remove pending edit fields
        await db.payment_proofs.update_one(
            {"id": proof_id},
            {"$set": {
                "pending_edit": False,
                "proof_history": proof_history
            },
            "$unset": {
                "new_gcs_path": "",
                "new_file_url": "",
                "edit_requested_by": "",
                "edit_requested_by_username": "",
                "edit_requested_at": "",
                "edit_notes": ""
            }}
        )
        
        # Also update pending_edit flag in wallet_transfers if applicable
        transfer_id = proof.get("transfer_id")
        proof_type = proof.get("proof_type")
        if transfer_id and proof_type:
            await db.wallet_transfers.update_one(
                {"id": transfer_id},
                {"$set": {f"{proof_type}_proof_pending_edit": False}}
            )
            logger.info(f"✅ Cleared pending_edit flag in wallet_transfers for transfer {transfer_id}")
        
        # Notify admin who requested edit
        admin_id = proof.get("edit_requested_by")
        if admin_id:
            notification = {
                "id": str(uuid.uuid4()),
                "title": "❌ Proof Edit Rejected",
                "message": f"Your proof edit request has been rejected by {current_super_admin.username}. Reason: {notes or 'No reason provided'}",
                "type": "proof_edit_rejected",
                "reference_id": proof_id,
                "is_read": False,
                "created_at": datetime.now(timezone.utc)
            }
            notification_dict = prepare_for_mongo(notification)
            await db.notifications.insert_one(notification_dict)
        
        # IMPORTANT: Save to admin_actions for history tracking
        request_id = proof.get("request_id")
        transfer_id = proof.get("transfer_id")
        account_id = proof.get("account_id")
        proof_type = proof.get("proof_type")
        
        action_record = {
            "id": str(uuid.uuid4()),
            "action_type": "proof_edit",
            "proof_type": proof_type,
            "proof_id": proof_id,
            "request_id": request_id,
            "account_id": account_id,
            "client_id": proof.get("client_id"),  # We need to get this from topup_request
            "edited_by": proof.get("edit_requested_by"),
            "edited_by_username": proof.get("edit_requested_by_username"),
            "super_admin_id": current_super_admin.id,
            "super_admin_username": current_super_admin.username,
            "status": "rejected",
            "notes": notes or "Rejected by Super Admin",
            "requested_at": proof.get("edit_requested_at"),
            "processed_at": datetime.now(timezone.utc),
            "created_at": datetime.now(timezone.utc)
        }
        
        # Get client_id from topup_request or wallet_transfer
        if transfer_id:
            wallet_transfer = await db.wallet_transfers.find_one({"id": transfer_id})
            if wallet_transfer:
                action_record["client_id"] = wallet_transfer.get("user_id")
        elif request_id:
            topup_request = await db.topup_requests.find_one({"id": request_id})
            if topup_request:
                action_record["client_id"] = topup_request.get("user_id")
        
        action_dict = prepare_for_mongo(action_record)
        await db.admin_actions.insert_one(action_dict)
        logger.info(f"✅ Saved proof edit rejection to admin_actions history")
        
        logger.info(f"❌ Super admin {current_super_admin.username} rejected proof edit for {proof_id}")
        
        return {"success": True, "message": "Proof edit rejected successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Error rejecting proof edit: {e}")
        raise HTTPException(status_code=500, detail="Error rejecting proof edit")

# ==================== END PROOF EDIT MANAGEMENT ====================

# ==================== REQUEST CLAIM/LOCK SYSTEM ====================

@api_router.post("/admin/requests/{request_type}/{request_id}/claim")
async def claim_request(
    request_type: str,  # "topup_request", "wallet_topup", "wallet_transfer", "withdrawal"
    request_id: str,
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Admin claims a request to work on it (prevents other admins from working on same request)"""
    try:
        # Determine which collection to use
        collection_map = {
            "topup_request": db.topup_requests,
            "wallet_topup": db.wallet_topup_requests,
            "wallet_transfer": db.wallet_transfers,
            "withdrawal": db.withdraw_requests
        }
        
        if request_type not in collection_map:
            raise HTTPException(status_code=400, detail="Invalid request type")
        
        collection = collection_map[request_type]
        
        # Get request
        request = await collection.find_one({"id": request_id})
        if not request:
            raise HTTPException(status_code=404, detail="Request not found")
        
        # Check if already claimed
        claimed_by = request.get("claimed_by")
        claimed_at = request.get("claimed_at")
        
        if claimed_by and claimed_by != current_admin.id:
            # Check if claim is expired (30 minutes)
            if claimed_at:
                if isinstance(claimed_at, str):
                    claimed_at = datetime.fromisoformat(claimed_at.replace('Z', '+00:00'))
                
                time_diff = datetime.now(timezone.utc) - claimed_at
                if time_diff.total_seconds() < 1800:  # 30 minutes
                    # Still claimed by another admin
                    admin = await db.admin_users.find_one({"id": claimed_by})
                    admin_name = admin.get("username") if admin else "Unknown Admin"
                    raise HTTPException(
                        status_code=409, 
                        detail=f"Request sedang dikerjakan oleh {admin_name}"
                    )
        
        # Claim the request
        update_data = {
            "claimed_by": current_admin.id,
            "claimed_by_username": current_admin.username,
            "claimed_at": datetime.now(timezone.utc).isoformat().replace('+00:00', '') + 'Z'
        }
        
        # Update status to processing if currently pending or uploaded
        current_status = request.get("status")
        if current_status in ["pending", "uploaded", "proof_uploaded"]:
            update_data["status"] = "processing"
        
        await collection.update_one(
            {"id": request_id},
            {"$set": update_data}
        )
        
        # Also update status in transactions collection for client visibility
        if "status" in update_data:
            # Find related transaction entry using reference_id
            transaction_query = {}
            if request_type == "wallet_transfer":
                transaction_query = {"reference_id": request_id, "type": "wallet_to_account_transfer"}
            elif request_type == "topup_request":
                transaction_query = {"reference_id": request_id, "type": "topup"}
            elif request_type == "wallet_topup":
                transaction_query = {"reference_id": request_id, "type": "wallet_topup"}
            elif request_type == "withdrawal":
                transaction_query = {"reference_id": request_id, "type": "withdraw_request"}
            
            if transaction_query:
                result = await db.transactions.update_many(
                    transaction_query,
                    {"$set": {"status": update_data["status"]}}
                )
                if result.modified_count > 0:
                    logger.info(f"✅ Updated {result.modified_count} transaction(s) to {update_data['status']}")
        
        logger.info(f"✅ Admin {current_admin.username} claimed {request_type} {request_id}")
        
        return {
            "success": True,
            "message": "Request berhasil diklaim",
            "claimed_by": current_admin.username
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Error claiming request: {e}")
        raise HTTPException(status_code=500, detail="Error claiming request")

@api_router.post("/admin/requests/{request_type}/{request_id}/release")
async def release_request(
    request_type: str,
    request_id: str,
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Admin releases a claimed request"""
    try:
        collection_map = {
            "topup_request": db.topup_requests,
            "wallet_topup": db.wallet_topup_requests,
            "wallet_transfer": db.wallet_transfers,
            "withdrawal": db.withdraw_requests
        }
        
        if request_type not in collection_map:
            raise HTTPException(status_code=400, detail="Invalid request type")
        
        collection = collection_map[request_type]
        
        # Get request
        request = await collection.find_one({"id": request_id})
        if not request:
            raise HTTPException(status_code=404, detail="Request not found")
        
        # Check if claimed by current admin
        if request.get("claimed_by") != current_admin.id:
            raise HTTPException(status_code=403, detail="You can only release requests claimed by you")
        
        # Release the request
        await collection.update_one(
            {"id": request_id},
            {"$unset": {
                "claimed_by": "",
                "claimed_by_username": "",
                "claimed_at": ""
            }}
        )
        
        logger.info(f"✅ Admin {current_admin.username} released {request_type} {request_id}")
        
        return {
            "success": True,
            "message": "Request berhasil direlease"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Error releasing request: {e}")
        raise HTTPException(status_code=500, detail="Error releasing request")

@api_router.post("/admin/requests/{request_type}/{request_id}/force-release")
async def force_release_request(
    request_type: str,
    request_id: str,
    current_super_admin: AdminUser = Depends(get_current_super_admin)
):
    """Super admin force releases a claimed request"""
    try:
        collection_map = {
            "topup_request": db.topup_requests,
            "wallet_topup": db.wallet_topup_requests,
            "wallet_transfer": db.wallet_transfers,
            "withdrawal": db.withdraw_requests
        }
        
        if request_type not in collection_map:
            raise HTTPException(status_code=400, detail="Invalid request type")
        
        collection = collection_map[request_type]
        
        # Get request
        request = await collection.find_one({"id": request_id})
        if not request:
            raise HTTPException(status_code=404, detail="Request not found")
        
        claimed_by = request.get("claimed_by")
        
        # Release the request
        await collection.update_one(
            {"id": request_id},
            {"$unset": {
                "claimed_by": "",
                "claimed_by_username": "",
                "claimed_at": ""
            }}
        )
        
        # Notify the admin who was working on it
        if claimed_by:
            notification = {
                "id": str(uuid.uuid4()),
                "title": "⚠️ Request Di-Release",
                "message": f"Request #{request_id[:8]} telah di-force release oleh Super Admin {current_super_admin.username}",
                "type": "force_release",
                "reference_id": request_id,
                "is_read": False,
                "created_at": datetime.now(timezone.utc)
            }
            notification_dict = prepare_for_mongo(notification)
            await db.notifications.insert_one(notification_dict)
        
        logger.info(f"✅ Super admin {current_super_admin.username} force released {request_type} {request_id}")
        
        return {
            "success": True,
            "message": "Request berhasil di-force release"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Error force releasing request: {e}")
        raise HTTPException(status_code=500, detail="Error force releasing request")

# ==================== END REQUEST CLAIM/LOCK SYSTEM ====================

@api_router.get("/admin/verification-files/{file_path:path}")
async def get_verification_file(
    file_path: str,
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get verification files (spend_limit_proof, budget_aspire_proof) for admin"""
    try:
        # Handle both relative and absolute paths
        full_path = Path(file_path)
        if not full_path.is_absolute():
            # If relative, prepend current working directory
            full_path = Path("/app") / file_path
        
        if not full_path.exists():
            logger.error(f"Verification file not found at: {full_path}")
            raise HTTPException(status_code=404, detail="File not found")
        
        # Determine content type
        file_name = full_path.name
        content_type = "image/jpeg"
        if file_name.lower().endswith('.png'):
            content_type = "image/png"
        elif file_name.lower().endswith('.pdf'):
            content_type = "application/pdf"
        elif file_name.lower().endswith('.webp'):
            content_type = "image/webp"
        
        return FileResponse(
            path=full_path,
            media_type=content_type,
            headers={
                "Access-Control-Allow-Origin": "*",
                "Access-Control-Allow-Methods": "GET",
                "Access-Control-Allow-Headers": "Authorization"
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error serving verification file: {e}")
        raise HTTPException(status_code=500, detail="Error serving file")


@api_router.get("/files/{file_path:path}")
async def get_gcs_file(
    file_path: str,
    authorization: Optional[str] = Header(None)
):
    """Get file from Google Cloud Storage - Generic endpoint for all GCS files
    
    Access Control:
    - Profile pictures: Accessible by authenticated users (admin or client)
    - Payment proofs: Admin only
    - Other files: Requires authentication
    """
    try:
        # Verify authentication if Authorization header is provided
        authenticated_user = None
        is_admin = False
        
        if authorization:
            try:
                # Try to extract token
                token = authorization.replace("Bearer ", "")
                payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
                user_type = payload.get("user_type")
                
                if user_type == "admin":
                    is_admin = True
                    authenticated_user = payload.get("sub")
                elif user_type == "client":
                    authenticated_user = payload.get("sub")
            except:
                pass  # Continue without auth for public files
        
        # For profile pictures in profile_pictures/ folder, allow authenticated users
        if file_path.startswith('profile_pictures/'):
            if not authenticated_user:
                raise HTTPException(status_code=401, detail="Authentication required for profile pictures")
        
        # For payment proofs and sensitive files, require admin access
        elif any(file_path.startswith(prefix) for prefix in ['payment_proofs/', 'wallet_payment_proofs/', 'verification_files/']):
            if not is_admin:
                raise HTTPException(status_code=403, detail="Admin access required for this file")
        
        # Initialize GCS client
        from google.cloud import storage
        storage_client = storage.Client()
        bucket_name = os.getenv("GCS_BUCKET_NAME", "rimuru-file-uploads")
        bucket = storage_client.bucket(bucket_name)
        
        # Get the blob
        blob = bucket.blob(file_path)
        
        if not blob.exists():
            logger.error(f"GCS file not found: {file_path}")
            raise HTTPException(status_code=404, detail="File not found in storage")
        
        # Download blob content
        file_content = blob.download_as_bytes()
        
        # Determine content type from blob or filename
        content_type = blob.content_type or "application/octet-stream"
        file_name = file_path.split('/')[-1]
        
        if not blob.content_type:
            # Infer from extension
            if file_name.lower().endswith(('.jpg', '.jpeg')):
                content_type = "image/jpeg"
            elif file_name.lower().endswith('.png'):
                content_type = "image/png"
            elif file_name.lower().endswith('.pdf'):
                content_type = "application/pdf"
            elif file_name.lower().endswith('.webp'):
                content_type = "image/webp"
        
        return Response(
            content=file_content,
            media_type=content_type,
            headers={
                "Access-Control-Allow-Origin": "*",
                "Access-Control-Allow-Methods": "GET",
                "Access-Control-Allow-Headers": "Authorization",
                "Content-Disposition": f'inline; filename="{file_name}"'
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error serving GCS file: {e}")
        raise HTTPException(status_code=500, detail=f"Error serving file: {str(e)}")

class TransferRequestStatusUpdate(BaseModel):
    status: str
    admin_notes: Optional[str] = None
    spend_limit_proof_url: Optional[str] = None
    budget_aspire_proof_url: Optional[str] = None

@api_router.put("/admin/transfer-requests/{request_id}/status", response_model=dict)
async def update_transfer_request_status(
    request_id: str,
    request_data: TransferRequestStatusUpdate,
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Update transfer request status (approve/reject)"""
    try:
        # Validate status
        if request_data.status not in ["approved", "rejected"]:
            raise HTTPException(status_code=400, detail="Invalid status. Must be 'approved' or 'rejected'")
            
        # Get transfer request
        transfer_request = await db.transfer_requests.find_one({"id": request_id})
        if not transfer_request:
            raise HTTPException(status_code=404, detail="Transfer request not found")
            
        if transfer_request["status"] != "pending":
            raise HTTPException(status_code=400, detail="Transfer request is not pending")
        
        # Update transfer request
        update_data = {
            "status": request_data.status,
            "admin_id": current_admin.id,
            "processed_at": datetime.now(timezone.utc),
            "admin_notes": request_data.admin_notes
        }
        
        # Add proof URLs if provided (for approved status)
        if request_data.spend_limit_proof_url:
            update_data["spend_limit_proof_url"] = request_data.spend_limit_proof_url
        if request_data.budget_aspire_proof_url:
            update_data["budget_aspire_proof_url"] = request_data.budget_aspire_proof_url
        
        await db.transfer_requests.update_one(
            {"id": request_id},
            {"$set": prepare_for_mongo(update_data)}
        )
        
        # If approved, process the transfer
        if request_data.status == "approved":
            # Get user and account details
            user = await db.users.find_one({"id": transfer_request["user_id"]})
            account = await db.ad_accounts.find_one({"id": transfer_request["account_id"]})
            
            if user and account:
                # Determine currency and wallet field
                currency = transfer_request["currency"]
                wallet_field = "wallet_balance_idr" if currency == "IDR" else "wallet_balance_usd"
                
                # Check wallet balance again
                current_wallet_balance = user.get(wallet_field, 0)
                if current_wallet_balance >= transfer_request["amount"]:
                    # Update wallet balance using precise decimal calculations
                    current_wallet_decimal = to_decimal(current_wallet_balance)
                    transfer_amount_decimal = to_decimal(transfer_request["amount"])
                    new_wallet_decimal = decimal_subtract(current_wallet_decimal, transfer_amount_decimal)
                    new_wallet_balance = to_float(decimal_round(new_wallet_decimal))
                    
                    await db.users.update_one(
                        {"id": transfer_request["user_id"]},
                        {"$set": {wallet_field: new_wallet_balance}}
                    )
                    
                    # Update account balance using precise decimal calculations
                    current_account_balance = account.get("balance", 0)
                    current_account_decimal = to_decimal(current_account_balance)
                    new_account_decimal = decimal_add(current_account_decimal, transfer_amount_decimal)
                    new_account_balance = to_float(decimal_round(new_account_decimal))
                    
                    await db.ad_accounts.update_one(
                        {"id": transfer_request["account_id"]},
                        {"$set": {"balance": new_account_balance}}
                    )
                    
                    # Create transaction record
                    transaction = Transaction(
                        user_id=transfer_request["user_id"],
                        type="approved_transfer",
                        amount=transfer_request["amount"],
                        currency=currency,
                        description=f"Approved transfer to {account['account_name']} ({account['platform']})",
                        status="completed",
                        metadata={
                            "transfer_request_id": request_id,
                            "account_id": transfer_request["account_id"],
                            "account_name": account["account_name"],
                            "platform": account["platform"],
                            "approved_by": current_admin.username,
                            "wallet_balance_before": current_wallet_balance,
                            "wallet_balance_after": new_wallet_balance,
                            "account_balance_before": current_account_balance,
                            "account_balance_after": new_account_balance
                        }
                    )
                    
                    await db.transactions.insert_one(prepare_for_mongo(transaction.dict()))
                    
                    # Update transfer request to completed
                    await db.transfer_requests.update_one(
                        {"id": request_id},
                        {"$set": {"status": "completed"}}
                    )
                    
                    # Create notification for client - transfer approved
                    currency_symbol = "Rp " if currency == "IDR" else "$"
                    formatted_amount = f"{transfer_request['amount']:,.0f}" if currency == "IDR" else f"{transfer_request['amount']:.2f}"
                    client_notification = {
                        "id": str(uuid.uuid4()),
                        "user_id": transfer_request["user_id"], 
                        "title": "Transfer Berhasil",
                        "message": f"Transfer {currency_symbol}{formatted_amount} ke {account['account_name']} telah berhasil diproses.",
                        "type": "transfer_approved",
                        "reference_id": request_id,
                        "is_read": False,
                        "created_at": datetime.now(timezone.utc)
                    }
                    
                    client_notification_dict = prepare_for_mongo(client_notification)
                    await db.client_notifications.insert_one(client_notification_dict)
                    
                    return {
                        "message": "Transfer request approved and processed successfully",
                        "status": "completed",
                        "amount_transferred": transfer_request["amount"],
                        "new_wallet_balance": new_wallet_balance,
                        "new_account_balance": new_account_balance
                    }
                else:
                    # Insufficient balance, mark as failed
                    await db.transfer_requests.update_one(
                        {"id": request_id},
                        {"$set": {"status": "failed", "admin_notes": "Insufficient wallet balance at processing time"}}
                    )
                    
                    # Create notification for client - transfer failed
                    currency = transfer_request["currency"]
                    currency_symbol = "Rp " if currency == "IDR" else "$"
                    formatted_amount = f"{transfer_request['amount']:,.0f}" if currency == "IDR" else f"{transfer_request['amount']:.2f}"
                    client_notification = {
                        "id": str(uuid.uuid4()),
                        "user_id": transfer_request["user_id"], 
                        "title": "Transfer Gagal",
                        "message": f"Transfer {currency_symbol}{formatted_amount} gagal diproses karena saldo wallet tidak mencukupi.",
                        "type": "transfer_failed", 
                        "reference_id": request_id,
                        "is_read": False,
                        "created_at": datetime.now(timezone.utc)
                    }
                    
                    client_notification_dict = prepare_for_mongo(client_notification)
                    await db.client_notifications.insert_one(client_notification_dict)
                    
                    return {
                        "message": "Transfer request approved but failed due to insufficient wallet balance",
                        "status": "failed"
                    }
        
        # Handle rejected status
        elif request_data.status == "rejected":
            # Create notification for client - transfer rejected
            user = await db.users.find_one({"id": transfer_request["user_id"]})
            account = await db.ad_accounts.find_one({"id": transfer_request["account_id"]})
            
            if user and account:
                currency = transfer_request["currency"]
                currency_symbol = "Rp " if currency == "IDR" else "$"
                formatted_amount = f"{transfer_request['amount']:,.0f}" if currency == "IDR" else f"{transfer_request['amount']:.2f}"
                client_notification = {
                    "id": str(uuid.uuid4()),
                    "user_id": transfer_request["user_id"], 
                    "title": "Transfer Ditolak",
                    "message": f"Transfer request {currency_symbol}{formatted_amount} ke {account['account_name']} telah ditolak oleh admin.",
                    "type": "transfer_rejected",
                    "reference_id": request_id,
                    "is_read": False,
                    "created_at": datetime.now(timezone.utc)
                }
                
                client_notification_dict = prepare_for_mongo(client_notification)
                await db.client_notifications.insert_one(client_notification_dict)
        
        return {
            "message": f"Transfer request {request_data.status} successfully",
            "status": request_data.status
        }
        
    except HTTPException:
        # Re-raise HTTPExceptions (like 400, 404) without modification
        raise
    except Exception as e:
        logger.error(f"Error updating transfer request status: {e}")
        raise HTTPException(status_code=500, detail="Error updating transfer request status")
    if existing_group:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Group name already exists"
        )
    
    # Create new group
    group = Group(
        name=group_data.name,
        user_id=current_user.id
    )
    
    group_dict = prepare_for_mongo(group.dict())
    result = await db.groups.insert_one(group_dict)
    
    if not result.inserted_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Failed to create group"
        )
    
    return group

@api_router.post("/profile/picture", response_model=dict)
async def upload_profile_picture(file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    # Validate file type
    allowed_types = ["image/jpeg", "image/jpg", "image/png", "image/webp"]
    if file.content_type not in allowed_types:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only JPEG, PNG, and WebP images are allowed"
        )
    
    # Upload to GCS
    gcs_data = await upload_to_gcs(file, folder="profile_pictures")
    
    # Validate file size (max 5MB)
    if gcs_data["file_size"] > 5 * 1024 * 1024:  # 5MB
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File size must be less than 5MB"
        )
    
    # Update user profile picture with GCS path
    profile_picture_gcs = gcs_data["gcs_path"]
    # Store as path format: /files/{folder}/{filename} (without /api prefix)
    # Frontend will add /api prefix when fetching
    profile_picture_path = f"/files/{profile_picture_gcs}"
    result = await db.users.update_one(
        {"id": current_user.id},
        {"$set": {
            "profile_picture": profile_picture_path,
            "profile_picture_gcs": profile_picture_gcs,  # Store raw GCS path for reference
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    if result.modified_count == 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Failed to update profile picture"
        )
    
    return {
        "message": "Profile picture uploaded successfully",
        "profile_picture": profile_picture_path,
        "storage": "gcs"
    }

# Duplicate endpoint removed - use /api/files/{file_path:path} at line 9099 instead
# (This endpoint was unauthenticated and conflicted with the authenticated version)

# Dashboard endpoints
@api_router.post("/accounts/request", response_model=dict)
async def request_ad_account(request: AdAccountRequest, current_user: User = Depends(get_current_user)):
    """Create ad account request with rate limiting and duplicate prevention"""
    
    # Rate limiting: Check if user has created request in last 5 seconds
    five_seconds_ago = datetime.now(timezone.utc) - timedelta(seconds=5)
    recent_request = await db.ad_account_requests.find_one({
        "user_id": current_user.id,
        "created_at": {"$gte": five_seconds_ago},
        "platform": request.platform,
        "account_name": request.account_name
    })
    
    if recent_request:
        raise HTTPException(
            status_code=429,
            detail="Request terlalu cepat. Mohon tunggu beberapa detik sebelum submit lagi."
        )
    
    # CRITICAL FIX: Check for duplicate account name for this user
    # Check in pending/approved requests
    duplicate_request = await db.ad_account_requests.find_one({
        "user_id": current_user.id,
        "platform": request.platform,
        "account_name": request.account_name,
        "status": {"$in": ["pending", "processing"]}  # Don't allow duplicate if still pending/processing
    })
    
    if duplicate_request:
        raise HTTPException(
            status_code=400,
            detail=f"Anda sudah memiliki request dengan nama '{request.account_name}' yang masih dalam proses. Mohon tunggu hingga diproses atau gunakan nama lain."
        )
    
    # Check in existing active accounts
    existing_account = await db.ad_accounts.find_one({
        "user_id": current_user.id,
        "platform": request.platform,
        "account_name": request.account_name,
        "status": {"$in": ["active", "inactive"]}  # Check active or inactive accounts
    })
    
    if existing_account:
        raise HTTPException(
            status_code=400,
            detail=f"Anda sudah memiliki akun dengan nama '{request.account_name}'. Mohon gunakan nama yang berbeda."
        )
    
    # Create ad account request (not actual account)
    request_data = {
        "id": str(uuid.uuid4()),
        "user_id": current_user.id,
        "platform": request.platform,
        "account_name": request.account_name,
        "status": "pending",
        "created_at": datetime.now(timezone.utc),
        # Platform-specific fields
        "gmt": request.gmt,
        "currency": request.currency,
        "delivery_method": request.delivery_method,
        "bm_id_or_email": request.bm_id_or_email,  # Legacy single BM
        "bm_ids": request.bm_ids if request.bm_ids else [],  # NEW: Array of BM IDs
        "email": request.email,  # For Google Ads
        "website": request.website,  # For Google Ads
        "bc_id": request.bc_id,  # For TikTok Ads
        "notes": request.notes,
        "group_id": request.group_id
    }
    
    request_dict = prepare_for_mongo(request_data)
    await db.ad_account_requests.insert_one(request_dict)
    
    # Create transaction record with more details for Facebook
    if request.platform == "facebook":
        description = f"Request Facebook ads account: {request.account_name} (GMT: {request.gmt}, Currency: {request.currency})"
    else:
        description = f"Request {request.platform} ads account: {request.account_name}"
        
    transaction = Transaction(
        user_id=current_user.id,
        type="account_request",
        amount=0.0,
        description=description
    )
    
    transaction_dict = prepare_for_mongo(transaction.dict())
    await db.transactions.insert_one(transaction_dict)
    
    # Create notification for admin
    platform_name = request.platform.title()
    if request.platform == "facebook":
        platform_name = "Facebook Ads"
    elif request.platform == "google":
        platform_name = "Google Ads"
    elif request.platform == "tiktok":
        platform_name = "TikTok Ads"
    
    # Create admin notification (idempotency handled inside create_localized_notification)
    await create_localized_notification(
        title_key="new_account_request",
        message_key="account_request_submitted",
        notification_type="account_request",
        lang="id",
        reference_id=request_data["id"],
        platform=platform_name,
        username=current_user.username
    )
    
    return {"message": "Ad account request submitted successfully", "request_id": request_data["id"]}

@api_router.post("/topup", response_model=dict)
async def create_topup_request(request: TopUpRequest, current_user: User = Depends(get_current_user)):
    # Generate unique reference code
    reference_code = f"RMR{str(uuid.uuid4())[:8].upper()}"
    
    # Use unique code from frontend if provided, otherwise generate new one
    # This ensures the code displayed to user matches the one saved in backend
    if request.unique_code is not None and request.currency == "IDR":
        unique_code = request.unique_code
        logger.info(f"✅ Using unique code from frontend: {unique_code}")
    else:
        # Generate 3-digit unique code (100-999) for payment verification
        import random
        unique_code = random.randint(100, 999)
        logger.info(f"✅ Generated new unique code: {unique_code}")
    
    # Calculate total with unique code (only for IDR bank transfers)
    if request.currency == "IDR":
        total_with_unique_code = request.total_amount + unique_code
    else:
        # For USD crypto, no unique code needed
        total_with_unique_code = request.total_amount
        unique_code = 0
    
    # Set bank/wallet details based on currency
    if request.currency == "IDR":
        bank_details = {
            "bank_name": "BRI",
            "bank_account": "057901002665566",
            "bank_holder": "PT RINAIYANTI CAHAYA INTERMA",
            "wallet_address": None,
            "wallet_name": None,
            "network": None
        }
    else:  # USD
        bank_details = {
            "bank_name": None,
            "bank_account": None,
            "bank_holder": None,
            "wallet_address": "TBCJiUoYGxBYBpqMNb9ZtuwXWJLuCwBPXa",
            "wallet_name": "BINANCE",
            "network": "USDT TRC20"
        }
    
    # Fetch account details for each account
    accounts_with_details = []
    for account in request.accounts:
        account_doc = await db.ad_accounts.find_one({"id": account.account_id})
        account_dict = account.dict()
        if account_doc:
            account_dict["account_name"] = account_doc.get("account_name", "Unknown")
            account_dict["account_platform"] = account_doc.get("platform", "Unknown")
            account_dict["platform_account_id"] = account_doc.get("account_id", "N/A")  # Platform's account ID (FB/Google/TikTok ID)
        else:
            account_dict["account_name"] = "Unknown"
            account_dict["account_platform"] = "Unknown"
            account_dict["platform_account_id"] = "N/A"
        accounts_with_details.append(account_dict)
    
    # Create top-up request record
    topup_request = TopUpRequestRecord(
        user_id=current_user.id,
        currency=request.currency,
        accounts=accounts_with_details,  # Store array of accounts with details
        total_amount=request.total_amount,
        total_fee=request.total_fee,
        unique_code=unique_code,
        total_with_unique_code=total_with_unique_code,
        reference_code=reference_code,
        **bank_details
    )
    
    topup_request_dict = prepare_for_mongo(topup_request.dict())
    await db.topup_requests.insert_one(topup_request_dict)
    
    # Create admin notification
    currency_symbol = "Rp " if request.currency == "IDR" else "$"
    formatted_amount = f"{currency_symbol}{request.total_amount:,.2f}"
    
    await create_localized_notification(
        title_key="new_topup_request",
        message_key="topup_request_submitted",
        notification_type="topup_request",
        lang="id",
        reference_id=topup_request.id,
        username=current_user.username,
        amount=formatted_amount
    )
    
    # Send email notification to all active admins
    try:
        admin_emails = await get_active_admin_emails()
        if admin_emails and accounts_with_details:
            from email_service import send_admin_new_topup_request_email
            # Get first account details for email
            first_account = accounts_with_details[0]
            send_admin_new_topup_request_email(
                admin_emails=admin_emails,
                client_name=current_user.name or current_user.username,
                amount=request.total_amount,
                currency=request.currency,
                platform=first_account.get("account_platform", "Unknown"),
                account_name=first_account.get("account_name", "Unknown")
            )
            logger.info(f"📧 Admin top-up notification emails sent to {len(admin_emails)} admins")
    except Exception as e:
        logger.error(f"❌ Failed to send admin top-up notification emails: {e}")
    
    # Prepare return bank details based on currency
    if request.currency == "IDR":
        return_details = {
            "type": "bank_transfer",
            "bank_name": "BRI", 
            "account_number": "057901002665566",
            "account_holder": "PT RINAIYANTI CAHAYA INTERMA",
            "subtotal": request.total_amount,
            "unique_code": unique_code,
            "total_transfer": total_with_unique_code,
            "currency": request.currency
        }
    else:  # USD
        return_details = {
            "type": "crypto_wallet",
            "wallet_address": "TBCJiUoYGxBYBpqMNb9ZtuwXWJLuCwBPXa",
            "wallet_name": "BINANCE",
            "network": "USDT TRC20",
            "amount": request.total_amount,
            "currency": request.currency
        }
    
    return {
        "message": "Top-up request created successfully",
        "request_id": topup_request.id,
        "reference_code": reference_code,
        "accounts": accounts_with_details,  # Return account details for display
        "transfer_details": return_details
    }

@api_router.put("/topup-request/{request_id}/cancel")
async def cancel_topup_request(
    request_id: str,
    current_user: User = Depends(get_current_user)
):
    """Cancel a pending top-up request"""
    try:
        # Get the top-up request
        request = await db.topup_requests.find_one({"id": request_id})
        if not request:
            raise HTTPException(status_code=404, detail="Top-up request not found")
        
        # Check if user owns this request
        if request["user_id"] != current_user.id:
            raise HTTPException(status_code=403, detail="Not authorized to cancel this request")
        
        # Check if request can be cancelled (only pending requests)
        if request["status"] != "pending":
            raise HTTPException(
                status_code=400, 
                detail=f"Cannot cancel request with status: {request['status']}"
            )
        
        # Update request status to cancelled
        await db.topup_requests.update_one(
            {"id": request_id},
            {"$set": {
                "status": "cancelled",
                "cancelled_at": datetime.now(timezone.utc),
                "cancelled_by": "user"
            }}
        )
        
        return {
            "success": True,
            "message": "Top-up request cancelled successfully"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error cancelling top-up request: {e}")
        raise HTTPException(status_code=500, detail="Error cancelling top-up request")

async def auto_cancel_expired_topup_requests():
    """Auto-cancel top-up requests that are older than 24 hours and still pending"""
    try:
        # Calculate 24 hours ago
        twenty_four_hours_ago = datetime.now(timezone.utc) - timedelta(hours=24)
        
        # Find all pending requests older than 24 hours
        expired_requests = await db.topup_requests.find({
            "status": "pending",
            "created_at": {"$lt": twenty_four_hours_ago}
        }).to_list(None)
        
        cancelled_count = 0
        for request in expired_requests:
            try:
                # Update request status to cancelled
                await db.topup_requests.update_one(
                    {"id": request["id"]},
                    {"$set": {
                        "status": "cancelled",
                        "cancelled_at": datetime.now(timezone.utc),
                        "cancelled_by": "system_auto"
                    }}
                )
                
                # Create notification for user about auto-cancellation
                user_notification = {
                    "id": str(uuid.uuid4()),
                    "user_id": request["user_id"], 
                    "title": "Top-up Dibatalkan Otomatis",
                    "message": f"Top-up request dengan kode {request.get('reference_code', request['id'][:8])} telah dibatalkan otomatis karena tidak ada pembayaran dalam 24 jam.",
                    "type": "warning",
                    "reference_id": request["id"],
                    "is_read": False,
                    "created_at": datetime.now(timezone.utc)
                }
                
                user_notification_dict = prepare_for_mongo(user_notification)
                await db.client_notifications.insert_one(user_notification_dict)
                
                # Send email notification to client about auto-cancellation
                try:
                    user = await db.users.find_one({"id": request["user_id"]})
                    if user and user.get("email"):
                        # Get account info
                        account = await db.ad_accounts.find_one({"id": request.get("account_id", "")})
                        account_name = account.get("account_name", "Unknown") if account else "Unknown"
                        platform = account.get("platform", "unknown") if account else "unknown"
                        
                        send_client_topup_auto_cancelled_email(
                            client_email=user["email"],
                            client_name=user.get("full_name") or user.get("username"),
                            amount=request.get("total_amount", 0),
                            currency=request.get("currency", "IDR"),
                            account_name=account_name,
                            platform=platform
                        )
                        logger.info(f"📧 Auto-cancelled top-up notification sent to {user['email']}")
                except Exception as e:
                    logger.error(f"Failed to send auto-cancelled email: {e}")
                
                cancelled_count += 1
                logger.info(f"Auto-cancelled expired top-up request: {request['id']}")
                
            except Exception as e:
                logger.error(f"Error auto-cancelling request {request['id']}: {e}")
        
        if cancelled_count > 0:
            logger.info(f"Auto-cancelled {cancelled_count} expired top-up requests")
        
        return cancelled_count
        
    except Exception as e:
        logger.error(f"Error in auto_cancel_expired_topup_requests: {e}")
        return 0


async def auto_cancel_expired_wallet_topup_requests():
    """Auto-cancel wallet top-up requests that are older than 24 hours and still pending without payment proof"""
    try:
        # Calculate 24 hours ago
        twenty_four_hours_ago = datetime.now(timezone.utc) - timedelta(hours=24)
        
        # Find all pending wallet top-up requests older than 24 hours without payment proof
        expired_requests = await db.wallet_topup_requests.find({
            "status": "pending",
            "created_at": {"$lt": twenty_four_hours_ago},
            "payment_proof_id": {"$exists": False}  # Only cancel if no proof uploaded
        }).to_list(None)
        
        cancelled_count = 0
        for request in expired_requests:
            try:
                # Update request status to cancelled
                await db.wallet_topup_requests.update_one(
                    {"id": request["id"]},
                    {"$set": {
                        "status": "cancelled",
                        "cancelled_at": datetime.now(timezone.utc),
                        "cancelled_by": "system_auto"
                    }}
                )
                
                # Create notification for user about auto-cancellation
                user_notification = {
                    "id": str(uuid.uuid4()),
                    "user_id": request["user_id"],
                    "title": "Wallet Top-Up Dibatalkan Otomatis",
                    "message": f"Wallet top-up {request.get('wallet_type', 'main').title()} {request.get('currency', 'IDR')} sebesar {request.get('amount', 0):,.0f} telah dibatalkan otomatis karena belum upload bukti pembayaran dalam 24 jam.",
                    "type": "warning",
                    "reference_id": request["id"],
                    "is_read": False,
                    "created_at": datetime.now(timezone.utc)
                }
                
                user_notification_dict = prepare_for_mongo(user_notification)
                await db.client_notifications.insert_one(user_notification_dict)
                
                # Send email notification to client about auto-cancellation
                try:
                    user = await db.users.find_one({"id": request["user_id"]})
                    if user and user.get("email"):
                        send_client_wallet_topup_auto_cancelled_email(
                            client_email=user["email"],
                            client_name=user.get("name") or user.get("username"),
                            amount=request.get("amount", 0),
                            currency=request.get("currency", "IDR"),
                            wallet_type=request.get("wallet_type", "main")
                        )
                        logger.info(f"📧 Auto-cancelled wallet top-up notification sent to {user['email']}")
                except Exception as e:
                    logger.error(f"Failed to send auto-cancelled wallet top-up email: {e}")
                
                cancelled_count += 1
                logger.info(f"Auto-cancelled expired wallet top-up request: {request['id']}")
                
            except Exception as e:
                logger.error(f"Error auto-cancelling wallet top-up request {request['id']}: {e}")
        
        if cancelled_count > 0:
            logger.info(f"Auto-cancelled {cancelled_count} expired wallet top-up requests")
        
        return cancelled_count
        
    except Exception as e:
        logger.error(f"Error in auto_cancel_expired_wallet_topup_requests: {e}")
        return 0


@api_router.post("/admin/auto-cancel-expired-topups")
async def trigger_auto_cancel_expired_topups(current_admin: AdminUser = Depends(get_current_admin)):
    """Manually trigger auto-cancel of expired top-up requests"""
    try:
        cancelled_count = await auto_cancel_expired_topup_requests()
        return {
            "success": True,
            "cancelled_count": cancelled_count,
            "message": f"Auto-cancelled {cancelled_count} expired top-up requests"
        }
    except Exception as e:
        logger.error(f"Error in manual auto-cancel trigger: {e}")
        raise HTTPException(status_code=500, detail="Error triggering auto-cancel")


@api_router.post("/admin/auto-cancel-expired-wallet-topups")
async def trigger_auto_cancel_expired_wallet_topups(current_admin: AdminUser = Depends(get_current_admin)):
    """Manually trigger auto-cancel of expired wallet top-up requests"""
    try:
        cancelled_count = await auto_cancel_expired_wallet_topup_requests()
        return {
            "success": True,
            "cancelled_count": cancelled_count,
            "message": f"Auto-cancelled {cancelled_count} expired wallet top-up requests"
        }
    except Exception as e:
        logger.error(f"Error in manual auto-cancel wallet trigger: {e}")
        raise HTTPException(status_code=500, detail="Error triggering auto-cancel for wallet top-ups")

@api_router.get("/topup-requests", response_model=List[dict])
async def get_user_topup_requests(current_user: User = Depends(get_current_user)):
    """Get user's own top-up requests"""
    requests = await db.topup_requests.find({
        "user_id": current_user.id
    }).sort("created_at", -1).to_list(100)
    
    result = []
    for req in requests:
        req = parse_from_mongo(req)
        
        # Get payment proof if exists
        payment_proof = None
        if req.get("payment_proof_id"):
            proof = await db.payment_proofs.find_one({"id": req["payment_proof_id"]})
            if proof:
                payment_proof = parse_from_mongo(proof)
        
        result.append({
            "id": req["id"],
            "reference_code": req.get("reference_code", "N/A"),
            "currency": req["currency"],
            "total_amount": req["total_amount"],
            "total_fee": req.get("total_fee", 0),
            "accounts_count": len(req.get("accounts", [])),
            "accounts": req.get("accounts", []),  # Include full accounts array for display
            "unique_code": req.get("unique_code", 0),
            "status": req["status"],
            "created_at": req["created_at"],
            "verified_at": req.get("verified_at"),
            "admin_notes": req.get("admin_notes"),
            "payment_proof": {
                "uploaded": payment_proof is not None,
                "uploaded_at": payment_proof["uploaded_at"] if payment_proof else None,
                "file_name": payment_proof["file_name"] if payment_proof else None
            }
        })
    
    return result

@api_router.post("/topup/{request_id}/upload-proof", response_model=dict)
async def upload_payment_proof(
    request_id: str,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    # Validate top-up request exists and belongs to user
    topup_request = await db.topup_requests.find_one({
        "id": request_id,
        "user_id": current_user.id
    })
    if not topup_request:
        raise HTTPException(status_code=404, detail="Top-up request not found")
    
    if topup_request["status"] not in ["pending", "proof_uploaded", "rejected"]:
        raise HTTPException(status_code=400, detail="Cannot upload proof for this request")
    
    # Validate file type and size
    allowed_types = ["image/jpeg", "image/png", "image/jpg", "application/pdf"]
    max_size = 10 * 1024 * 1024  # 10MB
    
    if file.content_type not in allowed_types:
        raise HTTPException(
            status_code=400, 
            detail="Only JPG, PNG, and PDF files are allowed"
        )
    
    # Upload to GCS
    gcs_data = await upload_to_gcs(file, folder="payment_proofs")
    
    if gcs_data["file_size"] > max_size:
        raise HTTPException(status_code=400, detail="File size must be less than 10MB")
    
    # Create payment proof record with GCS data
    payment_proof = PaymentProof(
        topup_request_id=request_id,
        user_id=current_user.id,
        file_name=gcs_data["file_name"],
        gcs_path=gcs_data["gcs_path"],
        gcs_bucket=gcs_data["gcs_bucket"],
        storage_type="gcs",
        file_size=gcs_data["file_size"],
        mime_type=gcs_data["mime_type"]
    )
    
    payment_proof_dict = prepare_for_mongo(payment_proof.dict())
    await db.payment_proofs.insert_one(payment_proof_dict)
    
    # Update top-up request status and link to proof
    await db.topup_requests.update_one(
        {"id": request_id},
        {
            "$set": {
                "status": "proof_uploaded",
                "payment_proof_id": payment_proof.id,
                "proof_uploaded_at": datetime.now(timezone.utc)  # Add timestamp
            }
        }
    )
    
    # Create admin notification
    await create_localized_notification(
        title_key="payment_proof_uploaded",
        message_key="proof_uploaded_message",
        notification_type="payment_proof",
        lang="id",
        reference_id=request_id,
        username=current_user.username,
        code=topup_request.get('reference_code', request_id)
    )
    
    return {
        "message": "Payment proof uploaded successfully",
        "status": "proof_uploaded"
    }

@api_router.get("/topup/{request_id}/status", response_model=dict)
async def get_topup_status(request_id: str, current_user: User = Depends(get_current_user)):
    topup_request = await db.topup_requests.find_one({
        "id": request_id,
        "user_id": current_user.id
    })
    if not topup_request:
        raise HTTPException(status_code=404, detail="Top-up request not found")
    
    topup_request = parse_from_mongo(topup_request)
    
    # Get payment proof if exists
    payment_proof = None
    if topup_request.get("payment_proof_id"):
        proof = await db.payment_proofs.find_one({"id": topup_request["payment_proof_id"]})
        if proof:
            payment_proof = parse_from_mongo(proof)
    
    # Prepare transfer details with unique code info
    if topup_request["currency"] == "IDR":
        transfer_details = {
            "type": "bank_transfer",
            "bank_name": topup_request.get("bank_name"),
            "account_number": topup_request.get("bank_account"),
            "account_holder": topup_request.get("bank_holder"),
            "subtotal": topup_request["total_amount"],
            "unique_code": topup_request.get("unique_code", 0),
            "total_transfer": topup_request.get("total_with_unique_code", topup_request["total_amount"])
        }
    else:
        transfer_details = {
            "type": "crypto_wallet",
            "wallet_address": topup_request.get("wallet_address"),
            "wallet_name": topup_request.get("wallet_name"),
            "network": topup_request.get("network"),
            "amount": topup_request["total_amount"]
        }
    
    return {
        "request_id": topup_request["id"],
        "status": topup_request["status"],
        "reference_code": topup_request.get("reference_code", "N/A"),
        "total_amount": topup_request["total_amount"],
        "currency": topup_request["currency"],
        "accounts": topup_request.get("accounts", []),  # Include accounts array
        "created_at": topup_request["created_at"],
        "verified_at": topup_request.get("verified_at"),
        "admin_notes": topup_request.get("admin_notes"),
        "transfer_details": transfer_details,
        "payment_proof": {
            "uploaded": payment_proof is not None,
            "uploaded_at": payment_proof["uploaded_at"] if payment_proof else None,
            "file_name": payment_proof["file_name"] if payment_proof else None
        }
    }

@api_router.get("/wallet-topup/{request_id}/status", response_model=dict)
async def get_wallet_topup_status(request_id: str, current_user: User = Depends(get_current_user)):
    """Get wallet top-up request status"""
    topup_request = await db.wallet_topup_requests.find_one({
        "id": request_id,
        "user_id": current_user.id
    })
    if not topup_request:
        raise HTTPException(status_code=404, detail="Wallet top-up request not found")
    
    topup_request = parse_from_mongo(topup_request)
    
    # Get payment proof if exists
    payment_proof = None
    if topup_request.get("payment_proof_id"):
        proof = await db.payment_proofs.find_one({"id": topup_request["payment_proof_id"]})
        if proof:
            payment_proof = parse_from_mongo(proof)
    
    # Prepare transfer details with unique code info
    if topup_request["currency"] == "IDR":
        transfer_details = {
            "type": "bank_transfer",
            "bank_name": topup_request.get("bank_name"),
            "account_number": topup_request.get("bank_account"),
            "account_holder": topup_request.get("bank_holder"),
            "amount": topup_request["amount"],
            "unique_code": topup_request.get("unique_code", 0),
            "total_transfer": topup_request.get("total_with_unique_code", topup_request["amount"])
        }
    else:
        transfer_details = {
            "type": "crypto_wallet",
            "wallet_address": topup_request.get("wallet_address"),
            "wallet_name": topup_request.get("wallet_name"),
            "network": topup_request.get("network"),
            "amount": topup_request["amount"]
        }
    
    return {
        "request_id": topup_request["id"],
        "status": topup_request["status"],
        "reference_code": topup_request["reference_code"],
        "amount": topup_request["amount"],
        "currency": topup_request["currency"],
        "wallet_type": topup_request["wallet_type"],
        "payment_method": topup_request["payment_method"],
        "created_at": topup_request["created_at"],
        "verified_at": topup_request.get("verified_at"),
        "admin_notes": topup_request.get("admin_notes"),
        "transfer_details": transfer_details,
        "payment_proof": {
            "uploaded": payment_proof is not None,
            "uploaded_at": payment_proof["uploaded_at"] if payment_proof else None,
            "file_name": payment_proof["file_name"] if payment_proof else None
        }
    }

@api_router.post("/wallet-topup/{request_id}/upload-proof")
async def upload_wallet_topup_proof(
    request_id: str,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """Upload payment proof for wallet top-up request - Store in Google Cloud Storage"""
    
    logger.info(f"🔵 WALLET TOPUP UPLOAD ENDPOINT CALLED: request_id={request_id}, user={current_user.username}, filename={file.filename}")
    
    # Validate wallet top-up request exists and belongs to user
    wallet_request = await db.wallet_topup_requests.find_one({
        "id": request_id,
        "user_id": current_user.id
    })
    
    if not wallet_request:
        raise HTTPException(status_code=404, detail="Wallet top-up request not found")
        
    if wallet_request["status"] not in ["pending", "proof_uploaded", "rejected"]:
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot upload proof for request with status: {wallet_request['status']}"
        )
    
    # Validate file type
    if file.content_type not in ['image/jpeg', 'image/png', 'image/jpg', 'application/pdf']:
        raise HTTPException(status_code=400, detail="Invalid file type. Only JPEG, PNG, and PDF files are allowed")
    
    # Validate file size (max 10MB)
    max_size = 10 * 1024 * 1024  # 10MB
    
    try:
        # Read file content
        content = await file.read()
        file_size = len(content)
        
        if file_size > max_size:
            raise HTTPException(status_code=400, detail="File size exceeds 10MB limit")
        
        # Upload to Google Cloud Storage
        from google.cloud import storage
        import os
        
        try:
            # Initialize GCS client
            os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = "/app/backend/gcs-service-account.json"
            storage_client = storage.Client()
            bucket_name = os.getenv("GCS_BUCKET_NAME", "rimuru-file-uploads")
            bucket = storage_client.bucket(bucket_name)
            
            # Create unique filename
            file_extension = file.filename.split('.')[-1] if '.' in file.filename else 'jpg'
            gcs_filename = f"wallet_payment_proofs/{request_id}_{uuid.uuid4().hex}.{file_extension}"
            
            # Upload to GCS
            blob = bucket.blob(gcs_filename)
            blob.upload_from_string(content, content_type=file.content_type)
            
            logger.info(f"✅ Uploaded to GCS: {gcs_filename}, size={file_size}")
            
            # Create payment proof record with GCS info
            proof_record = {
                "id": str(uuid.uuid4()),
                "user_id": current_user.id,
                "file_name": file.filename,
                "gcs_path": gcs_filename,
                "gcs_bucket": bucket_name,
                "storage_type": "gcs",
                "file_size": file_size,
                "mime_type": file.content_type,
                "uploaded_at": datetime.now(timezone.utc)
            }
            
        except Exception as gcs_error:
            logger.error(f"❌ GCS upload failed: {gcs_error}")
            raise HTTPException(status_code=500, detail=f"Failed to upload to cloud storage: {str(gcs_error)}")
        
        proof_record_dict = prepare_for_mongo(proof_record)
        await db.payment_proofs.insert_one(proof_record_dict)
        
        # Update wallet top-up request with proof ID and status
        await db.wallet_topup_requests.update_one(
            {"id": request_id},
            {
                "$set": {
                    "payment_proof_id": proof_record["id"],
                    "status": "proof_uploaded",
                    "proof_uploaded_at": datetime.now(timezone.utc)
                }
            }
        )
        
        # Create notification for admin
        notification = {
            "id": str(uuid.uuid4()),
            "title": "🔔 Wallet Top-Up Proof Uploaded",
            "message": f"User {current_user.username} uploaded payment proof for wallet top-up {wallet_request['currency']} {wallet_request['amount']:,.2f}",
            "type": "wallet_topup_proof_uploaded", 
            "reference_id": request_id,
            "is_read": False,
            "created_at": datetime.now(timezone.utc)
        }
        
        notification_dict = prepare_for_mongo(notification)
        await db.notifications.insert_one(notification_dict)
        
        # Send email notification to admins about wallet top-up proof uploaded
        try:
            admin_emails = await get_active_admin_emails()
            if admin_emails:
                send_admin_wallet_topup_proof_uploaded_email(
                    admin_emails=admin_emails,
                    client_name=current_user.full_name or current_user.username,
                    amount=wallet_request["amount"],
                    currency=wallet_request["currency"],
                    wallet_type=wallet_request["wallet_type"]
                )
                logger.info(f"📧 Wallet top-up proof uploaded notification sent to {len(admin_emails)} admins")
        except Exception as e:
            logger.error(f"Failed to send wallet top-up proof uploaded email: {e}")
        
        logger.info(f"Wallet topup proof uploaded to GCS: proof_id={proof_record['id']}, gcs_path={gcs_filename}, size={file_size}")
        
        return {
            "message": "Payment proof uploaded successfully",
            "proof_id": proof_record["id"],
            "status": "proof_uploaded",
            "storage": "gcs"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error uploading wallet top-up proof to GCS: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to upload payment proof: {str(e)}")

@api_router.put("/wallet-topup-request/{request_id}/cancel")
async def cancel_wallet_topup_request(
    request_id: str,
    current_user: User = Depends(get_current_user)
):
    """Cancel a pending wallet top-up request"""
    try:
        # Get the wallet top-up request
        request = await db.wallet_topup_requests.find_one({"id": request_id})
        if not request:
            raise HTTPException(status_code=404, detail="Wallet top-up request not found")
        
        # Check if user owns this request
        if request["user_id"] != current_user.id:
            raise HTTPException(status_code=403, detail="Not authorized to cancel this request")
        
        # Check if request can be cancelled (only pending requests)
        if request["status"] != "pending":
            raise HTTPException(
                status_code=400, 
                detail=f"Cannot cancel request with status: {request['status']}"
            )
        
        # Update request status to cancelled
        result = await db.wallet_topup_requests.update_one(
            {"id": request_id},
            {"$set": {
                "status": "cancelled",
                "cancelled_at": datetime.now(timezone.utc),
                "cancelled_by": "user"
            }}
        )
        
        logger.info(f"✅ Wallet topup cancelled: request_id={request_id}, matched={result.matched_count}, modified={result.modified_count}")
        
        # Verify update
        updated_request = await db.wallet_topup_requests.find_one({"id": request_id})
        logger.info(f"🔍 After cancel - status={updated_request.get('status')}, cancelled_at={updated_request.get('cancelled_at')}")
        
        return {
            "success": True,
            "message": "Wallet top-up request cancelled successfully"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error cancelling wallet top-up request: {e}")
        raise HTTPException(status_code=500, detail="Error cancelling wallet top-up request")

# Wallet endpoints
@api_router.get("/wallet/balances", response_model=dict)
async def get_wallet_balances(current_user: User = Depends(get_current_user)):
    """Get user wallet balances with pending amounts"""
    
    # Get current balances from user model
    main_idr = getattr(current_user, 'main_wallet_idr', 0.0)
    main_usd = getattr(current_user, 'main_wallet_usd', 0.0)
    withdrawal_idr = getattr(current_user, 'withdrawal_wallet_idr', 0.0)
    withdrawal_usd = getattr(current_user, 'withdrawal_wallet_usd', 0.0)
    
    # Calculate pending amounts for each wallet
    pending_transfers = await db.wallet_transfers.find({
        "user_id": current_user.id,
        "status": "pending"
    }).to_list(1000)
    
    logger.info(f"User {current_user.username} wallet balances: main_idr={main_idr}, main_usd={main_usd}, withdrawal_idr={withdrawal_idr}, withdrawal_usd={withdrawal_usd}")
    logger.info(f"User {current_user.username}: Found {len(pending_transfers)} pending transfers")
    for t in pending_transfers[:3]:  # Log first 3
        logger.info(f"  Pending: wallet={t.get('source_wallet_type')}, currency={t.get('currency')}, amount={t.get('amount')}, total={t.get('total')}")
    
    pending_main_idr = sum(t.get("total", t.get("amount", 0)) for t in pending_transfers 
                           if t.get("source_wallet_type") == "main" and t.get("currency") == "IDR")
    pending_main_usd = sum(t.get("total", t.get("amount", 0)) for t in pending_transfers 
                           if t.get("source_wallet_type") == "main" and t.get("currency") == "USD")
    pending_withdrawal_idr = sum(t.get("total", t.get("amount", 0)) for t in pending_transfers 
                                 if t.get("source_wallet_type") == "withdrawal" and t.get("currency") == "IDR")
    pending_withdrawal_usd = sum(t.get("total", t.get("amount", 0)) for t in pending_transfers 
                                 if t.get("source_wallet_type") == "withdrawal" and t.get("currency") == "USD")
    
    available_main_idr = main_idr - pending_main_idr
    logger.info(f"Calculated: main_idr={main_idr}, pending={pending_main_idr}, available={available_main_idr}")
    
    return {
        "main_wallet_idr": main_idr,
        "main_wallet_idr_pending": pending_main_idr,
        "main_wallet_idr_available": available_main_idr,
        "main_wallet_usd": main_usd,
        "main_wallet_usd_pending": pending_main_usd,
        "main_wallet_usd_available": main_usd - pending_main_usd,
        "withdrawal_wallet_idr": withdrawal_idr,
        "withdrawal_wallet_idr_pending": pending_withdrawal_idr,
        "withdrawal_wallet_idr_available": withdrawal_idr - pending_withdrawal_idr,
        "withdrawal_wallet_usd": withdrawal_usd,
        "withdrawal_wallet_usd_pending": pending_withdrawal_usd,
        "withdrawal_wallet_usd_available": withdrawal_usd - pending_withdrawal_usd
    }

@api_router.post("/wallet/topup", response_model=dict)
async def create_wallet_topup_request(
    wallet_type: str = Form(...),
    currency: str = Form(...),
    amount: float = Form(...),
    payment_method: str = Form(...),
    notes: str = Form(""),
    unique_code: int = Form(0),
    total_with_unique_code: float = Form(...),
    payment_proof: UploadFile = File(None),  # Made optional
    current_user: User = Depends(get_current_user)
):
    """Create a new wallet top-up request (payment proof is optional, can be uploaded later)"""
    
    # Validate inputs
    if wallet_type not in ["main", "withdrawal"]:
        raise HTTPException(status_code=400, detail="Invalid wallet type")
    
    if currency not in ["IDR", "USD"]:
        raise HTTPException(status_code=400, detail="Invalid currency")
    
    if amount <= 0:
        raise HTTPException(status_code=400, detail="Amount must be greater than 0")
    
    # Generate unique reference code
    reference_code = f"WLT{str(uuid.uuid4())[:8].upper()}"
    
    # Use unique code from frontend (already generated there for better UX)
    # If not provided, generate as fallback
    if unique_code == 0 and currency == "IDR":
        import random
        unique_code = random.randint(100, 999)
        total_with_unique_code = amount + unique_code
    
    # Set bank/wallet details based on payment method - using existing system data
    bank_details = {}
    if payment_method in ["bank_bri", "bank_transfer", "idr"]:  # IDR bank transfer
        bank_details = {
            "bank_name": "BRI",
            "bank_account": "057901002665566",
            "bank_holder": "PT RINAIYANTI CAHAYA INTERMA",
            "wallet_address": None,
            "wallet_name": None,
            "network": None
        }
    elif payment_method in ["usdt_trc20", "usdt", "crypto"]:  # USDT TRC20
        bank_details = {
            "bank_name": None,
            "bank_account": None, 
            "bank_holder": None,
            "wallet_address": "TBCJiUoYGxBYBpqMNb9ZtuwXWJLuCwBPXa",
            "wallet_name": "BINANCE",
            "network": "USDT TRC20"
        }
    
    # Upload payment proof to Google Cloud Storage (if provided)
    payment_proof_record = None
    if payment_proof and payment_proof.filename:
        try:
            from google.cloud import storage
            import os
            
            # Read file content
            content = await payment_proof.read()
            
            # Initialize GCS client
            os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = "/app/backend/gcs-service-account.json"
            storage_client = storage.Client()
            bucket_name = os.getenv("GCS_BUCKET_NAME", "rimuru-file-uploads")
            bucket = storage_client.bucket(bucket_name)
            
            # Generate unique filename for GCS
            file_extension = Path(payment_proof.filename).suffix
            gcs_filename = f"wallet_payment_proofs/{str(uuid.uuid4())[:12]}{file_extension}"
            
            # Upload to GCS
            blob = bucket.blob(gcs_filename)
            blob.upload_from_string(content, content_type=payment_proof.content_type)
            
            logger.info(f"✅ Wallet topup proof uploaded to GCS: {gcs_filename}, size={len(content)}")
            
            # Create payment proof record with GCS storage
            payment_proof_record = PaymentProof(
                topup_request_id="",  # Will be updated after wallet request is created
                user_id=current_user.id,
                file_name=payment_proof.filename,
                gcs_path=gcs_filename,
                gcs_bucket=bucket_name,
                storage_type="gcs",
                file_size=len(content),
                mime_type=payment_proof.content_type
            )
            
        except Exception as gcs_error:
            logger.error(f"❌ GCS upload failed: {gcs_error}")
            raise HTTPException(status_code=500, detail=f"Failed to upload proof to cloud storage: {str(gcs_error)}")
    
    # Create wallet top-up request
    wallet_request = WalletTopUpRecord(
        user_id=current_user.id,
        wallet_type=wallet_type,
        currency=currency,
        amount=amount,
        payment_method=payment_method,
        notes=notes,
        unique_code=unique_code,
        total_with_unique_code=total_with_unique_code,
        status="pending" if not payment_proof_record else "proof_uploaded",  # pending if no proof yet
        payment_proof_id=payment_proof_record.id if payment_proof_record else None,
        reference_code=reference_code,  # Add reference code to the record
        **bank_details
    )
    
    # Update payment proof with correct request ID (if exists)
    if payment_proof_record:
        payment_proof_record.topup_request_id = wallet_request.id
    
    # Save to database
    wallet_request_dict = prepare_for_mongo(wallet_request.dict())
    await db.wallet_topup_requests.insert_one(wallet_request_dict)
    
    # Save payment proof if it exists
    if payment_proof_record:
        payment_proof_dict = prepare_for_mongo(payment_proof_record.dict())
        await db.payment_proofs.insert_one(payment_proof_dict)
    
    # Create admin notification for wallet top-up request
    await create_localized_notification(
        title_key="wallet_topup_request",
        message_key="topup_request_submitted", 
        notification_type="wallet_topup_request",
        lang="id",
        reference_id=wallet_request.id,
        username=current_user.username,
        amount=f"{currency} {amount:,.2f}"
    )
    
    # Send email notification to all active admins
    try:
        admin_emails = await get_active_admin_emails()
        if admin_emails:
            from email_service import send_admin_new_topup_request_email
            send_admin_new_topup_request_email(
                admin_emails=admin_emails,
                client_name=current_user.name or current_user.username,
                amount=amount,
                currency=currency,
                platform="Wallet Top-Up",
                account_name=f"{wallet_type.capitalize()} Wallet"
            )
            logger.info(f"📧 Admin wallet top-up notification emails sent to {len(admin_emails)} admins")
    except Exception as e:
        logger.error(f"❌ Failed to send admin wallet top-up notification emails: {e}")
    
    return {
        "message": "Wallet top-up request submitted successfully",
        "id": wallet_request.id,
        "reference_code": reference_code,
        "status": "pending" if not payment_proof_record else "proof_uploaded",
        "total_with_unique_code": total_with_unique_code if currency == "IDR" else amount
    }

@api_router.get("/wallet/topup/{request_id}", response_model=dict)
async def get_wallet_topup_request(
    request_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get wallet top-up request details"""
    
    # Get request from database
    request = await db.wallet_topup_requests.find_one({"id": request_id, "user_id": current_user.id})
    
    if not request:
        raise HTTPException(status_code=404, detail="Wallet top-up request not found")
    
    # Remove MongoDB _id
    if '_id' in request:
        del request['_id']
    
    return request

@api_router.post("/wallet/topup/{request_id}/upload-proof", response_model=dict)
async def upload_wallet_topup_proof(
    request_id: str,
    payment_proof: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    """Upload payment proof for an existing wallet top-up request"""
    
    # Get request from database
    request = await db.wallet_topup_requests.find_one({"id": request_id, "user_id": current_user.id})
    
    if not request:
        raise HTTPException(status_code=404, detail="Wallet top-up request not found")
    
    # Check if request allows proof upload/re-upload
    if request.get('status') not in ['pending', 'proof_uploaded', 'rejected']:
        raise HTTPException(status_code=400, detail="Cannot upload proof for this request status")
    
    # Upload payment proof to Google Cloud Storage
    try:
        from google.cloud import storage
        import os
        
        # Read file content
        content = await payment_proof.read()
        
        # Initialize GCS client
        os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = "/app/backend/gcs-service-account.json"
        storage_client = storage.Client()
        bucket_name = os.getenv("GCS_BUCKET_NAME", "rimuru-file-uploads")
        bucket = storage_client.bucket(bucket_name)
        
        # Generate unique filename for GCS
        file_extension = Path(payment_proof.filename).suffix
        gcs_filename = f"wallet_payment_proofs/{str(uuid.uuid4())[:12]}{file_extension}"
        
        # Upload to GCS
        blob = bucket.blob(gcs_filename)
        blob.upload_from_string(content, content_type=payment_proof.content_type)
        
        logger.info(f"✅ Wallet topup proof uploaded to GCS: {gcs_filename}, size={len(content)}")
        
        # Create payment proof record with GCS storage
        payment_proof_record = PaymentProof(
            topup_request_id=request_id,
            user_id=current_user.id,
            file_name=payment_proof.filename,
            gcs_path=gcs_filename,
            gcs_bucket=bucket_name,
            storage_type="gcs",
            file_size=len(content),
            mime_type=payment_proof.content_type
        )
        
        # Save payment proof to database
        payment_proof_dict = prepare_for_mongo(payment_proof_record.dict())
        await db.payment_proofs.insert_one(payment_proof_dict)
        
        # Update wallet request with proof ID and status
        await db.wallet_topup_requests.update_one(
            {"id": request_id},
            {"$set": {
                "payment_proof_id": payment_proof_record.id,
                "status": "proof_uploaded",
                "proof_uploaded_at": datetime.now(timezone.utc),
                "updated_at": datetime.now(timezone.utc)
            }}
        )
        
        # Create notification for admin
        notification = {
            "id": str(uuid.uuid4()),
            "title": "🔔 Wallet Top-Up Proof Uploaded",
            "message": f"User {current_user.username} uploaded payment proof for wallet top-up {request['currency']} {request['amount']:,.2f}",
            "type": "wallet_topup_proof_uploaded", 
            "reference_id": request_id,
            "is_read": False,
            "created_at": datetime.now(timezone.utc)
        }
        
        notification_dict = prepare_for_mongo(notification)
        await db.notifications.insert_one(notification_dict)
        logger.info(f"✅ Admin notification created for wallet topup proof upload")
        
        # Send email notification to admins about wallet top-up proof uploaded
        try:
            admin_emails = await get_active_admin_emails()
            if admin_emails:
                send_admin_wallet_topup_proof_uploaded_email(
                    admin_emails=admin_emails,
                    client_name=current_user.name or current_user.username,
                    amount=request["amount"],
                    currency=request["currency"],
                    wallet_type=request["wallet_type"]
                )
                logger.info(f"📧 Wallet top-up proof uploaded notification sent to {len(admin_emails)} admins")
        except Exception as e:
            logger.error(f"Failed to send wallet top-up proof uploaded email: {e}")
        
        logger.info(f"✅ Payment proof uploaded for wallet request {request_id}")
        
        return {
            "message": "Payment proof uploaded successfully",
            "proof_id": payment_proof_record.id,
            "status": "proof_uploaded"
        }
        
    except Exception as gcs_error:
        logger.error(f"❌ GCS upload failed: {gcs_error}")
        raise HTTPException(status_code=500, detail=f"Failed to upload proof to cloud storage: {str(gcs_error)}")

# Wallet transfer models
class WalletToAccountTransfer(BaseModel):
    source_wallet_type: str  # main or withdrawal
    target_account_id: str
    currency: str  # IDR or USD
    amount: float
    notes: Optional[str] = None

class WalletToAccountTransferRecord(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    source_wallet_type: str
    target_account_id: str
    target_account_name: str
    currency: str
    amount: float
    fee: Optional[float] = 0
    total: Optional[float] = None
    notes: Optional[str] = None
    status: str = "completed"  # For wallet transfers, they're instant
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

@api_router.post("/wallet/transfer-to-account", response_model=dict)
async def transfer_wallet_to_account(
    transfer: WalletToAccountTransfer,
    current_user: User = Depends(get_current_user)
):
    """Transfer from wallet to ad account"""
    
    # Validate inputs
    if transfer.source_wallet_type not in ["main", "withdrawal"]:
        raise HTTPException(status_code=400, detail="Invalid wallet type")
    
    if transfer.currency not in ["IDR", "USD"]:
        raise HTTPException(status_code=400, detail="Invalid currency")
    
    if transfer.amount <= 0:
        raise HTTPException(status_code=400, detail="Amount must be greater than 0")
    
    # Check if target account exists and belongs to user
    ad_account = await db.ad_accounts.find_one({
        "id": transfer.target_account_id,
        "user_id": current_user.id,
        "status": "active"
    })
    if not ad_account:
        raise HTTPException(status_code=404, detail="Target ad account not found or not active")
    
    # Check if currencies match
    if ad_account.get("currency", "IDR") != transfer.currency:
        raise HTTPException(status_code=400, detail="Currency mismatch with target account")
    
    # For now, we'll simulate the wallet balance check and transfer
    # In a real implementation, this would check actual wallet balances from a wallet_balances collection
    
    # Simulate wallet balance (for demo purposes)
    # In production, you would fetch from actual wallet_balances collection
    simulated_balance = 1000000 if transfer.currency == "IDR" else 1000  # Assume sufficient balance for demo
    
    if transfer.amount > simulated_balance:
        raise HTTPException(status_code=400, detail="Insufficient wallet balance")
    
    # Create transfer record WITH PROPER FEE CALCULATION
    # Calculate fee based on target account's fee_percentage AND source wallet type
    fee = 0
    if transfer.source_wallet_type == "main" and ad_account.get("fee_percentage"):
        # Only main wallet transfers have fees, withdrawal wallet transfers are free
        fee = transfer.amount * (ad_account["fee_percentage"] / 100)
    # withdrawal wallet transfers always have fee = 0
    
    total = transfer.amount + fee
    
    transfer_record = WalletToAccountTransferRecord(
        user_id=current_user.id,
        source_wallet_type=transfer.source_wallet_type,
        target_account_id=transfer.target_account_id,
        target_account_name=ad_account.get("account_name", "Unknown"),
        currency=transfer.currency,
        amount=transfer.amount,
        fee=fee,  # CRITICAL FIX: Fee = 0 for withdrawal wallet, calculated for main wallet
        total=total,  # CRITICAL FIX: amount + fee
        notes=transfer.notes
    )
    
    # Save transfer record
    transfer_dict = prepare_for_mongo(transfer_record.dict())
    await db.wallet_transfers.insert_one(transfer_dict)
    
    # Update account balance
    current_balance = ad_account.get("balance", 0)
    new_balance = current_balance + transfer.amount
    
    await db.ad_accounts.update_one(
        {"id": transfer.target_account_id},
        {"$set": {"balance": new_balance}}
    )
    
    # Create notification for user
    await create_localized_notification(
        title_key="wallet_transfer_success",
        message_key="wallet_transfer_completed",
        notification_type="wallet_transfer",
        lang="id",
        reference_id=transfer_record.id,
        username=current_user.username,
        amount=f"{transfer.currency} {transfer.amount:,.2f}",
        account_name=ad_account.get("account_name", "Unknown")
    )
    
    # Send email notification to client about successful wallet-to-account transfer
    try:
        if current_user.email:
            from_account = f"{transfer.source_wallet_type.title()} Wallet"
            to_account = ad_account.get("account_name", "Unknown")
            send_client_transfer_request_success_email(
                client_email=current_user.email,
                client_name=current_user.name or current_user.username,
                amount=transfer.amount,
                currency=transfer.currency,
                from_account=from_account,
                to_account=to_account
            )
            logger.info(f"📧 Wallet-to-account transfer success email sent to {current_user.email}")
    except Exception as e:
        logger.error(f"Failed to send wallet-to-account transfer success email: {e}")
    
    # Create notification for admin about new wallet transfer request
    formatted_amount = f"Rp {transfer.amount:,.0f}" if transfer.currency == "IDR" else f"${transfer.amount:,.2f}"
    admin_notification = {
        "id": str(uuid.uuid4()),
        "title": f"🔄 New Wallet Transfer Request", 
        "message": f"User {current_user.username} created wallet transfer: {formatted_amount} from {transfer.source_wallet_type} wallet to {ad_account.get('account_name', 'Unknown')}",
        "type": "wallet_transfer_request",
        "reference_id": transfer_record.id,
        "is_read": False,
        "created_at": datetime.now(timezone.utc)
    }
    
    admin_notification_dict = prepare_for_mongo(admin_notification)
    await db.notifications.insert_one(admin_notification_dict)
    
    # Send email notification to admins about new wallet-to-account transfer request
    try:
        admin_emails = await get_active_admin_emails()
        if admin_emails:
            from_account = f"{transfer.source_wallet_type.title()} Wallet"
            to_account = ad_account.get("account_name", "Unknown")
            send_admin_transfer_request_created_email(
                admin_emails=admin_emails,
                client_name=current_user.name or current_user.username,
                amount=transfer.amount,
                currency=transfer.currency,
                from_account=from_account,
                to_account=to_account
            )
            logger.info(f"📧 Wallet-to-account transfer notification emails sent to {len(admin_emails)} admins")
    except Exception as e:
        logger.error(f"Failed to send wallet-to-account transfer notification email to admins: {e}")
    
    return {
        "message": "Transfer completed successfully", 
        "transfer_id": transfer_record.id,
        "new_account_balance": new_balance,
        "currency": transfer.currency
    }

# Multiple wallet transfer models
class WalletToAccountsTransfer(BaseModel):
    source_wallet_type: str  # main or withdrawal
    currency: str  # IDR or USD
    transfers: List[dict]  # [{ target_account_id, target_account_name, amount, fee, total }]
    notes: Optional[str] = None

@api_router.post("/wallet/transfer-to-accounts", response_model=dict)
async def transfer_wallet_to_multiple_accounts(
    transfer: WalletToAccountsTransfer,
    current_user: User = Depends(get_current_user)
):
    """Transfer from wallet to multiple ad accounts"""
    
    # Validate inputs
    if transfer.source_wallet_type not in ["main", "withdrawal"]:
        raise HTTPException(status_code=400, detail="Invalid wallet type")
    
    if transfer.currency not in ["IDR", "USD"]:
        raise HTTPException(status_code=400, detail="Invalid currency")
    
    if not transfer.transfers or len(transfer.transfers) == 0:
        raise HTTPException(status_code=400, detail="No transfers specified")
    
    # Calculate total amount needed including fees
    total_amount_needed = sum(t.get("total", 0) for t in transfer.transfers)
    
    if total_amount_needed <= 0:
        raise HTTPException(status_code=400, detail="Total amount must be greater than 0")
    
    # Get user's current wallet balance
    user = await db.users.find_one({"id": current_user.id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Determine which wallet to use
    wallet_field = f"{transfer.source_wallet_type}_wallet_{transfer.currency.lower()}"
    current_wallet_balance = user.get(wallet_field, 0)
    
    # Calculate total pending transfers (not yet approved/rejected)
    pending_transfers = await db.wallet_transfers.find({
        "user_id": current_user.id,
        "source_wallet_type": transfer.source_wallet_type,
        "currency": transfer.currency,
        "status": "pending"
    }).to_list(1000)
    
    total_pending_amount = sum(t.get("total", t.get("amount", 0)) for t in pending_transfers)
    
    # Available balance = Current balance - Pending transfers (reserved)
    available_balance = current_wallet_balance - total_pending_amount
    
    if total_amount_needed > available_balance:
        raise HTTPException(
            status_code=400, 
            detail=f"Saldo wallet tidak mencukupi. Dibutuhkan: {total_amount_needed:,.2f}, Tersedia: {available_balance:,.2f} (Saldo: {current_wallet_balance:,.2f}, Pending: {total_pending_amount:,.2f})"
        )
    
    # NOTE: Wallet will be deducted only when admin approves the transfer
    # Not deducted here to avoid balance issues if transfer is rejected
    # But we track pending transfers to prevent over-spending
    
    # Validate and process each transfer
    successful_transfers = []
    failed_transfers = []
    
    for t in transfer.transfers:
        try:
            account_id = t.get("target_account_id")
            amount = t.get("amount", 0)
            
            if amount <= 0:
                continue
            
            # Check if target account exists and belongs to user
            ad_account = await db.ad_accounts.find_one({
                "id": account_id,
                "user_id": current_user.id,
                "status": "active"
            })
            
            if not ad_account:
                failed_transfers.append({
                    "account_id": account_id,
                    "reason": "Account not found or not active"
                })
                continue
            
            # Check currency match
            if ad_account.get("currency", "IDR") != transfer.currency:
                failed_transfers.append({
                    "account_id": account_id,
                    "reason": "Currency mismatch"
                })
                continue
            
            # Create transfer record with pending status
            transfer_record = WalletToAccountTransferRecord(
                user_id=current_user.id,
                source_wallet_type=transfer.source_wallet_type,
                target_account_id=account_id,
                target_account_name=t.get("target_account_name", ad_account.get("account_name", "Unknown")),
                currency=transfer.currency,
                amount=amount,
                fee=t.get("fee", 0),
                total=t.get("total", amount),
                notes=transfer.notes,
                status="pending"  # Needs admin verification
            )
            
            # Save transfer record
            transfer_dict = prepare_for_mongo(transfer_record.dict())
            await db.wallet_transfers.insert_one(transfer_dict)
            
            # Create transaction record immediately with pending status for client monitoring
            transaction = Transaction(
                user_id=current_user.id,
                type="wallet_to_account_transfer",
                amount=amount,
                currency=transfer.currency,
                status="pending",  # Pending admin approval
                description=f"Transfer dari {transfer.source_wallet_type} wallet ke akun {transfer_record.target_account_name} (Menunggu Verifikasi Admin)",
                account_id=account_id,
                account_name=transfer_record.target_account_name,
                reference_id=transfer_record.id,
                fee=t.get("fee", 0),
                total_amount=t.get("total", amount)
            )
            
            transaction_dict = prepare_for_mongo(transaction.dict())
            await db.transactions.insert_one(transaction_dict)
            
            # DO NOT update account balance yet - wait for admin verification
            # Balance will be updated after admin approves and uploads proof
            
            successful_transfers.append({
                "transfer_id": transfer_record.id,
                "account_id": account_id,
                "account_name": transfer_record.target_account_name,
                "amount": amount,
                "fee": t.get("fee", 0),
                "status": "pending"
            })
            
        except Exception as e:
            failed_transfers.append({
                "account_id": t.get("target_account_id"),
                "reason": str(e)
            })
    
    # Create notification for admin to verify
    if successful_transfers:
        total_accounts = len(successful_transfers)
        total_amount = sum(t["amount"] for t in successful_transfers)
        
        # Notify user that request is submitted
        await create_localized_notification(
            title_key="wallet_transfer_submitted",
            message_key="wallet_transfer_pending_admin",
            notification_type="wallet_transfer",
            lang="id",
            reference_id=f"batch_{len(successful_transfers)}",
            username=current_user.username,
            amount=f"{transfer.currency} {total_amount:,.2f}",
            account_name=f"{total_accounts} accounts"
        )
        
        # Notify admin for verification
        logger.info(f"Creating admin notification for wallet transfer: user={current_user.username}, accounts={total_accounts}, amount={total_amount}")
        await create_localized_notification(
            title_key="wallet_transfer_request",
            message_key="wallet_transfer_needs_verification",
            notification_type="wallet_transfer_request",
            lang="id",
            reference_id=f"batch_{len(successful_transfers)}_{datetime.now(timezone.utc).timestamp()}",
            username=current_user.username,
            amount=f"{transfer.currency} {total_amount:,.2f}",
            account_name=f"{total_accounts} accounts"
        )
        logger.info(f"Admin notification created successfully")
    
    return {
        "message": f"Transfer request submitted for {len(successful_transfers)} accounts. Menunggu verifikasi admin.",
        "status": "pending",
        "successful_transfers": successful_transfers,
        "failed_transfers": failed_transfers,
        "total_amount": sum(t["amount"] for t in successful_transfers),
        "total_fee": sum(t["fee"] for t in successful_transfers),
        "currency": transfer.currency
    }

@api_router.get("/wallet/statement", response_model=List[dict])
async def get_wallet_statement(
    wallet_type: str = "main",  # main or withdrawal
    currency: str = "IDR",  # IDR or USD
    start_date: Optional[str] = None,  # YYYY-MM-DD
    end_date: Optional[str] = None,  # YYYY-MM-DD
    limit: int = 100,
    current_user: User = Depends(get_current_user)
):
    """Get wallet statement/history for client with optional date range filter"""
    
    # Validate inputs
    if wallet_type not in ["main", "withdrawal"]:
        raise HTTPException(status_code=400, detail="Invalid wallet type")
    
    if currency not in ["IDR", "USD"]:
        raise HTTPException(status_code=400, detail="Invalid currency")
    
    # Build date filter
    # Use GMT+7 (Asia/Jakarta) timezone for Indonesian users
    jakarta_tz = ZoneInfo("Asia/Jakarta")
    
    date_filter = {}
    date_filter_str = {}  # For string-based date fields
    
    if start_date:
        try:
            # Parse date as Jakarta time, then convert to UTC for database query
            start_dt = datetime.fromisoformat(start_date).replace(hour=0, minute=0, second=0, microsecond=0, tzinfo=jakarta_tz)
            # Convert to UTC for database comparison
            start_dt_utc = start_dt.astimezone(timezone.utc)
            date_filter["$gte"] = start_dt_utc
            # Also create string version for string-based fields
            date_filter_str["$gte"] = start_dt_utc.isoformat()
            logger.info(f"Start date filter: {start_date} Jakarta -> {start_dt_utc} UTC")
        except:
            raise HTTPException(status_code=400, detail="Invalid start_date format. Use YYYY-MM-DD")
    
    if end_date:
        try:
            # Parse date as Jakarta time, then convert to UTC for database query
            end_dt = datetime.fromisoformat(end_date).replace(hour=23, minute=59, second=59, microsecond=999999, tzinfo=jakarta_tz)
            # Convert to UTC for database comparison
            end_dt_utc = end_dt.astimezone(timezone.utc)
            date_filter["$lte"] = end_dt_utc
            # Also create string version for string-based fields
            date_filter_str["$lte"] = end_dt_utc.isoformat()
            logger.info(f"End date filter: {end_date} Jakarta -> {end_dt_utc} UTC")
        except:
            raise HTTPException(status_code=400, detail="Invalid end_date format. Use YYYY-MM-DD")
    
    # Get all transactions related to this wallet
    transactions = []
    
    # 1. Wallet top-ups (credit)
    topup_query = {
        "user_id": current_user.id,
        "wallet_type": wallet_type,
        "currency": currency,
        "status": "verified"
    }
    # Use string-based filter for wallet topups (created_at stored as string)
    if date_filter_str:
        topup_query["created_at"] = date_filter_str
    
    wallet_topups = await db.wallet_topup_requests.find(topup_query).sort("created_at", -1).to_list(limit)
    
    for topup in wallet_topups:
        topup = parse_from_mongo(topup)
        # Convert string date to datetime if needed
        created_at = topup["created_at"]
        if isinstance(created_at, str):
            created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
        # Ensure timezone aware
        if created_at.tzinfo is None:
            created_at = created_at.replace(tzinfo=timezone.utc)
        
        transactions.append({
            "id": topup["id"],
            "date": created_at,
            "type": "credit",
            "description": f"Top-Up {wallet_type.title()} Wallet",
            "amount": topup["amount"],
            "currency": currency,
            "reference": topup.get("reference_code", topup["id"][:8]),
            "status": "completed"
        })
    
    # 2. Account deletion balance transfers (credit to withdrawal wallet)
    if wallet_type == "withdrawal":
        deletion_query = {
            "user_id": current_user.id,
            "type": "account_deletion_balance_transfer",
            "currency": currency,
            "status": "completed"
        }
        if date_filter:
            deletion_query["created_at"] = date_filter
        
        deletion_transfers = await db.transactions.find(deletion_query).sort("created_at", -1).to_list(limit)
        
        for deletion in deletion_transfers:
            deletion = parse_from_mongo(deletion)
            created_at = deletion["created_at"]
            if isinstance(created_at, str):
                created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
            if created_at.tzinfo is None:
                created_at = created_at.replace(tzinfo=timezone.utc)
            
            transactions.append({
                "id": deletion["id"],
                "date": created_at,
                "type": "credit",
                "description": f"Transfer saldo dari akun {deletion.get('account_name', 'Unknown')} (dihapus)",
                "amount": deletion["amount"],
                "currency": currency,
                "reference": deletion["id"][:8],
                "status": "completed"
            })
        
        # 2b. Withdraw requests approved (credit to withdrawal wallet)
        withdraw_query = {
            "user_id": current_user.id,
            "currency": currency,
            "status": {"$in": ["approved", "completed"]}
        }
        if date_filter:
            withdraw_query["processed_at"] = date_filter
        
        withdraw_requests = await db.withdraw_requests.find(withdraw_query).sort("processed_at", -1).to_list(limit)
        
        for withdraw in withdraw_requests:
            withdraw = parse_from_mongo(withdraw)
            # Use processed_at when approved
            processed_at = withdraw.get("processed_at", withdraw.get("created_at"))
            if isinstance(processed_at, str):
                processed_at = datetime.fromisoformat(processed_at.replace('Z', '+00:00'))
            if processed_at.tzinfo is None:
                processed_at = processed_at.replace(tzinfo=timezone.utc)
            
            # Get account name
            account = await db.ad_accounts.find_one({"id": withdraw["account_id"]})
            account_name = account.get("account_name", "Unknown") if account else "Unknown"
            
            # Use verified amount if available, else requested amount
            credit_amount = withdraw.get("admin_verified_amount", withdraw.get("amount", 0))
            
            transactions.append({
                "id": withdraw["id"],
                "date": processed_at,
                "type": "credit",
                "description": f"Penarikan dari akun {account_name}",
                "amount": credit_amount,
                "currency": currency,
                "reference": withdraw["id"][:8],
                "status": "completed"
            })
    
    # 3. Wallet transfers (debit) - only approved ones
    transfer_query = {
        "user_id": current_user.id,
        "source_wallet_type": wallet_type,
        "currency": currency,
        "status": "approved"
    }
    if date_filter:
        # Use processed_at for transfers (stored as proper ISODate, not string)
        transfer_query["processed_at"] = date_filter
    
    wallet_transfers = await db.wallet_transfers.find(transfer_query).sort("processed_at", -1).to_list(limit)
    
    for transfer in wallet_transfers:
        transfer = parse_from_mongo(transfer)
        transfer_amount = transfer["amount"]
        transfer_total = transfer.get("total", transfer_amount)
        
        # Calculate fee: if fee field exists use it, otherwise calculate from total - amount
        transfer_fee = transfer.get("fee", 0)
        if transfer_fee == 0 and transfer_total > transfer_amount:
            transfer_fee = transfer_total - transfer_amount
        
        # Total debit = amount + fee (what was actually deducted from wallet)
        total_debit = transfer_amount + transfer_fee
        
        # DEBUG: Log transfer data
        logger.info(f"Wallet transfer: id={transfer['id'][:8]}, amount={transfer_amount}, fee={transfer_fee}, total_debit={total_debit}")
        
        # Use processed_at for transfers (proper ISODate object)
        # This is when transfer was actually approved and balance was deducted
        date_field = transfer.get("processed_at", transfer.get("created_at"))
        if isinstance(date_field, str):
            date_field = datetime.fromisoformat(date_field.replace('Z', '+00:00'))
        # Ensure timezone aware
        if date_field.tzinfo is None:
            date_field = date_field.replace(tzinfo=timezone.utc)
        
        # Add transfer transaction with TOTAL amount (amount + fee)
        transactions.append({
            "id": transfer["id"],
            "date": date_field,
            "type": "debit",
            "description": f"Transfer ke {transfer['target_account_name']}",
            "amount": total_debit,  # Show total debit (amount + fee)
            "currency": currency,
            "reference": transfer["id"][:8],
            "status": "completed",
            "fee": transfer_fee  # Show fee in separate column
        })
    
    # 4. Currency exchanges (both debit from source currency and credit to target currency)
    # Only show for main wallet since exchanges only work with main wallet
    if wallet_type == "main":
        # Query exchanges where this currency is involved (either from or to)
        exchange_query = {
            "user_id": current_user.id,
            "$or": [
                {"from_currency": currency},
                {"to_currency": currency}
            ]
        }
        if date_filter:
            exchange_query["created_at"] = date_filter
        
        currency_exchanges = await db.currency_exchanges.find(exchange_query).sort("created_at", -1).to_list(limit)
        
        for exchange in currency_exchanges:
            exchange = parse_from_mongo(exchange)
            created_at = exchange.get("created_at")
            if isinstance(created_at, str):
                created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
            if created_at.tzinfo is None:
                created_at = created_at.replace(tzinfo=timezone.utc)
            
            # If this currency is the source (from), it's a debit
            if exchange["from_currency"] == currency:
                transactions.append({
                    "id": exchange["id"],
                    "date": created_at,
                    "type": "debit",
                    "description": f"Tukar {currency} → {exchange['to_currency']} (Rate: {exchange['exchange_rate']:.6f})",
                    "amount": exchange["from_amount"],
                    "currency": currency,
                    "reference": exchange["id"][:8],
                    "status": "completed",
                    "fee": 0
                })
            # If this currency is the target (to), it's a credit
            elif exchange["to_currency"] == currency:
                transactions.append({
                    "id": exchange["id"],
                    "date": created_at,
                    "type": "credit",
                    "description": f"Tukar {exchange['from_currency']} → {currency} (Rate: {exchange['exchange_rate']:.6f})",
                    "amount": exchange["to_amount"],
                    "currency": currency,
                    "reference": exchange["id"][:8],
                    "status": "completed",
                    "fee": 0
                })
    
    # 5. Wallet statements (admin deductions, etc)
    wallet_stmt_query = {
        "user_id": current_user.id,
        "wallet_type": f"{wallet_type}_{currency.lower()}"
    }
    
    if date_filter_str:
        wallet_stmt_query["created_at"] = date_filter_str
    
    wallet_statements = await db.wallet_statements.find(wallet_stmt_query).to_list(length=None)
    
    for stmt in wallet_statements:
        # Parse date
        date_field = stmt.get("created_at")
        if isinstance(date_field, str):
            date_field = datetime.fromisoformat(date_field.replace('Z', '+00:00'))
        if date_field.tzinfo is None:
            date_field = date_field.replace(tzinfo=timezone.utc)
        
        # Determine type based on amount
        stmt_amount = abs(stmt.get("amount", 0))
        tx_type = "debit" if stmt.get("amount", 0) < 0 else "credit"
        
        transactions.append({
            "id": stmt["id"],
            "date": date_field,
            "type": tx_type,
            "description": stmt.get("description", "Wallet adjustment"),
            "amount": stmt_amount,
            "currency": stmt.get("currency", currency),
            "reference": stmt.get("reference_id", stmt["id"])[:8],
            "status": "completed",
            "fee": 0
        })
    
    # Sort by date descending
    transactions.sort(key=lambda x: x["date"], reverse=True)
    
    # Calculate running balance (from oldest to newest, then reverse)
    # Get current wallet balance
    user = await db.users.find_one({"id": current_user.id})
    wallet_field = f"{wallet_type}_wallet_{currency.lower()}"
    current_balance = user.get(wallet_field, 0)
    
    # Add running balance
    balance = current_balance
    for tx in transactions[:limit]:
        tx["balance_after"] = balance
        if tx["type"] == "credit":
            balance -= tx["amount"]
        else:  # debit
            balance += tx["amount"]
        tx["balance_before"] = balance
    
    return transactions[:limit]

@api_router.get("/wallet/statement/export/pdf")
async def export_wallet_statement_pdf(
    wallet_type: str = "main",
    currency: str = "IDR",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Export wallet statement as PDF"""
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    
    # Get statement data (reuse existing logic)
    params = {
        "wallet_type": wallet_type,
        "currency": currency,
        "limit": 1000  # Export all
    }
    if start_date:
        params["start_date"] = start_date
    if end_date:
        params["end_date"] = end_date
    
    # Get transactions using existing endpoint logic
    transactions = await get_wallet_statement(**params, current_user=current_user)
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []
    styles = getSampleStyleSheet()
    
    # Title
    title_text = f"Wallet Statement - {wallet_type.title()} Wallet ({currency})"
    story.append(Paragraph(title_text, styles['Title']))
    story.append(Spacer(1, 0.2*inch))
    
    # Period info
    period_text = "Periode: "
    if start_date and end_date:
        period_text += f"{start_date} s/d {end_date}"
    else:
        period_text += "Semua transaksi"
    story.append(Paragraph(period_text, styles['Normal']))
    story.append(Spacer(1, 0.3*inch))
    
    # Table data
    table_data = [['Tanggal', 'Tipe', 'Deskripsi', 'Jumlah', 'Saldo']]
    
    for tx in transactions:
        date_str = tx['date'].strftime('%Y-%m-%d %H:%M') if isinstance(tx['date'], datetime) else str(tx['date'])
        type_str = 'Kredit' if tx['type'] == 'credit' else 'Debit'
        desc_str = tx['description'][:40]  # Truncate long descriptions
        
        # Format amount
        if currency == 'USD':
            amount_str = f"${tx['amount']:,.2f}"
            balance_str = f"${tx.get('balance_after', 0):,.2f}"
        else:
            amount_str = f"Rp {tx['amount']:,.0f}"
            balance_str = f"Rp {tx.get('balance_after', 0):,.0f}"
        
        table_data.append([date_str, type_str, desc_str, amount_str, balance_str])
    
    # Create table
    table = Table(table_data, colWidths=[1.5*inch, 0.8*inch, 2.5*inch, 1.2*inch, 1.2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    
    story.append(table)
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=Wallet_Statement_{wallet_type}_{currency}.pdf"
        }
    )

@api_router.get("/wallet/statement/export/excel")
async def export_wallet_statement_excel(
    wallet_type: str = "main",
    currency: str = "IDR",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Export wallet statement as Excel"""
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Get statement data
    params = {
        "wallet_type": wallet_type,
        "currency": currency,
        "limit": 1000  # Export all
    }
    if start_date:
        params["start_date"] = start_date
    if end_date:
        params["end_date"] = end_date
    
    transactions = await get_wallet_statement(**params, current_user=current_user)
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Wallet Statement"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Add title
    ws['A1'] = f"Wallet Statement - {wallet_type.title()} Wallet ({currency})"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Add period
    period_text = "Periode: "
    if start_date and end_date:
        period_text += f"{start_date} s/d {end_date}"
    else:
        period_text += "Semua transaksi"
    ws['A2'] = period_text
    
    # Headers
    headers = ['Tanggal', 'Tipe', 'Deskripsi', 'Jumlah', 'Saldo Setelah', 'Referensi', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    row_num = 5
    for tx in transactions:
        date_str = tx['date'].strftime('%Y-%m-%d %H:%M') if isinstance(tx['date'], datetime) else str(tx['date'])
        
        ws.cell(row=row_num, column=1, value=date_str)
        ws.cell(row=row_num, column=2, value='Kredit' if tx['type'] == 'credit' else 'Debit')
        ws.cell(row=row_num, column=3, value=tx['description'])
        ws.cell(row=row_num, column=4, value=tx['amount'])
        ws.cell(row=row_num, column=5, value=tx.get('balance_after', 0))
        ws.cell(row=row_num, column=6, value=tx.get('reference', ''))
        ws.cell(row=row_num, column=7, value=tx.get('status', ''))
        
        row_num += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=Wallet_Statement_{wallet_type}_{currency}.xlsx"
        }
    )

@api_router.post("/withdraw", response_model=dict)
async def withdraw_balance(request: WithdrawRequest, current_user: User = Depends(get_current_user)):
    # Check ad account exists and belongs to user
    ad_account = await db.ad_accounts.find_one({"id": request.account_id, "user_id": current_user.id})
    if not ad_account:
        raise HTTPException(status_code=404, detail="Ad account not found")
    
    # Check if account is active
    if ad_account.get("status") != "active":
        raise HTTPException(status_code=400, detail="Akun harus dalam status aktif untuk melakukan penarikan")
    
    # Check if there's already a pending withdrawal request for this account
    existing_pending = await db.withdraw_requests.find_one({
        "user_id": current_user.id,
        "account_id": request.account_id,
        "status": {"$in": ["pending", "approved"]}
    })
    
    if existing_pending:
        raise HTTPException(status_code=400, detail="Sudah ada permintaan penarikan yang sedang diproses untuk akun ini. Tunggu hingga selesai atau lakukan top-up untuk mengaktifkan kembali akun.")
    
    # Check if account has been withdrawn before and not topped up since
    last_withdraw = await db.withdraw_requests.find_one({
        "user_id": current_user.id,
        "account_id": request.account_id,
        "status": "approved"
    }, sort=[("processed_at", -1)])
    
    # SIMPLIFIED LOGIC: Just check if account has balance > 0
    if ad_account["balance"] <= 0:
        raise HTTPException(status_code=400, detail="Akun tidak memiliki saldo yang cukup untuk ditarik.")
    
    # Basic validation - minimum amount
    if request.amount <= 0:
        raise HTTPException(status_code=400, detail="Jumlah penarikan harus lebih besar dari 0")
    
    # Create withdraw request record for admin approval
    withdraw_record = WithdrawRequestRecord(
        user_id=current_user.id,
        account_id=request.account_id,
        platform=ad_account["platform"],
        account_name=ad_account["account_name"],
        requested_amount=request.amount,
        currency=ad_account.get("currency", "IDR")
    )
    
    withdraw_dict = prepare_for_mongo(withdraw_record.dict())
    await db.withdraw_requests.insert_one(withdraw_dict)
    
    # Create transaction record in pending state
    transaction = Transaction(
        user_id=current_user.id,
        type="withdraw",
        amount=request.amount,
        currency=ad_account.get("currency", "IDR"),  # Use account's currency
        description=f"Withdraw request from {ad_account['platform']} account '{ad_account['account_name']}'",
        status="pending",
        reference_id=withdraw_record.id
    )
    
    transaction_dict = prepare_for_mongo(transaction.dict())
    await db.transactions.insert_one(transaction_dict)
    
    # Create notification for admins
    currency_symbol = "Rp " if withdraw_record.currency == "IDR" else "$"
    formatted_amount = f"{currency_symbol}{request.amount:,.2f}"
    
    admin_notification = {
        "id": str(uuid.uuid4()),
        "title": get_notification_text("new_withdraw_request", "id"),
        "message": get_notification_text("withdraw_request_message", "id",
                                       username=current_user.username,
                                       currency=currency_symbol,
                                       amount=f"{request.amount:,.2f}",
                                       platform=ad_account['platform']),
        "type": "withdraw_request",
        "reference_id": withdraw_record.id,
        "is_read": False,
        "created_at": datetime.now(timezone.utc)
    }
    
    admin_notification_dict = prepare_for_mongo(admin_notification)
    await db.notifications.insert_one(admin_notification_dict)
    
    # Send email notification to all active admins
    try:
        admin_emails = await get_active_admin_emails()
        if admin_emails:
            from email_service import send_admin_new_withdraw_request_email
            send_admin_new_withdraw_request_email(
                admin_emails=admin_emails,
                client_name=current_user.name or current_user.username,
                amount=request.amount,
                currency=withdraw_record.currency,
                account_name=ad_account['account_name']
            )
            logger.info(f"📧 Admin withdraw notification emails sent to {len(admin_emails)} admins")
    except Exception as e:
        logger.error(f"❌ Failed to send admin withdraw notification emails: {e}")
    
    return {
        "message": "Withdraw request submitted successfully. Waiting for admin approval.",
        "request_id": withdraw_record.id
    }

@api_router.get("/transactions", response_model=List[dict])
async def get_transactions(current_user: User = Depends(get_current_user)):
    transactions = await db.transactions.find({"user_id": current_user.id}).sort("created_at", -1).to_list(1000)
    return [parse_from_mongo(transaction) for transaction in transactions]

@api_router.get("/wallet-topup-requests", response_model=List[dict])
async def get_client_wallet_topup_requests(current_user: User = Depends(get_current_user)):
    """Get client's own wallet top-up requests"""
    requests = await db.wallet_topup_requests.find({
        "user_id": current_user.id
    }).sort("created_at", -1).to_list(1000)
    
    result = []
    for req in requests:
        req = parse_from_mongo(req)
        
        # Get payment proof if exists
        payment_proof = None
        if req.get("payment_proof_id"):
            proof = await db.payment_proofs.find_one({"id": req["payment_proof_id"]})
            if proof:
                payment_proof = parse_from_mongo(proof)
        
        # Format similar to regular top-up requests for compatibility
        result.append({
            "id": req["id"],
            "reference_code": req.get("reference_code", ""),
            "currency": req["currency"],
            "total_amount": req["amount"],
            "total_fee": 0,  # Wallet top-up has no fees
            "status": req["status"],
            "created_at": req["created_at"],
            "verified_at": req.get("verified_at"),
            "admin_notes": req.get("admin_notes"),
            "wallet_type": req["wallet_type"],
            "payment_method": req["payment_method"],
            "unique_code": req.get("unique_code", 0),
            "total_with_unique_code": req.get("total_with_unique_code", req["amount"]),
            "type": "wallet_topup",  # Identifier for frontend
            "payment_proof": {
                "uploaded": payment_proof is not None,
                "uploaded_at": payment_proof["uploaded_at"] if payment_proof else None,
                "file_name": payment_proof["file_name"] if payment_proof else None,
                "file_path": payment_proof["file_path"] if payment_proof else None
            },
            "accounts": []  # Empty for wallet top-up
        })
    
    return result

@api_router.get("/client/monthly-topup-amount")
async def get_monthly_topup_amount(current_user: User = Depends(get_current_user)):
    """Get total amount of top-ups (wallet + regular) for current month"""
    from datetime import datetime, timezone
    
    # Get current month and year
    now = datetime.now(timezone.utc)
    current_month = now.month
    current_year = now.year
    
    # Helper function to check if a date string is in current month
    def is_current_month(date_str):
        if not date_str:
            return False
        try:
            # Parse ISO format date string
            date = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
            return date.month == current_month and date.year == current_year
        except:
            return False
    
    # Fetch wallet top-up requests for current user
    wallet_requests = await db.wallet_topup_requests.find({
        "user_id": current_user.id,
        "status": "verified"  # Only count verified top-ups
    }).to_list(1000)
    
    # Fetch regular top-up requests for current user
    regular_requests = await db.topup_requests.find({
        "user_id": current_user.id,
        "status": "verified"  # Only count verified top-ups
    }).to_list(1000)
    
    # Calculate wallet top-ups for current month
    wallet_idr_total = 0
    wallet_usd_total = 0
    for req in wallet_requests:
        if is_current_month(req.get("created_at")):
            if req.get("currency") == "IDR":
                wallet_idr_total += req.get("amount", 0)
            elif req.get("currency") == "USD":
                wallet_usd_total += req.get("amount", 0)
    
    # Calculate regular top-ups for current month
    regular_idr_total = 0
    regular_usd_total = 0
    for req in regular_requests:
        if is_current_month(req.get("created_at")):
            if req.get("currency") == "IDR":
                regular_idr_total += req.get("total_amount", 0)
            elif req.get("currency") == "USD":
                regular_usd_total += req.get("total_amount", 0)
    
    # Combine totals
    total_idr = wallet_idr_total + regular_idr_total
    total_usd = wallet_usd_total + regular_usd_total
    
    return {
        "month": current_month,
        "year": current_year,
        "total_idr": total_idr,
        "total_usd": total_usd,
        "wallet_topup_idr": wallet_idr_total,
        "wallet_topup_usd": wallet_usd_total,
        "regular_topup_idr": regular_idr_total,
        "regular_topup_usd": regular_usd_total
    }


@api_router.post("/transactions/export/excel")
async def export_transactions_excel(
    request: dict,
    current_user: User = Depends(get_current_user)
):
    """Export transactions to Excel"""
    try:
        import pandas as pd
        from io import BytesIO
        
        # Get transactions data
        transactions = request.get('transactions', [])
        
        # Function to format transaction type to user-friendly text
        def format_transaction_type(trans_type):
            type_mapping = {
                'topup': 'Top Up',
                'withdraw': 'Penarikan',
                'withdraw_request': 'Penarikan',
                'approved_transfer': 'Transfer Disetujui',
                'transfer': 'Transfer',
                'account_request': 'Permintaan Akun',
                'balance_transfer': 'Transfer Saldo',
                'fee': 'Fee/Biaya'
            }
            clean_type = str(trans_type).lower().strip()
            return type_mapping.get(clean_type, clean_type.replace('_', ' ').title())
        
        def format_status(status):
            status_mapping = {
                'completed': 'Selesai',
                'pending': 'Menunggu',
                'approved': 'Disetujui',
                'rejected': 'Ditolak',
                'failed': 'Gagal',
                'cancelled': 'Dibatalkan'
            }
            clean_status = str(status).lower().strip()
            return status_mapping.get(clean_status, str(status).title())
        
        if not transactions:
            # Create empty dataframe
            df = pd.DataFrame(columns=['Tanggal', 'Jenis Transaksi', 'Akun', 'Jumlah', 'Currency', 'Status'])
        else:
            # Prepare data for Excel
            data = []
            for trans in transactions:
                data.append({
                    'Tanggal': trans.get('created_at', '')[:10] if trans.get('created_at') else '',
                    'Jenis Transaksi': format_transaction_type(trans.get('type', '')),
                    'Akun': trans.get('account_name', 'N/A'),
                    'Jumlah': trans.get('amount', 0),
                    'Currency': trans.get('currency', 'IDR'),
                    'Status': format_status(trans.get('status', ''))
                })
            
            df = pd.DataFrame(data)
        
        # Create Excel buffer
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Transactions', index=False)
            
            # Auto-adjust column widths
            worksheet = writer.sheets['Transactions']
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except (TypeError, AttributeError):
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        buffer.seek(0)
        
        return Response(
            content=buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=transactions.xlsx"}
        )
        
    except Exception as e:
        logger.error(f"Error exporting transactions to Excel: {e}")
        raise HTTPException(status_code=500, detail="Failed to export Excel")

@api_router.get("/dashboard/stats", response_model=dict)
async def get_dashboard_stats(current_user: User = Depends(get_current_user)):
    # Get user's accounts count
    accounts_count = await db.ad_accounts.count_documents({"user_id": current_user.id})
    
    # Get total balance across all accounts
    accounts = await db.ad_accounts.find({"user_id": current_user.id}).to_list(1000)
    total_ads_balance = sum(account.get("balance", 0) for account in accounts)
    
    # Get recent transactions count
    recent_transactions = await db.transactions.count_documents({
        "user_id": current_user.id,
        "created_at": {"$gte": datetime.now(timezone.utc) - timedelta(days=30)}
    })
    
    return {
        "wallet_balance_idr": current_user.wallet_balance_idr,  # Legacy
        "wallet_balance_usd": current_user.wallet_balance_usd,  # Legacy
        "main_wallet_idr": getattr(current_user, 'main_wallet_idr', 0.0),
        "main_wallet_usd": getattr(current_user, 'main_wallet_usd', 0.0),
        "withdrawal_wallet_idr": getattr(current_user, 'withdrawal_wallet_idr', 0.0),
        "withdrawal_wallet_usd": getattr(current_user, 'withdrawal_wallet_usd', 0.0),
        "total_ads_balance": total_ads_balance,
        "accounts_count": accounts_count,
        "recent_transactions": recent_transactions
    }

# Currency Exchange Endpoints
@api_router.get("/exchange-rate/{from_currency}/{to_currency}")
async def get_current_exchange_rate(from_currency: str, to_currency: str):
    """Get real-time exchange rate"""
    rate = await get_exchange_rate(from_currency.upper(), to_currency.upper())
    return {
        "from_currency": from_currency.upper(),
        "to_currency": to_currency.upper(),
        "rate": rate,
        "timestamp": datetime.now(timezone.utc).isoformat()
    }

@api_router.post("/exchange", response_model=dict)
async def exchange_currency(request: ExchangeRequest, current_user: User = Depends(get_current_user)):
    """Exchange between IDR and USD"""
    try:
        # Validate currencies
        if request.from_currency.upper() not in ["IDR", "USD"] or request.to_currency.upper() not in ["IDR", "USD"]:
            raise HTTPException(status_code=400, detail="Only IDR and USD currencies are supported")
        
        if request.from_currency.upper() == request.to_currency.upper():
            raise HTTPException(status_code=400, detail="Cannot exchange to the same currency")
        
        if request.amount <= 0:
            raise HTTPException(status_code=400, detail="Amount must be greater than 0")
        
        # Get current exchange rate
        rate = await get_exchange_rate(request.from_currency.upper(), request.to_currency.upper())
        to_amount = request.amount * rate
        
        # Check sufficient balance
        current_balance = 0
        if request.from_currency.upper() == "IDR":
            current_balance = current_user.main_wallet_idr
        else:
            current_balance = current_user.main_wallet_usd
        
        if current_balance < request.amount:
            raise HTTPException(status_code=400, detail="Insufficient balance")
        
        # Perform exchange using precise decimal calculations
        update_data = {"updated_at": datetime.now(timezone.utc)}
        
        if request.from_currency.upper() == "IDR":
            # Subtract from IDR wallet
            current_idr_decimal = to_decimal(current_user.main_wallet_idr)
            request_amount_decimal = to_decimal(request.amount)
            new_idr_decimal = decimal_subtract(current_idr_decimal, request_amount_decimal)
            update_data["main_wallet_idr"] = to_float(decimal_round(new_idr_decimal))
            
            # Add to USD wallet
            current_usd_decimal = to_decimal(current_user.main_wallet_usd)
            to_amount_decimal = to_decimal(to_amount)
            new_usd_decimal = decimal_add(current_usd_decimal, to_amount_decimal)
            update_data["main_wallet_usd"] = to_float(decimal_round(new_usd_decimal))
        else:
            # Subtract from USD wallet
            current_usd_decimal = to_decimal(current_user.main_wallet_usd)
            request_amount_decimal = to_decimal(request.amount)
            new_usd_decimal = decimal_subtract(current_usd_decimal, request_amount_decimal)
            update_data["main_wallet_usd"] = to_float(decimal_round(new_usd_decimal))
            
            # Add to IDR wallet
            current_idr_decimal = to_decimal(current_user.main_wallet_idr)
            to_amount_decimal = to_decimal(to_amount)
            new_idr_decimal = decimal_add(current_idr_decimal, to_amount_decimal)
            update_data["main_wallet_idr"] = to_float(decimal_round(new_idr_decimal))
        
        await db.users.update_one({"id": current_user.id}, {"$set": update_data})
        
        # Create exchange record
        exchange_record = CurrencyExchange(
            user_id=current_user.id,
            from_currency=request.from_currency.upper(),
            to_currency=request.to_currency.upper(),
            from_amount=request.amount,
            to_amount=to_amount,
            exchange_rate=rate
        )
        
        exchange_dict = prepare_for_mongo(exchange_record.dict())
        await db.currency_exchanges.insert_one(exchange_dict)
        
        return {
            "message": "Currency exchange successful",
            "from_amount": request.amount,
            "to_amount": to_amount,
            "exchange_rate": rate,
            "new_balance_idr": update_data.get("main_wallet_idr", current_user.main_wallet_idr),
            "new_balance_usd": update_data.get("main_wallet_usd", current_user.main_wallet_usd)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Exchange error: {e}")
        raise HTTPException(status_code=500, detail="Exchange failed")

@api_router.get("/exchange-history")
async def get_exchange_history(current_user: User = Depends(get_current_user)):
    """Get user's currency exchange history"""
    try:
        exchanges = await db.currency_exchanges.find(
            {"user_id": current_user.id}
        ).sort("created_at", -1).limit(50).to_list(50)
        
        result = []
        for exchange in exchanges:
            exchange = parse_from_mongo(exchange)
            result.append(exchange)
        
        return result
    except Exception as e:
        logger.error(f"Error fetching exchange history: {e}")
        return []

# Share Account endpoints
@api_router.post("/accounts/share", response_model=dict)
async def request_account_share(request: ShareRequestCreate, current_user: User = Depends(get_current_user)):
    """Request to share an existing account to different BM/Email/BC ID"""
    # Verify the account belongs to the user
    account = await db.ad_accounts.find_one({"id": request.account_id, "user_id": current_user.id})
    if not account:
        raise HTTPException(status_code=404, detail="Account not found or doesn't belong to you")
    
    # Validate platform-specific fields
    platform = account["platform"]
    if platform == "facebook" and (not request.target_bm_email or len(request.target_bm_email) == 0):
        raise HTTPException(status_code=400, detail="BM ID or Email is required for Facebook account sharing")
    elif platform == "google" and (not request.target_email or len(request.target_email) == 0):
        raise HTTPException(status_code=400, detail="Target email is required for Google Ads account sharing")
    elif platform == "tiktok" and (not request.target_bc_id or len(request.target_bc_id) == 0):
        raise HTTPException(status_code=400, detail="Target BC ID is required for TikTok Ads account sharing")
    
    # Create share request
    share_request = ShareRequest(
        user_id=current_user.id,
        account_id=request.account_id,
        platform=platform,
        account_name=account["account_name"],
        target_bm_email=request.target_bm_email,
        target_email=request.target_email,
        target_bc_id=request.target_bc_id,
        notes=request.notes
    )
    
    share_dict = prepare_for_mongo(share_request.dict())
    await db.share_requests.insert_one(share_dict)
    
    # Create admin notification for new share request
    platform_name = platform.title()
    if platform == "facebook":
        platform_name = "Facebook Ads"
        target_count = len(request.target_bm_email) if request.target_bm_email else 0
        target_info = f" to {target_count} BM ID(s)"
    elif platform == "google":
        platform_name = "Google Ads"
        target_count = len(request.target_email) if request.target_email else 0
        target_info = f" to {target_count} Email(s)"
    elif platform == "tiktok":
        platform_name = "TikTok Ads"
        target_count = len(request.target_bc_id) if request.target_bc_id else 0
        target_info = f" to {target_count} BC ID(s)"
    
    await create_notification(
        title=f"New {platform_name} Share Request",
        message=f"User {current_user.username} requested to share {platform_name} account '{account['account_name']}'{target_info}",
        notification_type="share_request",
        reference_id=share_request.id
    )
    
    # Send email notification to admins about new share request
    try:
        admin_emails = await get_active_admin_emails()
        if admin_emails:
            send_admin_new_share_request_email(
                admin_emails=admin_emails,
                client_name=current_user.full_name or current_user.username,
                platform=platform,
                account_name=account["account_name"],
                target_count=target_count
            )
            logger.info(f"📧 Share request notification emails sent to {len(admin_emails)} admins")
    except Exception as e:
        logger.error(f"Failed to send share request email to admins: {e}")
    
    return {"message": "Share request submitted successfully", "request_id": share_request.id}

@api_router.get("/accounts/share-requests", response_model=List[ShareRequestResponse])
async def get_user_share_requests(current_user: User = Depends(get_current_user)):
    """Get share requests for the current user"""
    try:
        requests = await db.share_requests.find({"user_id": current_user.id}).sort("created_at", -1).to_list(length=None)
        
        result = []
        for request in requests:
            try:
                parsed_request = parse_from_mongo(request)
                
                # Convert old string format to array for backward compatibility
                if parsed_request.get("target_bm_email") and isinstance(parsed_request["target_bm_email"], str):
                    parsed_request["target_bm_email"] = [parsed_request["target_bm_email"]]
                if parsed_request.get("target_email") and isinstance(parsed_request["target_email"], str):
                    parsed_request["target_email"] = [parsed_request["target_email"]]
                if parsed_request.get("target_bc_id") and isinstance(parsed_request["target_bc_id"], str):
                    parsed_request["target_bc_id"] = [parsed_request["target_bc_id"]]
                
                result.append(ShareRequestResponse(**parsed_request))
            except Exception as e:
                logger.error(f"Error parsing share request {request.get('id')}: {str(e)}")
                continue
        
        return result
    except Exception as e:
        logger.error(f"Error fetching share requests: {str(e)}")
        return []

@api_router.get("/accounts", response_model=List[dict])
async def get_user_accounts(current_user: User = Depends(get_current_user)):
    """Get active ad accounts for the current user for top-up/withdraw"""
    accounts = await db.ad_accounts.find({
        "user_id": current_user.id,
        "status": {"$in": ["active", "sharing"]}  # Include active and sharing accounts
    }).sort("created_at", -1).to_list(length=None)
    
    logger.info(f"[get_user_accounts] User {current_user.username} has {len(accounts)} active/sharing accounts")
    
    # Get all groups for the user to map group names
    groups = await db.account_groups.find({
        "user_id": current_user.id
    }).to_list(length=None)
    
    # Create a mapping of group_id to group_name
    group_map = {}
    for group in groups:
        group = parse_from_mongo(group)
        group_map[group["id"]] = group["name"]
    
    result = []
    for account in accounts:
        account = parse_from_mongo(account)
        # NEW SIMPLIFIED WITHDRAWAL ELIGIBILITY RULE:
        # If account has balance > 0 AND no pending withdrawals = can withdraw
        can_withdraw = True
        last_topup_date = account.get("last_topup_date")
        
        # Check for pending/approved withdrawals (still block if there's pending withdrawal)
        pending_withdraws = await db.withdraw_requests.find({
            "account_id": account["id"],
            "status": {"$in": ["pending", "approved"]}
        }).to_list(None)
        
        if pending_withdraws:
            can_withdraw = False
        else:
            # Simple rule: If balance > 0, can withdraw
            account_balance = account.get("balance", 0)
            if account_balance <= 0:
                can_withdraw = False
        
        account_data = {
            "id": account["id"],
            "platform": account["platform"],
            "account_name": account["account_name"],
            "account_id": account.get("account_id"),
            "status": account.get("status", "active"),
            "balance": account.get("balance", 0),
            "fee_percentage": account.get("fee_percentage", 0),
            "currency": account.get("currency"),  # Add currency field for withdraw functionality
            "created_at": account["created_at"],
            "updated_at": account.get("updated_at"),
            "admin_notes": account.get("admin_notes"),
            "group_id": account.get("group_id"),
            "group_name": group_map.get(account.get("group_id")) if account.get("group_id") else None,
            "can_withdraw": can_withdraw,
            "last_topup_date": last_topup_date.isoformat() if last_topup_date and isinstance(last_topup_date, datetime) else last_topup_date
        }
        result.append(account_data)
    
    logger.info(f"[get_user_accounts] Returning {len(result)} accounts to user {current_user.username}")
    return result

# Account Groups Management
@api_router.get("/account-groups", response_model=List[dict])
async def get_user_account_groups(current_user: User = Depends(get_current_user)):
    """Get all account groups for the current user"""
    try:
        groups = await db.account_groups.find({
            "user_id": current_user.id
        }).sort("created_at", -1).to_list(length=None)
        
        result = []
        for group in groups:
            group = parse_from_mongo(group)
            # Count accounts in this group
            account_count = await db.ad_accounts.count_documents({
                "user_id": current_user.id,
                "group_id": group["id"]
            })
            
            result.append({
                "id": group["id"],
                "name": group["name"],
                "description": group.get("description", ""),
                "account_ids": group.get("account_ids", []),
                "account_count": account_count,
                "created_at": group["created_at"],
                "updated_at": group.get("updated_at")
            })
        
        return result
    except Exception as e:
        logger.error(f"Error fetching account groups: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch account groups")

@api_router.post("/account-groups", response_model=dict)
async def create_account_group(
    group_data: AccountGroupCreate,
    current_user: User = Depends(get_current_user)
):
    """Create a new account group"""
    try:
        # Check if group name already exists for this user
        existing_group = await db.account_groups.find_one({
            "user_id": current_user.id,
            "name": group_data.name
        })
        
        if existing_group:
            raise HTTPException(status_code=400, detail="Group name already exists")
        
        # Create new group
        group = AccountGroup(
            user_id=current_user.id,
            name=group_data.name,
            description=group_data.description,
            account_ids=group_data.accounts
        )
        
        group_dict = prepare_for_mongo(group.dict())
        await db.account_groups.insert_one(group_dict)
        
        # Update accounts to belong to this group
        if group_data.accounts:
            await db.ad_accounts.update_many(
                {
                    "user_id": current_user.id,
                    "id": {"$in": group_data.accounts}
                },
                {"$set": {"group_id": group.id}}
            )
        
        # Count accounts in this group
        account_count = await db.ad_accounts.count_documents({
            "user_id": current_user.id,
            "group_id": group.id
        })
        
        # Return group object with same format as GET endpoint
        return {
            "id": group.id,
            "name": group.name,
            "description": group.description,
            "account_ids": group.account_ids,
            "account_count": account_count,
            "created_at": group.created_at.isoformat() if isinstance(group.created_at, datetime) else group.created_at,
            "updated_at": group.updated_at.isoformat() if group.updated_at and isinstance(group.updated_at, datetime) else group.updated_at
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating account group: {e}")
        raise HTTPException(status_code=500, detail="Failed to create account group")

@api_router.put("/account-groups/{group_id}", response_model=dict)
async def update_account_group(
    group_id: str,
    group_data: AccountGroupUpdate,
    current_user: User = Depends(get_current_user)
):
    """Update an existing account group"""
    try:
        # Check if group exists and belongs to user
        existing_group = await db.account_groups.find_one({
            "id": group_id,
            "user_id": current_user.id
        })
        
        if not existing_group:
            raise HTTPException(status_code=404, detail="Group not found")
        
        # Check if new name conflicts with other groups
        if group_data.name:
            name_conflict = await db.account_groups.find_one({
                "user_id": current_user.id,
                "name": group_data.name,
                "id": {"$ne": group_id}
            })
            
            if name_conflict:
                raise HTTPException(status_code=400, detail="Group name already exists")
        
        # Prepare update data
        update_data = {"updated_at": datetime.now(timezone.utc)}
        
        if group_data.name is not None:
            update_data["name"] = group_data.name
        if group_data.description is not None:
            update_data["description"] = group_data.description
        if group_data.accounts is not None:
            update_data["account_ids"] = group_data.accounts
            
            # Remove group assignment from all current accounts
            await db.ad_accounts.update_many(
                {
                    "user_id": current_user.id,
                    "group_id": group_id
                },
                {"$unset": {"group_id": ""}}
            )
            
            # Assign new accounts to this group
            if group_data.accounts:
                await db.ad_accounts.update_many(
                    {
                        "user_id": current_user.id,
                        "id": {"$in": group_data.accounts}
                    },
                    {"$set": {"group_id": group_id}}
                )
        
        # Update the group
        await db.account_groups.update_one(
            {"id": group_id},
            {"$set": prepare_for_mongo(update_data)}
        )
        
        return {"message": "Group updated successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating account group: {e}")
        raise HTTPException(status_code=500, detail="Failed to update account group")

@api_router.delete("/account-groups/{group_id}", response_model=dict)
async def delete_account_group(
    group_id: str,
    current_user: User = Depends(get_current_user)
):
    """Delete an account group"""
    try:
        # Check if group exists and belongs to user
        existing_group = await db.account_groups.find_one({
            "id": group_id,
            "user_id": current_user.id
        })
        
        if not existing_group:
            raise HTTPException(status_code=404, detail="Group not found")
        
        # Remove group assignment from all accounts in this group
        await db.ad_accounts.update_many(
            {
                "user_id": current_user.id,
                "group_id": group_id
            },
            {"$unset": {"group_id": ""}}
        )
        
        # Delete the group
        await db.account_groups.delete_one({"id": group_id})
        
        return {"message": "Group deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting account group: {e}")
        raise HTTPException(status_code=500, detail="Failed to delete account group")

@api_router.put("/account-groups/{group_id}/accounts", response_model=dict)
async def add_accounts_to_group(
    group_id: str,
    request_data: AddAccountsToGroup,
    current_user: User = Depends(get_current_user)
):
    """Add accounts to a group"""
    account_ids = request_data.account_ids
    try:
        # Check if group exists and belongs to user
        existing_group = await db.account_groups.find_one({
            "id": group_id,
            "user_id": current_user.id
        })
        
        if not existing_group:
            raise HTTPException(status_code=404, detail="Group not found")
        
        # Verify all accounts belong to the user
        user_accounts = await db.ad_accounts.count_documents({
            "user_id": current_user.id,
            "id": {"$in": account_ids}
        })
        
        if user_accounts != len(account_ids):
            raise HTTPException(status_code=400, detail="Some accounts do not exist or do not belong to you")
        
        # Add accounts to the group
        await db.ad_accounts.update_many(
            {
                "user_id": current_user.id,
                "id": {"$in": account_ids}
            },
            {"$set": {"group_id": group_id}}
        )
        
        # Update group's account_ids list
        current_account_ids = existing_group.get("account_ids", [])
        updated_account_ids = list(set(current_account_ids + account_ids))
        
        await db.account_groups.update_one(
            {"id": group_id},
            {"$set": {
                "account_ids": updated_account_ids,
                "updated_at": datetime.now(timezone.utc)
            }}
        )
        
        return {"message": f"Added {len(account_ids)} accounts to group successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error adding accounts to group: {e}")
        raise HTTPException(status_code=500, detail="Failed to add accounts to group")

@api_router.get("/account-groups/{group_id}/accounts", response_model=dict)
async def get_group_accounts(
    group_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get all accounts in a specific group"""
    try:
        # Check if group exists and belongs to user
        group = await db.account_groups.find_one({
            "id": group_id,
            "user_id": current_user.id
        })
        
        if not group:
            raise HTTPException(status_code=404, detail="Group not found")
        
        # Get accounts in this group
        accounts_cursor = db.ad_accounts.find({
            "user_id": current_user.id,
            "group_id": group_id
        })
        
        accounts = []
        async for account in accounts_cursor:
            account = parse_from_mongo(account)
            accounts.append({
                "id": account["id"],
                "account_name": account["account_name"],
                "platform": account["platform"],
                "status": account["status"],
                "currency": account.get("currency", "IDR"),
                "balance": account.get("balance", 0),
                "created_at": account.get("created_at")
            })
        
        return {
            "group": {
                "id": group["id"],
                "name": group["name"],
                "description": group.get("description")
            },
            "accounts": accounts
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching group accounts: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch group accounts")

@api_router.get("/account-groups/{group_id}/available-accounts", response_model=dict)
async def get_available_accounts_for_group(
    group_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get accounts that can be added to a group (not already in this group)"""
    try:
        # Check if group exists and belongs to user
        group = await db.account_groups.find_one({
            "id": group_id,
            "user_id": current_user.id
        })
        
        if not group:
            raise HTTPException(status_code=404, detail="Group not found")
        
        # Get accounts that are not in this group
        accounts_cursor = db.ad_accounts.find({
            "user_id": current_user.id,
            "$or": [
                {"group_id": {"$ne": group_id}},
                {"group_id": {"$exists": False}},
                {"group_id": None}
            ]
        })
        
        accounts = []
        async for account in accounts_cursor:
            account = parse_from_mongo(account)
            accounts.append({
                "id": account["id"],
                "account_name": account["account_name"],
                "platform": account["platform"],
                "status": account["status"],
                "currency": account.get("currency", "IDR"),
                "balance": account.get("balance", 0),
                "created_at": account.get("created_at")
            })
        
        return {
            "group": {
                "id": group["id"],
                "name": group["name"],
                "description": group.get("description")
            },
            "accounts": accounts
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching available accounts: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch available accounts")

@api_router.delete("/account-groups/{group_id}/accounts/{account_id}", response_model=dict)
async def remove_account_from_group(
    group_id: str,
    account_id: str,
    current_user: User = Depends(get_current_user)
):
    """Remove an account from a group"""
    try:
        # Check if group exists and belongs to user
        group = await db.account_groups.find_one({
            "id": group_id,
            "user_id": current_user.id
        })
        
        if not group:
            raise HTTPException(status_code=404, detail="Group not found")
        
        # Check if account exists and belongs to user
        account = await db.ad_accounts.find_one({
            "id": account_id,
            "user_id": current_user.id,
            "group_id": group_id
        })
        
        if not account:
            raise HTTPException(status_code=404, detail="Account not found in this group")
        
        # Remove account from group
        await db.ad_accounts.update_one(
            {"id": account_id, "user_id": current_user.id},
            {"$unset": {"group_id": ""}}
        )
        
        # Update group's account_ids list
        current_account_ids = group.get("account_ids", [])
        updated_account_ids = [aid for aid in current_account_ids if aid != account_id]
        
        await db.account_groups.update_one(
            {"id": group_id},
            {"$set": {
                "account_ids": updated_account_ids,
                "updated_at": datetime.now(timezone.utc)
            }}
        )
        
        return {"message": "Account removed from group successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error removing account from group: {e}")
        raise HTTPException(status_code=500, detail="Failed to remove account from group")

@api_router.get("/accounts/{account_id}/details", response_model=dict)
async def get_account_details(
    account_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get detailed information about a specific account"""
    try:
        # Find the account - search by both internal id and external account_id
        account = await db.ad_accounts.find_one({
            "$or": [
                {"id": account_id},
                {"account_id": account_id}
            ],
            "user_id": current_user.id
        })
        
        if not account:
            raise HTTPException(status_code=404, detail="Account not found")
        
        account = parse_from_mongo(account)
        
        # Get group information if account belongs to a group
        group_info = None
        if account.get("group_id"):
            group = await db.account_groups.find_one({
                "id": account["group_id"],
                "user_id": current_user.id
            })
            if group:
                group = parse_from_mongo(group)
                group_info = {
                    "id": group["id"],
                    "name": group["name"],
                    "description": group.get("description")
                }
        
        # Get recent transactions for this account - ONLY VERIFIED
        # Use the internal account id for querying
        internal_account_id = account["id"]
        
        # Query supports BOTH old (accounts array) and new (single account_id) structures
        recent_topups = await db.topup_requests.find({
            "user_id": current_user.id,
            "$or": [
                {"account_id": internal_account_id},  # New structure: single account_id field
                {"accounts.account_id": internal_account_id}  # Old structure: accounts array
            ],
            "status": {"$in": ["verified", "completed", "approved"]}  # Include all approved statuses
        }).sort("verified_at", -1).limit(10).to_list(length=None)
        
        # Also get wallet transfers to this account
        wallet_transfers = await db.wallet_transfers.find({
            "user_id": current_user.id,
            "target_account_id": internal_account_id,  # Use internal id
            "status": "approved"
        }).sort("processed_at", -1).limit(10).to_list(length=None)
        
        topup_history = []
        
        # Add regular topups
        for topup in recent_topups:
            topup = parse_from_mongo(topup)
            
            # For multi-account topups (accounts array structure), find the specific account amount
            if "accounts" in topup and isinstance(topup["accounts"], list):
                # Find this specific account in the accounts array
                account_topup = next(
                    (acc for acc in topup["accounts"] if acc.get("account_id") == internal_account_id),
                    None
                )
                if account_topup:
                    # Use the account-specific amount (WITHOUT fee)
                    account_amount = account_topup.get("amount", 0)
                    
                    topup_history.append({
                        "id": topup["id"],
                        "amount": account_amount,  # Use account-specific amount only (no fee)
                        "currency": topup["currency"],
                        "status": topup["status"],
                        "created_at": topup["created_at"],
                        "verified_at": topup.get("verified_at"),
                        "type": "topup"
                    })
            else:
                # Single account topup (old structure or direct account_id field)
                topup_history.append({
                    "id": topup["id"],
                    "amount": topup.get("total_amount", topup.get("amount", 0)),
                    "currency": topup["currency"],
                    "status": topup["status"],
                    "created_at": topup["created_at"],
                    "verified_at": topup.get("verified_at"),
                    "type": "topup"
                })
        
        # Add wallet transfers
        for transfer in wallet_transfers:
            transfer = parse_from_mongo(transfer)
            topup_history.append({
                "id": transfer["id"],
                "amount": transfer["amount"],
                "currency": transfer["currency"],
                "status": "verified",  # Approved transfers treated as verified
                "created_at": transfer["created_at"],
                "verified_at": transfer.get("processed_at"),
                "type": "wallet_transfer"
            })
        
        # Sort combined list by verified_at
        topup_history.sort(key=lambda x: x.get("verified_at") or x.get("created_at"), reverse=True)
        topup_history = topup_history[:5]  # Keep only 5 most recent
        
        # Get recent withdraw requests
        recent_withdraws = await db.withdrawal_requests.find({
            "user_id": current_user.id,
            "account_id": internal_account_id  # Use internal id
        }).sort("created_at", -1).limit(5).to_list(length=None)
        
        withdraw_history = []
        for withdraw in recent_withdraws:
            withdraw = parse_from_mongo(withdraw)
            withdraw_history.append({
                "id": withdraw["id"],
                "requested_amount": withdraw["requested_amount"],
                "verified_amount": withdraw.get("admin_verified_amount"),
                "currency": withdraw["currency"],
                "status": withdraw["status"],
                "created_at": withdraw["created_at"]
            })
        
        return {
            "id": account["id"],
            "platform": account["platform"],
            "account_name": account["account_name"],
            "account_id": account.get("account_id"),
            "status": account.get("status"),
            "balance": account.get("balance", 0),
            "currency": account.get("currency", "IDR"),  # Added currency field
            "fee_percentage": account.get("fee_percentage", 0),
            "created_at": account["created_at"],
            "updated_at": account.get("updated_at"),
            "admin_notes": account.get("admin_notes"),
            "group": group_info,
            "recent_topups": topup_history,
            "recent_withdraws": withdraw_history
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching account details: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch account details")

# Admin Share Request Management
@api_router.get("/admin/share-requests", response_model=List[dict])
async def get_admin_share_requests(
    status: Optional[str] = None,
    platform: Optional[str] = None,
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get all share requests for admin management"""
    filter_query = {}
    if status:
        filter_query["status"] = status
    if platform:
        filter_query["platform"] = platform
    
    requests = await db.share_requests.find(filter_query).sort("created_at", -1).to_list(length=None)
    
    result = []
    for request in requests:
        request = parse_from_mongo(request)
        
        # Get user info
        user = await db.users.find_one({"id": request["user_id"]})
        user_info = None
        if user:
            user_info = {
                "username": user.get("username"),
                "email": user.get("email")
            }
        
        # Get ad account info to get the real account ID
        ad_account_info = None
        if request.get("account_id"):
            ad_account = await db.ad_accounts.find_one({"id": request["account_id"]})
            if ad_account:
                ad_account_info = {
                    "real_account_id": ad_account.get("account_id"),  # The real ad account ID (FB/Google/TikTok ID)
                    "account_name": ad_account.get("account_name"),
                    "platform": ad_account.get("platform")
                }
        
        # Get processed by info if available
        processed_by = None
        if request.get("processed_by"):
            admin = await db.admin_users.find_one({"id": request["processed_by"]})
            if admin:
                admin = parse_from_mongo(admin)
                processed_by = {
                    "id": admin.get("id"),
                    "username": admin.get("username"),
                    "name": admin.get("name", admin.get("username"))
                }
        
        result.append({
            **request,
            "user": user_info,
            "ad_account": ad_account_info,
            "processed_by": processed_by
        })
    
    return result

@api_router.put("/admin/share-requests/{request_id}/status", response_model=dict)
async def update_share_request_status(
    request_id: str,
    status_data: RequestStatusUpdate,  # Reuse the same model
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Update share request status"""
    request = await db.share_requests.find_one({"id": request_id})
    if not request:
        raise HTTPException(status_code=404, detail="Share request not found")
    
    update_data = {
        "status": status_data.status,
        "processed_by": current_admin.id,
        "processed_at": datetime.now(timezone.utc).isoformat()
    }
    
    if status_data.admin_notes:
        update_data["admin_notes"] = status_data.admin_notes
    
    await db.share_requests.update_one(
        {"id": request_id},
        {"$set": update_data}
    )
    
    # Create client notification
    if status_data.status in ["approved", "completed", "rejected"]:
        platform_name = request["platform"].title()
        if request["platform"] == "facebook":
            platform_name = "Facebook Ads"
        elif request["platform"] == "google":
            platform_name = "Google Ads"
        elif request["platform"] == "tiktok":
            platform_name = "TikTok Ads"
        
        if status_data.status == "approved":
            title = f"🎉 {platform_name} Share Request Approved"
            message = f"Your {platform_name} share request for '{request['account_name']}' has been approved and is being processed."
            notif_type = "approval"
        elif status_data.status == "completed":
            title = f"✅ {platform_name} Share Completed"
            message = f"Your {platform_name} account '{request['account_name']}' has been successfully shared."
            notif_type = "completed"
        else:  # rejected
            title = f"❌ {platform_name} Share Request Rejected"
            message = f"Your {platform_name} share request for '{request['account_name']}' has been rejected. Please contact support for more information."
            notif_type = "rejection"
        
        client_notification = {
            "id": str(uuid.uuid4()),
            "user_id": request["user_id"],
            "title": title,
            "message": message,
            "type": notif_type,
            "reference_id": request_id,
            "is_read": False,
            "created_at": datetime.now(timezone.utc)
        }
        
        client_notification_dict = prepare_for_mongo(client_notification)
        await db.client_notifications.insert_one(client_notification_dict)
        
        # Send email notification to client about share request status
        try:
            user = await db.users.find_one({"id": request["user_id"]})
            if user and user.get("email"):
                if status_data.status == "approved":
                    send_client_share_request_approved_email(
                        client_email=user["email"],
                        client_name=user.get("full_name") or user.get("username"),
                        platform=request["platform"],
                        account_name=request["account_name"]
                    )
                    logger.info(f"📧 Share request approval email sent to {user['email']}")
                elif status_data.status == "rejected":
                    send_client_share_request_rejected_email(
                        client_email=user["email"],
                        client_name=user.get("full_name") or user.get("username"),
                        platform=request["platform"],
                        account_name=request["account_name"],
                        reason=status_data.admin_notes or ""
                    )
                    logger.info(f"📧 Share request rejection email sent to {user['email']}")
        except Exception as e:
            logger.error(f"Failed to send share request status email: {e}")
    
    return {"message": "Share request status updated successfully"}

@api_router.put("/admin/share-requests/bulk-update", response_model=dict)
async def bulk_update_share_request_status(
    bulk_data: BulkRequestUpdate,  # Reuse the existing BulkRequestUpdate model
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Bulk update share request status"""
    updated_count = 0
    failed_updates = []
    
    for request_id in bulk_data.request_ids:
        try:
            request = await db.share_requests.find_one({"id": request_id})
            if not request:
                failed_updates.append(f"Share request {request_id} not found")
                continue
            
            update_data = {
                "status": bulk_data.status,
                "processed_by": current_admin.id,
                "processed_at": datetime.now(timezone.utc)
            }
            
            if bulk_data.admin_notes:
                update_data["admin_notes"] = bulk_data.admin_notes
            
            await db.share_requests.update_one(
                {"id": request_id},
                {"$set": update_data}
            )
            
            # Create client notification for status changes
            if bulk_data.status in ["approved", "completed", "rejected"]:
                platform_name = request["platform"].title()
                if request["platform"] == "facebook":
                    platform_name = "Facebook Ads"
                elif request["platform"] == "google":
                    platform_name = "Google Ads"
                elif request["platform"] == "tiktok":
                    platform_name = "TikTok Ads"
                
                if bulk_data.status == "approved":
                    title = f"🎉 {platform_name} Share Request Approved"
                    message = f"Your {platform_name} share request for '{request['account_name']}' has been approved and is being processed."
                    notif_type = "approval"
                elif bulk_data.status == "completed":
                    title = f"✅ {platform_name} Share Completed"
                    message = f"Your {platform_name} account '{request['account_name']}' has been successfully shared."
                    notif_type = "completed"
                else:  # rejected
                    title = f"❌ {platform_name} Share Request Rejected"
                    message = f"Your {platform_name} share request for '{request['account_name']}' has been rejected. Please contact support for more information."
                    notif_type = "rejection"
                
                client_notification = {
                    "id": str(uuid.uuid4()),
                    "user_id": request["user_id"],
                    "title": title,
                    "message": message,
                    "type": notif_type,
                    "reference_id": request_id,
                    "is_read": False,
                    "created_at": datetime.now(timezone.utc)
                }
                
                client_notification_dict = prepare_for_mongo(client_notification)
                await db.client_notifications.insert_one(client_notification_dict)
                
                # Send email notification to client about share request status (bulk update)
                try:
                    user = await db.users.find_one({"id": request["user_id"]})
                    if user and user.get("email"):
                        if bulk_data.status == "approved":
                            send_client_share_request_approved_email(
                                client_email=user["email"],
                                client_name=user.get("full_name") or user.get("username"),
                                platform=request["platform"],
                                account_name=request["account_name"]
                            )
                            logger.info(f"📧 Bulk share request approval email sent to {user['email']}")
                        elif bulk_data.status == "rejected":
                            send_client_share_request_rejected_email(
                                client_email=user["email"],
                                client_name=user.get("full_name") or user.get("username"),
                                platform=request["platform"],
                                account_name=request["account_name"],
                                reason=bulk_data.admin_notes or ""
                            )
                            logger.info(f"📧 Bulk share request rejection email sent to {user['email']}")
                except Exception as e:
                    logger.error(f"Failed to send bulk share request status email: {e}")
            
            updated_count += 1
        
        except Exception as e:
            logger.error(f"Error updating share request {request_id}: {e}")
            failed_updates.append(f"Error updating share request {request_id}: {str(e)}")
    
    # Create admin notification for bulk update
    await create_notification(
        title=f"📋 Bulk Share Request Update Completed",
        message=f"Admin {current_admin.full_name} bulk updated {updated_count} share requests to '{bulk_data.status}' status",
        notification_type="bulk_update",
        reference_id=None
    )
    
    response_message = f"Successfully updated {updated_count} share requests"
    if failed_updates:
        response_message += f". {len(failed_updates)} updates failed."
    
    return {
        "message": response_message,
        "updated_count": updated_count,
        "failed_count": len(failed_updates),
        "failed_updates": failed_updates
    }

# Client Notification endpoints
@api_router.get("/client/notifications", response_model=List[ClientNotificationResponse])
async def get_client_notifications(
    unread_only: bool = False,
    limit: int = 50,
    current_user: User = Depends(get_current_user)
):
    """Get notifications for client"""
    filter_query = {"user_id": current_user.id}
    if unread_only:
        filter_query["is_read"] = False
    
    notifications_cursor = db.client_notifications.find(filter_query).sort("created_at", -1).limit(limit)
    notifications = await notifications_cursor.to_list(length=None)
    
    # Handle missing is_read field for backward compatibility
    parsed_notifications = []
    for notification in notifications:
        parsed_notif = parse_from_mongo(notification)
        if 'is_read' not in parsed_notif:
            parsed_notif['is_read'] = False
        
        # Fix datetime format issue - remove trailing 'Z' if present
        if 'created_at' in parsed_notif and isinstance(parsed_notif['created_at'], str):
            parsed_notif['created_at'] = parsed_notif['created_at'].rstrip('Z')
        
        try:
            parsed_notifications.append(ClientNotificationResponse(**parsed_notif))
        except Exception as e:
            logger.error(f"Failed to parse notification {parsed_notif.get('id')}: {e}")
            continue
    
    return parsed_notifications

@api_router.get("/client/notifications/unread-count", response_model=dict)
async def get_client_unread_notification_count(current_user: User = Depends(get_current_user)):
    """Get count of unread notifications for client"""
    count = await db.client_notifications.count_documents({"user_id": current_user.id, "is_read": False})
    return {"count": count}

@api_router.put("/client/notifications/{notification_id}/read", response_model=dict)
async def mark_client_notification_read(
    notification_id: str,
    current_user: User = Depends(get_current_user)
):
    """Mark client notification as read"""
    result = await db.client_notifications.update_one(
        {"id": notification_id, "user_id": current_user.id},
        {"$set": {"is_read": True}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Notification not found")
    
    return {"message": "Notification marked as read"}

@api_router.put("/client/notifications/mark-all-read", response_model=dict)
async def mark_all_client_notifications_read(current_user: User = Depends(get_current_user)):
    """Mark all client notifications as read"""
    await db.client_notifications.update_many(
        {"user_id": current_user.id},
        {"$set": {"is_read": True}}
    )
    return {"message": "All notifications marked as read"}

# ============================================================================
# ADMIN ACTIONS ENDPOINTS (For Super Admin Approval System)
# ============================================================================

@api_router.post("/admin/client-actions/topup-wallet", response_model=dict)
async def create_admin_topup_wallet_action(
    client_id: str = Form(...),
    wallet_type: str = Form(...),
    amount: float = Form(...),
    notes: str = Form(None),
    payment_proof: UploadFile = File(...),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Admin creates wallet top-up request for client (requires super admin approval)"""
    try:
        # Validate client exists
        client = await db.users.find_one({"id": client_id})
        if not client:
            raise HTTPException(status_code=404, detail="Client not found")
        
        # Upload payment proof to GCS
        proof_data = await upload_to_gcs(payment_proof, folder="admin_actions/payment_proofs")
        
        # Create admin action record
        action_record = AdminActionRecord(
            action_type="topup_wallet",
            client_id=client_id,
            client_username=client.get("username"),
            admin_id=current_admin.id,
            admin_username=current_admin.username,
            wallet_type=wallet_type,
            amount=amount,
            currency="IDR" if "idr" in wallet_type.lower() else "USD",
            notes=notes,
            payment_proof_gcs=proof_data["gcs_path"],
            status="pending"
        )
        
        # Save to database
        await db.admin_actions.insert_one(action_record.dict())
        
        # Notify super admins
        super_admins = await db.admin_users.find({"is_super_admin": True}).to_list(length=None)
        for super_admin in super_admins:
            notification = {
                "id": str(uuid.uuid4()),
                "user_id": super_admin["id"],
                "title": "🔔 New Admin Action Request",
                "message": f"Admin {current_admin.username} requested wallet top-up for client {client.get('username')} (Amount: {amount})",
                "type": "admin_action_pending",
                "data": {"action_id": action_record.id, "action_type": "topup_wallet"},
                "created_at": datetime.now(timezone.utc).isoformat(),
                "is_read": False
            }
            await db.notifications.insert_one(notification)
        
        logger.info(f"✅ Admin action created: topup_wallet by {current_admin.username} for client {client_id}")
        
        return {
            "message": "Top-up request created successfully. Waiting for super admin approval.",
            "action_id": action_record.id
        }
        
    except Exception as e:
        logger.error(f"❌ Failed to create admin topup action: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/admin/client-actions/deduct-wallet", response_model=dict)
async def create_admin_deduct_wallet_action(
    client_id: str = Form(...),
    wallet_type: str = Form(...),
    amount: float = Form(...),
    reason: str = Form(...),
    proof_file: UploadFile = File(...),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Admin creates wallet deduction request for client (requires super admin approval)"""
    try:
        # Validate client exists
        client = await db.users.find_one({"id": client_id})
        if not client:
            raise HTTPException(status_code=404, detail="Client not found")
        
        # Upload proof file to GCS
        proof_data = await upload_to_gcs(proof_file, folder="wallet_deductions/proofs")
        
        # Create deduction record
        deduction_record = WalletDeductionRecord(
            client_id=client_id,
            client_name=client.get("name", client.get("username")),
            admin_id=current_admin.id,
            admin_username=current_admin.username,
            wallet_type=wallet_type,
            amount=amount,
            reason=reason,
            proof_file_id=proof_data["gcs_path"],  # Use gcs_path as file_id
            proof_file_url=proof_data["gcs_path"],
            status="pending"
        )
        
        # Save to database
        await db.wallet_deduction_requests.insert_one(deduction_record.dict())
        
        # Notify super admins
        super_admins = await db.admin_users.find({"is_super_admin": True}).to_list(length=None)
        currency = "IDR" if "idr" in wallet_type.lower() else "USD"
        currency_symbol = "Rp " if currency == "IDR" else "$"
        formatted_amount = f"{currency_symbol}{amount:,.0f}" if currency == "IDR" else f"{currency_symbol}{amount:,.2f}"
        
        for super_admin in super_admins:
            notification = {
                "id": str(uuid.uuid4()),
                "user_id": super_admin["id"],
                "title": "⚠️ Wallet Deduction Request",
                "message": f"Admin {current_admin.username} requests to deduct {formatted_amount} from {client.get('name', client.get('username'))}'s {wallet_type.upper()} wallet",
                "type": "wallet_deduction_pending",
                "data": {"deduction_id": deduction_record.id, "client_id": client_id},
                "created_at": datetime.now(timezone.utc).isoformat(),
                "is_read": False
            }
            await db.notifications.insert_one(notification)
        
        logger.info(f"✅ Wallet deduction request created by {current_admin.username} for client {client_id}: {formatted_amount} from {wallet_type}")
        
        return {
            "message": "Wallet deduction request created successfully. Waiting for super admin approval.",
            "deduction_id": deduction_record.id
        }
        
    except Exception as e:
        logger.error(f"❌ Failed to create wallet deduction request: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/admin/client-actions/withdraw-account", response_model=dict)
async def create_admin_withdraw_account_action(
    client_id: str = Form(...),
    account_id: str = Form(...),
    amount: float = Form(...),
    currency: str = Form(...),
    notes: str = Form(None),
    real_balance_proof: UploadFile = File(...),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Admin creates account withdrawal request for client (requires super admin approval)"""
    try:
        # Validate client and account
        client = await db.users.find_one({"id": client_id})
        if not client:
            raise HTTPException(status_code=404, detail="Client not found")
        
        account = await db.ad_accounts.find_one({"id": account_id, "user_id": client_id})
        if not account:
            raise HTTPException(status_code=404, detail="Account not found")
        
        # Upload proof to GCS
        proof_data = await upload_to_gcs(real_balance_proof, folder="admin_actions/balance_proofs")
        
        # Create admin action record
        action_record = AdminActionRecord(
            action_type="withdraw_account",
            client_id=client_id,
            client_username=client.get("username"),
            admin_id=current_admin.id,
            admin_username=current_admin.username,
            account_id=account_id,
            amount=amount,
            currency=currency,
            notes=notes,
            real_balance_proof_gcs=proof_data["gcs_path"],
            status="pending"
        )
        
        # Save to database
        await db.admin_actions.insert_one(action_record.dict())
        
        # Notify super admins
        super_admins = await db.admin_users.find({"is_super_admin": True}).to_list(length=None)
        for super_admin in super_admins:
            notification = {
                "id": str(uuid.uuid4()),
                "user_id": super_admin["id"],
                "title": "🔔 New Admin Action Request",
                "message": f"Admin {current_admin.username} requested account withdrawal for client {client.get('username')} (Amount: {currency} {amount})",
                "type": "admin_action_pending",
                "data": {"action_id": action_record.id, "action_type": "withdraw_account"},
                "created_at": datetime.now(timezone.utc).isoformat(),
                "is_read": False
            }
            await db.notifications.insert_one(notification)
        
        logger.info(f"✅ Admin action created: withdraw_account by {current_admin.username} for client {client_id}")
        
        return {
            "message": "Withdrawal request created successfully. Waiting for super admin approval.",
            "action_id": action_record.id
        }
        
    except Exception as e:
        logger.error(f"❌ Failed to create admin withdraw action: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/admin/client-actions/transfer-wallet-to-account", response_model=dict)
async def create_admin_transfer_wallet_to_account_action(
    client_id: str = Form(...),
    from_wallet: str = Form(...),
    to_account_id: str = Form(...),
    amount: float = Form(...),
    currency: str = Form(...),
    notes: str = Form(None),
    spending_limit_proof: UploadFile = File(...),
    budget_aspire_proof: UploadFile = File(...),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Admin creates wallet-to-account transfer request for client (requires super admin approval)"""
    try:
        # Validate client and account
        client = await db.users.find_one({"id": client_id})
        if not client:
            raise HTTPException(status_code=404, detail="Client not found")
        
        account = await db.ad_accounts.find_one({"id": to_account_id, "user_id": client_id})
        if not account:
            raise HTTPException(status_code=404, detail="Account not found")
        
        # Upload proofs to GCS
        spending_proof_data = await upload_to_gcs(spending_limit_proof, folder="admin_actions/spending_limits")
        budget_proof_data = await upload_to_gcs(budget_aspire_proof, folder="admin_actions/budget_aspire")
        
        # Create admin action record
        action_record = AdminActionRecord(
            action_type="transfer_wallet_to_account",
            client_id=client_id,
            client_username=client.get("username"),
            admin_id=current_admin.id,
            admin_username=current_admin.username,
            from_wallet=from_wallet,
            to_account_id=to_account_id,
            amount=amount,
            currency=currency,
            notes=notes,
            spending_limit_proof_gcs=spending_proof_data["gcs_path"],
            budget_aspire_proof_gcs=budget_proof_data["gcs_path"],
            status="pending"
        )
        
        # Save to database
        await db.admin_actions.insert_one(action_record.dict())
        
        # Notify super admins
        super_admins = await db.admin_users.find({"is_super_admin": True}).to_list(length=None)
        for super_admin in super_admins:
            notification = {
                "id": str(uuid.uuid4()),
                "user_id": super_admin["id"],
                "title": "🔔 New Admin Action Request",
                "message": f"Admin {current_admin.username} requested wallet transfer for client {client.get('username')} (Amount: {currency} {amount})",
                "type": "admin_action_pending",
                "data": {"action_id": action_record.id, "action_type": "transfer_wallet_to_account"},
                "created_at": datetime.now(timezone.utc).isoformat(),
                "is_read": False
            }
            await db.notifications.insert_one(notification)
        
        logger.info(f"✅ Admin action created: transfer_wallet_to_account by {current_admin.username} for client {client_id}")
        
        return {
            "message": "Transfer request created successfully. Waiting for super admin approval.",
            "action_id": action_record.id
        }
        
    except Exception as e:
        logger.error(f"❌ Failed to create admin transfer action: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Super Admin Endpoints
@api_router.get("/super-admin/pending-actions", response_model=List[dict])
async def get_pending_admin_actions(current_super_admin: AdminUser = Depends(get_current_super_admin)):
    """Get all pending admin actions for super admin approval"""
    try:
        pending_actions = await db.admin_actions.find({"status": "pending"}).sort("created_at", -1).to_list(length=None)
        
        # Parse and enrich data
        result = []
        for action in pending_actions:
            action_data = parse_from_mongo(action)
            
            # Add client info
            client = await db.users.find_one({"id": action_data["client_id"]})
            if client:
                action_data["client_name"] = client.get("name") or client.get("display_name")
            
            # Add account info if applicable
            if action_data.get("account_id"):
                account = await db.ad_accounts.find_one({"id": action_data["account_id"]})
                if account:
                    action_data["account_name"] = account.get("account_name")
                    action_data["platform"] = account.get("platform")
            
            if action_data.get("to_account_id"):
                account = await db.ad_accounts.find_one({"id": action_data["to_account_id"]})
                if account:
                    action_data["account_name"] = account.get("account_name")
                    action_data["platform"] = account.get("platform")
            
            result.append(action_data)
        
        return result
        
    except Exception as e:
        logger.error(f"❌ Failed to get pending actions: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/super-admin/actions-history", response_model=dict)
async def get_admin_actions_history(
    page: int = 1,
    limit: int = 10,
    status: str = None,
    current_super_admin: AdminUser = Depends(get_current_super_admin)
):
    """Get completed/rejected admin actions history with pagination"""
    try:
        # Build query - exclude pending status
        query = {"status": {"$in": ["approved", "rejected"]}}
        if status and status in ["approved", "rejected"]:
            query = {"status": status}
        
        # Calculate pagination
        skip = (page - 1) * limit
        
        # Get total count from both collections
        total_actions = await db.admin_actions.count_documents(query)
        total_history = await db.admin_actions_history.count_documents(query)
        total = total_actions + total_history
        
        # Get paginated actions from both collections
        actions_from_admin_actions = await db.admin_actions.find(query).sort("processed_at", -1).to_list(length=None)
        actions_from_history = await db.admin_actions_history.find(query).sort("processed_at", -1).to_list(length=None)
        
        # Combine and sort
        all_actions = actions_from_admin_actions + actions_from_history
        all_actions.sort(key=lambda x: x.get("processed_at", ""), reverse=True)
        
        # Apply pagination
        actions = all_actions[skip:skip+limit]
        
        # Parse and enrich data
        result = []
        for action in actions:
            action_data = parse_from_mongo(action)
            
            # Add client info
            client = await db.users.find_one({"id": action_data["client_id"]})
            if client:
                action_data["client_name"] = client.get("name") or client.get("display_name")
                action_data["client_username"] = client.get("username")
            
            # Add admin who processed
            if action_data.get("super_admin_id"):
                admin = await db.admins.find_one({"id": action_data["super_admin_id"]})
                if admin:
                    action_data["processed_by_name"] = admin.get("name") or admin.get("username")
                else:
                    # Fallback to super_admin_username if admin not found
                    action_data["processed_by_name"] = action_data.get("super_admin_username", "Unknown")
            
            # Add account info if applicable
            if action_data.get("account_id"):
                account = await db.ad_accounts.find_one({"id": action_data["account_id"]})
                if account:
                    action_data["account_name"] = account.get("account_name")
                    action_data["platform"] = account.get("platform")
            
            if action_data.get("to_account_id"):
                account = await db.ad_accounts.find_one({"id": action_data["to_account_id"]})
                if account:
                    action_data["to_account_name"] = account.get("account_name")
                    action_data["to_platform"] = account.get("platform")
            
            result.append(action_data)
        
        return {
            "actions": result,
            "total": total,
            "page": page,
            "limit": limit,
            "total_pages": (total + limit - 1) // limit
        }
        
    except Exception as e:
        logger.error(f"❌ Failed to get actions history: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.put("/super-admin/actions/{action_id}/approve", response_model=dict)
async def approve_admin_action(
    action_id: str,
    approval_data: AdminActionApproval,
    current_super_admin: AdminUser = Depends(get_current_super_admin)
):
    """Super admin approves admin action and processes transaction automatically"""
    try:
        # Get action record
        action = await db.admin_actions.find_one({"id": action_id})
        if not action:
            raise HTTPException(status_code=404, detail="Action not found")
        
        if action["status"] != "pending":
            raise HTTPException(status_code=400, detail="Action already processed")
        
        action_type = action["action_type"]
        client_id = action["client_id"]
        amount = action["amount"]
        currency = action["currency"]
        
        # Process based on action type
        if approval_data.action == "approve":
            if action_type == "topup_wallet":
                # Top up wallet
                wallet_field = action["wallet_type"]
                wallet_key = f"{wallet_field}"  # e.g., "main_idr" maps to "main_wallet_idr"
                
                # Map wallet_type to actual field name
                field_mapping = {
                    "main_idr": "main_wallet_idr",
                    "main_usd": "main_wallet_usd",
                    "withdrawal_idr": "withdrawal_wallet_idr",
                    "withdrawal_usd": "withdrawal_wallet_usd"
                }
                
                actual_field = field_mapping.get(wallet_field)
                if not actual_field:
                    raise HTTPException(status_code=400, detail="Invalid wallet type")
                
                # Update wallet balance
                await db.users.update_one(
                    {"id": client_id},
                    {"$inc": {actual_field: amount}}
                )
                
                # Create transaction record
                transaction = {
                    "id": str(uuid.uuid4()),
                    "user_id": client_id,
                    "type": "admin_topup",
                    "amount": amount,
                    "currency": currency,
                    "wallet_type": wallet_field,
                    "description": f"Admin top-up by {action['admin_username']} (Approved by {current_super_admin.username})",
                    "status": "completed",
                    "admin_id": action["admin_id"],
                    "super_admin_id": current_super_admin.id,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.transactions.insert_one(transaction)
                
                # Notify client
                client_notification = {
                    "id": str(uuid.uuid4()),
                    "user_id": client_id,
                    "title": "✅ Wallet Top-Up Completed",
                    "message": f"Admin has topped up your {wallet_field.replace('_', ' ').title()} wallet with {currency} {amount:,.2f}",
                    "type": "admin_topup_completed",
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "is_read": False
                }
                await db.client_notifications.insert_one(client_notification)
                
                # Send email notification for wallet top-up completion
                try:
                    user = await db.users.find_one({"id": client_id})
                    if user and user.get("email"):
                        send_client_super_admin_completion_email(
                            client_email=user["email"],
                            client_name=user.get("full_name") or user.get("username"),
                            action_type='wallet_topup',
                            amount=amount,
                            currency=currency,
                            details=f"Wallet {wallet_field.replace('_', ' ').title()} telah ditambahkan oleh admin"
                        )
                        logger.info(f"📧 Super admin wallet top-up completion email sent to {user['email']}")
                except Exception as e:
                    logger.error(f"Failed to send super admin completion email: {e}")
            
            elif action_type == "withdraw_account":
                # Withdraw from account to withdrawal wallet
                account_id = action["account_id"]
                
                # Deduct from account balance
                await db.ad_accounts.update_one(
                    {"id": account_id},
                    {"$inc": {"balance": -amount}}
                )
                
                # Add to withdrawal wallet
                wallet_field = "withdrawal_wallet_idr" if currency == "IDR" else "withdrawal_wallet_usd"
                await db.users.update_one(
                    {"id": client_id},
                    {"$inc": {wallet_field: amount}}
                )
                
                # Create transaction record
                transaction = {
                    "id": str(uuid.uuid4()),
                    "user_id": client_id,
                    "type": "admin_withdraw",
                    "amount": amount,
                    "currency": currency,
                    "account_id": account_id,
                    "description": f"Admin withdrawal by {action['admin_username']} (Approved by {current_super_admin.username})",
                    "status": "completed",
                    "admin_id": action["admin_id"],
                    "super_admin_id": current_super_admin.id,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.transactions.insert_one(transaction)
                
                # Notify client
                account = await db.ad_accounts.find_one({"id": account_id})
                client_notification = {
                    "id": str(uuid.uuid4()),
                    "user_id": client_id,
                    "title": "✅ Account Withdrawal Completed",
                    "message": f"Admin has withdrawn {currency} {amount:,.2f} from account {account.get('account_name', account_id)} to your withdrawal wallet",
                    "type": "admin_withdraw_completed",
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "is_read": False
                }
                await db.client_notifications.insert_one(client_notification)
                
                # Send email notification for withdrawal completion
                try:
                    user = await db.users.find_one({"id": client_id})
                    if user and user.get("email"):
                        send_client_super_admin_completion_email(
                            client_email=user["email"],
                            client_name=user.get("full_name") or user.get("username"),
                            action_type='withdrawal',
                            amount=amount,
                            currency=currency,
                            details=f"Penarikan dari akun {account.get('account_name', 'Unknown')} ke withdrawal wallet telah selesai"
                        )
                        logger.info(f"📧 Super admin withdrawal completion email sent to {user['email']}")
                except Exception as e:
                    logger.error(f"Failed to send super admin completion email: {e}")
            
            elif action_type == "transfer_wallet_to_account":
                # Transfer from wallet to account
                from_wallet = action["from_wallet"]
                to_account_id = action["to_account_id"]
                
                # Deduct from wallet
                field_mapping = {
                    "main_idr": "main_wallet_idr",
                    "main_usd": "main_wallet_usd",
                    "withdrawal_idr": "withdrawal_wallet_idr",
                    "withdrawal_usd": "withdrawal_wallet_usd"
                }
                
                wallet_field = field_mapping.get(from_wallet)
                if not wallet_field:
                    raise HTTPException(status_code=400, detail="Invalid wallet type")
                
                await db.users.update_one(
                    {"id": client_id},
                    {"$inc": {wallet_field: -amount}}
                )
                
                # Add to account balance
                await db.ad_accounts.update_one(
                    {"id": to_account_id},
                    {"$inc": {"balance": amount}}
                )
                
                # Create transaction record
                transaction = {
                    "id": str(uuid.uuid4()),
                    "user_id": client_id,
                    "type": "admin_transfer",
                    "amount": amount,
                    "currency": currency,
                    "from_wallet": from_wallet,
                    "to_account_id": to_account_id,
                    "description": f"Admin transfer by {action['admin_username']} (Approved by {current_super_admin.username})",
                    "status": "completed",
                    "admin_id": action["admin_id"],
                    "super_admin_id": current_super_admin.id,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.transactions.insert_one(transaction)
                
                # Notify client
                account = await db.ad_accounts.find_one({"id": to_account_id})
                client_notification = {
                    "id": str(uuid.uuid4()),
                    "user_id": client_id,
                    "title": "✅ Wallet Transfer Completed",
                    "message": f"Admin has transferred {currency} {amount:,.2f} from your {from_wallet.replace('_', ' ').title()} to account {account.get('account_name', to_account_id)}",
                    "type": "admin_transfer_completed",
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "is_read": False
                }
                await db.client_notifications.insert_one(client_notification)
                
                # Send email notification for wallet transfer completion
                try:
                    user = await db.users.find_one({"id": client_id})
                    if user and user.get("email"):
                        send_client_super_admin_completion_email(
                            client_email=user["email"],
                            client_name=user.get("full_name") or user.get("username"),
                            action_type='wallet_transfer',
                            amount=amount,
                            currency=currency,
                            details=f"Transfer dari {from_wallet.replace('_', ' ').title()} ke akun {account.get('account_name', 'Unknown')} telah selesai"
                        )
                        logger.info(f"📧 Super admin wallet transfer completion email sent to {user['email']}")
                except Exception as e:
                    logger.error(f"Failed to send super admin completion email: {e}")
            
            # Update action status
            await db.admin_actions.update_one(
                {"id": action_id},
                {"$set": {
                    "status": "approved",
                    "super_admin_id": current_super_admin.id,
                    "super_admin_username": current_super_admin.username,
                    "approval_notes": approval_data.notes,
                    "processed_at": datetime.now(timezone.utc).isoformat()
                }}
            )
            
            # Notify requesting admin
            admin_notification = {
                "id": str(uuid.uuid4()),
                "user_id": action["admin_id"],
                "title": "✅ Action Approved",
                "message": f"Your {action_type.replace('_', ' ')} request has been approved by {current_super_admin.username}",
                "type": "admin_action_approved",
                "data": {"action_id": action_id},
                "created_at": datetime.now(timezone.utc).isoformat(),
                "is_read": False
            }
            await db.notifications.insert_one(admin_notification)
            
            logger.info(f"✅ Admin action approved: {action_id} by super admin {current_super_admin.username}")
            
            return {"message": "Action approved and processed successfully"}
        
        else:  # reject
            # Update action status
            await db.admin_actions.update_one(
                {"id": action_id},
                {"$set": {
                    "status": "rejected",
                    "super_admin_id": current_super_admin.id,
                    "super_admin_username": current_super_admin.username,
                    "approval_notes": approval_data.notes,
                    "processed_at": datetime.now(timezone.utc).isoformat()
                }}
            )
            
            # Notify requesting admin
            admin_notification = {
                "id": str(uuid.uuid4()),
                "user_id": action["admin_id"],
                "title": "❌ Action Rejected",
                "message": f"Your {action_type.replace('_', ' ')} request has been rejected by {current_super_admin.username}. Reason: {approval_data.notes or 'No reason provided'}",
                "type": "admin_action_rejected",
                "data": {"action_id": action_id},
                "created_at": datetime.now(timezone.utc).isoformat(),
                "is_read": False
            }
            await db.notifications.insert_one(admin_notification)
            
            logger.info(f"✅ Admin action rejected: {action_id} by super admin {current_super_admin.username}")
            
            return {"message": "Action rejected successfully"}
        
    except Exception as e:
        logger.error(f"❌ Failed to process admin action approval: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Wallet Deduction Endpoints (Super Admin)
@api_router.get("/super-admin/pending-wallet-deductions", response_model=List[dict])
async def get_pending_wallet_deductions(current_super_admin: AdminUser = Depends(get_current_super_admin)):
    """Get all pending wallet deduction requests"""
    try:
        pending = await db.wallet_deduction_requests.find({"status": "pending"}).sort("created_at", -1).to_list(length=None)
        
        result = []
        for deduction in pending:
            deduction_data = parse_from_mongo(deduction)
            result.append(deduction_data)
        
        return result
    except Exception as e:
        logger.error(f"❌ Failed to get pending wallet deductions: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.put("/super-admin/wallet-deductions/{deduction_id}/approve", response_model=dict)
async def approve_wallet_deduction(
    deduction_id: str,
    notes: str = None,
    current_super_admin: AdminUser = Depends(get_current_super_admin)
):
    """Super admin approves wallet deduction"""
    try:
        # Get deduction record
        deduction = await db.wallet_deduction_requests.find_one({"id": deduction_id})
        if not deduction:
            raise HTTPException(status_code=404, detail="Deduction request not found")
        
        if deduction["status"] != "pending":
            raise HTTPException(status_code=400, detail="Deduction already processed")
        
        client_id = deduction["client_id"]
        wallet_type = deduction["wallet_type"]
        amount = deduction["amount"]
        
        # Map wallet_type to actual field name
        field_mapping = {
            "main_idr": "main_wallet_idr",
            "main_usd": "main_wallet_usd",
            "withdrawal_idr": "withdrawal_wallet_idr",
            "withdrawal_usd": "withdrawal_wallet_usd"
        }
        
        actual_field = field_mapping.get(wallet_type)
        if not actual_field:
            raise HTTPException(status_code=400, detail="Invalid wallet type")
        
        currency = "IDR" if "idr" in wallet_type.lower() else "USD"
        
        # Deduct from wallet (no validation - can go negative as per requirements)
        await db.users.update_one(
            {"id": client_id},
            {"$inc": {actual_field: -amount}}
        )
        
        # Create transaction record
        transaction = {
            "id": str(uuid.uuid4()),
            "user_id": client_id,
            "type": "admin_deduction",
            "amount": -amount,  # Negative for deduction
            "currency": currency,
            "wallet_type": wallet_type,
            "description": f"Wallet deduction by admin {deduction['admin_username']} (Approved by {current_super_admin.username}). Reason: {deduction['reason']}",
            "status": "completed",
            "admin_id": deduction["admin_id"],
            "super_admin_id": current_super_admin.id,
            "reference_id": deduction_id,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.transactions.insert_one(transaction)
        
        # Create wallet statement entry
        wallet_statement = {
            "id": str(uuid.uuid4()),
            "user_id": client_id,
            "wallet_type": wallet_type,
            "type": "deduction",
            "amount": -amount,  # Negative for deduction
            "currency": currency,
            "description": f"Pengurangan saldo oleh admin. Alasan: {deduction['reason']}",
            "reference_type": "admin_deduction",
            "reference_id": deduction_id,
            "admin_id": deduction["admin_id"],
            "super_admin_id": current_super_admin.id,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.wallet_statements.insert_one(wallet_statement)
        
        # Update deduction status
        await db.wallet_deduction_requests.update_one(
            {"id": deduction_id},
            {"$set": {
                "status": "approved",
                "super_admin_id": current_super_admin.id,
                "super_admin_username": current_super_admin.username,
                "approval_notes": notes,
                "processed_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        
        # Add to actions history
        action_history = {
            "id": str(uuid.uuid4()),
            "action_type": "wallet_deduction",
            "client_id": client_id,
            "client_name": deduction["client_name"],
            "admin_id": deduction["admin_id"],
            "admin_username": deduction["admin_username"],
            "super_admin_id": current_super_admin.id,
            "super_admin_username": current_super_admin.username,
            "status": "approved",
            "amount": amount,  # Add to top level for consistent display
            "currency": currency,  # Add to top level
            "details": {
                "wallet_type": wallet_type,
                "amount": amount,
                "currency": currency,
                "reason": deduction["reason"],
                "proof_file_url": deduction.get("proof_file_url")
            },
            "approval_notes": notes,
            "created_at": deduction["created_at"],
            "processed_at": datetime.now(timezone.utc).isoformat()
        }
        await db.admin_actions_history.insert_one(action_history)
        
        # Notify client
        currency_symbol = "Rp " if currency == "IDR" else "$"
        formatted_amount = f"{currency_symbol}{amount:,.0f}" if currency == "IDR" else f"{currency_symbol}{amount:,.2f}"
        
        client_notification = {
            "id": str(uuid.uuid4()),
            "user_id": client_id,
            "title": "⚠️ Pengurangan Saldo Wallet",
            "message": f"Saldo {wallet_type.replace('_', ' ').upper()} Anda telah dikurangi {formatted_amount}. Alasan: {deduction['reason']}",
            "type": "wallet_deduction",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "is_read": False
        }
        await db.client_notifications.insert_one(client_notification)
        
        # Notify requesting admin
        admin_notification = {
            "id": str(uuid.uuid4()),
            "user_id": deduction["admin_id"],
            "title": "✅ Wallet Deduction Approved",
            "message": f"Your wallet deduction request for {deduction['client_name']} has been approved ({formatted_amount})",
            "type": "wallet_deduction_approved",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "is_read": False
        }
        await db.notifications.insert_one(admin_notification)
        
        logger.info(f"✅ Wallet deduction approved: {deduction_id} by super admin {current_super_admin.username}")
        
        return {"message": "Wallet deduction approved and processed successfully"}
        
    except Exception as e:
        logger.error(f"❌ Failed to approve wallet deduction: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.put("/super-admin/wallet-deductions/{deduction_id}/reject", response_model=dict)
async def reject_wallet_deduction(
    deduction_id: str,
    notes: str = None,
    current_super_admin: AdminUser = Depends(get_current_super_admin)
):
    """Super admin rejects wallet deduction"""
    try:
        # Get deduction record
        deduction = await db.wallet_deduction_requests.find_one({"id": deduction_id})
        if not deduction:
            raise HTTPException(status_code=404, detail="Deduction request not found")
        
        if deduction["status"] != "pending":
            raise HTTPException(status_code=400, detail="Deduction already processed")
        
        # Update deduction status
        await db.wallet_deduction_requests.update_one(
            {"id": deduction_id},
            {"$set": {
                "status": "rejected",
                "super_admin_id": current_super_admin.id,
                "super_admin_username": current_super_admin.username,
                "approval_notes": notes,
                "processed_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        
        # Add to actions history
        currency = "IDR" if "idr" in deduction["wallet_type"].lower() else "USD"
        action_history = {
            "id": str(uuid.uuid4()),
            "action_type": "wallet_deduction",
            "client_id": deduction["client_id"],
            "client_name": deduction["client_name"],
            "admin_id": deduction["admin_id"],
            "admin_username": deduction["admin_username"],
            "super_admin_id": current_super_admin.id,
            "super_admin_username": current_super_admin.username,
            "status": "rejected",
            "amount": deduction["amount"],  # Add to top level
            "currency": currency,  # Add to top level
            "details": {
                "wallet_type": deduction["wallet_type"],
                "amount": deduction["amount"],
                "currency": currency,
                "reason": deduction["reason"],
                "proof_file_url": deduction.get("proof_file_url")
            },
            "approval_notes": notes,
            "created_at": deduction["created_at"],
            "processed_at": datetime.now(timezone.utc).isoformat()
        }
        await db.admin_actions_history.insert_one(action_history)
        
        # Notify requesting admin
        currency_symbol = "Rp " if currency == "IDR" else "$"
        formatted_amount = f"{currency_symbol}{deduction['amount']:,.0f}" if currency == "IDR" else f"{currency_symbol}{deduction['amount']:,.2f}"
        
        admin_notification = {
            "id": str(uuid.uuid4()),
            "user_id": deduction["admin_id"],
            "title": "❌ Wallet Deduction Rejected",
            "message": f"Your wallet deduction request for {deduction['client_name']} has been rejected ({formatted_amount}). Reason: {notes or 'No reason provided'}",
            "type": "wallet_deduction_rejected",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "is_read": False
        }
        await db.notifications.insert_one(admin_notification)
        
        logger.info(f"✅ Wallet deduction rejected: {deduction_id} by super admin {current_super_admin.username}")
        
        return {"message": "Wallet deduction rejected successfully"}
        
    except Exception as e:
        logger.error(f"❌ Failed to reject wallet deduction: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ==========================
# Admin Ad Copy Management
# ==========================

@api_router.get("/admin/ad-copies")
async def admin_get_all_ad_copies(
    current_admin: AdminUser = Depends(get_current_admin),
    search: Optional[str] = None,
    goal: Optional[str] = None,
    user_id: Optional[str] = None
):
    """
    Admin endpoint: Get all saved ad copies from all clients
    Supports search by label/product_name/username, filter by goal and user_id
    """
    try:
        # Build query
        query = {}
        
        # Add user filter if specified
        if user_id:
            query['user_id'] = user_id
        
        # Add search filter
        if search:
            # Get matching users
            users_cursor = db.users.find(
                {'$or': [
                    {'username': {'$regex': search, '$options': 'i'}},
                    {'email': {'$regex': search, '$options': 'i'}},
                    {'name': {'$regex': search, '$options': 'i'}}
                ]},
                {'id': 1}
            )
            matching_user_ids = [user['id'] for user in await users_cursor.to_list(length=None)]
            
            # Search in ad copies or matching users
            query['$or'] = [
                {'label': {'$regex': search, '$options': 'i'}},
                {'product_name': {'$regex': search, '$options': 'i'}},
                {'user_id': {'$in': matching_user_ids}}
            ]
        
        # Add goal filter
        if goal and goal != 'all':
            query['goal'] = goal
        
        # Fetch ad copies, sorted by most recent first
        ad_copies_cursor = db.saved_ad_copies.find(
            query,
            {'_id': 0}
        ).sort('created_at', -1)
        
        ad_copies = await ad_copies_cursor.to_list(length=None)
        
        # Enrich with user information
        enriched_ad_copies = []
        for ad_copy in ad_copies:
            # Get user info
            user = await db.users.find_one({'id': ad_copy['user_id']})
            if user:
                ad_copy['user_info'] = {
                    'username': user.get('username'),
                    'email': user.get('email'),
                    'name': user.get('name', user.get('username'))
                }
            else:
                ad_copy['user_info'] = {
                    'username': 'Unknown',
                    'email': 'N/A',
                    'name': 'Unknown User'
                }
            enriched_ad_copies.append(ad_copy)
        
        return {
            'success': True,
            'ad_copies': enriched_ad_copies,
            'total': len(enriched_ad_copies)
        }
    
    except Exception as e:
        logger.error(f"Error fetching ad copies for admin: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch ad copies: {str(e)}")

# Include the router in the main app
app.include_router(api_router)

@app.get("/")
async def root():
    return {"message": "Ad Manager Pro API is running"}

@app.get("/api/test-gcs")
async def test_gcs():
    """Test GCS connection and configuration"""
    try:
        from gcs_storage import get_gcs_storage
        import os
        
        # Get environment variables
        env_vars = {
            "GCS_BUCKET_NAME": os.environ.get("GCS_BUCKET_NAME"),
            "GCS_PROJECT_ID": os.environ.get("GCS_PROJECT_ID"),
            "GOOGLE_APPLICATION_CREDENTIALS": os.environ.get("GOOGLE_APPLICATION_CREDENTIALS"),
            "credentials_file_exists": os.path.exists("/app/backend/gcs-service-account.json")
        }
        
        # Try to initialize GCS
        gcs = get_gcs_storage()
        
        return {
            "status": "success",
            "message": "GCS initialized successfully",
            "config": env_vars,
            "bucket": gcs.bucket_name,
            "project": gcs.project_id
        }
    except Exception as e:
        return {
            "status": "error",
            "message": str(e),
            "config": env_vars if 'env_vars' in locals() else None
        }

# Mount static files for profile pictures
import os
uploads_dir = Path("/app/uploads")
uploads_dir.mkdir(exist_ok=True)
app.mount("/uploads", StaticFiles(directory="/app/uploads"), name="uploads")



# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Admin endpoint to clear all data except user logins (for testing purposes)
@app.post("/api/admin/clear-database")
async def clear_database(current_admin: AdminUser = Depends(require_super_admin)):
    """
    Clear all data from database except user login data (admin_users and users collections)
    Only super admin can execute this
    """
    try:
        collections_to_clear = [
            'ad_account_requests',
            'client_notifications',
            'admin_notifications',
            'transfer_requests',
            'notifications',
            'withdraw_requests',
            'ad_accounts',
            'share_requests',
            'account_groups',
            'topup_requests',
            'currency_exchanges',
            'groups',
            'system_settings',
            'transactions',
            'wallet_transfers',
            'payment_proofs',
            'wallet_topup_requests'
        ]
        
        deleted_count = {}
        
        for col_name in collections_to_clear:
            result = await db[col_name].delete_many({})
            deleted_count[col_name] = result.deleted_count
        
        # Get remaining counts
        remaining = {
            'users': await db.users.count_documents({}),
            'admin_users': await db.admin_users.count_documents({})
        }
        
        return {
            "success": True,
            "message": "Database cleared successfully",
            "deleted": deleted_count,
            "preserved": remaining
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to clear database: {str(e)}")

# Get all clients for deletion selection
@app.get("/api/admin/clients-list")
async def get_clients_list(current_admin: AdminUser = Depends(require_super_admin)):
    """
    Get list of all clients with basic info for deletion selection
    Only super admin can access this
    """
    try:
        clients = await db.users.find({}, {
            'id': 1,
            'username': 1,
            'email': 1,
            'name': 1,
            'display_name': 1,
            'created_at': 1,
            'wallet_balance_idr': 1,
            'wallet_balance_usd': 1
        }).to_list(length=None)
        
        # Parse from mongo
        for client in clients:
            if '_id' in client:
                del client['_id']
        
        return clients
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to fetch clients: {str(e)}")

# Database Cleaner PIN Verification
class PINVerification(BaseModel):
    pin: str

class ChangePINRequest(BaseModel):
    old_pin: str
    new_pin: str
    confirm_pin: str

# Helper function to get current PIN (from DB or .env)
async def get_current_pin():
    """Get current PIN from database or fallback to .env"""
    try:
        # Try to get from database first
        settings = await db.admin_settings.find_one({"setting_key": "database_cleaner_pin"})
        if settings and settings.get("setting_value"):
            return settings["setting_value"]
    except:
        pass
    
    # Fallback to .env
    return os.environ.get('DATABASE_CLEANER_PIN', '123456')

# Helper function to update PIN in database
async def update_pin_in_db(new_pin: str):
    """Update PIN in database"""
    await db.admin_settings.update_one(
        {"setting_key": "database_cleaner_pin"},
        {"$set": {
            "setting_key": "database_cleaner_pin",
            "setting_value": new_pin,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }},
        upsert=True
    )

@app.post("/api/admin/database-cleaner/verify-pin")
async def verify_database_cleaner_pin(
    pin_data: PINVerification,
    current_admin: AdminUser = Depends(require_super_admin)
):
    """
    Verify PIN for Database Cleaner access
    Only super admin can access this
    """
    try:
        correct_pin = await get_current_pin()
        
        if pin_data.pin == correct_pin:
            return {
                "success": True,
                "message": "PIN verified successfully"
            }
        else:
            raise HTTPException(status_code=401, detail="Invalid PIN")
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error verifying PIN: {e}")
        raise HTTPException(status_code=500, detail="Failed to verify PIN")

@app.post("/api/admin/database-cleaner/change-pin")
async def change_database_cleaner_pin(
    pin_data: ChangePINRequest,
    current_admin: AdminUser = Depends(require_super_admin)
):
    """
    Change PIN for Database Cleaner
    Only super admin can change PIN
    """
    try:
        # Validate new PIN
        if len(pin_data.new_pin) != 6 or not pin_data.new_pin.isdigit():
            raise HTTPException(status_code=400, detail="PIN harus 6 digit angka")
        
        # Check if new PIN matches confirmation
        if pin_data.new_pin != pin_data.confirm_pin:
            raise HTTPException(status_code=400, detail="PIN baru dan konfirmasi tidak sama")
        
        # Verify old PIN
        correct_pin = await get_current_pin()
        if pin_data.old_pin != correct_pin:
            raise HTTPException(status_code=401, detail="PIN lama salah")
        
        # Check if new PIN is same as old PIN
        if pin_data.new_pin == pin_data.old_pin:
            raise HTTPException(status_code=400, detail="PIN baru harus berbeda dengan PIN lama")
        
        # Update PIN in database
        await update_pin_in_db(pin_data.new_pin)
        
        return {
            "success": True,
            "message": "PIN berhasil diubah"
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error changing PIN: {e}")
        raise HTTPException(status_code=500, detail="Gagal mengubah PIN")

# ==================== DATABASE BACKUP & RESTORE ====================

@app.post("/api/admin/database/backup")
async def create_database_backup(
    pin_data: PINVerification,
    current_admin: AdminUser = Depends(require_super_admin)
):
    """
    Create manual database backup
    Requires PIN verification
    """
    try:
        # Verify PIN
        correct_pin = await get_current_pin()
        if pin_data.pin != correct_pin:
            raise HTTPException(status_code=401, detail="Invalid PIN")
        
        # Create backup
        result = await create_backup(db, backup_type="manual")
        
        if not result.get("success"):
            raise HTTPException(status_code=500, detail=result.get("error", "Failed to create backup"))
        
        # Remove _id from metadata
        metadata = result.get("metadata", {})
        metadata.pop('_id', None)
        
        return {
            "success": True,
            "message": "Backup berhasil dibuat",
            "backup_id": result["backup_id"],
            "filename": result["filename"],
            "gcs_url": result.get("gcs_url"),
            "metadata": metadata
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating backup: {e}")
        raise HTTPException(status_code=500, detail="Gagal membuat backup")

@app.get("/api/admin/database/backups")
async def list_backups(
    current_admin: AdminUser = Depends(require_super_admin),
    limit: int = 50
):
    """
    Get list of available backups
    """
    try:
        backups = await get_backup_history(db, limit=limit)
        return {
            "success": True,
            "backups": backups
        }
        
    except Exception as e:
        logger.error(f"Error getting backups: {e}")
        raise HTTPException(status_code=500, detail="Gagal mengambil daftar backup")

@app.get("/api/admin/database/restore-history")
async def list_restore_history(
    current_admin: AdminUser = Depends(require_super_admin),
    limit: int = 20
):
    """
    Get restore history
    """
    try:
        history = await db.restore_history.find().sort("restore_date", -1).limit(limit).to_list(length=limit)
        
        result = []
        for item in history:
            item.pop('_id', None)
            result.append(item)
        
        return {
            "success": True,
            "history": result
        }
        
    except Exception as e:
        logger.error(f"Error getting restore history: {e}")
        raise HTTPException(status_code=500, detail="Gagal mengambil restore history")

@app.get("/api/admin/database/backup/{backup_id}/download")
async def download_backup(
    backup_id: str,
    current_admin: AdminUser = Depends(require_super_admin)
):
    """
    Download backup file
    """
    try:
        # Get backup metadata
        backup = await db.backup_history.find_one({"backup_id": backup_id})
        
        if not backup:
            raise HTTPException(status_code=404, detail="Backup not found")
        
        local_path = backup.get("local_path")
        filename = backup.get("filename")
        
        if not local_path or not os.path.exists(local_path):
            raise HTTPException(status_code=404, detail="Backup file not found")
        
        return FileResponse(
            path=local_path,
            filename=filename,
            media_type="application/gzip"
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error downloading backup: {e}")
        raise HTTPException(status_code=500, detail="Gagal download backup")

class RestoreRequest(BaseModel):
    backup_id: str
    pin: str
    selected_collections: Optional[List[str]] = None

@app.post("/api/admin/database/restore")
async def restore_database(
    request_data: RestoreRequest,
    current_admin: AdminUser = Depends(require_super_admin)
):
    """
    Restore database from backup
    Requires PIN verification
    """
    try:
        # Verify PIN
        correct_pin = await get_current_pin()
        if request_data.pin != correct_pin:
            raise HTTPException(status_code=401, detail="Invalid PIN")
        
        # Get backup metadata
        backup = await db.backup_history.find_one({"backup_id": request_data.backup_id})
        
        if not backup:
            raise HTTPException(status_code=404, detail="Backup not found")
        
        local_path = backup.get("local_path")
        
        if not local_path or not os.path.exists(local_path):
            raise HTTPException(status_code=404, detail="Backup file not found")
        
        # Restore
        result = await restore_backup(
            db,
            local_path,
            selected_collections=request_data.selected_collections
        )
        
        if not result.get("success"):
            raise HTTPException(status_code=500, detail=result.get("error", "Failed to restore"))
        
        return {
            "success": True,
            "message": "Database berhasil di-restore",
            "results": result.get("results")
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error restoring database: {e}")
        raise HTTPException(status_code=500, detail="Gagal restore database")

@app.delete("/api/admin/database/backup/{backup_id}")
async def delete_backup(
    backup_id: str,
    pin_data: PINVerification,
    current_admin: AdminUser = Depends(require_super_admin)
):
    """
    Delete backup
    Requires PIN verification
    """
    try:
        # Verify PIN
        correct_pin = await get_current_pin()
        if pin_data.pin != correct_pin:
            raise HTTPException(status_code=401, detail="Invalid PIN")
        
        # Get backup
        backup = await db.backup_history.find_one({"backup_id": backup_id})
        
        if not backup:
            raise HTTPException(status_code=404, detail="Backup not found")
        
        # Delete local file
        local_path = backup.get("local_path")
        if local_path and os.path.exists(local_path):
            os.remove(local_path)
        
        # Delete from database
        await db.backup_history.delete_one({"backup_id": backup_id})
        
        return {
            "success": True,
            "message": "Backup berhasil dihapus"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting backup: {e}")
        raise HTTPException(status_code=500, detail="Gagal menghapus backup")

# ==================== MAINTENANCE MODE ====================

class MaintenanceModeUpdate(BaseModel):
    enabled: bool
    message: Optional[str] = "System sedang dalam maintenance. Silakan coba lagi nanti."
    estimated_completion: Optional[str] = None

@app.get("/api/maintenance/status")
async def get_maintenance_status():
    """
    Get maintenance mode status (public endpoint)
    """
    try:
        settings = await db.admin_settings.find_one({"setting_key": "maintenance_mode"})
        
        if not settings:
            return {
                "enabled": False,
                "message": None,
                "estimated_completion": None
            }
        
        return {
            "enabled": settings.get("enabled", False),
            "message": settings.get("message", "System sedang dalam maintenance."),
            "estimated_completion": settings.get("estimated_completion"),
            "activated_at": settings.get("activated_at"),
            "activated_by": settings.get("activated_by")
        }
        
    except Exception as e:
        logger.error(f"Error getting maintenance status: {e}")
        return {"enabled": False}

@app.post("/api/admin/maintenance/toggle")
async def toggle_maintenance_mode(
    data: MaintenanceModeUpdate,
    current_admin: AdminUser = Depends(require_super_admin)
):
    """
    Toggle maintenance mode (Super Admin only)
    """
    try:
        maintenance_settings = {
            "setting_key": "maintenance_mode",
            "enabled": data.enabled,
            "message": data.message,
            "estimated_completion": data.estimated_completion,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        if data.enabled:
            maintenance_settings["activated_at"] = datetime.now(timezone.utc).isoformat()
            maintenance_settings["activated_by"] = current_admin.username
        else:
            maintenance_settings["deactivated_at"] = datetime.now(timezone.utc).isoformat()
            maintenance_settings["deactivated_by"] = current_admin.username
        
        await db.admin_settings.update_one(
            {"setting_key": "maintenance_mode"},
            {"$set": maintenance_settings},
            upsert=True
        )
        
        status = "diaktifkan" if data.enabled else "dinonaktifkan"
        logger.info(f"Maintenance mode {status} by {current_admin.username}")
        
        return {
            "success": True,
            "message": f"Maintenance mode berhasil {status}",
            "enabled": data.enabled
        }
        
    except Exception as e:
        logger.error(f"Error toggling maintenance mode: {e}")
        raise HTTPException(status_code=500, detail="Gagal mengubah maintenance mode")

# Delete specific clients and all their related data (with PIN)
class DeleteClientsRequest(BaseModel):
    client_ids: List[str]
    pin: str

@app.post("/api/admin/delete-clients")


# ===== LANDING PAGE BUILDER HELPER FUNCTION =====
async def generate_landing_page_content(product_name: str, product_description: str):
    """Generate AI content for landing page using Emergent LLM"""
    try:
        import litellm
        import os
        import json
        
        emergent_key = os.environ.get("EMERGENT_LLM_KEY")
        if not emergent_key:
            return {"copy_blocks": {}, "layout_map": {}}
        
        # Configure for emergent key
        api_base = os.getenv("INTEGRATION_PROXY_URL", "https://integrations.emergentagent.com")
        
        prompt = f"""Create marketing copy for a landing page IN INDONESIAN LANGUAGE.
Product: {product_name}
Description: {product_description}

IMPORTANT: All content MUST be in Bahasa Indonesia (Indonesian language).

Return JSON with:
{{
  "hero_headline": "Headline utama yang kuat dan menarik (dalam Bahasa Indonesia)",
  "hero_description": "Deskripsi persuasif 2-3 kalimat (dalam Bahasa Indonesia)",
  "subheadline": "Text badge/label pendukung (dalam Bahasa Indonesia)",
  "cta_primary": "Text call-to-action utama (dalam Bahasa Indonesia)",
  "cta_headline": "Headline section CTA (dalam Bahasa Indonesia)",
  "cta_subheadline": "Deskripsi section CTA (dalam Bahasa Indonesia)",
  "social_proof": [{{"name": "Nama Customer", "quote": "Testimonial singkat dalam Bahasa Indonesia"}}],
  "urgency": "Text penawaran terbatas (dalam Bahasa Indonesia)"
}}

REMEMBER: Use ONLY Indonesian language for all text content!"""
        
        response = litellm.completion(
            model="openai/gpt-4o",
            messages=[{"role": "user", "content": prompt}],
            api_key=emergent_key,
            api_base=api_base + "/llm",
            custom_llm_provider="openai",
            temperature=0.7
        )
        
        content = response.choices[0].message.content
        if '```json' in content:
            content = content.split('```json')[1].split('```')[0].strip()
        
        return {"copy_blocks": json.loads(content), "layout_map": {}}
    except Exception as e:
        logger.error(f"AI generation failed: {e}")
        return {"copy_blocks": {}, "layout_map": {}}

# ===== LANDING PAGE BUILDER ENDPOINTS =====

# CREATE Landing Page
@app.post("/api/landing-pages")
async def create_landing_page(
    data: LandingPageCreate,
    current_user: User = Depends(get_current_user)
):
    try:
        # Generate AI content
        ai_content = await generate_landing_page_content(
            data.product_name,
            data.product_description
        )
        
        # Create landing page document
        landing_page = {
            "id": str(uuid.uuid4()),
            "user_id": current_user.id,
            "username": current_user.username,
            "template_id": data.template_id,
            "product_name": data.product_name,
            "product_description": data.product_description,
            "pricing_mode": data.pricing_mode,
            "product_price": data.product_price,
            "product_original_price": data.product_original_price,
            "currency": data.currency,
            "pricing_packages": data.pricing_packages,
            "benefits": data.benefits,
            "hero_image": data.hero_image,
            "gallery_images": data.gallery_images,
            "testimonials": data.testimonials,
            "primary_color": data.primary_color,
            "accent_color": data.accent_color,
            "font_heading": data.font_heading,
            "font_body": data.font_body,
            "facebook_pixel_id": data.facebook_pixel_id,
            "tiktok_pixel_id": data.tiktok_pixel_id,
            "ga_measurement_id": data.ga_measurement_id,
            "whatsapp_numbers": data.whatsapp_numbers,
            "whatsapp_number": data.whatsapp_number,
            "cta_event_name": data.cta_event_name,
            "seo_title": data.seo_title or data.product_name,
            "seo_description": data.seo_description or data.product_description,
            "seo_keywords": data.seo_keywords,
            "slug": data.slug,
            "copy_blocks": ai_content["copy_blocks"],
            "layout_map": ai_content["layout_map"],
            "product_details": data.product_details,  # CRITICAL FIX: Save product_details directly
            "status": "draft",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.landing_pages.insert_one(landing_page)
        
        return {
            "success": True,
            "landing_page_id": landing_page["id"],
            "slug": landing_page["slug"]
        }
    except Exception as e:
        logger.error(f"Create landing page error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# GET All Landing Pages
@app.get("/api/landing-pages")
async def get_landing_pages(current_user: User = Depends(get_current_user)):
    try:
        pages = await db.landing_pages.find({
            "user_id": current_user.id
        }).sort("created_at", -1).to_list(length=None)
        
        # Remove MongoDB _id to avoid serialization issues
        for page in pages:
            if "_id" in page:
                del page["_id"]
        
        return {"landing_pages": pages}
    except Exception as e:
        logger.error(f"Get landing pages error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# GET Single Landing Page
@app.get("/api/landing-pages/{landing_page_id}")
async def get_landing_page(
    landing_page_id: str,
    current_user: User = Depends(get_current_user)
):
    try:
        page = await db.landing_pages.find_one({
            "id": landing_page_id,
            "user_id": current_user.id
        })
        
        if not page:
            raise HTTPException(status_code=404, detail="Landing page not found")
        
        # Remove MongoDB _id to avoid serialization issues
        if "_id" in page:
            del page["_id"]
        
        return page
    except Exception as e:
        logger.error(f"Get landing page error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# UPDATE Landing Page
@app.put("/api/landing-pages/{landing_page_id}")
async def update_landing_page(
    landing_page_id: str,
    data: LandingPageUpdate,
    current_user: User = Depends(get_current_user)
):
    try:
        # Check if page exists
        page = await db.landing_pages.find_one({
            "id": landing_page_id,
            "user_id": current_user.id
        })
        
        if not page:
            raise HTTPException(status_code=404, detail="Landing page not found")
        
        # Regenerate AI content if product info changed
        if data.product_name or data.product_description:
            product_name = data.product_name or page.get("product_name")
            product_description = data.product_description or page.get("product_description")
            
            ai_content = await generate_landing_page_content(product_name, product_description)
            data_dict = data.model_dump(exclude_none=True)
            data_dict["copy_blocks"] = ai_content["copy_blocks"]
            data_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
        else:
            data_dict = data.model_dump(exclude_none=True)
            data_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
        
        # Update
        await db.landing_pages.update_one(
            {"id": landing_page_id, "user_id": current_user.id},
            {"$set": data_dict}
        )
        
        return {"success": True, "message": "Landing page updated"}
    except Exception as e:
        logger.error(f"Update landing page error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# DELETE Landing Page
@app.delete("/api/landing-pages/{landing_page_id}")
async def delete_landing_page(
    landing_page_id: str,
    current_user: User = Depends(get_current_user)
):
    try:
        result = await db.landing_pages.delete_one({
            "id": landing_page_id,
            "user_id": current_user.id
        })
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Landing page not found")
        
        return {"success": True, "message": "Landing page deleted"}
    except Exception as e:
        logger.error(f"Delete landing page error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# PUBLISH Landing Page
@app.post("/api/landing-pages/{landing_page_id}/publish")
async def publish_landing_page(
    landing_page_id: str,
    current_user: User = Depends(get_current_user)
):
    try:
        page = await db.landing_pages.find_one({
            "id": landing_page_id,
            "user_id": current_user.id
        })
        
        if not page:
            raise HTTPException(status_code=404, detail="Landing page not found")
        
        await db.landing_pages.update_one(
            {"id": landing_page_id},
            {"$set": {"status": "published"}}
        )
        
        return {
            "success": True,
            "message": "Landing page published",
            "url": f"/{page['slug']}",
            "slug": page['slug']
        }
    except Exception as e:
        logger.error(f"Publish landing page error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# PUBLIC VIEW Landing Page (no auth required)
@app.get("/api/landing-pages/public/{slug}")
async def get_public_landing_page(slug: str):
    try:
        page = await db.landing_pages.find_one({
            "slug": slug,
            "status": "published"
        })
        
        if not page:
            raise HTTPException(status_code=404, detail="Landing page not found")
        
        # Remove MongoDB _id to avoid serialization issues
        if "_id" in page:
            del page["_id"]
        
        return page
    except Exception as e:
        logger.error(f"Get public landing page error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ADMIN: Get All Landing Pages from All Clients
@app.get("/api/admin/landing-pages")
async def admin_get_all_landing_pages(
    current_user: User = Depends(get_current_admin)
):
    try:
        # Get all landing pages with user info
        pages = await db.landing_pages.find().sort("created_at", -1).to_list(length=None)
        
        # Enrich with user information
        enriched_pages = []
        for page in pages:
            # Remove MongoDB _id
            if "_id" in page:
                del page["_id"]
            
            # Get user info
            user = await db.users.find_one({"id": page["user_id"]})
            if user:
                page["client_name"] = user.get("name", user.get("username", "Unknown"))
                page["client_email"] = user.get("email", "")
                page["client_username"] = user.get("username", "")
            
            enriched_pages.append(page)
        
        return {"landing_pages": enriched_pages}
    except Exception as e:
        logger.error(f"Admin get all landing pages error: {e}")
        raise HTTPException(status_code=500, detail=str(e))
        logger.error(f"Get public landing page error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# UPLOAD Image for Landing Page
@app.post("/api/landing-pages/upload-image")
async def upload_landing_page_image(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user)
):
    try:
        from google.cloud import storage
        import os
        
        # Validate file type
        if not file.content_type.startswith("image/"):
            raise HTTPException(status_code=400, detail="File must be an image")
        
        # Read file content
        content = await file.read()
        
        # Initialize GCS client (credentials loaded from .env GOOGLE_APPLICATION_CREDENTIALS)
        storage_client = storage.Client()
        bucket_name = os.getenv("GCS_BUCKET_NAME", "rimuru-file-uploads")
        bucket = storage_client.bucket(bucket_name)
        
        # Generate filename
        file_extension = os.path.splitext(file.filename)[1]
        gcs_filename = f"landing_page_images/{str(uuid.uuid4())[:12]}{file_extension}"
        
        # Upload to GCS
        blob = bucket.blob(gcs_filename)
        blob.upload_from_string(content, content_type=file.content_type)
        
        # Generate public URL
        public_url = f"https://storage.googleapis.com/{bucket_name}/{gcs_filename}"
        
        return {"success": True, "url": public_url}
    except Exception as e:
        logger.error(f"Upload image error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# AI HELPER for generating content
@app.post("/api/landing-pages/ai-helper")
async def landing_page_ai_helper(
    request: dict,
    current_user: User = Depends(get_current_user)
):
    try:
        helper_type = request.get("type")  # "benefits", "testimonials", "seo", "pricing_packages"
        product_name = request.get("product_name", "")
        product_description = request.get("product_description", "")
        base_price = request.get("base_price", 0)
        currency = request.get("currency", "IDR")
        
        if not helper_type or not product_name:
            raise HTTPException(status_code=400, detail="Type and product_name required")
        
        import litellm
        import os
        import json
        
        emergent_key = os.environ.get("EMERGENT_LLM_KEY")
        if not emergent_key:
            raise HTTPException(status_code=500, detail="AI service not configured")
        
        # Configure for emergent key
        api_base = os.getenv("INTEGRATION_PROXY_URL", "https://integrations.emergentagent.com")
        
        # Different prompts based on type
        if helper_type == "benefits":
            prompt = f"""Generate 5-7 key benefits for this product. Return ONLY valid JSON array:
["Benefit 1", "Benefit 2", "Benefit 3", ...]

Product: {product_name}
Description: {product_description}"""
        
        elif helper_type == "testimonials":
            prompt = f"""Generate 3 realistic customer testimonials. Return ONLY valid JSON:
[
  {{"name": "Customer Name", "quote": "Testimonial text"}},
  ...
]

Product: {product_name}
Description: {product_description}"""
        
        elif helper_type == "seo":
            prompt = f"""Generate SEO content for this product. Return ONLY valid JSON:
{{
  "title": "SEO title max 60 chars",
  "description": "Meta description max 160 chars",
  "keywords": ["keyword1", "keyword2", "keyword3", "keyword4", "keyword5"]
}}

Product: {product_name}
Description: {product_description}"""
        
        elif helper_type == "pricing_packages":
            prompt = f"""Generate 3 pricing packages for this product with tiered pricing. Return ONLY valid JSON array:
[
  {{
    "name": "Package name (e.g., 'Beli 1 Pcs')",
    "price": actual_price_number,
    "original_price": original_price_number_higher_than_price,
    "description": "Short package description (max 50 chars)",
    "features": ["Feature 1", "Feature 2", "Feature 3"],
    "badge": "Label text like 'HEMAT 15%' or empty string",
    "is_highlighted": false,
    "cta_text": "CTA button text like 'Beli Sekarang'"
  }}
]

Product: {product_name}
Description: {product_description}
Base Price: {base_price} {currency}

Requirements:
- Create 3 packages with increasing quantities
- Calculate bulk discount pricing (10-30% discount)
- Highlight middle package (is_highlighted: true)
- Use Indonesian language"""
        
        else:
            raise HTTPException(status_code=400, detail="Invalid type. Must be: benefits, testimonials, seo, or pricing_packages")
        
        # Call AI
        response = litellm.completion(
            model="openai/gpt-4o",
            messages=[{"role": "user", "content": prompt}],
            api_key=emergent_key,
            api_base=api_base + "/llm",
            custom_llm_provider="openai",
            temperature=0.7
        )
        
        # Parse response
        content = response.choices[0].message.content
        
        # Extract JSON from response if wrapped
        if '```json' in content:
            content = content.split('```json')[1].split('```')[0].strip()
        elif '```' in content:
            content = content.split('```')[1].split('```')[0].strip()
        
        data = json.loads(content)
        
        return {"success": True, "data": data}
    except Exception as e:
        logger.error(f"AI helper error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

async def delete_clients(
    request_data: DeleteClientsRequest,
    current_admin: AdminUser = Depends(require_super_admin)
):
    """
    Delete specific clients and all their related data
    Only super admin can execute this with valid PIN
    """
    try:
        # Verify PIN first
        correct_pin = await get_current_pin()
        if request_data.pin != correct_pin:
            raise HTTPException(status_code=401, detail="Invalid PIN")
        
        client_ids = request_data.client_ids
        if not client_ids:
            raise HTTPException(status_code=400, detail="No client IDs provided")
        
        deleted_summary = {
            'clients_deleted': 0,
            'data_deleted': {}
        }
        
        # Delete users
        user_result = await db.users.delete_many({'id': {'$in': client_ids}})
        deleted_summary['clients_deleted'] = user_result.deleted_count
        
        # Delete all related data for these clients
        # Ad accounts
        accounts_result = await db.ad_accounts.delete_many({'user_id': {'$in': client_ids}})
        deleted_summary['data_deleted']['ad_accounts'] = accounts_result.deleted_count
        
        # Ad account requests
        requests_result = await db.ad_account_requests.delete_many({'user_id': {'$in': client_ids}})
        deleted_summary['data_deleted']['ad_account_requests'] = requests_result.deleted_count
        
        # Top-up requests
        topup_result = await db.topup_requests.delete_many({'user_id': {'$in': client_ids}})
        deleted_summary['data_deleted']['topup_requests'] = topup_result.deleted_count
        
        # Wallet top-up requests
        wallet_topup_result = await db.wallet_topup_requests.delete_many({'user_id': {'$in': client_ids}})
        deleted_summary['data_deleted']['wallet_topup_requests'] = wallet_topup_result.deleted_count
        
        # Transfer requests
        transfer_result = await db.transfer_requests.delete_many({'user_id': {'$in': client_ids}})
        deleted_summary['data_deleted']['transfer_requests'] = transfer_result.deleted_count
        
        # Wallet transfers
        wallet_transfer_result = await db.wallet_transfers.delete_many({'user_id': {'$in': client_ids}})
        deleted_summary['data_deleted']['wallet_transfers'] = wallet_transfer_result.deleted_count
        
        # Withdraw requests
        withdraw_result = await db.withdraw_requests.delete_many({'user_id': {'$in': client_ids}})
        deleted_summary['data_deleted']['withdraw_requests'] = withdraw_result.deleted_count
        
        # Transactions
        transactions_result = await db.transactions.delete_many({'user_id': {'$in': client_ids}})
        deleted_summary['data_deleted']['transactions'] = transactions_result.deleted_count
        
        # Share requests (as requester or receiver)
        share_result = await db.share_requests.delete_many({
            '$or': [
                {'requester_id': {'$in': client_ids}},
                {'receiver_id': {'$in': client_ids}}
            ]
        })
        deleted_summary['data_deleted']['share_requests'] = share_result.deleted_count
        
        # Client notifications
        notif_result = await db.client_notifications.delete_many({'user_id': {'$in': client_ids}})
        deleted_summary['data_deleted']['client_notifications'] = notif_result.deleted_count
        
        # Currency exchanges
        exchange_result = await db.currency_exchanges.delete_many({'user_id': {'$in': client_ids}})
        deleted_summary['data_deleted']['currency_exchanges'] = exchange_result.deleted_count
        
        return {
            "success": True,
            "message": f"Successfully deleted {deleted_summary['clients_deleted']} client(s) and all related data",
            "summary": deleted_summary
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to delete clients: {str(e)}")

# ==========================
# AI Ad-Copy Generator Endpoint
# ==========================

class AdCopyGenerateRequest(BaseModel):
    product_name: str
    description: str
    goal: str  # Purchase, Leads, Awareness

class AdCopyResponse(BaseModel):
    success: bool
    data: Optional[dict] = None
    message: Optional[str] = None

@app.post("/api/generate-ad-copy", response_model=AdCopyResponse)
async def generate_ad_copy(
    request: AdCopyGenerateRequest,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """
    Generate AI-powered ad copy for Meta (Facebook & Instagram) in Indonesian
    """
    try:
        # Verify client authentication
        current_user = await get_current_user(credentials)
        
        # Import emergentintegrations
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        
        # Get API key from environment
        api_key = os.getenv('EMERGENT_LLM_KEY')
        if not api_key:
            raise HTTPException(status_code=500, detail="LLM API key not configured")
        
        # Create system message for Meta Performance Copywriter
        system_message = """You are an expert AI Meta Performance Copywriter specializing in creating high-converting ad copy for Facebook and Instagram in Indonesian language.

Core Principles:
1. Readability: Write clear, engaging copy that resonates with Indonesian audiences
2. Relevance: Match copy to product, audience, and campaign goal
3. Policy Compliance: Strictly avoid content that violates Meta's advertising policies
4. Hook Structure: Follow Hook → Value → Proof → CTA structure
5. Avoid Misleading Claims: Use qualified language (can help, up to, designed for)
6. Generate Both Text Lengths: Short (≤125 chars) and Standard (≤280 chars)
7. Natural Indonesian: Use conversational tone with 1-2 emojis where appropriate

Your output MUST be valid JSON only, following this exact structure:
{
  "primary_text_short": ["text1", "text2", "text3"],
  "primary_text_standard": ["text1", "text2", "text3"],
  "headlines": ["headline1", "headline2", "headline3", "headline4", "headline5"],
  "descriptions": ["desc1", "desc2", "desc3"],
  "hooks": ["hook1", "hook2", "hook3"],
  "ctas": ["call_to_action1", "call_to_action2", "call_to_action3"],
  "ugc_scripts": [
    {
      "scenario": "scenario_title1",
      "script": "Pure spoken dialogue here - NO scene directions, NO 'Shot 1', NO camera notes. Just what avatar will SAY."
    },
    {
      "scenario": "scenario_title2",
      "script": "Another spoken dialogue - direct copywriting that flows naturally when spoken aloud."
    }
  ]
}

**CRITICAL for ugc_scripts:**
- "script" field = PURE DIALOGUE/COPYWRITING only
- Avatar will read this EXACTLY as written
- NO scene directions like "Shot 1 -", "B-roll", etc
- Write as if YOU are the person speaking to camera
- Example: "Hai! Aku mau kasih tau produk favorit aku nih. Serum ini bikin kulit aku glowing banget dalam seminggu!"
- NOT: "Shot 1 - Person holding serum: 'Hai! Aku mau kasih tau...'"

IMPORTANT: Return ONLY valid JSON, no markdown, no explanations, no code blocks."""

        # Create developer message with production rules
        developer_message = f"""Production Rules:
- Generate ad copy in Bahasa Indonesia (natural, conversational)
- Generate 3 variants for each text type (primary_text_short, primary_text_standard, headlines, hooks, ctas)
- At least 5 headlines, 3 descriptions
- Variants must be at least 60% different from each other
- Use qualified claims (dapat membantu, hingga, dirancang untuk)
- Default tone: confident and helpful, without boasting
- Optional: Use 1-2 relevant emojis per variant
- **UGC Scripts: Keep VERY SHORT (max 100 words / ~20 seconds when spoken)**
  * Focus on 3-5 quick shots/scenes only
  * Direct, punchy dialogue
  * Perfect for short-form video (Instagram Reels, TikTok style)
- Strictly follow Meta advertising policies:
  * No personal attributes (body weight, health conditions, disabilities)
  * No absolute claims without qualification
  * No before/after transformations
  * No sensitive targeting language
  * No misleading urgency or scarcity
- Return ONLY valid JSON format, no markdown formatting"""

        # Create user message with campaign data
        user_message_text = f"""Campaign Information:
Product Name: {request.product_name}
Description: {request.description}
Campaign Goal: {request.goal}

Based on this information, generate comprehensive ad copy variations for Meta ads targeting Indonesian audiences. 
Automatically determine the best angles (problem-solution, social-proof, value-stack, rational) based on the product description and goal.
Generate copy that will perform well for the stated goal ({request.goal}).

Return valid JSON ONLY."""

        # Initialize LLM Chat with GPT-5 (better quality for ad copy)
        session_id = f"adcopy_{current_user.id}_{datetime.now(timezone.utc).isoformat()}"
        chat = LlmChat(
            api_key=api_key,
            session_id=session_id,
            system_message=system_message
        ).with_model("openai", "gpt-5")
        
        # Add developer context as a system-level message (we'll send it with user message)
        full_user_message = f"{developer_message}\n\n{user_message_text}"
        
        # Create user message
        user_msg = UserMessage(text=full_user_message)
        
        # Send message with retry mechanism for 502 errors
        max_retries = 2  # Reduced from 3 to make it faster
        retry_delay = 1  # Start with shorter delay
        last_error = None
        
        for attempt in range(max_retries):
            try:
                # Try with timeout of 30 seconds per attempt
                response = await asyncio.wait_for(
                    chat.send_message(user_msg),
                    timeout=30.0
                )
                break  # Success, exit retry loop
            except asyncio.TimeoutError:
                last_error = "Request timeout"
                if attempt < max_retries - 1:
                    logging.warning(f"GPT-5 timeout, retrying... (attempt {attempt + 1}/{max_retries})")
                    await asyncio.sleep(retry_delay)
                    retry_delay *= 2
                    continue
                else:
                    logging.error(f"GPT-5 failed after {max_retries} attempts due to timeout")
                    raise HTTPException(status_code=504, detail="AI sedang lambat, coba lagi dalam beberapa saat")
            except Exception as e:
                error_msg = str(e)
                last_error = error_msg
                if ('502' in error_msg or '503' in error_msg or '504' in error_msg) and attempt < max_retries - 1:
                    logging.warning(f"GPT-5 returned error {error_msg}, retrying... (attempt {attempt + 1}/{max_retries})")
                    await asyncio.sleep(retry_delay)
                    retry_delay *= 2
                    continue
                else:
                    # If all retries failed or not a retryable error
                    logging.error(f"Failed after {attempt + 1} attempts: {error_msg}")
                    if 'budget' in error_msg.lower() or 'insufficient' in error_msg.lower():
                        raise HTTPException(status_code=402, detail="Saldo Emergent LLM Key habis. Silakan top up di Profile → Universal Key")
                    raise HTTPException(status_code=500, detail=f"AI generation error: {error_msg}")
        
        # Parse JSON response
        import json
        try:
            # Clean response if it has markdown code blocks
            response_text = response.strip()
            if response_text.startswith("```json"):
                response_text = response_text[7:]
            if response_text.startswith("```"):
                response_text = response_text[3:]
            if response_text.endswith("```"):
                response_text = response_text[:-3]
            response_text = response_text.strip()
            
            ad_copy_data = json.loads(response_text)
            
            # Validate required fields
            required_fields = ['primary_text_short', 'primary_text_standard', 'headlines', 
                             'descriptions', 'hooks', 'ctas', 'ugc_scripts']
            for field in required_fields:
                if field not in ad_copy_data:
                    raise ValueError(f"Missing required field: {field}")
            
            return {
                "success": True,
                "data": ad_copy_data,
                "message": "Ad copy generated successfully"
            }
            
        except json.JSONDecodeError as e:
            logging.error(f"Failed to parse LLM response as JSON: {str(e)}")
            logging.error(f"Response was: {response}")
            raise HTTPException(
                status_code=500, 
                detail="Failed to parse AI response. Please try again."
            )
    
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error generating ad copy: {str(e)}")
        logging.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Failed to generate ad copy: {str(e)}")

# ==========================
# Saved Ad Copies CRUD Endpoints
# ==========================

class SaveAdCopyRequest(BaseModel):
    label: str  # User-friendly name for the ad copy
    product_name: str
    description: str
    goal: str
    generated_content: dict  # The actual ad copy content

class UpdateAdCopyRequest(BaseModel):
    label: Optional[str] = None
    generated_content: Optional[dict] = None

@app.post("/api/ad-copies")
async def save_ad_copy(
    request: SaveAdCopyRequest,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """
    Save generated ad copy for future reference
    Limit: 100 saved ad copies per client
    """
    try:
        current_user = await get_current_user(credentials)
        user_id = current_user.id  # Fixed: Use .id instead of .get('user_id')
        
        # Check if user has reached the limit of 100 saved ad copies
        existing_count = await db.saved_ad_copies.count_documents({'user_id': user_id})
        if existing_count >= 100:
            raise HTTPException(
                status_code=400, 
                detail="Anda sudah mencapai limit maksimal 100 saved ad copies. Hapus beberapa untuk menambah yang baru."
            )
        
        # Create ad copy document
        ad_copy_doc = {
            'ad_copy_id': str(uuid.uuid4()),
            'user_id': user_id,
            'username': current_user.username,  # Fixed: Use .username instead of .get('username')
            'label': request.label,
            'product_name': request.product_name,
            'description': request.description,
            'goal': request.goal,
            'generated_content': request.generated_content,
            'created_at': datetime.now(timezone.utc).isoformat(),
            'updated_at': datetime.now(timezone.utc).isoformat()
        }
        
        await db.saved_ad_copies.insert_one(ad_copy_doc)
        
        return {
            'success': True,
            'message': 'Ad copy berhasil disimpan',
            'ad_copy_id': ad_copy_doc['ad_copy_id']
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error saving ad copy: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to save ad copy: {str(e)}")

@app.get("/api/ad-copies")
async def get_saved_ad_copies(
    credentials: HTTPAuthorizationCredentials = Depends(security),
    search: Optional[str] = None,
    goal: Optional[str] = None
):
    """
    Get all saved ad copies for the current user
    Supports search by label/product_name and filter by goal
    """
    try:
        current_user = await get_current_user(credentials)
        user_id = current_user.id  # Fixed: Use .id instead of .get('user_id')
        
        # Build query
        query = {'user_id': user_id}
        
        # Add search filter
        if search:
            query['$or'] = [
                {'label': {'$regex': search, '$options': 'i'}},
                {'product_name': {'$regex': search, '$options': 'i'}}
            ]
        
        # Add goal filter
        if goal and goal != 'all':
            query['goal'] = goal
        
        # Fetch ad copies, sorted by most recent first
        ad_copies_cursor = db.saved_ad_copies.find(
            query,
            {'_id': 0}
        ).sort('created_at', -1)
        
        ad_copies = await ad_copies_cursor.to_list(length=None)
        
        return {
            'success': True,
            'ad_copies': ad_copies,
            'total': len(ad_copies)
        }
    
    except Exception as e:
        logging.error(f"Error fetching saved ad copies: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch saved ad copies: {str(e)}")

@app.get("/api/ad-copies/{ad_copy_id}")
async def get_ad_copy_detail(
    ad_copy_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """
    Get detailed information about a specific saved ad copy
    """
    try:
        current_user = await get_current_user(credentials)
        user_id = current_user.id  # Fixed
        
        ad_copy = await db.saved_ad_copies.find_one(
            {'ad_copy_id': ad_copy_id, 'user_id': user_id},
            {'_id': 0}
        )
        
        if not ad_copy:
            raise HTTPException(status_code=404, detail="Ad copy tidak ditemukan")
        
        return {
            'success': True,
            'ad_copy': ad_copy
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error fetching ad copy detail: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch ad copy detail: {str(e)}")

@app.put("/api/ad-copies/{ad_copy_id}")
async def update_ad_copy(
    ad_copy_id: str,
    request: UpdateAdCopyRequest,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """
    Update a saved ad copy (label or content)
    """
    try:
        current_user = await get_current_user(credentials)
        user_id = current_user.id  # Fixed
        
        # Check if ad copy exists and belongs to user
        ad_copy = await db.saved_ad_copies.find_one(
            {'ad_copy_id': ad_copy_id, 'user_id': user_id}
        )
        
        if not ad_copy:
            raise HTTPException(status_code=404, detail="Ad copy tidak ditemukan")
        
        # Build update document
        update_doc = {
            'updated_at': datetime.now(timezone.utc).isoformat()
        }
        
        if request.label is not None:
            update_doc['label'] = request.label
        
        if request.generated_content is not None:
            update_doc['generated_content'] = request.generated_content
        
        # Update ad copy
        await db.saved_ad_copies.update_one(
            {'ad_copy_id': ad_copy_id, 'user_id': user_id},
            {'$set': update_doc}
        )
        
        return {
            'success': True,
            'message': 'Ad copy berhasil diupdate'
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error updating ad copy: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to update ad copy: {str(e)}")

@app.delete("/api/ad-copies/{ad_copy_id}")
async def delete_ad_copy(
    ad_copy_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """
    Delete a saved ad copy
    """
    try:
        current_user = await get_current_user(credentials)
        user_id = current_user.id  # Fixed
        
        # Delete ad copy
        result = await db.saved_ad_copies.delete_one(
            {'ad_copy_id': ad_copy_id, 'user_id': user_id}
        )
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Ad copy tidak ditemukan")
        
        return {
            'success': True,
            'message': 'Ad copy berhasil dihapus'
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error deleting ad copy: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to delete ad copy: {str(e)}")

# ==========================
# AI Ad Image Generator Endpoints
# ==========================

# ==========================
# Video generation endpoints removed
# ==========================

# ==========================
# BitShip Shipping Integration Endpoints
# ==========================

from biteship_client import get_biteship_client

# Calculate shipping rates (PUBLIC - no auth required)
@app.post("/api/shipping/calculate-rates")
async def calculate_shipping_rates(
    origin_postal_code: int,
    destination_postal_code: int,
    weight: int,  # in grams
    length: int = 10,  # cm
    width: int = 10,  # cm
    height: int = 10,  # cm
    value: int = 100000,  # item value in IDR
    couriers: str = "jne,jnt,sicepat,anteraja"
):
    """Calculate shipping rates from multiple couriers"""
    try:
        biteship = get_biteship_client()
        
        items = [{
            "name": "Product",
            "value": value,
            "weight": weight,
            "length": length,
            "width": width,
            "height": height,
            "quantity": 1
        }]
        
        result = await biteship.get_rates(
            origin_postal_code=origin_postal_code,
            destination_postal_code=destination_postal_code,
            couriers=couriers,
            items=items
        )
        
        return {"success": True, "data": result}
    except Exception as e:
        logger.error(f"Error calculating shipping rates: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# Get available couriers (PUBLIC)
@app.get("/api/shipping/couriers")
async def get_available_couriers():
    """Get list of available courier services"""
    try:
        biteship = get_biteship_client()
        result = await biteship.get_couriers()
        return {"success": True, "data": result}
    except Exception as e:
        logger.error(f"Error retrieving couriers: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# Track order by waybill (PUBLIC)
@app.get("/api/shipping/track/{waybill_id}/{courier_code}")
async def track_shipment(waybill_id: str, courier_code: str):
    """Track shipment using waybill ID and courier code"""
    try:
        biteship = get_biteship_client()
        result = await biteship.track_order(waybill_id, courier_code)
        return {"success": True, "data": result}
    except Exception as e:
        logger.error(f"Error tracking shipment: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# ==========================
# Order Management Endpoints
# ==========================

# Pydantic models for orders
class OrderItem(BaseModel):
    product_name: str
    quantity: int
    unit_price: float

class CreateOrderRequest(BaseModel):
    landing_page_id: str
    customer_name: str
    customer_phone: str
    customer_address: str
    customer_city: str
    customer_postal_code: int
    quantity: int
    unit_price: float
    courier_company: str
    courier_type: str
    courier_service_name: str
    shipping_cost: float
    estimated_delivery: str
    payment_method: str  # 'cod' or 'transfer'
    notes: Optional[str] = None


# Create order from landing page (PUBLIC - no auth)
@app.post("/api/orders/create")
async def create_order(order_data: CreateOrderRequest):
    """Create a new order from landing page checkout"""
    try:
        # Get landing page details
        landing_page = await db.landing_pages.find_one({"id": order_data.landing_page_id})
        if not landing_page:
            raise HTTPException(status_code=404, detail="Landing page not found")
        
        # Check if product is enabled
        if not landing_page.get("product_details", {}).get("is_enabled", True):
            raise HTTPException(status_code=400, detail="Product is not available")
        
        # Check stock
        current_stock = landing_page.get("product_details", {}).get("stock_quantity", 0)
        if current_stock < order_data.quantity:
            raise HTTPException(status_code=400, detail=f"Insufficient stock. Available: {current_stock}")
        
        # Generate order number
        import random
        import string
        order_number = f"ORD-{''.join(random.choices(string.digits, k=8))}"
        
        # Calculate totals
        subtotal = order_data.unit_price * order_data.quantity
        total = subtotal + order_data.shipping_cost
        
        # Create order document
        order = {
            "id": str(uuid.uuid4()),
            "order_number": order_number,
            "landing_page_id": order_data.landing_page_id,
            "product_name": landing_page.get("product_name", "Product"),
            "product_slug": landing_page.get("slug", ""),
            "merchant_id": landing_page.get("user_id", ""),
            
            # Customer info
            "customer_name": order_data.customer_name,
            "customer_phone": order_data.customer_phone,
            "customer_address": order_data.customer_address,
            "customer_city": order_data.customer_city,
            "customer_postal_code": order_data.customer_postal_code,
            
            # Order details
            "quantity": order_data.quantity,
            "unit_price": order_data.unit_price,
            "subtotal": subtotal,
            "shipping_cost": order_data.shipping_cost,
            "total": total,
            
            # Shipping
            "courier_company": order_data.courier_company,
            "courier_type": order_data.courier_type,
            "courier_service_name": order_data.courier_service_name,
            "estimated_delivery": order_data.estimated_delivery,
            
            # Payment
            "payment_method": order_data.payment_method,
            "payment_status": "pending",
            "payment_proof_url": None,
            
            # Status
            "order_status": "pending",  # pending -> confirmed -> processing -> shipped -> delivered
            
            # BitShip
            "biteship_order_id": None,
            "waybill_id": None,
            "tracking_url": None,
            
            # Metadata
            "notes": order_data.notes,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        # Insert order
        await db.orders.insert_one(order)
        
        # Update stock (reduce)
        new_stock = current_stock - order_data.quantity
        await db.landing_pages.update_one(
            {"id": order_data.landing_page_id},
            {
                "$set": {
                    "product_details.stock_quantity": new_stock,
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
            }
        )
        
        logger.info(f"✅ Order created: {order_number} for merchant {order['merchant_id']}")
        
        return {
            "success": True,
            "order_number": order_number,
            "order_id": order["id"],
            "message": "Order created successfully"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating order: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# Track order by order number (PUBLIC)
@app.get("/api/orders/track/{order_number}")
async def track_order_public(order_number: str):
    """Track order status by order number (public endpoint)"""
    try:
        order = await db.orders.find_one({"order_number": order_number})
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")
        
        # Get tracking info from BitShip if available
        tracking_data = None
        if order.get("waybill_id") and order.get("courier_company"):
            try:
                biteship = get_biteship_client()
                tracking_result = await biteship.track_order(
                    order["waybill_id"],
                    order["courier_company"]
                )
                tracking_data = tracking_result
            except Exception as e:
                logger.error(f"Error fetching tracking: {str(e)}")
        
        # Remove sensitive internal fields
        order.pop("_id", None)
        order.pop("merchant_id", None)
        
        return {
            "success": True,
            "order": order,
            "tracking": tracking_data
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error tracking order: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# Get merchant's orders
@app.get("/api/merchant/orders")
async def get_merchant_orders(
    current_user: User = Depends(get_current_user),
    status: Optional[str] = None,
    limit: int = 50,
    skip: int = 0
):
    """Get all orders for current merchant"""
    try:
        query = {"merchant_id": current_user.id}
        
        if status:
            query["order_status"] = status
        
        orders = await db.orders.find(query).sort("created_at", -1).skip(skip).limit(limit).to_list(length=limit)
        total = await db.orders.count_documents(query)
        
        # Clean up MongoDB _id
        for order in orders:
            order.pop("_id", None)
        
        return {
            "success": True,
            "orders": orders,
            "total": total,
            "limit": limit,
            "skip": skip
        }
        
    except Exception as e:
        logger.error(f"Error retrieving orders: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# Get single order detail
@app.get("/api/merchant/orders/{order_id}")
async def get_order_detail(
    order_id: str,
    current_user: User = Depends(get_current_user)
):
    """Get detailed information for a specific order"""
    try:
        order = await db.orders.find_one({
            "id": order_id,
            "merchant_id": current_user.id
        })
        
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")
        
        order.pop("_id", None)
        
        return {"success": True, "order": order}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error retrieving order: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# Update order status
@app.put("/api/merchant/orders/{order_id}/status")
async def update_order_status(
    order_id: str,
    new_status: str,
    current_user: User = Depends(get_current_user)
):
    """Update order status"""
    try:
        valid_statuses = ["pending", "confirmed", "processing", "shipped", "delivered", "cancelled"]
        if new_status not in valid_statuses:
            raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {', '.join(valid_statuses)}")
        
        order = await db.orders.find_one({
            "id": order_id,
            "merchant_id": current_user.id
        })
        
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")
        
        await db.orders.update_one(
            {"id": order_id},
            {
                "$set": {
                    "order_status": new_status,
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
            }
        )
        
        logger.info(f"✅ Order {order['order_number']} status updated to {new_status}")
        
        return {"success": True, "message": "Order status updated"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating order status: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# Process shipping via BitShip
@app.post("/api/merchant/orders/{order_id}/process-shipping")
async def process_order_shipping(
    order_id: str,
    current_user: User = Depends(get_current_user)
):
    """Create shipping order via BitShip and get waybill number"""
    try:
        order = await db.orders.find_one({
            "id": order_id,
            "merchant_id": current_user.id
        })
        
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")
        
        if order.get("biteship_order_id"):
            raise HTTPException(status_code=400, detail="Shipping already processed")
        
        # Get landing page for origin info
        landing_page = await db.landing_pages.find_one({"id": order["landing_page_id"]})
        if not landing_page:
            raise HTTPException(status_code=404, detail="Landing page not found")
        
        shipping_origin = landing_page.get("product_details", {}).get("shipping_origin", {})
        product_details = landing_page.get("product_details", {})
        
        # Create BitShip order
        biteship = get_biteship_client()
        
        items = [{
            "name": order["product_name"],
            "value": int(order["unit_price"]),
            "quantity": order["quantity"],
            "weight": product_details.get("weight", 500),
            "length": product_details.get("length", 10),
            "width": product_details.get("width", 10),
            "height": product_details.get("height", 10)
        }]
        
        biteship_result = await biteship.create_order(
            shipper_contact_name=shipping_origin.get("contact_name", "Merchant"),
            shipper_contact_phone=shipping_origin.get("contact_phone", getattr(current_user, 'phone_number', None) or "081234567890"),
            shipper_contact_email=current_user.email,
            shipper_organization=getattr(current_user, 'name', None) or current_user.username,
            origin_contact_name=shipping_origin.get("contact_name", "Merchant"),
            origin_contact_phone=shipping_origin.get("contact_phone", "081234567890"),
            origin_address=shipping_origin.get("address", ""),
            origin_postal_code=int(shipping_origin.get("postal_code", 12345)),
            destination_contact_name=order["customer_name"],
            destination_contact_phone=order["customer_phone"],
            destination_address=order["customer_address"],
            destination_postal_code=order["customer_postal_code"],
            courier_company=order["courier_company"],
            courier_type=order["courier_type"],
            items=items,
            reference_id=order["order_number"]
        )
        
        # Update order with BitShip info
        await db.orders.update_one(
            {"id": order_id},
            {
                "$set": {
                    "biteship_order_id": biteship_result.get("id"),
                    "waybill_id": biteship_result.get("waybill_id"),
                    "tracking_url": biteship_result.get("courier", {}).get("link"),
                    "order_status": "shipped",
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
            }
        )
        
        logger.info(f"✅ Shipping processed for order {order['order_number']}: {biteship_result.get('waybill_id')}")
        
        return {
            "success": True,
            "waybill_id": biteship_result.get("waybill_id"),
            "tracking_url": biteship_result.get("courier", {}).get("link"),
            "message": "Shipping order created successfully"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error processing shipping: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# Get merchant's products (landing pages with product enabled)
@app.get("/api/merchant/products")
async def get_merchant_products(
    current_user: User = Depends(get_current_user),
    limit: int = 50,
    skip: int = 0
):
    """Get all products (landing pages) for current merchant"""
    try:
        query = {
            "user_id": current_user.id,
            "product_details.is_enabled": True
        }
        
        products = await db.landing_pages.find(query).sort("created_at", -1).skip(skip).limit(limit).to_list(length=limit)
        total = await db.landing_pages.count_documents(query)
        
        # Clean up and add order count
        for product in products:
            product.pop("_id", None)
            # Get order count for this product
            order_count = await db.orders.count_documents({"landing_page_id": product["id"]})
            product["order_count"] = order_count
        
        return {
            "success": True,
            "products": products,
            "total": total
        }
        
    except Exception as e:
        logger.error(f"Error retrieving products: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# Update product stock
@app.put("/api/landing-pages/{page_id}/stock")
async def update_product_stock(
    page_id: str,
    stock_quantity: int,
    current_user: User = Depends(get_current_user)
):
    """Update product stock quantity"""
    try:
        # Verify ownership
        landing_page = await db.landing_pages.find_one({
            "id": page_id,
            "user_id": current_user.id
        })
        
        if not landing_page:
            raise HTTPException(status_code=404, detail="Landing page not found")
        
        # Update stock
        await db.landing_pages.update_one(
            {"id": page_id},
            {
                "$set": {
                    "product_details.stock_quantity": stock_quantity,
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
            }
        )
        
        logger.info(f"✅ Updated stock for {page_id}: {stock_quantity}")
        
        return {"success": True, "message": "Stock updated successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating stock: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()